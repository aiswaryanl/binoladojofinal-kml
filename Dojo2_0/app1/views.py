from tokenize import Comment
import datetime



from django.conf import settings
from django.shortcuts import render
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

# Create your views here.

from functools import cache
from django.shortcuts import get_list_or_404, render
from .serializers import CompanyLogoSerializer, KeyEventSerializer, MasterTableSerializer, RegisterSerializer, ScoreSerializer, SimpleScoreSerializer
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
#(2) Views for the User Login Views for the User Login
from django.shortcuts import get_list_or_404, render
from .serializers import RegisterSerializer
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
# #Create your views here.
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth import login
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.exceptions import ValidationError
from .serializers import LoginSerializer

from django.shortcuts import render

# views.py
from django.shortcuts import render, get_object_or_404
from django.contrib.auth import login
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import AllowAny
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.exceptions import ValidationError

from .serializers import LoginSerializer
from .models import User

def dojo_app(request):
    return render(request, 'index.html')


from django.contrib.auth import get_user_model, login
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework import status
from rest_framework_simplejwt.tokens import RefreshToken

User = get_user_model()

class LoginAPIView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        print("Login request data:", request.data)  # ok for debug, remove in production

        serializer = LoginSerializer(data=request.data, context={'request': request})
        if not serializer.is_valid():
            # return validation/auth errors
            return Response(
                {"message": "Authentication failed", "errors": serializer.errors},
                status=status.HTTP_400_BAD_REQUEST
            )

        # serializer is valid -> get user by id
        user_id = serializer.validated_data.get('user_id')
        try:
            user = User.objects.get(pk=user_id)
        except User.DoesNotExist:
            return Response(
                {"message": "Authentication failed", "errors": "User not found"},
                status=status.HTTP_401_UNAUTHORIZED
            )

        # Log user in (optional depending on token-only approach)
        login(request, user)

        # Generate JWT tokens
        refresh = RefreshToken.for_user(user)
        access_token = str(refresh.access_token)

        # Build user payload safely (use getattr with defaults)
        user_payload = {
            'email': getattr(user, 'email', None),
            'first_name': getattr(user, 'first_name', None),
            'last_name': getattr(user, 'last_name', None),
            'employeeid': getattr(user, 'employeeid', None),
            # Prefer returning primitive values (ids or names) for related objects:
            'role': getattr(getattr(user, 'role', None), 'name', None),
            'hq': getattr(getattr(user, 'hq', None), 'name', getattr(user, 'hq', None)),
            'factory': getattr(getattr(user, 'factory', None), 'name', getattr(user, 'factory', None)),
            'department': getattr(getattr(user, 'department', None), 'name', getattr(user, 'department', None)),
            'status': getattr(user, 'status', None),
        }

        return Response({
            'message': 'Login successful',
            'access_token': access_token,
            'refresh_token': str(refresh),
            'user': user_payload
        }, status=status.HTTP_200_OK)





#(1) Views for the User Register

from django.db import IntegrityError
from django.shortcuts import render
from .serializers import RegisterSerializer
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
from .models import CompanyLogo, KeyEvent, MasterTable, Score, TestSession, User

class RegisterView(generics.GenericAPIView):
    serializer_class = RegisterSerializer

    def post(self, request):
        user_data = request.data
        serializer = self.serializer_class(data=user_data)

        try:
            serializer.is_valid(raise_exception=True)

            # Check if user already exists by email
            if User.objects.filter(email=user_data.get("email")).exists():
                return Response({
                    "message": "Registration failed",
                    "errors": {"email": "This email is already registered."}
                }, status=status.HTTP_400_BAD_REQUEST)

            # Check if employee ID already exists
            if User.objects.filter(employeeid=user_data.get("employeeid")).exists():
                return Response({
                    "message": "Registration failed",
                    "errors": {"employeeid": "This employee ID is already in use."}
                }, status=status.HTTP_400_BAD_REQUEST)

            # Save the user
            serializer.save()

            return Response({
                "message": "User registered successfully!"
            }, status=status.HTTP_201_CREATED)

        except ValidationError as e:
            # Handle specific validation errors
            return Response({
                "message": "Validation failed",
                "errors": e.detail
            }, status=status.HTTP_400_BAD_REQUEST)

        except IntegrityError:
            # Handle database integrity errors (like duplicate entries)
            return Response({
                "message": "Database error",
                "errors": {"detail": "Duplicate entry or constraint violation."}
            }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            # Handle unexpected errors
            return Response({
                "message": "Unexpected error occurred",
                "errors": {"detail": str(e)}
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.decorators import action
from .models import User, Role
from .serializers import RegisterSerializer, RoleSerializer

class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all().select_related("role")
    serializer_class = RegisterSerializer

    def get_permissions(self):
        if self.action in ["create"]:
            return [AllowAny()]  # registration allowed for unauthenticated
        return [IsAuthenticated()]

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def me(self, request):
        """Return current user info"""
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)

class RoleViewSet(viewsets.ModelViewSet):
    queryset = Role.objects.all()
    serializer_class = RoleSerializer

    def get_permissions(self):
        if self.action in ["create", "list"]:
            return [AllowAny()]  # Allow unauthenticated users to create and list roles
        return [IsAuthenticated()]


#(3) Views for the User Logout

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from .serializers import LogoutSerializer

class LogoutAPIView(APIView):
    """
    User Logout API View
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = LogoutSerializer(data=request.data)
        if serializer.is_valid():
            refresh_token = serializer.validated_data["refresh_token"]

            try:
                token = RefreshToken(refresh_token)
                token.blacklist()  # Blacklist the refresh token
                return Response({"message": "Logout successful"}, status=status.HTTP_200_OK)
            except Exception as e:
                return Response({"error": "Invalid token"}, status=status.HTTP_400_BAD_REQUEST)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    





from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q
from .models import Hq, Factory, Department, Line, SubLine, Station
from .serializers import (
    HqSerializer, FactorySerializer, DepartmentSerializer,
    LineSerializer, SubLineSerializer, StationSerializer
)

class HqViewSet(viewsets.ModelViewSet):
    queryset = Hq.objects.all()
    serializer_class = HqSerializer
    
    def get_queryset(self):
        queryset = Hq.objects.all().order_by('hq_name')
        return queryset

class FactoryViewSet(viewsets.ModelViewSet):
    queryset = Factory.objects.all()
    serializer_class = FactorySerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['hq']
    
    def get_queryset(self):
        queryset = Factory.objects.all().select_related('hq').order_by('factory_name')
        hq_id = self.request.query_params.get('hq', None)
        if hq_id is not None:
            queryset = queryset.filter(hq_id=hq_id)
        return queryset

class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['factory', 'hq']
    
    def get_queryset(self):
        queryset = Department.objects.all().select_related('factory', 'hq').order_by('department_name')
        factory_id = self.request.query_params.get('factory', None)
        hq_id = self.request.query_params.get('hq', None)
        
        if factory_id is not None:
            queryset = queryset.filter(factory_id=factory_id)
        elif hq_id is not None:
            queryset = queryset.filter(hq_id=hq_id)
            
        return queryset

class LineViewSet(viewsets.ModelViewSet):
    queryset = Line.objects.all()
    serializer_class = LineSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['department', 'factory', 'hq']
    
    def get_queryset(self):
        queryset = Line.objects.all().select_related('department', 'factory', 'hq').order_by('line_name')
        department_id = self.request.query_params.get('department', None)
        factory_id = self.request.query_params.get('factory', None)
        hq_id = self.request.query_params.get('hq', None)
        
        if department_id is not None:
            queryset = queryset.filter(department_id=department_id)
        elif factory_id is not None:
            queryset = queryset.filter(factory_id=factory_id)
        elif hq_id is not None:
            queryset = queryset.filter(hq_id=hq_id)
            
        return queryset

class SubLineViewSet(viewsets.ModelViewSet):
    queryset = SubLine.objects.all()
    serializer_class = SubLineSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['line', 'department', 'factory', 'hq']
    
    def get_queryset(self):
        queryset = SubLine.objects.all().select_related('line', 'department', 'factory', 'hq').order_by('subline_name')
        line_id = self.request.query_params.get('line', None)
        department_id = self.request.query_params.get('department', None)
        factory_id = self.request.query_params.get('factory', None)
        hq_id = self.request.query_params.get('hq', None)
        
        if line_id is not None:
            queryset = queryset.filter(line_id=line_id)
        elif department_id is not None:
            queryset = queryset.filter(department_id=department_id)
        elif factory_id is not None:
            queryset = queryset.filter(factory_id=factory_id)
        elif hq_id is not None:
            queryset = queryset.filter(hq_id=hq_id)
            
        return queryset

class StationViewSet(viewsets.ModelViewSet):
    queryset = Station.objects.all()
    serializer_class = StationSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['subline', 'line', 'department', 'factory', 'hq']
    
    def get_queryset(self):
        queryset = Station.objects.all().select_related('subline', 'line', 'department', 'factory', 'hq').order_by('station_name')
        subline_id = self.request.query_params.get('subline', None)
        line_id = self.request.query_params.get('line', None)
        department_id = self.request.query_params.get('department', None)
        factory_id = self.request.query_params.get('factory', None)
        hq_id = self.request.query_params.get('hq', None)
        
        if subline_id is not None:
            queryset = queryset.filter(subline_id=subline_id)
        elif line_id is not None:
            queryset = queryset.filter(line_id=line_id)
        elif department_id is not None:
            queryset = queryset.filter(department_id=department_id)
        elif factory_id is not None:
            queryset = queryset.filter(factory_id=factory_id)
        elif hq_id is not None:
            queryset = queryset.filter(hq_id=hq_id)
            
        return queryset

# Add this to your views.py

from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from .models import HierarchyStructure
from .serializers import HierarchyStructureSerializer

# views.py - Updated HierarchyStructureViewSet
from rest_framework.decorators import api_view

@api_view(['GET'])
def get_all_departments(request):
    departments = Department.objects.all()
    serializer = DepartmentReadSerializer(departments, many=True)
    return Response({
        'departments': serializer.data,
    }, status=status.HTTP_200_OK)



class HierarchyByDepartmentView(APIView):
    """
    Fetch hierarchy by department_id with flexible nesting:
    department → line → subline → station
    department → line → station
    department → subline → station
    department → station
    """

    def get(self, request, *args, **kwargs):
        department_id = request.query_params.get("department_id")
        if not department_id:
            return Response(
                {"error": "department_id parameter is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            department_id = int(department_id)
            structures = HierarchyStructure.objects.filter(department_id=department_id)

            if not structures.exists():
                return Response(
                    {"error": f"No hierarchy found for department id {department_id}"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            department = structures.first().department
            department_data = {
                "department_id": department.department_id,
                "department_name": department.department_name,
                "lines": {},
                "sublines": {},
                "stations": {},
            }

            # build hierarchy flexibly
            for structure in structures:
                line = structure.line
                subline = structure.subline
                station = structure.station

                if line:  # if line exists
                    if line.line_id not in department_data["lines"]:
                        department_data["lines"][line.line_id] = {
                            "line_id": line.line_id,
                            "line_name": line.line_name,
                            "sublines": {},
                            "stations": {},
                        }

                    if subline:  # subline under line
                        if subline.subline_id not in department_data["lines"][line.line_id]["sublines"]:
                            department_data["lines"][line.line_id]["sublines"][subline.subline_id] = {
                                "subline_id": subline.subline_id,
                                "subline_name": subline.subline_name,
                                "stations": {},
                            }

                        if station:
                            department_data["lines"][line.line_id]["sublines"][subline.subline_id]["stations"][
                                station.station_id
                            ] = {
                                "station_id": station.station_id,
                                "station_name": station.station_name,
                            }
                    else:  # station directly under line
                        if station:
                            department_data["lines"][line.line_id]["stations"][station.station_id] = {
                                "station_id": station.station_id,
                                "station_name": station.station_name,
                            }

                elif subline:  # no line, but subline exists
                    if subline.subline_id not in department_data["sublines"]:
                        department_data["sublines"][subline.subline_id] = {
                            "subline_id": subline.subline_id,
                            "subline_name": subline.subline_name,
                            "stations": {},
                        }

                    if station:
                        department_data["sublines"][subline.subline_id]["stations"][station.station_id] = {
                            "station_id": station.station_id,
                            "station_name": station.station_name,
                        }

                elif station:  # no line, no subline → station directly under department
                    department_data["stations"][station.station_id] = {
                        "station_id": station.station_id,
                        "station_name": station.station_name,
                    }

            # convert dicts to lists
            department_data["lines"] = list(department_data["lines"].values())
            for line in department_data["lines"]:
                line["sublines"] = list(line["sublines"].values())
                line["stations"] = list(line["stations"].values())
                for subline in line["sublines"]:
                    subline["stations"] = list(subline["stations"].values())

            department_data["sublines"] = list(department_data["sublines"].values())
            for subline in department_data["sublines"]:
                subline["stations"] = list(subline["stations"].values())

            department_data["stations"] = list(department_data["stations"].values())

            return Response(department_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
from rest_framework.decorators import api_view 
from rest_framework.response import Response 
from .models import HierarchyStructure 
 
@api_view(['GET']) 
def get_hierarchy_structures(request): 
    structures = HierarchyStructure.objects.all() 
 
    merged_data = {} 
 
    for s in structures: 
        # key for grouping (structure_name + hq_id + factory_id) 
        key = (s.structure_name, s.hq.hq_id if s.hq else None, s.factory.factory_id if s.factory else None) 
 
        if key not in merged_data: 
            merged_data[key] = { 
                "structure_id": s.structure_id,  # you can choose the first id or min id 
                "structure_name": s.structure_name, 
                "hq": s.hq.hq_id if s.hq else None, 
                "hq_name": f"{s.hq.hq_name} (ID: {s.hq.hq_id})" if s.hq else None, 
                "factory": s.factory.factory_id if s.factory else None, 
                "factory_name": f"{s.factory.factory_name} (ID: {s.factory.factory_id})" if s.factory else None, 
                "structure_data": { 
                    "hq_name": s.hq.hq_name if s.hq else None, 
                    "factory_name": s.factory.factory_name if s.factory else None, 
                    "departments": [] 
                } 
            } 
 
        structure_data = merged_data[key]["structure_data"] 
 
        # Departments 
        if s.department: 
            dept_obj = next( 
                (d for d in structure_data["departments"] if d["id"] == s.department.department_id), 
                None 
            ) 
            if not dept_obj: 
                dept_obj = { 
                    "id": s.department.department_id, 
                    "department_name": s.department.department_name, 
                    "lines": [],
                    "stations": []  # Add stations array for departments
                } 
                structure_data["departments"].append(dept_obj) 
 
            # Handle stations directly under department (when no line/subline)
            if s.station and not s.line and not s.subline:
                if not any(st["id"] == s.station.station_id for st in dept_obj["stations"]):
                    dept_obj["stations"].append({
                        "id": s.station.station_id,
                        "station_name": s.station.station_name
                    })
 
            # Lines 
            if s.line: 
                line_obj = next( 
                    (l for l in dept_obj["lines"] if l["id"] == s.line.line_id), 
                    None 
                ) 
                if not line_obj: 
                    line_obj = { 
                        "id": s.line.line_id, 
                        "line_name": s.line.line_name, 
                        "sublines": [],
                        "stations": []  # Add stations array for lines
                    } 
                    dept_obj["lines"].append(line_obj) 
 
                # Handle stations directly under line (when no subline)
                if s.station and not s.subline:
                    if not any(st["id"] == s.station.station_id for st in line_obj["stations"]):
                        line_obj["stations"].append({
                            "id": s.station.station_id,
                            "station_name": s.station.station_name
                        })
 
                # Sublines 
                if s.subline: 
                    subline_obj = next( 
                        (sl for sl in line_obj["sublines"] if sl["id"] == s.subline.subline_id), 
                        None 
                    ) 
                    if not subline_obj: 
                        subline_obj = { 
                            "id": s.subline.subline_id, 
                            "subline_name": s.subline.subline_name, 
                            "stations": [] 
                        } 
                        line_obj["sublines"].append(subline_obj) 
 
                    # Stations under sublines
                    if s.station: 
                        if not any(st["id"] == s.station.station_id for st in subline_obj["stations"]): 
                            subline_obj["stations"].append({ 
                                "id": s.station.station_id, 
                                "station_name": s.station.station_name 
                            }) 
 
    # return merged list 
    return Response(list(merged_data.values()))




from rest_framework import viewsets, status
from rest_framework.response import Response
from django.db import transaction
from .models import HierarchyStructure
from .serializers import HierarchyStructureSerializer
import logging

logger = logging.getLogger(__name__)

class HierarchyStructureViewSet(viewsets.ModelViewSet):
    queryset = HierarchyStructure.objects.all().select_related(
        "hq", "factory", "department", "line", "subline", "station"
    )
    serializer_class = HierarchyStructureSerializer

    def create(self, request, *args, **kwargs):
        """Create hierarchy records from nested structure_data"""
        data = request.data
        print("📥 Incoming POST data:", data)

        structure_name = data.get('structure_name')
        structure_data = data.get('structure_data', {})
        hq_id = data.get('hq')
        factory_id = data.get('factory')

        if not structure_name:
            return Response({'error': 'structure_name is required'}, status=status.HTTP_400_BAD_REQUEST)

        return self._create_hierarchy_records(structure_name, structure_data, hq_id, factory_id)

    def update(self, request, *args, **kwargs):
        """Update hierarchy structure by deleting old and creating new records"""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        data = request.data
        
        print(f"📥 Incoming PUT data for structure_id {instance.structure_id}:", data)

        # Get the structure_name from the instance or data
        structure_name = data.get('structure_name', instance.structure_name)
        structure_data = data.get('structure_data', {})
        hq_id = data.get('hq', instance.hq_id if instance.hq else None)
        factory_id = data.get('factory', instance.factory_id if instance.factory else None)

        if 'structure_data' in data:
            # If structure_data is provided, delete old records and create new ones
            print(f"🔄 Updating hierarchy structure: {structure_name}")
            return self._create_hierarchy_records(structure_name, structure_data, hq_id, factory_id)
        else:
            # If no structure_data, do regular update
            serializer = self.get_serializer(instance, data=data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data)

    def _create_hierarchy_records(self, structure_name, structure_data, hq_id, factory_id):
        """Helper method to create hierarchy records (used by both create and update)"""
        try:
            with transaction.atomic():
                # Delete existing records with the same structure_name
                deleted_count = HierarchyStructure.objects.filter(structure_name=structure_name).count()
                HierarchyStructure.objects.filter(structure_name=structure_name).delete()
                print(f"🗑 Deleted {deleted_count} existing records for structure: {structure_name}")

                created_records = []

                # Iterate over departments → lines → sublines → stations
                departments = structure_data.get('departments', [])
                for department in departments:
                    dept_id = department.get('id')
                    print(f"🏢 Processing department {dept_id}")

                    # ✅ Case 1: stations directly under department
                    dept_stations = department.get('stations', [])
                    if dept_stations:
                        print(f"📌 Found {len(dept_stations)} stations directly under department {dept_id}")
                        for station in dept_stations:
                            station_id = station.get('id')
                            print(f"💾 Creating record: dept={dept_id}, station={station_id}")
                            
                            try:
                                record = HierarchyStructure.objects.create(
                                    structure_name=structure_name,
                                    hq_id=hq_id,
                                    factory_id=factory_id,
                                    department_id=dept_id,
                                    line_id=None,
                                    subline_id=None,
                                    station_id=station_id
                                )
                                created_records.append(record)
                                print(f"✅ Successfully created record with ID: {record.structure_id}")
                                
                                # Verify the record was actually saved
                                verification = HierarchyStructure.objects.filter(structure_id=record.structure_id).first()
                                if verification:
                                    print(f"✅ Verification: Record {record.structure_id} exists in DB with station_id={verification.station_id}")
                                else:
                                    print(f"❌ Verification failed: Record {record.structure_id} not found in DB")
                                    
                            except Exception as create_error:
                                print(f"❌ Error creating record: {str(create_error)}")
                                raise create_error

                    # ✅ Case 2: lines processing
                    lines = department.get('lines', [])
                    for line in lines:
                        line_id = line.get('id')
                        
                        # ✅ Case 2a: stations directly under line
                        line_stations = line.get('stations', [])
                        if line_stations:
                            print(f"📌 Found {len(line_stations)} stations directly under line {line_id}")
                            for station in line_stations:
                                station_id = station.get('id')
                                print(f"💾 Creating record: dept={dept_id}, line={line_id}, station={station_id}")
                                
                                try:
                                    record = HierarchyStructure.objects.create(
                                        structure_name=structure_name,
                                        hq_id=hq_id,
                                        factory_id=factory_id,
                                        department_id=dept_id,
                                        line_id=line_id,
                                        subline_id=None,
                                        station_id=station_id
                                    )
                                    created_records.append(record)
                                    print(f"✅ Successfully created line-station record with ID: {record.structure_id}")
                                except Exception as create_error:
                                    print(f"❌ Error creating line-station record: {str(create_error)}")
                                    raise create_error

                        # ✅ Case 2b: sublines processing
                        sublines = line.get('sublines', [])
                        if not sublines:
                            # Create line-only record if no sublines and no stations
                            if not line_stations:
                                print(f"📌 Saving line {line_id} under department {dept_id}")
                                try:
                                    record = HierarchyStructure.objects.create(
                                        structure_name=structure_name,
                                        hq_id=hq_id,
                                        factory_id=factory_id,
                                        department_id=dept_id,
                                        line_id=line_id,
                                        subline_id=None,
                                        station_id=None
                                    )
                                    created_records.append(record)
                                except Exception as create_error:
                                    print(f"❌ Error creating line record: {str(create_error)}")
                                    raise create_error
                        else:
                            for subline in sublines:
                                subline_id = subline.get('id')
                                stations = subline.get('stations', [])

                                if not stations:
                                    # Create subline-only record if no stations
                                    print(f"📌 Saving subline {subline_id} under line {line_id}")
                                    try:
                                        record = HierarchyStructure.objects.create(
                                            structure_name=structure_name,
                                            hq_id=hq_id,
                                            factory_id=factory_id,
                                            department_id=dept_id,
                                            line_id=line_id,
                                            subline_id=subline_id,
                                            station_id=None
                                        )
                                        created_records.append(record)
                                    except Exception as create_error:
                                        print(f"❌ Error creating subline record: {str(create_error)}")
                                        raise create_error
                                else:
                                    for station in stations:
                                        station_id = station.get('id')
                                        print(f"📌 Saving station {station_id} under subline {subline_id}")
                                        try:
                                            record = HierarchyStructure.objects.create(
                                                structure_name=structure_name,
                                                hq_id=hq_id,
                                                factory_id=factory_id,
                                                department_id=dept_id,
                                                line_id=line_id,
                                                subline_id=subline_id,
                                                station_id=station_id
                                            )
                                            created_records.append(record)
                                        except Exception as create_error:
                                            print(f"❌ Error creating station record: {str(create_error)}")
                                            raise create_error

                print(f"🎯 Total created records: {len(created_records)}")
                
                # Double-check all records exist before serializing
                for record in created_records:
                    db_record = HierarchyStructure.objects.filter(structure_id=record.structure_id).first()
                    if not db_record:
                        raise Exception(f"Record {record.structure_id} not found in database after creation")

                # Serialize all created records
                serializer = self.get_serializer(created_records, many=True)
                print("✅ Created hierarchy records:", len(serializer.data))
                
                # Final verification before returning
                final_count = HierarchyStructure.objects.filter(structure_name=structure_name).count()
                print(f"🔍 Final count in DB for structure '{structure_name}': {final_count}")
                
                return Response(serializer.data, status=status.HTTP_201_CREATED)

        except Exception as e:
            print(f"❌ Exception occurred: {str(e)}")
            logger.error(f"Failed to create hierarchy structure: {str(e)}")
            return Response({'error': 'Failed to create hierarchy structure', 'details': str(e)},
                            status=status.HTTP_400_BAD_REQUEST)

    def list(self, request, *args, **kwargs):
        """Return hierarchy records (including dept-level stations)"""
        queryset = self.filter_queryset(self.get_queryset())

        # Show:
        #   - Full hierarchy (dept + line + subline + station)
        #   - Dept + station (when no line/subline exist)
        #   - Line + station (when no subline exist)
        queryset = queryset.filter(
            station__isnull=False,
            department__isnull=False
        )

        # Unique structures
        seen_structures = set()
        unique_structures = []
        for record in queryset:
            if record.structure_name not in seen_structures:
                unique_structures.append(record)
                seen_structures.add(record.structure_name)

        serializer = self.get_serializer(unique_structures, many=True)
        return Response(serializer.data)

    def get_queryset(self):
        """Filter queryset based on query params"""
        queryset = super().get_queryset()
        for field in ['hq', 'factory', 'department', 'line', 'subline', 'station', 'structure_name']:
            value = self.request.query_params.get(field)
            if value:
                queryset = queryset.filter({f"{field}_id" if field != 'structure_name' else field: value})
        return queryset

    def _find_first_station(self, structure_data):
        """Helper to find first station in nested data"""
        departments = structure_data.get('departments', [])
        for department in departments:
            dept_id = department.get('id')

            # Handle case: station directly under department
            for station in department.get('stations', []):
                return {'department_id': dept_id, 'line_id': None,
                        'subline_id': None, 'station_id': station.get('id')}

            # Handle case: station directly under line
            for line in department.get('lines', []):
                line_id = line.get('id')
                for station in line.get('stations', []):
                    return {'department_id': dept_id, 'line_id': line_id,
                            'subline_id': None, 'station_id': station.get('id')}

                # Normal flow: line → subline → station
                for subline in line.get('sublines', []):
                    subline_id = subline.get('id')
                    stations = subline.get('stations', [])
                    if stations:
                        station_id = stations[0].get('id')
                        return {'department_id': dept_id, 'line_id': line_id,
                                'subline_id': subline_id, 'station_id': station_id}
        return None


from rest_framework.decorators import action

# ------------------ Mastertable Views ------------------
# import pandas as pd
# from django.http import HttpResponse, JsonResponse
# from django.views.decorators.csrf import csrf_exempt
# from django.utils.decorators import method_decorator
# from rest_framework import viewsets, status
# from rest_framework.decorators import action
# from rest_framework.response import Response
# from rest_framework.parsers import MultiPartParser, FormParser ,JSONParser
# from io import BytesIO
# import json
# from datetime import datetime
# from .models import MasterTable, Department
# from .serializers import MasterTableSerializer

# class MasterTableViewSet(viewsets.ModelViewSet):
#     queryset = MasterTable.objects.all()
#     serializer_class = MasterTableSerializer
#     parser_classes = (MultiPartParser, FormParser, JSONParser)

#     @action(detail=False, methods=['get'])
#     def download_template(self, request):
#         """Download Excel template with headers only"""
#         try:
#             # Create DataFrame with column headers matching the model fields
#             headers = [
#                 'emp_id',
#                 'first_name', 
#                 'last_name',
#                 'department_name',  # We'll use department name in template
#                 'designation',
#                 'date_of_joining',
#                 'birth_date',
#                 'sex',
#                 'email',
#                 'phone'
#             ]
            
#             # Create empty DataFrame with headers
#             df = pd.DataFrame(columns=headers)
            
#             # Add example row to show format
#             example_row = {
#                 'emp_id': 'EMP001',
#                 'first_name': 'John',
#                 'last_name': 'Doe',
#                 'department_name': 'IT Department',
#                 'designation': 'Senior Developer',
#                 'date_of_joining': '2024-01-15',
#                 'birth_date': '1990-05-20',
#                 'sex': 'M',
#                 'email': 'john.doe@company.com',
#                 'phone': '+1234567890'
#             }
            
#             # Add example row and then clear it (keeps formatting)
#             df.loc[0] = example_row
#             df = df.iloc[0:0]  # Remove the example row, keep structure
            
#             # Create Excel file in memory
#             output = BytesIO()
#             with pd.ExcelWriter(output, engine='openpyxl') as writer:
#                 df.to_excel(writer, sheet_name='Employee Template', index=False)
                
#                 # Get the workbook and worksheet
#                 workbook = writer.book
#                 worksheet = writer.sheets['Employee Template']
                
#                 # Add instructions in a separate sheet
#                 instructions_df = pd.DataFrame({
#                     'Instructions': [
#                         '1. Fill in employee details in the Employee Template sheet',
#                         '2. emp_id: Must be unique (e.g., EMP001, EMP002)',
#                         '3. first_name, last_name: Employee names',
#                         '4. department_name: Exact department name (case sensitive)',
#                         '5. date_of_joining: Format YYYY-MM-DD (e.g., 2024-01-15)',
#                         '6. birth_date: Format YYYY-MM-DD (optional)',
#                         '7. sex: M for Male, F for Female, O for Other',
#                         '8. email: Must be unique and valid email format',
#                         '9. phone: Include country code (e.g., +1234567890)',
#                         '10. designation: Job title (e.g., Software Engineer)',
#                         '',
#                         'Available Departments:'
#                     ]
#                 })
                
#                 # Add available departments to instructions
#                 departments = Department.objects.all().values_list('department_name', flat=True)
#                 for dept in departments:
#                     instructions_df = pd.concat([
#                         instructions_df,
#                         pd.DataFrame({'Instructions': [f'- {dept}']})
#                     ], ignore_index=True)
                
#                 instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
            
#             output.seek(0)
            
#             # Create response
#             response = HttpResponse(
#                 output.getvalue(),
#                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#             )
#             response['Content-Disposition'] = 'attachment; filename=employee_template.xlsx'
            
#             return response
            
#         except Exception as e:
#             return Response(
#                 {'error': f'Failed to generate template: {str(e)}'}, 
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR
#             )
        
#     @action(detail=False, methods=['post'])
#     def upload_excel(self, request):
#         """Upload Excel file and create or update employee records"""
#         if 'file' not in request.FILES:
#             return Response(
#                 {'error': 'No file provided'}, 
#                 status=status.HTTP_400_BAD_REQUEST
#             )
        
#         file = request.FILES['file']
        
#         if not file.name.endswith(('.xlsx', '.xls')):
#             return Response(
#                 {'error': 'Invalid file format. Please upload Excel file (.xlsx or .xls)'}, 
#                 status=status.HTTP_400_BAD_REQUEST
#             )
        
#         try:
#             df = pd.read_excel(file, sheet_name='Employee Template')
            
#             required_columns = ['emp_id', 'first_name', 'email', 'date_of_joining']
#             missing_columns = [col for col in required_columns if col not in df.columns]
            
#             if missing_columns:
#                 return Response(
#                     {'error': f'Missing required columns: {", ".join(missing_columns)}'}, 
#                     status=status.HTTP_400_BAD_REQUEST
#                 )
            
#             df = df.dropna(subset=['emp_id'])
            
#             if df.empty:
#                 return Response(
#                     {'error': 'No valid employee data found in the file'}, 
#                     status=status.HTTP_400_BAD_REQUEST
#                 )
            
#             created_employees = []
#             updated_employees = []
#             errors = []
            
#             for index, row in df.iterrows():
#                 try:
#                     department = None
#                     if pd.notna(row.get('department_name')):
#                         try:
#                             department = Department.objects.get(
#                                 department_name=str(row['department_name']).strip()
#                             )
#                         except Department.DoesNotExist:
#                             errors.append({
#                                 'row': index + 2,
#                                 'emp_id': row.get('emp_id', 'N/A'),
#                                 'error': f'Department "{row["department_name"]}" not found'
#                             })
#                             continue
                    
#                     emp_id = str(row['emp_id']).strip()
                    
#                     def clean_field(val):
#                         if pd.isna(val):
#                             return ''
#                         if isinstance(val, str) and val.strip() == '':
#                             return ''
#                         return val
                    
#                     employee_data = {
#                         'emp_id': emp_id,
#                         'first_name': clean_field(row.get('first_name', '')),
#                         'last_name': clean_field(row.get('last_name', '')),
#                         'department': department.department_id if department else None,
#                         'email': str(row['email']).strip().lower(),
#                         'date_of_joining': pd.to_datetime(row['date_of_joining']).date(),
#                         'birth_date': pd.to_datetime(row['birth_date']).date() if pd.notna(row.get('birth_date')) else None,
#                         'sex': str(row.get('sex', '')).upper() if pd.notna(row.get('sex')) else '',
#                         'phone': clean_field(row.get('phone', '')),
#                         'designation': clean_field(row.get('designation', '')),
#                     }
                    
#                     if employee_data['sex'] and employee_data['sex'] not in ['M', 'F', 'O', '']:
#                         errors.append({
#                             'row': index + 2,
#                             'emp_id': emp_id,
#                             'error': f'Invalid sex value "{employee_data["sex"]}". Use M, F, O, or leave blank'
#                         })
#                         continue
                    
#                     try:
#                         employee = MasterTable.objects.get(emp_id=emp_id)
                        
#                         # Check if all relevant fields are filled (non-empty)
#                         fields_to_check = ['first_name', 'last_name', 'department', 'email', 'date_of_joining', 'birth_date', 'sex', 'phone','designation']
#                         all_filled = all([
#                             getattr(employee, f) not in [None, '', []]
#                             for f in fields_to_check
#                         ])
                        
#                         if all_filled:
#                             errors.append({
#                                 'row': index + 2,
#                                 'emp_id': emp_id,
#                                 'error': 'Same emp_id cannot be uploaded again as all fields are already filled.'
#                             })
#                             continue
                        
#                         # Else update only blank fields with new data
#                         updated_fields = []
#                         for field in employee_data:
#                             new_val = employee_data[field]
#                             old_val = getattr(employee, field, None)
#                             if (old_val in [None, '', []]) and new_val not in [None, '', []]:
#                                 setattr(employee, field, new_val)
#                                 updated_fields.append(field)
#                         if updated_fields:
#                             employee.save(update_fields=updated_fields)
#                             updated_employees.append({
#                                 'emp_id': employee.emp_id,
#                                 'name': f"{employee.first_name} {employee.last_name}".strip(),
#                                 'email': employee.email
#                             })
#                     except MasterTable.DoesNotExist:
#                         serializer = MasterTableSerializer(data=employee_data)
#                         if serializer.is_valid():
#                             employee = serializer.save()
#                             created_employees.append({
#                                 'emp_id': employee.emp_id,
#                                 'name': f"{employee.first_name} {employee.last_name}".strip(),
#                                 'email': employee.email
#                             })
#                         else:
#                             error_messages = []
#                             for field, field_errors in serializer.errors.items():
#                                 if isinstance(field_errors, list):
#                                     error_messages.extend([f"{field}: {error}" for error in field_errors])
#                                 else:
#                                     error_messages.append(f"{field}: {field_errors}")
#                             errors.append({
#                                 'row': index + 2,
#                                 'emp_id': emp_id,
#                                 'error': '; '.join(error_messages)
#                             })
#                 except Exception as e:
#                     errors.append({
#                         'row': index + 2,
#                         'emp_id': row.get('emp_id', 'N/A'),
#                         'error': str(e)
#                     })
            
#             response_data = {
#                 'message': f'Upload completed. {len(created_employees)} employees created, {len(updated_employees)} employees updated.',
#                 'created_count': len(created_employees),
#                 'updated_count': len(updated_employees),
#                 'error_count': len(errors),
#                 'created_employees': created_employees,
#                 'updated_employees': updated_employees
#             }
            
#             if errors:
#                 response_data['errors'] = errors
            
#             return Response(response_data, status=status.HTTP_201_CREATED)
        
#         except Exception as e:
#             return Response(
#                 {'error': f'Failed to process file: {str(e)}'},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR
#             )
        
#     @action(detail=False, methods=['get'], url_path='by-employee-code/(?P<emp_id>[^/.]+)')
#     def retrieve_by_employee_code(self, request, emp_id=None):
#         try:
#            employee = MasterTable.objects.get(emp_id=emp_id)
#            serializer = self.get_serializer(employee)
#            return Response(serializer.data)
#         except MasterTable.DoesNotExist:
#            return Response({"error": "Employee not found"}, status=404)

# import pandas as pd
# from django.http import HttpResponse
# from rest_framework import viewsets, status
# from rest_framework.decorators import action
# from rest_framework.response import Response
# from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
# from io import BytesIO
# from .models import MasterTable, Department
# from .serializers import MasterTableSerializer


# class MasterTableViewSet(viewsets.ModelViewSet):
#     queryset = MasterTable.objects.all()
#     serializer_class = MasterTableSerializer
#     parser_classes = (MultiPartParser, FormParser, JSONParser)

#     # --------------------------------------------------------------------- #
#     # 1. DOWNLOAD TEMPLATE
#     # --------------------------------------------------------------------- #
#     @action(detail=False, methods=['get'])
#     def download_template(self, request):
#         """Download Excel template with headers + instructions + example row"""
#         try:
#             headers = [
#                 'emp_id', 'first_name', 'last_name', 'department_name',
#                 'designation', 'date_of_joining', 'birth_date',
#                 'sex', 'email', 'phone'
#             ]

#             df = pd.DataFrame(columns=headers)

#             # Example row (will be removed but keeps formatting)
#             example = {
#                 'emp_id': 'EMP001',
#                 'first_name': 'John',
#                 'last_name': 'Doe',
#                 'department_name': 'IT Department',
#                 'designation': 'DRIVER',
#                 'date_of_joining': '2024-01-15',
#                 'birth_date': '1990-05-20',
#                 'sex': 'M',
#                 'email': 'john.doe@company.com',
#                 'phone': '+1234567890'
#             }
#             df.loc[0] = example
#             df = df.iloc[0:0]  # Remove example, keep structure

#             output = BytesIO()
#             with pd.ExcelWriter(output, engine='openpyxl') as writer:
#                 df.to_excel(writer, sheet_name='Employee Template', index=False)

#                 # Instructions sheet
#                 instructions = [
#                     '1. Fill in employee details in the Employee Template sheet',
#                     '2. emp_id: Must be unique (e.g., EMP001)',
#                     '3. first_name: Required',
#                     '4. last_name: Optional',
#                     '5. department_name: Exact name (case sensitive)',
#                     '6. designation: Optional. Use exact values from list below',
#                     '7. date_of_joining: YYYY-MM-DD (optional)',
#                     '8. birth_date: YYYY-MM-DD (optional)',
#                     '9. sex: M, F, O or blank',
#                     '10. email: Optional but must be valid & unique if provided',
#                     '11. phone: Optional but must be unique if provided',
#                     '',
#                     'Allowed Designations:'
#                 ] + [f'- {label}' for value, label in MasterTable.DESIGNATION_CHOICES[1:]]

#                 instructions_df = pd.DataFrame({'Instructions': instructions})
#                 instructions_df.to_excel(writer, sheet_name='Instructions', index=False)

#                 # List departments
#                 depts = Department.objects.values_list('department_name', flat=True)
#                 dept_list = pd.DataFrame({'Available Departments': list(depts)})
#                 dept_list.to_excel(writer, sheet_name='Departments', index=False)

#             output.seek(0)
#             response = HttpResponse(
#                 output.getvalue(),
#                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#             )
#             response['Content-Disposition'] = 'attachment; filename=employee_template.xlsx'
#             return response

#         except Exception as e:
#             return Response(
#                 {'error': f'Failed to generate template: {str(e)}'},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR
#             )

#     # --------------------------------------------------------------------- #
#     # 2. UPLOAD EXCEL – FULLY OPTIONAL FIELDS
#     # --------------------------------------------------------------------- #
#     @action(detail=False, methods=['post'])
#     def upload_excel(self, request):
#         """Upload Excel and create/update employees – ALL FIELDS OPTIONAL except emp_id"""
#         if 'file' not in request.FILES:
#             return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

#         file = request.FILES['file']
#         if not file.name.lower().endswith(('.xlsx', '.xls')):
#             return Response(
#                 {'error': 'Invalid file format. Use .xlsx or .xls'},
#                 status=status.HTTP_400_BAD_REQUEST
#             )

#         try:
#             df = pd.read_excel(file, sheet_name='Employee Template')

#             # Only emp_id is required
#             if 'emp_id' not in df.columns:
#                 return Response(
#                     {'error': 'Missing required column: emp_id'},
#                     status=status.HTTP_400_BAD_REQUEST
#                 )

#             df = df.dropna(subset=['emp_id'])
#             if df.empty:
#                 return Response({'error': 'No data found'}, status=status.HTTP_400_BAD_REQUEST)

#             # Helpers
#             def clean(val):
#                 if pd.isna(val):
#                     return ''
#                 if isinstance(val, str):
#                     s = val.strip()
#                     return s if s else ''
#                 return val

#             def is_filled(v):
#                 if v is None: return False
#                 if isinstance(v, str): return len(v.strip()) > 0
#                 return True

#             created, updated, errors = [], [], []

#             for idx, row in df.iterrows():
#                 try:
#                     emp_id = str(row['emp_id']).strip()

#                     # Department lookup
#                     department = None
#                     dept_name = clean(row.get('department_name'))
#                     if dept_name:
#                         try:
#                             department = Department.objects.get(department_name=dept_name)
#                         except Department.DoesNotExist:
#                             errors.append({
#                                 'row': idx + 2,
#                                 'emp_id': emp_id,
#                                 'error': f'Department "{dept_name}" not found'
#                             })
#                             continue

#                     # Date fields (optional)
#                     doj_raw = row.get('date_of_joining')
#                     date_of_joining = (
#                         pd.to_datetime(doj_raw, errors='coerce', dayfirst=True).date()
#                         if pd.notna(doj_raw) else None
#                     )

#                     birth_raw = row.get('birth_date')
#                     birth_date = (
#                         pd.to_datetime(birth_raw, errors='coerce', dayfirst=True).date()
#                         if pd.notna(birth_raw) else None
#                     )

#                     # Phone: treat '0' or blank as None
#                     phone_raw = clean(row.get('phone'))
#                     phone = None if (not phone_raw or phone_raw == '0') else phone_raw

#                     # Build data
#                     data = {
#                         'emp_id': emp_id,
#                         'first_name': clean(row.get('first_name')),
#                         'last_name': clean(row.get('last_name')) or '',
#                         'department': department.pk if department else None,
#                         'email': clean(row.get('email')),
#                         'date_of_joining': date_of_joining,
#                         'birth_date': birth_date,
#                         'sex': (str(row.get('sex', '')).strip().upper() or None),
#                         'phone': phone,
#                         'designation': clean(row.get('designation')),
#                     }

#                     # Validate sex
#                     if data['sex'] and data['sex'] not in ['M', 'F', 'O']:
#                         errors.append({
#                             'row': idx + 2,
#                             'emp_id': emp_id,
#                             'error': f'Invalid sex: {data["sex"]}. Use M/F/O or leave blank'
#                         })
#                         continue

#                     # Validate designation
#                     if data['designation']:
#                         allowed = [c[0] for c in MasterTable.DESIGNATION_CHOICES]
#                         if data['designation'] not in allowed:
#                             errors.append({
#                                 'row': idx + 2,
#                                 'emp_id': emp_id,
#                                 'error': f'Invalid designation: {data["designation"]}'
#                             })
#                             continue

#                     # Uniqueness checks (only if value provided)
#                     unique_errors = []
#                     if data['email'] and MasterTable.objects.filter(email=data['email']).exclude(emp_id=emp_id).exists():
#                         unique_errors.append('email already used')
#                     if data['phone'] and MasterTable.objects.filter(phone=data['phone']).exclude(emp_id=emp_id).exists():
#                         unique_errors.append('phone already used')
#                     if unique_errors:
#                         errors.append({
#                             'row': idx + 2,
#                             'emp_id': emp_id,
#                             'error': '; '.join(unique_errors)
#                         })
#                         continue

#                     # CREATE or UPDATE
#                     try:
#                         employee = MasterTable.objects.get(emp_id=emp_id)

#                         new_fields = []
#                         for field, new_val in data.items():
#                             old_val = getattr(employee, field)
#                             if not is_filled(old_val) and is_filled(new_val):
#                                 setattr(employee, field, new_val)
#                                 new_fields.append(field)

#                         if not new_fields:
#                             errors.append({
#                                 'row': idx + 2,
#                                 'emp_id': emp_id,
#                                 'error': 'No new data to update'
#                             })
#                             continue

#                         employee.save(update_fields=new_fields)
#                         updated.append({
#                             'emp_id': emp_id,
#                             'name': f"{employee.first_name or ''} {employee.last_name or ''}".strip(),
#                             'email': employee.email or ''
#                         })

#                     except MasterTable.DoesNotExist:
#                         serializer = MasterTableSerializer(data=data)
#                         if serializer.is_valid():
#                             employee = serializer.save()
#                             created.append({
#                                 'emp_id': emp_id,
#                                 'name': f"{employee.first_name or ''} {employee.last_name or ''}".strip(),
#                                 'email': employee.email or ''
#                             })
#                         else:
#                             err_msgs = []
#                             for field, msgs in serializer.errors.items():
#                                 err_msgs.append(f"{field}: {'; '.join(msgs)}")
#                             errors.append({
#                                 'row': idx + 2,
#                                 'emp_id': emp_id,
#                                 'error': '; '.join(err_msgs)
#                             })

#                 except Exception as e:
#                     errors.append({
#                         'row': idx + 2,
#                         'emp_id': row.get('emp_id', 'N/A'),
#                         'error': str(e)
#                     })

#             # Final response
#             response_data = {
#                 'message': f'Upload complete. {len(created)} created, {len(updated)} updated.',
#                 'created_count': len(created),
#                 'updated_count': len(updated),
#                 'error_count': len(errors),
#                 'created_employees': created,
#                 'updated_employees': updated,
#             }
#             if errors:
#                 response_data['errors'] = errors

#             return Response(response_data, status=status.HTTP_201_CREATED)

#         except Exception as e:
#             return Response(
#                 {'error': f'Failed to process file: {str(e)}'},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR
#             )

#     # --------------------------------------------------------------------- #
#     # 3. GET BY EMP ID
#     # --------------------------------------------------------------------- #
#     @action(detail=False, methods=['get'], url_path='by-employee-code/(?P<emp_id>[^/.]+)')
#     def retrieve_by_employee_code(self, request, emp_id=None):
#         try:
#             employee = MasterTable.objects.get(emp_id=emp_id)
#             serializer = self.get_serializer(employee)
#             return Response(serializer.data)
#         except MasterTable.DoesNotExist:
#             return Response({"error": "Employee not found"}, status=404)



# --------------------------------------------------------------
#  MasterTableViewSet – UPDATED for sub_department choices
# --------------------------------------------------------------
# class MasterTableViewSet(viewsets.ModelViewSet):
#     queryset = MasterTable.objects.all()
#     serializer_class = MasterTableSerializer
#     parser_classes = (MultiPartParser, FormParser, JSONParser)

#     # --------------------------------------------------------------------- #
#     # 1. DOWNLOAD TEMPLATE
#     # --------------------------------------------------------------------- #
#     @action(detail=False, methods=['get'])
#     def download_template(self, request):
#         """Download Excel template with headers + instructions + example row"""
#         try:
#             headers = [
#                 'emp_id', 'first_name', 'last_name', 'department_name',
#                 'sub_department',          # <-- NEW
#                 'designation', 'date_of_joining', 'birth_date',
#                 'sex', 'email', 'phone'
#             ]

#             df = pd.DataFrame(columns=headers)

#             # Example row (keeps formatting)
#             example = {
#                 'emp_id': 'EMP001',
#                 'first_name': 'John',
#                 'last_name': 'Doe',
#                 'department_name': 'IT Department',
#                 'sub_department': 'Power Folding',      # <-- NEW
#                 'designation': 'DRIVER',
#                 'date_of_joining': '2024-01-15',
#                 'birth_date': '1990-05-20',
#                 'sex': 'M',
#                 'email': 'john.doe@company.com',
#                 'phone': '+1234567890'
#             }
#             df.loc[0] = example
#             df = df.iloc[0:0]               # remove example row

#             output = BytesIO()
#             with pd.ExcelWriter(output, engine='openpyxl') as writer:
#                 df.to_excel(writer, sheet_name='Employee Template', index=False)

#                 # -------------------------------------------------
#                 # Instructions sheet – now includes sub-department
#                 # -------------------------------------------------
#                 instructions = [
#                     '1. Fill in employee details in the Employee Template sheet',
#                     '2. emp_id: Must be unique (e.g., EMP001)',
#                     '3. first_name: Required',
#                     '4. last_name: Optional',
#                     '5. department_name: Exact name (case sensitive)',
#                     '6. sub_department: Optional – choose from the list below',
#                     '7. designation: Optional – choose from the list below',
#                     '8. date_of_joining: YYYY-MM-DD (optional)',
#                     '9. birth_date: YYYY-MM-DD (optional)',
#                     '10. sex: M, F, O or blank',
#                     '11. email: Optional but must be valid & unique if provided',
#                     '12. phone: Optional but must be unique if provided',
#                     '',
#                     'Allowed Sub-Departments:'
#                 ] + [f'- {label}' for value, label in MasterTable.SUB_DEPARTMENT_CHOICES[1:]]
#                 instructions += ['', 'Allowed Designations:'] \
#                               + [f'- {label}' for value, label in MasterTable.DESIGNATION_CHOICES[1:]]

#                 instructions_df = pd.DataFrame({'Instructions': instructions})
#                 instructions_df.to_excel(writer, sheet_name='Instructions', index=False)

#                 # List departments (unchanged)
#                 depts = Department.objects.values_list('department_name', flat=True)
#                 dept_list = pd.DataFrame({'Available Departments': list(depts)})
#                 dept_list.to_excel(writer, sheet_name='Departments', index=False)

#             output.seek(0)
#             response = HttpResponse(
#                 output.getvalue(),
#                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#             )
#             response['Content-Disposition'] = 'attachment; filename=employee_template.xlsx'
#             return response

#         except Exception as e:
#             return Response(
#                 {'error': f'Failed to generate template: {str(e)}'},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR
#             )

#     # --------------------------------------------------------------------- #
#     # 2. UPLOAD EXCEL – FULLY OPTIONAL FIELDS
#     # --------------------------------------------------------------------- #
#     @action(detail=False, methods=['post'])
#     def upload_excel(self, request):
#         """Upload Excel and create/update employees – ALL FIELDS OPTIONAL except emp_id"""
#         if 'file' not in request.FILES:
#             return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

#         file = request.FILES['file']
#         if not file.name.lower().endswith(('.xlsx', '.xls')):
#             return Response(
#                 {'error': 'Invalid file format. Use .xlsx or .xls'},
#                 status=status.HTTP_400_BAD_REQUEST
#             )

#         try:
#             df = pd.read_excel(file, sheet_name='Employee Template')

#             # Required column
#             if 'emp_id' not in df.columns:
#                 return Response(
#                     {'error': 'Missing required column: emp_id'},
#                     status=status.HTTP_400_BAD_REQUEST
#                 )

#             df = df.dropna(subset=['emp_id'])
#             if df.empty:
#                 return Response({'error': 'No data found'}, status=status.HTTP_400_BAD_REQUEST)

#             # -----------------------------------------------------------------
#             # Helper utilities
#             # -----------------------------------------------------------------
#             def clean(val):
#                 if pd.isna(val):
#                     return ''
#                 if isinstance(val, str):
#                     return val.strip()
#                 return val

#             def is_filled(v):
#                 if v is None:
#                     return False
#                 if isinstance(v, str):
#                     return len(v.strip()) > 0
#                 return True

#             # -----------------------------------------------------------------
#             # Choice lists (for fast look-up)
#             # -----------------------------------------------------------------
#             allowed_designations = [c[0] for c in MasterTable.DESIGNATION_CHOICES]
#             allowed_sub_depts    = [c[0] for c in MasterTable.SUB_DEPARTMENT_CHOICES]

#             created, updated, errors = [], [], []

#             for idx, row in df.iterrows():
#                 try:
#                     emp_id = str(row['emp_id']).strip()

#                     # ------------------- Department -------------------
#                     department = None
#                     dept_name = clean(row.get('department_name'))
#                     if dept_name:
#                         try:
#                             department = Department.objects.get(department_name=dept_name)
#                         except Department.DoesNotExist:
#                             errors.append({
#                                 'row': idx + 2,
#                                 'emp_id': emp_id,
#                                 'error': f'Department "{dept_name}" not found'
#                             })
#                             continue

#                     # ------------------- Dates -------------------
#                     doj_raw = row.get('date_of_joining')
#                     date_of_joining = (
#                         pd.to_datetime(doj_raw, errors='coerce', dayfirst=True).date()
#                         if pd.notna(doj_raw) else None
#                     )

#                     birth_raw = row.get('birth_date')
#                     birth_date = (
#                         pd.to_datetime(birth_raw, errors='coerce', dayfirst=True).date()
#                         if pd.notna(birth_raw) else None
#                     )

#                     # ------------------- Phone -------------------
#                     phone_raw = clean(row.get('phone'))
#                     phone = None if (not phone_raw or phone_raw == '0') else phone_raw

#                     # ------------------- Sub-Department -------------------
#                     sub_dept_raw = clean(row.get('sub_department'))
#                     sub_department = None
#                     if sub_dept_raw:
#                         if sub_dept_raw not in allowed_sub_depts:
#                             errors.append({
#                                 'row': idx + 2,
#                                 'emp_id': emp_id,
#                                 'error': f'Invalid sub_department: {sub_dept_raw}'
#                             })
#                             continue
#                         sub_department = sub_dept_raw

#                     # ------------------- Build data dict -------------------
#                     data = {
#                         'emp_id': emp_id,
#                         'first_name': clean(row.get('first_name')),
#                         'last_name': clean(row.get('last_name')) or '',
#                         'department': department.pk if department else None,
#                         'sub_department': sub_department,                # <-- NEW
#                         'email': clean(row.get('email')),
#                         'date_of_joining': date_of_joining,
#                         'birth_date': birth_date,
#                         'sex': (str(row.get('sex', '')).strip().upper() or None),
#                         'phone': phone,
#                         'designation': clean(row.get('designation')),
#                     }

#                     # ------------------- SEX validation -------------------
#                     if data['sex'] and data['sex'] not in ['M', 'F', 'O']:
#                         errors.append({
#                             'row': idx + 2,
#                             'emp_id': emp_id,
#                             'error': f'Invalid sex: {data["sex"]}. Use M/F/O or leave blank'
#                         })
#                         continue

#                     # ------------------- DESIGNATION validation -------------------
#                     if data['designation']:
#                         if data['designation'] not in allowed_designations:
#                             errors.append({
#                                 'row': idx + 2,
#                                 'emp_id': emp_id,
#                                 'error': f'Invalid designation: {data["designation"]}'
#                             })
#                             continue

#                     # ------------------- Uniqueness (email / phone) -------------------
#                     unique_errors = []
#                     if data['email'] and MasterTable.objects.filter(email=data['email']).exclude(emp_id=emp_id).exists():
#                         unique_errors.append('email already used')
#                     if data['phone'] and MasterTable.objects.filter(phone=data['phone']).exclude(emp_id=emp_id).exists():
#                         unique_errors.append('phone already used')
#                     if unique_errors:
#                         errors.append({
#                             'row': idx + 2,
#                             'emp_id': emp_id,
#                             'error': '; '.join(unique_errors)
#                         })
#                         continue

#                     # ------------------- CREATE or UPDATE -------------------
#                     try:
#                         employee = MasterTable.objects.get(emp_id=emp_id)

#                         # Determine which fields actually changed
#                         new_fields = []
#                         for field, new_val in data.items():
#                             old_val = getattr(employee, field)
#                             if not is_filled(old_val) and is_filled(new_val):
#                                 setattr(employee, field, new_val)
#                                 new_fields.append(field)

#                         if not new_fields:
#                             errors.append({
#                                 'row': idx + 2,
#                                 'emp_id': emp_id,
#                                 'error': 'No new data to update'
#                             })
#                             continue

#                         employee.save(update_fields=new_fields)
#                         updated.append({
#                             'emp_id': emp_id,
#                             'name': f"{employee.first_name or ''} {employee.last_name or ''}".strip(),
#                             'email': employee.email or '',
#                             'sub_department': employee.sub_department or ''
#                         })

#                     except MasterTable.DoesNotExist:
#                         serializer = MasterTableSerializer(data=data)
#                         if serializer.is_valid():
#                             employee = serializer.save()
#                             created.append({
#                                 'emp_id': emp_id,
#                                 'name': f"{employee.first_name or ''} {employee.last_name or ''}".strip(),
#                                 'email': employee.email or '',
#                                 'sub_department': employee.sub_department or ''
#                             })
#                         else:
#                             err_msgs = []
#                             for field, msgs in serializer.errors.items():
#                                 err_msgs.append(f"{field}: {'; '.join(msgs)}")
#                             errors.append({
#                                 'row': idx + 2,
#                                 'emp_id': emp_id,
#                                 'error': '; '.join(err_msgs)
#                             })

#                 except Exception as e:
#                     errors.append({
#                         'row': idx + 2,
#                         'emp_id': row.get('emp_id', 'N/A'),
#                         'error': str(e)
#                     })

#             # ------------------- Final response -------------------
#             response_data = {
#                 'message': f'Upload complete. {len(created)} created, {len(updated)} updated.',
#                 'created_count': len(created),
#                 'updated_count': len(updated),
#                 'error_count': len(errors),
#                 'created_employees': created,
#                 'updated_employees': updated,
#             }
#             if errors:
#                 response_data['errors'] = errors

#             return Response(response_data, status=status.HTTP_201_CREATED)

#         except Exception as e:
#             return Response(
#                 {'error': f'Failed to process file: {str(e)}'},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR
#             )

#     # --------------------------------------------------------------------- #
#     # 3. GET BY EMP ID (unchanged – just kept for completeness)
#     # --------------------------------------------------------------------- #
#     @action(detail=False, methods=['get'], url_path='by-employee-code/(?P<emp_id>[^/.]+)')
#     def retrieve_by_employee_code(self, request, emp_id=None):
#         try:
#             employee = MasterTable.objects.get(emp_id=emp_id)
#             serializer = self.get_serializer(employee)
#             return Response(serializer.data)
#         except MasterTable.DoesNotExist:
#             return Response({"error": "Employee not found"}, status=404)




# from django.http import HttpResponse
# from rest_framework import viewsets, status
# from rest_framework.decorators import action
# from rest_framework.response import Response
# from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
# from io import BytesIO
# import pandas as pd
# from .models import MasterTable, Department
# from .serializers import MasterTableSerializer


# class MasterTableViewSet(viewsets.ModelViewSet):
#     queryset = MasterTable.objects.all()
#     serializer_class = MasterTableSerializer
#     parser_classes = (MultiPartParser, FormParser, JSONParser)

#     # --------------------------------------------------------------------- #
#     # 1. DOWNLOAD TEMPLATE – UPDATED: NO EXTRA SHEET + GROUPED SUB-DEPTS
#     # --------------------------------------------------------------------- #
#     @action(detail=False, methods=['get'])
#     def download_template(self, request):
#         """Download Excel template with headers + full instructions (no extra sheets)"""
#         try:
#             # ------------------- Column Headers -------------------
#             headers = [
#                 'emp_id', 'first_name', 'last_name', 'department_name',
#                 'sub_department', 'designation', 'date_of_joining',
#                 'birth_date', 'sex', 'email', 'phone'
#             ]

#             df = pd.DataFrame(columns=headers)

#             # ------------------- Example Row (keeps formatting) -------------------
#             example = {
#                 'emp_id': 'EMP001',
#                 'first_name': 'John',
#                 'last_name': 'Doe',
#                 'department_name': 'Production',
#                 'sub_department': 'ACTUATOR',
#                 'designation': 'DRIVER',
#                 'date_of_joining': '2024-01-15',
#                 'birth_date': '1990-05-20',
#                 'sex': 'M',
#                 'email': 'john.doe@company.com',
#                 'phone': '+1234567890'
#             }
#             df.loc[0] = example
#             df = df.iloc[0:0]  # Remove row, keep structure

#             # ------------------- Write Excel -------------------
#             output = BytesIO()
#             with pd.ExcelWriter(output, engine='openpyxl') as writer:
#                 df.to_excel(writer, sheet_name='Employee Template', index=False)

#                 # ------------------- INSTRUCTIONS SHEET -------------------
#                 instructions = [
#                     '1. Fill in employee details in the "Employee Template" sheet',
#                     '2. emp_id: Must be unique (e.g., EMP001)',
#                     '3. first_name: Required',
#                     '4. last_name: Optional',
#                     '5. department_name: Exact name (case sensitive) – see list below',
#                     '6. sub_department: Optional – choose from the list that matches the department',
#                     '7. designation: Optional – choose from the list below',
#                     '8. date_of_joining: YYYY-MM-DD (optional)',
#                     '9. birth_date: YYYY-MM-DD (optional)',
#                     '10. sex: M, F, O or blank',
#                     '11. email: Optional but must be valid & unique if provided',
#                     '12. phone: Optional but must be unique if provided',
#                     '',
#                     '--- AVAILABLE DEPARTMENTS ---',
#                     '- Production',
#                     '- Quality',
#                     '- HR',
#                     '- IT',
#                     '',
#                     '--- PRODUCTION SUB-DEPARTMENTS ---',
#                     '- ACTUATOR',
#                     '- E/MIRROR',
#                     '- IMM',
#                     '- MIRROR',
#                     '- OFF LINE',
#                     '- PAINT SHOP',
#                     '- Power Folding',
#                     '',
#                     '--- QUALITY SUB-DEPARTMENTS ---',
#                     '- IMM PQA',
#                     '- PAINT SHOP PQA',
#                     '- E/MIRROR PQA',
#                     '- OFF LINE PQA',
#                     '- IQC',
#                     '- IFC',
#                     '- QA DISP. PDI',
#                     '- QUALITY ASSURANCE',
#                     '',
#                     '--- ALLOWED DESIGNATIONS ---',
#                 ] + [f'- {label}' for _, label in MasterTable.DESIGNATION_CHOICES[1:]]

#                 instructions_df = pd.DataFrame({'Instructions': instructions})
#                 instructions_df.to_excel(writer, sheet_name='Instructions', index=False)

#             # ------------------- Return File -------------------
#             output.seek(0)
#             response = HttpResponse(
#                 output.getvalue(),
#                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#             )
#             response['Content-Disposition'] = 'attachment; filename=employee_template.xlsx'
#             return response

#         except Exception as e:
#             return Response(
#                 {'error': f'Failed to generate template: {str(e)}'},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR
#             )

#     # --------------------------------------------------------------------- #
#     # 2. UPLOAD EXCEL – FULLY OPTIONAL FIELDS
#     # --------------------------------------------------------------------- #
#     @action(detail=False, methods=['post'])
#     def upload_excel(self, request):
#         """Upload Excel and create/update employees – ALL FIELDS OPTIONAL except emp_id"""
#         if 'file' not in request.FILES:
#             return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

#         file = request.FILES['file']
#         if not file.name.lower().endswith(('.xlsx', '.xls')):
#             return Response(
#                 {'error': 'Invalid file format. Use .xlsx or .xls'},
#                 status=status.HTTP_400_BAD_REQUEST
#             )

#         try:
#             df = pd.read_excel(file, sheet_name='Employee Template')

#             # Required column
#             if 'emp_id' not in df.columns:
#                 return Response(
#                     {'error': 'Missing required column: emp_id'},
#                     status=status.HTTP_400_BAD_REQUEST
#                 )

#             df = df.dropna(subset=['emp_id'])
#             if df.empty:
#                 return Response({'error': 'No data found'}, status=status.HTTP_400_BAD_REQUEST)

#             # -----------------------------------------------------------------
#             # Helper utilities
#             # -----------------------------------------------------------------
#             def clean(val):
#                 if pd.isna(val):
#                     return ''
#                 if isinstance(val, str):
#                     return val.strip()
#                 return val

#             def is_filled(v):
#                 if v is None:
#                     return False
#                 if isinstance(v, str):
#                     return len(v.strip()) > 0
#                 return True

#             # -----------------------------------------------------------------
#             # Choice lists (for fast look-up)
#             # -----------------------------------------------------------------
#             allowed_designations = [c[0] for c in MasterTable.DESIGNATION_CHOICES]
#             allowed_sub_depts    = [c[0] for c in MasterTable.SUB_DEPARTMENT_CHOICES]

#             created, updated, errors = [], [], []

#             for idx, row in df.iterrows():
#                 try:
#                     emp_id = str(row['emp_id']).strip()

#                     # ------------------- Department -------------------
#                     department = None
#                     dept_name = clean(row.get('department_name'))
#                     if dept_name:
#                         try:
#                             department = Department.objects.get(department_name=dept_name)
#                         except Department.DoesNotExist:
#                             errors.append({
#                                 'row': idx + 2,
#                                 'emp_id': emp_id,
#                                 'error': f'Department "{dept_name}" not found'
#                             })
#                             continue

#                     # ------------------- Dates -------------------
#                     doj_raw = row.get('date_of_joining')
#                     date_of_joining = (
#                         pd.to_datetime(doj_raw, errors='coerce', dayfirst=True).date()
#                         if pd.notna(doj_raw) else None
#                     )

#                     birth_raw = row.get('birth_date')
#                     birth_date = (
#                         pd.to_datetime(birth_raw, errors='coerce', dayfirst=True).date()
#                         if pd.notna(birth_raw) else None
#                     )

#                     # ------------------- Phone -------------------
#                     phone_raw = clean(row.get('phone'))
#                     phone = None if (not phone_raw or phone_raw == '0') else phone_raw

#                     # ------------------- Sub-Department -------------------
#                     sub_dept_raw = clean(row.get('sub_department'))
#                     sub_department = None
#                     if sub_dept_raw:
#                         if sub_dept_raw not in allowed_sub_depts:
#                             errors.append({
#                                 'row': idx + 2,
#                                 'emp_id': emp_id,
#                                 'error': f'Invalid sub_department: {sub_dept_raw}'
#                             })
#                             continue
#                         sub_department = sub_dept_raw

#                     # ------------------- Build data dict -------------------
#                     data = {
#                         'emp_id': emp_id,
#                         'first_name': clean(row.get('first_name')),
#                         'last_name': clean(row.get('last_name')) or '',
#                         'department': department.pk if department else None,
#                         'sub_department': sub_department,
#                         'email': clean(row.get('email')),
#                         'date_of_joining': date_of_joining,
#                         'birth_date': birth_date,
#                         'sex': (str(row.get('sex', '')).strip().upper() or None),
#                         'phone': phone,
#                         'designation': clean(row.get('designation')),
#                     }

#                     # ------------------- SEX validation -------------------
#                     if data['sex'] and data['sex'] not in ['M', 'F', 'O']:
#                         errors.append({
#                             'row': idx + 2,
#                             'emp_id': emp_id,
#                             'error': f'Invalid sex: {data["sex"]}. Use M/F/O or leave blank'
#                         })
#                         continue

#                     # ------------------- DESIGNATION validation -------------------
#                     if data['designation']:
#                         if data['designation'] not in allowed_designations:
#                             errors.append({
#                                 'row': idx + 2,
#                                 'emp_id': emp_id,
#                                 'error': f'Invalid designation: {data["designation"]}'
#                             })
#                             continue

#                     # ------------------- Uniqueness (email / phone) -------------------
#                     unique_errors = []
#                     if data['email'] and MasterTable.objects.filter(email=data['email']).exclude(emp_id=emp_id).exists():
#                         unique_errors.append('email already used')
#                     if data['phone'] and MasterTable.objects.filter(phone=data['phone']).exclude(emp_id=emp_id).exists():
#                         unique_errors.append('phone already used')
#                     if unique_errors:
#                         errors.append({
#                             'row': idx + 2,
#                             'emp_id': emp_id,
#                             'error': '; '.join(unique_errors)
#                         })
#                         continue

#                     # ------------------- CREATE or UPDATE -------------------
#                     try:
#                         employee = MasterTable.objects.get(emp_id=emp_id)

#                         # Determine which fields actually changed
#                         new_fields = []
#                         for field, new_val in data.items():
#                             old_val = getattr(employee, field)
#                             if not is_filled(old_val) and is_filled(new_val):
#                                 setattr(employee, field, new_val)
#                                 new_fields.append(field)

#                         if not new_fields:
#                             errors.append({
#                                 'row': idx + 2,
#                                 'emp_id': emp_id,
#                                 'error': 'No new data to update'
#                             })
#                             continue

#                         employee.save(update_fields=new_fields)
#                         updated.append({
#                             'emp_id': emp_id,
#                             'name': f"{employee.first_name or ''} {employee.last_name or ''}".strip(),
#                             'email': employee.email or '',
#                             'sub_department': employee.sub_department or ''
#                         })

#                     except MasterTable.DoesNotExist:
#                         serializer = MasterTableSerializer(data=data)
#                         if serializer.is_valid():
#                             employee = serializer.save()
#                             created.append({
#                                 'emp_id': emp_id,
#                                 'name': f"{employee.first_name or ''} {employee.last_name or ''}".strip(),
#                                 'email': employee.email or '',
#                                 'sub_department': employee.sub_department or ''
#                             })
#                         else:
#                             err_msgs = []
#                             for field, msgs in serializer.errors.items():
#                                 err_msgs.append(f"{field}: {'; '.join(msgs)}")
#                             errors.append({
#                                 'row': idx + 2,
#                                 'emp_id': emp_id,
#                                 'error': '; '.join(err_msgs)
#                             })

#                 except Exception as e:
#                     errors.append({
#                         'row': idx + 2,
#                         'emp_id': row.get('emp_id', 'N/A'),
#                         'error': str(e)
#                     })

#             # ------------------- Final response -------------------
#             response_data = {
#                 'message': f'Upload complete. {len(created)} created, {len(updated)} updated.',
#                 'created_count': len(created),
#                 'updated_count': len(updated),
#                 'error_count': len(errors),
#                 'created_employees': created,
#                 'updated_employees': updated,
#             }
#             if errors:
#                 response_data['errors'] = errors

#             return Response(response_data, status=status.HTTP_201_CREATED)

#         except Exception as e:
#             return Response(
#                 {'error': f'Failed to process file: {str(e)}'},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR
#             )

#     # --------------------------------------------------------------------- #
#     # 3. GET BY EMP ID
#     # --------------------------------------------------------------------- #
#     @action(detail=False, methods=['get'], url_path='by-employee-code/(?P<emp_id>[^/.]+)')
#     def retrieve_by_employee_code(self, request, emp_id=None):
#         try:
#             employee = MasterTable.objects.get(emp_id=emp_id)
#             serializer = self.get_serializer(employee)
#             return Response(serializer.data)
#         except MasterTable.DoesNotExist:
#             return Response({"error": "Employee not found"}, status=404)

# import pandas as pd
# from django.http import HttpResponse
# from django.db import transaction
# from rest_framework import viewsets, status
# from rest_framework.decorators import action
# from rest_framework.response import Response
# from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
# from io import BytesIO
# from .models import MasterTable, Department
# from .serializers import MasterTableSerializer


# class MasterTableViewSet(viewsets.ModelViewSet):
#     queryset = MasterTable.objects.select_related("department").all()
#     serializer_class = MasterTableSerializer
#     parser_classes = (MultiPartParser, FormParser, JSONParser)

#     # --------------------------------------------------------------------- #
#     # 1. LIST DEPARTMENTS (for frontend)
#     # --------------------------------------------------------------------- #
#     @action(detail=False, methods=["get"])
#     def list_departments(self, request):
#         """Return all departments for frontend dropdown"""
#         depts = Department.objects.values("id", "department_name").order_by("department_name")
#         return Response(list(depts))

#     # --------------------------------------------------------------------- #
#     # 2. DOWNLOAD TEMPLATE – DYNAMIC FROM DB
#     # --------------------------------------------------------------------- #
#     @action(detail=False, methods=["get"])
#     def download_template(self, request):
#         """Generate Excel template with real departments + grouped sub-depts"""
#         try:
#             headers = [
#                 "emp_id", "first_name", "last_name", "department_name",
#                 "sub_department", "designation", "date_of_joining",
#                 "birth_date", "sex", "email", "phone",
#             ]
#             df = pd.DataFrame(columns=headers)

#             # Fetch real departments
#             departments = list(Department.objects.values_list('department_name', flat=True))
#             example_dept = departments[0] if departments else "Production"
#             example_sub = ""
#             if example_dept == "Production":
#                 example_sub = "ACTUATOR"
#             elif example_dept == "Quality":
#                 example_sub = "IQC"

#             # Example row (keeps formatting)
#             example = {
#                 "emp_id": "EMP001",
#                 "first_name": "John",
#                 "last_name": "Doe",
#                 "department_name": example_dept,
#                 "sub_department": example_sub,
#                 "designation": "DRIVER",
#                 "date_of_joining": "2024-01-15",
#                 "birth_date": "1990-05-20",
#                 "sex": "M",
#                 "email": "john.doe@company.com",
#                 "phone": "+1234567890",
#             }
#             df.loc[0] = example
#             df = df.iloc[0:0]  # Remove row

#             output = BytesIO()
#             with pd.ExcelWriter(output, engine="openpyxl") as writer:
#                 df.to_excel(writer, sheet_name="Employee Template", index=False)

#                 # ------------------- INSTRUCTIONS -------------------
#                 instructions = [
#                     "1. Fill ONLY the columns you want to set.",
#                     "2. emp_id is **mandatory** and must be unique.",
#                     "3. All other columns are optional.",
#                     "4. department_name must match an existing department (case-sensitive).",
#                     "5. sub_department: only valid for Production or Quality.",
#                     "6. designation must be one of the allowed values.",
#                     "7. Dates → YYYY-MM-DD.",
#                     "8. sex → M / F / O (or leave blank).",
#                     "9. email / phone → unique if provided.",
#                     "",
#                     "AVAILABLE DEPARTMENTS (from database):",
#                 ] + [f"- {d}" for d in departments]

#                 # Sub-department groups
#                 production_subs = [
#                     "ACTUATOR", "E/MIRROR", "IMM", "MIRROR",
#                     "OFF LINE", "PAINT SHOP", "Power Folding"
#                 ]
#                 quality_subs = [
#                     "IMM PQA", "PAINT SHOP PQA", "E/MIRROR PQA",
#                     "OFF LINE PQA", "IQC", "IFC",
#                     "QA DISP. PDI", "QUALITY ASSURANCE"
#                 ]

#                 if "Production" in departments:
#                     instructions += ["", "PRODUCTION SUB-DEPARTMENTS:"]
#                     instructions += [f"- {s}" for s in production_subs]

#                 if "Quality" in departments:
#                     instructions += ["", "QUALITY SUB-DEPARTMENTS:"]
#                     instructions += [f"- {s}" for s in quality_subs]

#                 # Designations
#                 instructions += ["", "ALLOWED DESIGNATIONS:"]
#                 instructions += [f"- {label}" for _, label in MasterTable.DESIGNATION_CHOICES[1:]]

#                 pd.DataFrame({"Instructions": instructions}).to_excel(
#                     writer, sheet_name="Instructions", index=False
#                 )

#             output.seek(0)
#             response = HttpResponse(
#                 output.getvalue(),
#                 content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#             )
#             response["Content-Disposition"] = 'attachment; filename="employee_template.xlsx"'
#             return response

#         except Exception as e:
#             return Response(
#                 {"error": f"Failed to generate template: {str(e)}"},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR,
#             )

#     # --------------------------------------------------------------------- #
#     # 3. UPLOAD EXCEL – FULL VALIDATION
#     # --------------------------------------------------------------------- #
#     @action(detail=False, methods=["post"])
#     def upload_excel(self, request):
#         """Upload Excel and create/update employees"""
#         if "file" not in request.FILES:
#             return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

#         file = request.FILES["file"]
#         if not file.name.lower().endswith((".xlsx", ".xls")):
#             return Response(
#                 {"error": "Only .xlsx or .xls files are allowed"},
#                 status=status.HTTP_400_BAD_REQUEST,
#             )

#         try:
#             df = pd.read_excel(file, sheet_name="Employee Template")
#         except Exception as e:
#             return Response(
#                 {"error": f"Could not read Excel file: {e}"},
#                 status=status.HTTP_400_BAD_REQUEST,
#             )

#         # Basic validation
#         if "emp_id" not in df.columns:
#             return Response(
#                 {"error": "Column 'emp_id' is required"},
#                 status=status.HTTP_400_BAD_REQUEST,
#             )
#         df = df.dropna(subset=["emp_id"])
#         if df.empty:
#             return Response({"error": "No rows with emp_id found"}, status=status.HTTP_400_BAD_REQUEST)

#         # Helpers
#         def clean(val):
#             if pd.isna(val):
#                 return ""
#             return str(val).strip()

#         def filled(v):
#             return bool(v) if not isinstance(v, str) else bool(v.strip())

#         # Choice sets
#         allowed_designations = {c[0] for c in MasterTable.DESIGNATION_CHOICES}
#         allowed_sub_depts = {c[0] for c in MasterTable.SUB_DEPARTMENT_CHOICES}

#         created, updated, errors = [], [], []

#         for idx, row in df.iterrows():
#             try:
#                 emp_id = clean(row["emp_id"])
#                 if not emp_id:
#                     errors.append({"row": idx + 2, "emp_id": "N/A", "error": "emp_id is empty"})
#                     continue

#                 # Department lookup
#                 department = None
#                 dept_name = clean(row.get("department_name"))
#                 if dept_name:
#                     try:
#                         department = Department.objects.get(department_name=dept_name)
#                     except Department.DoesNotExist:
#                         errors.append(
#                             {"row": idx + 2, "emp_id": emp_id, "error": f"Department '{dept_name}' not found"}
#                         )
#                         continue

#                 # Dates
#                 def parse_date(val):
#                     if pd.isna(val):
#                         return None
#                     try:
#                         return pd.to_datetime(val, errors="coerce").date()
#                     except Exception:
#                         return None

#                 date_of_joining = parse_date(row.get("date_of_joining"))
#                 birth_date = parse_date(row.get("birth_date"))

#                 # Phone
#                 phone_raw = clean(row.get("phone"))
#                 phone = None if not phone_raw or phone_raw == "0" else phone_raw

#                 # Sub-department
#                 sub_dept_raw = clean(row.get("sub_department"))
#                 sub_department = None
#                 if sub_dept_raw:
#                     if sub_dept_raw not in allowed_sub_depts:
#                         errors.append(
#                             {"row": idx + 2, "emp_id": emp_id, "error": f"Invalid sub_department '{sub_dept_raw}'"}
#                         )
#                         continue
#                     # Optional: enforce sub-dept only for Production/Quality
#                     if dept_name and dept_name not in ["Production", "Quality"]:
#                         errors.append(
#                             {"row": idx + 2, "emp_id": emp_id, "error": "sub_department only allowed for Production or Quality"}
#                         )
#                         continue
#                     sub_department = sub_dept_raw

#                 # Build data
#                 data = {
#                     "emp_id": emp_id,
#                     "first_name": clean(row.get("first_name")),
#                     "last_name": clean(row.get("last_name")) or "",
#                     "department": department.pk if department else None,
#                     "sub_department": sub_department,
#                     "designation": clean(row.get("designation")),
#                     "email": clean(row.get("email")),
#                     "phone": phone,
#                     "date_of_joining": date_of_joining,
#                     "birth_date": birth_date,
#                     "sex": (clean(row.get("sex")).upper() or None),
#                 }

#                 # Sex validation
#                 if data["sex"] and data["sex"] not in {"M", "F", "O"}:
#                     errors.append(
#                         {"row": idx + 2, "emp_id": emp_id, "error": f"Invalid sex '{data['sex']}'"}
#                     )
#                     continue

#                 # Designation validation
#                 if data["designation"] and data["designation"] not in allowed_designations:
#                     errors.append(
#                         {"row": idx + 2, "emp_id": emp_id, "error": f"Invalid designation '{data['designation']}'"}
#                     )
#                     continue

#                 # Uniqueness
#                 uniq_errs = []
#                 if data["email"] and MasterTable.objects.filter(email=data["email"]).exclude(emp_id=emp_id).exists():
#                     uniq_errs.append("email already used")
#                 if data["phone"] and MasterTable.objects.filter(phone=data["phone"]).exclude(emp_id=emp_id).exists():
#                     uniq_errs.append("phone already used")
#                 if uniq_errs:
#                     errors.append(
#                         {"row": idx + 2, "emp_id": emp_id, "error": "; ".join(uniq_errs)}
#                     )
#                     continue

#                 # CREATE or UPDATE
#                 with transaction.atomic():
#                     try:
#                         employee = MasterTable.objects.get(emp_id=emp_id)
#                         changed = []
#                         for field, new_val in data.items():
#                             old_val = getattr(employee, field)
#                             if filled(new_val) and (old_val != new_val):
#                                 setattr(employee, field, new_val)
#                                 changed.append(field)
#                         if not changed:
#                             errors.append(
#                                 {"row": idx + 2, "emp_id": emp_id, "error": "No new data to update"}
#                             )
#                             continue
#                         employee.save(update_fields=changed)
#                         updated.append({
#                             "emp_id": emp_id,
#                             "name": f"{employee.first_name or ''} {employee.last_name or ''}".strip(),
#                             "email": employee.email or "",
#                         })
#                     except MasterTable.DoesNotExist:
#                         serializer = MasterTableSerializer(data=data)
#                         if serializer.is_valid():
#                             employee = serializer.save()
#                             created.append({
#                                 "emp_id": emp_id,
#                                 "name": f"{employee.first_name or ''} {employee.last_name or ''}".strip(),
#                                 "email": employee.email or "",
#                             })
#                         else:
#                             msgs = [f"{f}: {'; '.join(e)}" for f, e in serializer.errors.items()]
#                             errors.append(
#                                 {"row": idx + 2, "emp_id": emp_id, "error": "; ".join(msgs)}
#                             )

#             except Exception as exc:
#                 errors.append({"row": idx + 2, "emp_id": emp_id, "error": str(exc)})

#         # Final response
#         resp = {
#             "message": f"Processed {len(df)} rows. {len(created)} created, {len(updated)} updated.",
#             "created_count": len(created),
#             "updated_count": len(updated),
#             "error_count": len(errors),
#             "created_employees": created,
#             "updated_employees": updated,
#         }
#         if errors:
#             resp["errors"] = errors
#         return Response(resp, status=status.HTTP_201_CREATED)

#     # --------------------------------------------------------------------- #
#     # 4. GET BY EMP ID
#     # --------------------------------------------------------------------- #
#     @action(detail=False, methods=["get"], url_path=r"by-employee-code/(?P<emp_id>[^/.]+)")
#     def retrieve_by_employee_code(self, request, emp_id=None):
#         try:
#             employee = MasterTable.objects.select_related("department").get(emp_id=emp_id)
#             serializer = self.get_serializer(employee)
#             return Response(serializer.data)
#         except MasterTable.DoesNotExist:
#             return Response({"error": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)



# # views.py
# import pandas as pd
# from django.http import HttpResponse
# from django.db import transaction
# from rest_framework import viewsets, status
# from rest_framework.decorators import action
# from rest_framework.response import Response
# from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
# from io import BytesIO
# from .models import MasterTable, Department, Line
# from .serializers import MasterTableSerializer


# class MasterTableViewSet(viewsets.ModelViewSet):
#     queryset = MasterTable.objects.select_related("department", "sub_department").all()
#     serializer_class = MasterTableSerializer
#     parser_classes = (MultiPartParser, FormParser, JSONParser)

#     # --------------------------------------------------------------------- #
#     # 1. LIST DEPARTMENTS (for frontend)
#     # --------------------------------------------------------------------- #
#     @action(detail=False, methods=["get"])
#     def list_departments(self, request):
#         depts = Department.objects.values("id", "department_name").order_by("department_name")
#         return Response(list(depts))

#     # --------------------------------------------------------------------- #
#     # 2. LIST SUB-DEPARTMENTS (NEW!)
#     # --------------------------------------------------------------------- #
#     @action(detail=False, methods=["get"])
#     def list_sub_departments(self, request):
#         """Return all sub-departments (Lines) grouped by department"""
#         lines = Line.objects.select_related('department').all()
#         result = {}
#         for line in lines:
#             dept_name = line.department.department_name if line.department else "Other"
#             if dept_name not in result:
#                 result[dept_name] = []
#             result[dept_name].append({
#                 "id": line.line_id,
#                 "name": line.line_name
#             })
#         return Response(result)
    
#     @action(detail=False, methods=["get"])
#     def list_stations(self, request):
#         """Return all stations grouped by department and line"""
#         stations = Station.objects.select_related('department', 'line', 'subline').all()
#         result = {}
        
#         for station in stations:
#             # Determine grouping key
#             if station.line:
#                 dept_name = station.line.department.department_name if station.line.department else "Other"
#                 line_name = station.line.line_name
#                 key = f"{dept_name} > {line_name}"
#             elif station.department:
#                 dept_name = station.department.department_name
#                 key = dept_name
#             else:
#                 key = "Other"
            
#             if key not in result:
#                 result[key] = []
            
#             result[key].append({
#                 "id": station.station_id,
#                 "name": station.station_name
#             })
        
#         return Response(result)

#     # --------------------------------------------------------------------- #
#     # 3. DOWNLOAD TEMPLATE – DYNAMIC FROM DB
#     # --------------------------------------------------------------------- #
#     @action(detail=False, methods=["get"])
#     def download_template(self, request):
#         try:
#             headers = [
#                 "emp_id", "first_name", "last_name", "department_name",
#                 "sub_department", "designation", "date_of_joining",
#                 "birth_date", "sex", "email", "phone",
#             ]
#             df = pd.DataFrame(columns=headers)

#             # Fetch real data
#             departments = list(Department.objects.values_list('department_name', flat=True))
#             lines = Line.objects.select_related('department').all()

#             # Group sub-depts by department
#             sub_dept_by_dept = {}
#             for line in lines:
#                 dept_name = line.department.department_name if line.department else "Other"
#                 sub_dept_by_dept.setdefault(dept_name, []).append(line.line_name)

#             # Example row
#             example_dept = departments[0] if departments else "Production"
#             example_sub = sub_dept_by_dept.get(example_dept, [""])[0] if sub_dept_by_dept.get(example_dept) else ""

#             example = {
#                 "emp_id": "EMP001",
#                 "first_name": "John",
#                 "last_name": "Doe",
#                 "department_name": example_dept,
#                 "sub_department": example_sub,
#                 "designation": "DRIVER",
#                 "date_of_joining": "2024-01-15",
#                 "birth_date": "1990-05-20",
#                 "sex": "M",
#                 "email": "john.doe@company.com",
#                 "phone": "+1234567890",
#             }
#             df.loc[0] = example
#             df = df.iloc[0:0]  # Clear example row

#             output = BytesIO()
#             with pd.ExcelWriter(output, engine="openpyxl") as writer:
#                 df.to_excel(writer, sheet_name="Employee Template", index=False)

#                 # ------------------- INSTRUCTIONS -------------------
#                 instructions = [
#                     "1. Fill ONLY the columns you want to set.",
#                     "2. emp_id is **mandatory** and must be unique.",
#                     "3. All other columns are optional.",
#                     "4. department_name must match an existing department (case-sensitive).",
#                     "5. sub_department must be a valid name from the list below.",
#                     "6. designation must be one of the allowed values.",
#                     "7. Dates → YYYY-MM-DD.",
#                     "8. sex → M / F / O (or leave blank).",
#                     "9. email / phone → unique if provided.",
#                     "",
#                     "AVAILABLE DEPARTMENTS (from database):",
#                 ] + [f"- {d}" for d in departments]

#                 # Dynamic sub-departments per department
#                 instructions += ["", "SUB-DEPARTMENTS BY DEPARTMENT:"]
#                 for dept_name, subs in sub_dept_by_dept.items():
#                     instructions += [f"", f"{dept_name}:"]
#                     instructions += [f"  - {s}" for s in subs]

#                 # Designations
#                 instructions += ["", "ALLOWED DESIGNATIONS:"]
#                 instructions += [f"- {label}" for _, label in MasterTable.DESIGNATION_CHOICES[1:]]

#                 pd.DataFrame({"Instructions": instructions}).to_excel(
#                     writer, sheet_name="Instructions", index=False
#                 )

#             output.seek(0)
#             response = HttpResponse(
#                 output.getvalue(),
#                 content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#             )
#             response["Content-Disposition"] = 'attachment; filename="employee_template.xlsx"'
#             return response

#         except Exception as e:
#             return Response(
#                 {"error": f"Failed to generate template: {str(e)}"},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR,
#             )
        
   
#     # --------------------------------------------------------------------- #
#     # 4. UPLOAD EXCEL – FULL VALIDATION WITH Line FK
#     # --------------------------------------------------------------------- #
#     @action(detail=False, methods=["post"])
#     def upload_excel(self, request):
#         if "file" not in request.FILES:
#             return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

#         file = request.FILES["file"]
#         if not file.name.lower().endswith((".xlsx", ".xls")):
#             return Response(
#                 {"error": "Only .xlsx or .xls files are allowed"},
#                 status=status.HTTP_400_BAD_REQUEST,
#             )

#         try:
#             df = pd.read_excel(file, sheet_name="Employee Template")
#         except Exception as e:
#             return Response(
#                 {"error": f"Could not read Excel file: {e}"},
#                 status=status.HTTP_400_BAD_REQUEST,
#             )

#         if "emp_id" not in df.columns:
#             return Response(
#                 {"error": "Column 'emp_id' is required"},
#                 status=status.HTTP_400_BAD_REQUEST,
#             )
#         df = df.dropna(subset=["emp_id"])
#         if df.empty:
#             return Response({"error": "No rows with emp_id found"}, status=status.HTTP_400_BAD_REQUEST)

#         def clean(val):
#             if pd.isna(val):
#                 return ""
#             return str(val).strip()

#         def filled(v):
#             return bool(v) if not isinstance(v, str) else bool(v.strip())

#         # Load valid choices from DB
#         allowed_designations = {c[0] for c in MasterTable.DESIGNATION_CHOICES}
#         valid_sub_dept_names = set(Line.objects.values_list('line_name', flat=True))
#         department_map = {d.department_name: d for d in Department.objects.all()}
#         line_map = {l.line_name: l for l in Line.objects.select_related('department').all()}

#         created, updated, errors = [], [], []

#         for idx, row in df.iterrows():
#             try:
#                 emp_id = clean(row["emp_id"])
#                 if not emp_id:
#                     errors.append({"row": idx + 2, "emp_id": "N/A", "error": "emp_id is empty"})
#                     continue

#                 # Department
#                 department = None
#                 dept_name = clean(row.get("department_name"))
#                 if dept_name:
#                     if dept_name not in department_map:
#                         errors.append(
#                             {"row": idx + 2, "emp_id": emp_id, "error": f"Department '{dept_name}' not found"}
#                         )
#                         continue
#                     department = department_map[dept_name]

#                 # Sub-department (by name → Line object)
#                 sub_dept_raw = clean(row.get("sub_department"))
#                 sub_department = None
#                 if sub_dept_raw:
#                     if sub_dept_raw not in line_map:
#                         errors.append(
#                             {"row": idx + 2, "emp_id": emp_id, "error": f"Invalid sub_department '{sub_dept_raw}'"}
#                         )
#                         continue
#                     line = line_map[sub_dept_raw]
#                     # Optional: enforce department match
#                     if department and line.department and line.department != department:
#                         errors.append(
#                             {"row": idx + 2, "emp_id": emp_id,
#                              "error": f"sub_department '{sub_dept_raw}' belongs to {line.department.department_name}, not {dept_name}"}
#                         )
#                         continue
#                     sub_department = line

#                 # Dates
#                 def parse_date(val):
#                     if pd.isna(val):
#                         return None
#                     try:
#                         return pd.to_datetime(val, errors="coerce").date()
#                     except:
#                         return None

#                 date_of_joining = parse_date(row.get("date_of_joining"))
#                 birth_date = parse_date(row.get("birth_date"))

#                 # Phone
#                 phone_raw = clean(row.get("phone"))
#                 phone = None if not phone_raw or phone_raw == "0" else phone_raw

#                 # Build data
#                 data = {
#                     "emp_id": emp_id,
#                     "first_name": clean(row.get("first_name")),
#                     "last_name": clean(row.get("last_name")) or "",
#                     "department": department.pk if department else None,
#                     "sub_department": sub_department.pk if sub_department else None,
#                     "designation": clean(row.get("designation")),
#                     "email": clean(row.get("email")),
#                     "phone": phone,
#                     "date_of_joining": date_of_joining,
#                     "birth_date": birth_date,
#                     "sex": (clean(row.get("sex")).upper() or None),
#                 }

#                 # Validation
#                 if data["sex"] and data["sex"] not in {"M", "F", "O"}:
#                     errors.append(
#                         {"row": idx + 2, "emp_id": emp_id, "error": f"Invalid sex '{data['sex']}'"}
#                     )
#                     continue

#                 if data["designation"] and data["designation"] not in allowed_designations:
#                     errors.append(
#                         {"row": idx + 2, "emp_id": emp_id, "error": f"Invalid designation '{data['designation']}'"}
#                     )
#                     continue

#                 # Uniqueness
#                 uniq_errs = []
#                 if data["email"] and MasterTable.objects.filter(email=data["email"]).exclude(emp_id=emp_id).exists():
#                     uniq_errs.append("email already used")
#                 if data["phone"] and MasterTable.objects.filter(phone=data["phone"]).exclude(emp_id=emp_id).exists():
#                     uniq_errs.append("phone already used")
#                 if uniq_errs:
#                     errors.append(
#                         {"row": idx + 2, "emp_id": emp_id, "error": "; ".join(uniq_errs)}
#                     )
#                     continue

#                 # CREATE or UPDATE
#                 with transaction.atomic():
#                     try:
#                         employee = MasterTable.objects.get(emp_id=emp_id)
#                         changed = []
#                         for field, new_val in data.items():
#                             old_val = getattr(employee, field)
#                             if filled(new_val) and (old_val != new_val):
#                                 setattr(employee, field, new_val)
#                                 changed.append(field)
#                         if not changed:
#                             errors.append(
#                                 {"row": idx + 2, "emp_id": emp_id, "error": "No new data to update"}
#                             )
#                             continue
#                         employee.save(update_fields=changed)
#                         updated.append({
#                             "emp_id": emp_id,
#                             "name": f"{employee.first_name or ''} {employee.last_name or ''}".strip(),
#                             "email": employee.email or "",
#                         })
#                     except MasterTable.DoesNotExist:
#                         serializer = MasterTableSerializer(data=data)
#                         if serializer.is_valid():
#                             employee = serializer.save()
#                             created.append({
#                                 "emp_id": emp_id,
#                                 "name": f"{employee.first_name or ''} {employee.last_name or ''}".strip(),
#                                 "email": employee.email or "",
#                             })
#                         else:
#                             msgs = [f"{f}: {'; '.join(e)}" for f, e in serializer.errors.items()]
#                             errors.append(
#                                 {"row": idx + 2, "emp_id": emp_id, "error": "; ".join(msgs)}
#                             )

#             except Exception as exc:
#                 errors.append({"row": idx + 2, "emp_id": emp_id, "error": str(exc)})

#         resp = {
#             "message": f"Processed {len(df)} rows. {len(created)} created, {len(updated)} updated.",
#             "created_count": len(created),
#             "updated_count": len(updated),
#             "error_count": len(errors),
#             "created_employees": created,
#             "updated_employees": updated,
#         }
#         if errors:
#             resp["errors"] = errors
#         return Response(resp, status=status.HTTP_201_CREATED)

#     # --------------------------------------------------------------------- #
#     # 5. GET BY EMP ID
#     # --------------------------------------------------------------------- #
#     @action(detail=False, methods=["get"], url_path=r"by-employee-code/(?P<emp_id>[^/.]+)")
#     def retrieve_by_employee_code(self, request, emp_id=None):
#         try:
#             employee = MasterTable.objects.select_related("department", "sub_department").get(emp_id=emp_id)
#             serializer = self.get_serializer(employee)
#             return Response(serializer.data)
#         except MasterTable.DoesNotExist:
#             return Response({"error": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)



# views.py
import pandas as pd
from django.http import HttpResponse
from django.db import transaction
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from io import BytesIO
# 1. ADDED Station to imports
from .models import MasterTable, Department, Line, Station
from .serializers import MasterTableSerializer


class MasterTableViewSet(viewsets.ModelViewSet):
    queryset = MasterTable.objects.select_related("department", "sub_department", "station").all()
    serializer_class = MasterTableSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)
    
    

    # ... (list_departments, list_sub_departments, list_stations remain the same) ...
    @action(detail=False, methods=["get"])
    def list_departments(self, request):
        depts = Department.objects.values("id", "department_name").order_by("department_name")
        return Response(list(depts))

    @action(detail=False, methods=["get"])
    def list_sub_departments(self, request):
        lines = Line.objects.select_related('department').all()
        result = {}
        for line in lines:
            dept_name = line.department.department_name if line.department else "Other"
            if dept_name not in result:
                result[dept_name] = []
            result[dept_name].append({
                "id": line.line_id,
                "name": line.line_name
            })
        return Response(result)
    
    @action(detail=False, methods=["get"])
    def list_stations(self, request):
        stations = Station.objects.select_related('department', 'line', 'subline').all()
        result = {}
        for station in stations:
            if station.line:
                dept_name = station.line.department.department_name if station.line.department else "Other"
                line_name = station.line.line_name
                key = f"{dept_name} > {line_name}"
            elif station.department:
                dept_name = station.department.department_name
                key = dept_name
            else:
                key = "Other"
            
            if key not in result:
                result[key] = []
            result[key].append({
                "id": station.station_id,
                "name": station.station_name
            })
        return Response(result)

    # --------------------------------------------------------------------- #
    # 3. DOWNLOAD TEMPLATE
    # --------------------------------------------------------------------- #
    @action(detail=False, methods=["get"])
    def download_template(self, request):
        try:
            # 2. ADDED "station" to headers
            headers = [
                "emp_id", "first_name", "last_name", "department_name",
                "sub_department", "station", "designation", "date_of_joining",
                "birth_date", "sex", "email", "phone",
            ]
            df = pd.DataFrame(columns=headers)

            # Fetch real data for instructions
            departments = list(Department.objects.values_list('department_name', flat=True))
            lines = Line.objects.select_related('department').all()

            sub_dept_by_dept = {}
            for line in lines:
                dept_name = line.department.department_name if line.department else "Other"
                sub_dept_by_dept.setdefault(dept_name, []).append(line.line_name)

            # Example row
            example_dept = departments[0] if departments else "Production"
            example_sub = sub_dept_by_dept.get(example_dept, [""])[0] if sub_dept_by_dept.get(example_dept) else ""

            example = {
                "emp_id": "EMP001",
                "first_name": "John",
                "last_name": "Doe",
                "department_name": example_dept,
                "sub_department": example_sub,
                "station": "Station-1", # Example station
                "designation": "DRIVER",
                "date_of_joining": "2024-01-15",
                "birth_date": "1990-05-20",
                "sex": "M",
                "email": "john.doe@company.com",
                "phone": "+1234567890",
            }
            df.loc[0] = example
            df = df.iloc[0:0] 

            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Employee Template", index=False)

                instructions = [
                    "1. Fill ONLY the columns you want to set.",
                    "2. emp_id is **mandatory**.",
                    "3. department_name must match DB.",
                    "4. sub_department (Line) must match DB.",
                    "5. station must match DB Station Name.",
                    "",
                    "AVAILABLE DEPARTMENTS:",
                ] + [f"- {d}" for d in departments]

                pd.DataFrame({"Instructions": instructions}).to_excel(
                    writer, sheet_name="Instructions", index=False
                )

            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = 'attachment; filename="employee_template.xlsx"'
            return response

        except Exception as e:
            return Response(
                {"error": f"Failed to generate template: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        
    # --------------------------------------------------------------------- #
    # 4. UPLOAD EXCEL
    # --------------------------------------------------------------------- #
    @action(detail=False, methods=["post"])
    def upload_excel(self, request):
        if "file" not in request.FILES:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES["file"]
        if not file.name.lower().endswith((".xlsx", ".xls")):
            return Response({"error": "Only .xlsx or .xls files are allowed"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(file, sheet_name="Employee Template")
        except Exception as e:
            return Response({"error": f"Could not read Excel file: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        if "emp_id" not in df.columns:
            return Response({"error": "Column 'emp_id' is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        df = df.dropna(subset=["emp_id"])
        if df.empty:
            return Response({"error": "No rows with emp_id found"}, status=status.HTTP_400_BAD_REQUEST)

        def clean(val):
            if pd.isna(val): return ""
            return str(val).strip()

        def filled(v):
            return bool(v) if not isinstance(v, str) else bool(v.strip())

        allowed_designations = {c[0] for c in MasterTable.DESIGNATION_CHOICES}
        
        # Maps for lookups
        department_map = {d.department_name: d for d in Department.objects.all()}
        line_map = {l.line_name: l for l in Line.objects.select_related('department').all()}
        
        # 3. ADDED Station Map
        station_map = {s.station_name: s for s in Station.objects.all()}

        created, updated, errors = [], [], []

        for idx, row in df.iterrows():
            try:
                emp_id = clean(row["emp_id"])
                if not emp_id:
                    errors.append({"row": idx + 2, "emp_id": "N/A", "error": "emp_id is empty"})
                    continue

                # Department
                department = None
                dept_name = clean(row.get("department_name"))
                if dept_name:
                    if dept_name not in department_map:
                        errors.append({"row": idx + 2, "emp_id": emp_id, "error": f"Department '{dept_name}' not found"})
                        continue
                    department = department_map[dept_name]

                # Sub-department
                sub_dept_raw = clean(row.get("sub_department"))
                sub_department = None
                if sub_dept_raw:
                    if sub_dept_raw not in line_map:
                        errors.append({"row": idx + 2, "emp_id": emp_id, "error": f"Invalid sub_department '{sub_dept_raw}'"})
                        continue
                    line = line_map[sub_dept_raw]
                    sub_department = line

                # 4. ADDED Station Logic
                station_raw = clean(row.get("station"))
                station_obj = None
                if station_raw:
                    if station_raw not in station_map:
                        errors.append({"row": idx + 2, "emp_id": emp_id, "error": f"Invalid station '{station_raw}'"})
                        continue
                    station_obj = station_map[station_raw]

                # Dates
                def parse_date(val):
                    if pd.isna(val): return None
                    try: return pd.to_datetime(val, errors="coerce").date()
                    except: return None

                date_of_joining = parse_date(row.get("date_of_joining"))
                birth_date = parse_date(row.get("birth_date"))
                phone = clean(row.get("phone"))
                if phone == "0": phone = None

                # Build data
                data = {
                    "emp_id": emp_id,
                    "first_name": clean(row.get("first_name")),
                    "last_name": clean(row.get("last_name")) or "",
                    "department": department.pk if department else None,
                    "sub_department": sub_department.pk if sub_department else None,
                    "station": station_obj.pk if station_obj else None, # <--- 5. ADDED Station to data
                    "designation": clean(row.get("designation")),
                    "email": clean(row.get("email")),
                    "phone": phone,
                    "date_of_joining": date_of_joining,
                    "birth_date": birth_date,
                    "sex": (clean(row.get("sex")).upper() or None),
                }

                # Validation
                if data["sex"] and data["sex"] not in {"M", "F", "O"}:
                    errors.append({"row": idx + 2, "emp_id": emp_id, "error": f"Invalid sex '{data['sex']}'"})
                    continue

                if data["designation"] and data["designation"] not in allowed_designations:
                    errors.append({"row": idx + 2, "emp_id": emp_id, "error": f"Invalid designation '{data['designation']}'"})
                    continue

                # Uniqueness checks
                uniq_errs = []
                if data["email"] and MasterTable.objects.filter(email=data["email"]).exclude(emp_id=emp_id).exists():
                    uniq_errs.append("email already used")
                if data["phone"] and MasterTable.objects.filter(phone=data["phone"]).exclude(emp_id=emp_id).exists():
                    uniq_errs.append("phone already used")
                if uniq_errs:
                    errors.append({"row": idx + 2, "emp_id": emp_id, "error": "; ".join(uniq_errs)})
                    continue

                # Save
                with transaction.atomic():
                    try:
                        employee = MasterTable.objects.get(emp_id=emp_id)
                        changed = []
                        for field, new_val in data.items():
                            old_val = getattr(employee, field)
                            # Handle FK comparisons
                            if field in ['department', 'sub_department', 'station'] and old_val:
                                old_val = old_val.pk
                            
                            if filled(new_val) and (old_val != new_val):
                                setattr(employee, field, new_val)
                                changed.append(field)
                                
                        if not changed:
                            errors.append({"row": idx + 2, "emp_id": emp_id, "error": "No new data to update"})
                            continue
                        employee.save(update_fields=changed)
                        updated.append({"emp_id": emp_id})
                    except MasterTable.DoesNotExist:
                        serializer = MasterTableSerializer(data=data)
                        if serializer.is_valid():
                            serializer.save()
                            created.append({"emp_id": emp_id})
                        else:
                            msgs = [f"{f}: {'; '.join(e)}" for f, e in serializer.errors.items()]
                            errors.append({"row": idx + 2, "emp_id": emp_id, "error": "; ".join(msgs)})

            except Exception as exc:
                errors.append({"row": idx + 2, "emp_id": emp_id, "error": str(exc)})

        resp = {
            "message": f"Processed {len(df)} rows. {len(created)} created, {len(updated)} updated.",
            "created_count": len(created),
            "updated_count": len(updated),
            "errors": errors,
        }
        return Response(resp, status=status.HTTP_201_CREATED if not errors else status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path=r"by-employee-code/(?P<emp_id>[^/.]+)")
    def retrieve_by_employee_code(self, request, emp_id=None):
        try:
            employee = MasterTable.objects.select_related("department", "sub_department", "station").get(emp_id=emp_id)
            serializer = self.get_serializer(employee)
            return Response(serializer.data)
        except MasterTable.DoesNotExist:
            return Response({"error": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)

    def get_queryset(self):
        queryset = MasterTable.objects.all()
        
        # ✅ Match EXACTLY what frontend sends: ?emp_id=kl09
        emp_id = self.request.query_params.get('emp_id', None)
        
        if emp_id is not None:
            queryset = queryset.filter(emp_id=emp_id)
        
        return queryset

        






# Level 0

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.views import APIView
from .models import UserRegistration,HumanBodyQuestions,HumanBodyCheckSession,HumanBodyCheckSheet
from .serializers import UserRegistrationSerializer,UserWithBodyCheckSerializer,HumanBodyCheckSessionSerializer,HumanBodyQuestionsSerializer,UserUpdateSerializer

from rest_framework import filters
class UserRegistrationViewSet(viewsets.ModelViewSet):
    queryset = UserRegistration.objects.all()
    serializer_class = UserRegistrationSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['first_name', 'last_name', 'email', 'phone_number']
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    lookup_field = 'temp_id'
    
    def get_serializer_class(self):
        if self.request.method in ['PATCH', 'PUT']:
            return UserUpdateSerializer
        return UserRegistrationSerializer

class HumanBodyQuestionsViewSet(viewsets.ModelViewSet):
    queryset = HumanBodyQuestions.objects.all()
    serializer_class = HumanBodyQuestionsSerializer

    @action(detail=False, methods=['post'])
    def add_question(self, request):
        
        question_text = request.data.get('question_text', '').strip()
        if not question_text:
            return Response({"error": "question_text is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        question, created = HumanBodyQuestions.objects.get_or_create(
            question_text=question_text,
        )
        
        if created:
            serializer = self.get_serializer(question)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response({"error": "Question already exists"}, status=status.HTTP_400_BAD_REQUEST)


class BodyCheckSubmissionView(APIView):
    
    
    def get(self, request):
        
        temp_id = request.query_params.get('temp_id')
        if temp_id:
            
            try:
                user = UserRegistration.objects.get(temp_id=temp_id)
                
                session = HumanBodyCheckSession.objects.filter(user=user).order_by('-created_at').first()
                if not session:
                    
                    session = HumanBodyCheckSession.objects.filter(temp_id=temp_id).order_by('-created_at').first()
            except UserRegistration.DoesNotExist:
                
                session = HumanBodyCheckSession.objects.filter(temp_id=temp_id).order_by('-created_at').first()
            
            if session:
                serializer = HumanBodyCheckSessionSerializer(session)
                return Response([serializer.data])
            return Response([])
        
        
        sessions = HumanBodyCheckSession.objects.all()
        serializer = HumanBodyCheckSessionSerializer(sessions, many=True)
        return Response(serializer.data)
    
    def post(self, request):
        
        temp_id = request.data.get('temp_id')
        check_data = request.data.get('checkData', {})
        
        if not temp_id:
            return Response({"error": "temp_id is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = UserRegistration.objects.get(temp_id=temp_id)
            
            
            existing_session = HumanBodyCheckSession.objects.filter(user=user).first()
            if existing_session:
                return Response(
                    {"error": "Body check already exists for this user"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            
            existing_temp_session = HumanBodyCheckSession.objects.filter(temp_id=temp_id).first()
            if existing_temp_session:
                return Response(
                    {"error": "Body check already exists for this user"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            
            session = HumanBodyCheckSession.objects.create(temp_id=temp_id, user=user)
            
           
            for item_id, item_data in check_data.items():
                description = item_data.get('description', '')
                answer = item_data.get('status', 'pending')
                
                if description and answer != '':
                   
                    question, created = HumanBodyQuestions.objects.get_or_create(
                        question_text=description
                    )
                    
                   
                    HumanBodyCheckSheet.objects.create(
                        session=session,
                        question=question,
                        answer=answer
                    )
            
            serializer = HumanBodyCheckSessionSerializer(session)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
            
        except UserRegistration.DoesNotExist:
            return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)


class UserBodyCheckListView(APIView):
    
    def get(self, request):
        users = UserRegistration.objects.all()
        serializer = UserWithBodyCheckSerializer(users, many=True)
        return Response(serializer.data)
    
    def patch(self, request):
        """Mark user as added to master table"""
        temp_id = request.data.get('temp_id')
        
        if not temp_id:
            return Response({"error": "temp_id is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = UserRegistration.objects.get(temp_id=temp_id)
            
            # Check if user has passed status
            latest_session = HumanBodyCheckSession.objects.filter(user=user).order_by('-created_at').first()
            if not latest_session:
                latest_session = HumanBodyCheckSession.objects.filter(temp_id=temp_id).order_by('-created_at').first()
            
            if not latest_session or latest_session.overall_status != 'pass':
                return Response(
                    {"error": "User must have passed status to be added to master table"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            from django.utils import timezone
            if not user.is_added_to_master:
                user.is_added_to_master = True
                user.added_to_master_at = timezone.now()
            
            # Save or update emp_id if provided
            emp_id_val = request.data.get('emp_id')
            if emp_id_val:
                user.emp_id = emp_id_val
                
            user.save()
            
            # Return updated user list
            users = UserRegistration.objects.all()
            serializer = UserWithBodyCheckSerializer(users, many=True)
            return Response(serializer.data)
            
        except UserRegistration.DoesNotExist:
            return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)




# views.py
from rest_framework import viewsets
from .models import (
    Level, Days, SubTopic, SubTopicContent, TrainingContent, Evaluation
)
from .serializers import ( LevelSerializer, DaysSerializer,DaysWriteSerializer, 
     SubTopicListSerializer, SubTopicAdminSerializer, SubTopicSerializer, SubTopicContentSerializer,
    TrainingContentSerializer, EvaluationSerializer
)

def _int_or_none(v):
    try:
        return int(str(v).strip("/"))
    except (TypeError, ValueError):
        return None


class LevelViewSet(viewsets.ModelViewSet):
    queryset = Level.objects.all()
    serializer_class = LevelSerializer


class DaysViewSet(viewsets.ModelViewSet):
    queryset = Days.objects.all().select_related("level")
    # serializer_class = DaysSerializer

    def get_serializer_class(self):
        if self.action in ["create", "update", "partial_update"]:
            return DaysWriteSerializer
        return DaysSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        level = _int_or_none(self.request.query_params.get("level"))
        if level is not None:
            qs = qs.filter(level__level_id=level)
        return qs


from rest_framework import viewsets
from .models import Level, Days, SubTopic, SubTopicContent, TrainingContent, Evaluation
from .serializers import (
    LevelSerializer, DaysSerializer,
    SubTopicSerializer, SubTopicListSerializer, SubTopicAdminSerializer,
    SubTopicContentSerializer, TrainingContentSerializer, EvaluationSerializer
)

class SubTopicViewSet(viewsets.ModelViewSet):
    queryset = SubTopic.objects.all().select_related("days", "level")
      
    def get_serializer_class(self):
        if self.action in ["list", "retrieve"]:
            return SubTopicListSerializer  # read shape for frontend
        return SubTopicAdminSerializer  # subtopic_name, days, level
     
    def get_queryset(self):
        qs = super().get_queryset()
        level = self.request.query_params.get("level")   # /subtopics/?level=1
        days_id = self.request.query_params.get("days")  # /subtopics/?days=3
        
        if level is not None:
            qs = qs.filter(level__level_id=level)
        if days_id is not None:
            qs = qs.filter(days__days_id=days_id)
        return qs
    queryset = SubTopic.objects.all()
    serializer_class = SubTopicSerializer


class SubTopicContentViewSet(viewsets.ModelViewSet):
    queryset = SubTopicContent.objects.all()
    serializer_class = SubTopicContentSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        subtopic = self.request.query_params.get("subtopic")  # /subtopic-contents/?subtopic=12
        if subtopic is not None:
            qs = qs.filter(subtopic__subtopic_id=subtopic)
        return qs
    
    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx["request"] = self.request
        return ctx
    

from rest_framework import viewsets, parsers


class TrainingContentViewSet(viewsets.ModelViewSet):
    queryset = TrainingContent.objects.all()
    serializer_class = TrainingContentSerializer

    parser_classes = [parsers.MultiPartParser, parsers.FormParser, parsers.JSONParser]

    # def get_queryset(self):
    #     qs = super().get_queryset()
    #     stc = self.request.query_params.get("subtopiccontent")  # /training-contents/?subtopiccontent=55
    #     if stc:
    #         qs = qs.filter(subtopiccontent__subtopiccontent_id=stc)
    #     return qs

    def get_queryset(self):
        qs = super().get_queryset()
        # stc = self.request.query_params.get("subtopiccontent") or self.request.query_params.get("subtopic_content")
        stc = _int_or_none(self.request.query_params.get("subtopiccontent") or self.request.query_params.get("subtopic_content"))
        # if stc is not None:
        #     try:
        #         stc = int(str(stc).strip("/"))
        #         qs = qs.filter(subtopiccontent__subtopiccontent_id=stc)
        #     except ValueError:
        #         pass
        # return qs
        if stc is not None:
            qs = qs.filter(subtopiccontent__subtopiccontent_id=stc)
        return qs


    def get_serializer_context(self):
        # so training_file returns absolute URL
        ctx = super().get_serializer_context()
        ctx["request"] = self.request
        return ctx


class EvaluationViewSet(viewsets.ModelViewSet):
    queryset = Evaluation.objects.all()
    serializer_class = EvaluationSerializer

from rest_framework import viewsets, status
from rest_framework.response import Response
from .models import ProductionPlan
from .serializers import ProductionPlanSerializer

class ProductionPlanViewSet(viewsets.ModelViewSet):
    queryset = ProductionPlan.objects.all()
    serializer_class = ProductionPlanSerializer

    def create(self, request, *args, **kwargs):
        # Handle bulk creation
        if isinstance(request.data, list):
            serializer = self.get_serializer(data=request.data, many=True)
            serializer.is_valid(raise_exception=True)
            self.perform_bulk_create(serializer)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return super().create(request, *args, **kwargs)

    def perform_bulk_create(self, serializer):
        serializer.save()

    def get_queryset(self):
        queryset = super().get_queryset()
        params = self.request.query_params

        # Filtering based on new structure
        hq = params.get('hq')
        factory = params.get('factory')
        department = params.get('department')
        line = params.get('line')
        subline = params.get('subline')
        station = params.get('station')
        year = params.get('year')
        month = params.get('month')

        if hq:
            queryset = queryset.filter(hq_id=hq)
        if factory:
            queryset = queryset.filter(factory_id=factory)
        if department:
            queryset = queryset.filter(department_id=department)
        if line:
            queryset = queryset.filter(line_id=line)
        if subline:
            queryset = queryset.filter(subline_id=subline)
        if station:
            queryset = queryset.filter(station_id=station)
        if year:
            queryset = queryset.filter(year=year)
        if month:
            queryset = queryset.filter(month=month)

        return queryset



from rest_framework import viewsets
from .models import QuestionPaper
from .serializers import QuestionPaperSerializer
from .filters import QuestionPaperFilter


class QuestionPaperViewSet(viewsets.ModelViewSet):
    # queryset = QuestionPaper.objects.all().order_by("-created_at") # We'll improve this
    serializer_class = QuestionPaperSerializer
    
    # --- ADD THESE TWO LINES ---
    filter_backends = [DjangoFilterBackend]
    filterset_class = QuestionPaperFilter
    # ---------------------------

    def get_queryset(self):
        """
        Optimizing the queryset by pre-fetching related objects.
        This prevents N+1 query problems and makes your API much faster.
        """
        return QuestionPaper.objects.select_related(
            'department', 'line', 'subline', 'station', 'level'
        ).order_by("-created_at")

    @action(detail=True, methods=["get"])
    def questions(self, request, pk=None):
        paper = self.get_object()
        paper_serializer = self.get_serializer(paper)
    
        questions = TemplateQuestion.objects.filter(question_paper_id=pk,is_active=True).order_by('order','id')
        question_serializer = TemplateQuestionSerializer(
            questions, many=True, context={'request': request}
        )
    
        return Response({
            "question_paper": paper_serializer.data,
            "questions": question_serializer.data
        })
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            # If validation fails, print the errors to the console
            print("--- VALIDATION ERRORS ---")
            print(serializer.errors)
            print("--- REQUEST DATA ---")
            print(request.data)
            print("-------------------------")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        # If validation succeeds, continue with the normal creation process
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
    

    @action(detail=True, methods=['get'])
    def applicable_stations(self, request, pk=None):
        """
        For a given question paper, returns a list of all stations it applies to.
        """
        try:
            paper = self.get_object()
        except QuestionPaper.DoesNotExist:
            return Response({"error": "Question paper not found"}, status=status.HTTP_404_NOT_FOUND)

        # Case 1: Paper is assigned to a specific Station
        if paper.station:
            queryset = Station.objects.filter(pk=paper.station.pk)

        # Case 2: Paper is assigned to a SubLine (applies to all stations under it)
        elif paper.subline:
            queryset = Station.objects.filter(subline=paper.subline)

        # Case 3: Paper is assigned to a Line (applies to all stations under it)
        elif paper.line:
            # We need to find all sublines for this line, then all stations for those sublines
            queryset = Station.objects.filter(subline__line=paper.line)

        # Case 4: Paper is assigned to a Department (applies to all stations under it)
        elif paper.department:
            # This is your "bruhh" scenario. Find all stations under the department.
            queryset = Station.objects.filter(subline__line__department=paper.department)

        # Case 5: Paper has no assignment (should not happen with good validation)
        else:
            queryset = Station.objects.none()

        # Serialize the results and return them
        # Assuming you have a StationSerializer
        serializer = StationSerializer(queryset.order_by('station_name'), many=True)
        return Response(serializer.data)
    


from .filters import QuestionPaperFilter

from rest_framework import viewsets
from .models import StationLevelQuestionPaper
from .serializers import StationLevelQuestionPaperSerializer


class StationLevelQuestionPaperViewSet(viewsets.ModelViewSet):
    queryset = StationLevelQuestionPaper.objects.all()
    serializer_class = StationLevelQuestionPaperSerializer





import io
import os
import zipfile
from lxml import etree

import pandas as pd
from django.http import HttpResponse
from django.db import transaction
from django.core.files.base import ContentFile

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import TemplateQuestion, QuestionPaper
from .serializers import TemplateQuestionSerializer


def extract_images_from_xlsx(file_obj):
    """
    Extract images from XLSX regardless of Excel version / paste method.
    Returns {(row, col): ContentFile}
    """
    images = {}

    with zipfile.ZipFile(file_obj) as z:
        drawing_files = [
            f for f in z.namelist()
            if f.startswith("xl/drawings/drawing")
        ]
        if not drawing_files:
            return images

        ns = {
            "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }

        for drawing in drawing_files:
            root = etree.XML(z.read(drawing))

            rels_path = drawing.replace(
                "xl/drawings/", "xl/drawings/_rels/"
            ) + ".rels"
            rels_root = etree.XML(z.read(rels_path))

            rels_map = {
                rel.attrib["Id"]: rel.attrib["Target"]
                for rel in rels_root
            }

            for anchor in root.findall(".//xdr:twoCellAnchor", ns):
                from_cell = anchor.find("xdr:from", ns)
                row = int(from_cell.find("xdr:row", ns).text) + 1
                col = int(from_cell.find("xdr:col", ns).text) + 1

                blip = anchor.find(".//a:blip", ns)
                if blip is None:
                    continue

                embed = blip.attrib.get(
                    "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"
                )
                if not embed or embed not in rels_map:
                    continue

                img_path = os.path.normpath(
                    "xl/drawings/" + rels_map[embed]
                ).replace("\\", "/")

                if img_path not in z.namelist():
                    continue

                images[(row, col)] = ContentFile(
                    z.read(img_path),
                    name=f"r{row}_c{col}.png"
                )

    return images


class TemplateQuestionViewSet(viewsets.ModelViewSet):
    queryset = TemplateQuestion.objects.all()
    serializer_class = TemplateQuestionSerializer
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def get_queryset(self):
        qs = super().get_queryset()
        qp = self.request.query_params.get("question_paper")
        if qp:
            qs = qs.filter(question_paper_id=qp)
        return qs.order_by("order", "id")



    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Advanced Template"

        C_QUES = "E0F2FE"
        C_OPTA = "DCFCE7"
        C_OPTB = "FEF3C7"
        C_OPTC = "F3E8FF"
        C_OPTD = "FCE7F3"
        C_ANS  = "FFE4E6"

        columns_config = [
            ("question", 40, C_QUES),
            ("question_lang2", 30, C_QUES),
            ("question_image", 50, C_QUES),
            ("option_a", 30, C_OPTA),
            ("option_a_lang2", 20, C_OPTA),
            ("option_a_image", 40, C_OPTA),
            ("option_b", 30, C_OPTB),
            ("option_b_lang2", 20, C_OPTB),
            ("option_b_image", 40, C_OPTB),
            ("option_c", 30, C_OPTC),
            ("option_c_lang2", 20, C_OPTC),
            ("option_c_image", 40, C_OPTC),
            ("option_d", 30, C_OPTD),
            ("option_d_lang2", 20, C_OPTD),
            ("option_d_image", 40, C_OPTD),
            ("correct_answer", 30, C_ANS),
        ]

        header_font = Font(bold=True, size=12)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for idx, (name, width, color) in enumerate(columns_config, 1):
            cell = ws.cell(row=1, column=idx, value=name)
            cell.font = header_font
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(idx)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )


    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request):
        print("=== START BULK UPLOAD (IMAGE SAFE) ===")

        file_obj = request.FILES.get('file')
        question_paper_id = request.data.get('question_paper_id')

        if not file_obj or not question_paper_id:
            return Response({"error": "file and question_paper_id required"}, status=400)

        wb = load_workbook(file_obj)
        ws = wb.active

        images_map = extract_images_from_xlsx(file_obj)
        print(f"DEBUG: Extracted {len(images_map)} images")

        header_map = {
            str(cell.value).strip().lower(): cell.col_idx
            for cell in ws[1] if cell.value
        }

        created = 0
        errors = []

        for row_idx in range(2, ws.max_row + 1):

            has_text = any(
                ws.cell(row=row_idx, column=c).value
                for c in range(1, ws.max_column + 1)
            )
            has_image = any(
                (row_idx, c) in images_map
                for c in range(1, ws.max_column + 1)
            )

            if not has_text and not has_image:
                continue

            def get_val(key):
                col = header_map.get(key)
                if not col:
                    return ""
                val = ws.cell(row=row_idx, column=col).value
                return str(val).strip() if val else ""

            def get_img(img_key, txt_key):
                col = header_map.get(img_key)
                if col and (row_idx, col) in images_map:
                    return images_map[(row_idx, col)]
                col = header_map.get(txt_key)
                if col and (row_idx, col) in images_map:
                    return images_map[(row_idx, col)]
                return None

            data = {
                "question_paper": question_paper_id,
                "question": get_val("question"),
                "question_lang2": get_val("question_lang2"),
                "option_a": get_val("option_a"),
                "option_a_lang2": get_val("option_a_lang2"),
                "option_b": get_val("option_b"),
                "option_b_lang2": get_val("option_b_lang2"),
                "option_c": get_val("option_c"),
                "option_c_lang2": get_val("option_c_lang2"),
                "option_d": get_val("option_d"),
                "option_d_lang2": get_val("option_d_lang2"),
                "correct_answer": get_val("correct_answer"),
            }

            if img := get_img("question_image", "question"):
                data["question_image"] = img

            for opt in ["a", "b", "c", "d"]:
                img = get_img(f"option_{opt}_image", f"option_{opt}")
                if img:
                    data[f"option_{opt}_image"] = img

            serializer = self.get_serializer(data=data)
            if serializer.is_valid():
                serializer.save()
                created += 1
            else:
                errors.append({"row": row_idx, "errors": serializer.errors})


        if created > 0 and len(errors) == 0:
            status_msg = "Upload Success"
        elif created > 0 and len(errors) > 0:
            status_msg = "Partial Success"
        else:
            status_msg = "Upload Failed"

            
        return Response({
            "status": status_msg,
            "created_count": created,
            "error_count": len(errors),
            "errors": errors
        }, status=201 if created else 400)



    @action(detail=False, methods=['post'], url_path='bulk-action')
    def bulk_action(self, request):
        action_name = request.data.get('action')
        ids = request.data.get('ids', [])
        if isinstance(ids, str):
            ids = [ids]
        elif not isinstance(ids, list):
            ids = []
    
        ids = [int(i) for i in ids if str(i).isdigit()]
    
        if not ids:
            return Response({"error": "ids required"}, status=400)
    
        qs = self.get_queryset().filter(id__in=ids)
    
        if action_name == 'archive':
            count = qs.update(is_active=False)
            return Response({"message": f"Archived {count} questions."})
    
        elif action_name == 'restore':
            count = qs.update(is_active=True)
            return Response({"message": f"Restored {count} questions."})
    
        elif action_name == 'delete':
            count = qs.count()
            qs.delete()
            return Response({"message": f"Deleted {count} questions."})
    
        elif action_name == 'duplicate':
            with transaction.atomic():
                new_ids = []
                for q in qs:
                    q.pk = None
                    q.id = None
                    q.order = 0
                    q.is_active = True
                    q.save()
                    new_ids.append(q.id)
            return Response({"message": f"Duplicated {len(new_ids)} questions."})
    
        return Response({"error": "Unknown action"}, status=400)












from rest_framework import viewsets
from rest_framework.parsers import JSONParser, FormParser, MultiPartParser
from .models import Hq, Factory, Department, Line, SubLine, Station
from .serializers import HqSerializer, FactorySerializer, DepartmentSerializer, LineSerializer, SubLineSerializer, StationSerializer

# ------------------ ViewSets ------------------
class HqViewSet(viewsets.ModelViewSet):
    queryset = Hq.objects.all()
    serializer_class = HqSerializer
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def get_serializer(self, *args, **kwargs):
        if hasattr(self.request, "data"):
            print("Incoming data:", self.request.data)
        return super().get_serializer(*args, **kwargs)

    def create(self, request, *args, **kwargs):
        # Handle CORS preflight
        if request.method == 'OPTIONS':
            response = Response()
            origin = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Origin'] = origin
            response['Access-Control-Allow-Methods'] = 'POST, PUT, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type, X-CSRFToken, X-Requested-With'
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Access-Control-Max-Age'] = '86400'
            return response

        try:
            print("[HQ CREATE] headers:", dict(request.headers))
            print("[HQ CREATE] content_type:", request.content_type)
            print("[HQ CREATE] data:", request.data)
            print("[HQ CREATE] raw body:", request.body)
        except Exception:
            pass
        resp = super().create(request, *args, **kwargs)
        # Add CORS headers
        origin = request.headers.get('Origin', '*')
        resp['Access-Control-Allow-Origin'] = origin
        resp['Access-Control-Allow-Credentials'] = 'true'
        resp['Vary'] = 'Origin'
        return resp

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        # Handle CORS preflight
        if request.method == 'OPTIONS':
            response = Response()
            origin = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Origin'] = origin
            response['Access-Control-Allow-Methods'] = 'PUT, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type, X-CSRFToken, X-Requested-With'
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Access-Control-Max-Age'] = '86400'
            return response

        try:
            print(f"[HQ UPDATE partial={partial}] headers:", dict(request.headers))
            print("[HQ UPDATE] content_type:", request.content_type)
            print("[HQ UPDATE] data:", request.data)
            print("[HQ UPDATE] raw body:", request.body)
        except Exception:
            pass
        resp = super().update(request, *args, **kwargs)
        # Add CORS headers
        origin = request.headers.get('Origin', '*')
        resp['Access-Control-Allow-Origin'] = origin
        resp['Access-Control-Allow-Credentials'] = 'true'
        resp['Vary'] = 'Origin'
        return resp


class FactoryViewSet(viewsets.ModelViewSet):
    queryset = Factory.objects.all()
    serializer_class = FactorySerializer
    
    def create(self, request, *args, **kwargs):
        print("\n=== Incoming Request ===")
        print("Method:", request.method)
        print("Headers:", dict(request.headers))
        print("Content-Type:", request.content_type)
        print("Raw data:", request.data)
        print("User:", request.user)
        print("Authenticated:", request.user.is_authenticated)
        
        # Handle CORS preflight
        if request.method == 'OPTIONS':
            print("Handling OPTIONS request")
            response = Response()
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type, X-CSRFToken, X-Requested-With'
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Access-Control-Max-Age'] = '86400'  # 24 hours
            return response
            
        # Make a mutable copy of the QueryDict
        data = request.data.copy()
        print("Data copy:", data)
        
        try:
            # Call parent's create method
            response = super().create(request, *args, **kwargs)
            
            # Add CORS headers to the response
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Vary'] = 'Origin'
            
            print("Factory created successfully!")
            return response
            
        except Exception as e:
            print("Error creating factory:", str(e))
            response = Response(
                {"error": "Failed to create factory", "details": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            return response
                                            

class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer

    def create(self, request, *args, **kwargs):
        print("\n=== Incoming Department Request ===")
        print("Method:", request.method)
        print("Headers:", dict(request.headers))
        print("Data:", request.data)
        
        # Handle CORS preflight
        if request.method == 'OPTIONS':
            print("Handling OPTIONS request")
            response = Response()
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type, X-CSRFToken, X-Requested-With'
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Access-Control-Max-Age'] = '86400'  # 24 hours
            return response
            
        try:
            # Call parent's create method
            response = super().create(request, *args, **kwargs)
            
            # Add CORS headers to the response
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Vary'] = 'Origin'
            
            print("Department created successfully!")
            return response
            
        except Exception as e:
            print("Error creating department:", str(e))
            response = Response(
                {"error": "Failed to create department", "details": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            return response


class LineViewSet(viewsets.ModelViewSet):
    queryset = Line.objects.all()
    serializer_class = LineSerializer

    def create(self, request, *args, **kwargs):
        print("\n=== Incoming Line Request ===")
        print("Method:", request.method)
        print("Headers:", dict(request.headers))
        print("Data:", request.data)
        
        # Handle CORS preflight
        if request.method == 'OPTIONS':
            print("Handling OPTIONS request")
            response = Response()
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type, X-CSRFToken, X-Requested-With'
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Access-Control-Max-Age'] = '86400'  # 24 hours
            return response
            
        try:
            # Call parent's create method
            response = super().create(request, *args, **kwargs)
            
            # Add CORS headers to the response
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Vary'] = 'Origin'
            
            print("Line created successfully!")
            return response
            
        except Exception as e:
            print("Error creating line:", str(e))
            response = Response(
                {"error": "Failed to create line", "details": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            return response


class SubLineViewSet(viewsets.ModelViewSet):
    queryset = SubLine.objects.all()
    serializer_class = SubLineSerializer

    def create(self, request, *args, **kwargs):
        print("\n=== Incoming SubLine Request ===")
        print("Method:", request.method)
        print("Headers:", dict(request.headers))
        print("Data:", request.data)
        
        # Handle CORS preflight
        if request.method == 'OPTIONS':
            print("Handling OPTIONS request")
            response = Response()
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type, X-CSRFToken, X-Requested-With'
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Access-Control-Max-Age'] = '86400'  # 24 hours
            return response
            
        try:
            # Call parent's create method
            response = super().create(request, *args, **kwargs)
            
            # Add CORS headers to the response
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Vary'] = 'Origin'
            
            print("SubLine created successfully!")
            return response
            
        except Exception as e:
            print("Error creating subline:", str(e))
            response = Response(
                {"error": "Failed to create subline", "details": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            return response


class StationViewSet(viewsets.ModelViewSet):
    queryset = Station.objects.all()
    serializer_class = StationSerializer

    def create(self, request, *args, **kwargs):
        print("\n=== Incoming Station Request ===")
        print("Method:", request.method)
        print("Headers:", dict(request.headers))
        print("Data:", request.data)
        
        # Handle CORS preflight
        if request.method == 'OPTIONS':
            print("Handling OPTIONS request")
            response = Response()
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type, X-CSRFToken, X-Requested-With'
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Access-Control-Max-Age'] = '86400'  # 24 hours
            return response
            
        try:
            # Call parent's create method
            response = super().create(request, *args, **kwargs)
            
            # Add CORS headers to the response
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Vary'] = 'Origin'
            
            print("Station created successfully!")
            return response
            
        except Exception as e:
            print("Error creating station:", str(e))
            response = Response(
                {"error": "Failed to create station", "details": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            return response


# ------------------ AR/VR Views ------------------
from rest_framework import viewsets
from .models import ARVRTrainingContent
from .serializers import ARVRTrainingContentSerializer

class ARVRTrainingContentViewSet(viewsets.ModelViewSet):
    queryset = ARVRTrainingContent.objects.all()
    serializer_class = ARVRTrainingContentSerializer


# --------------------------
# Level 2 Process Dojo
# --------------------------

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import LevelWiseTrainingContent, TrainingTopic
from .serializers import LevelWiseTrainingContentSerializer, TrainingTopicSerializer

class LevelWiseTrainingContentViewSet(viewsets.ModelViewSet):
    queryset = LevelWiseTrainingContent.objects.all()
    serializer_class = LevelWiseTrainingContentSerializer

    # Custom endpoint: filter by level, station, and topic
    @action(detail=False, methods=["get"])
    def by_level_station_topic(self, request):
        level_id = request.query_params.get("level_id")
        station_id = request.query_params.get("station_id")
        topic_id = request.query_params.get("topic_id")

        queryset = self.queryset
        if level_id:
            queryset = queryset.filter(level_id=level_id)
        if station_id:
            queryset = queryset.filter(station_id=station_id)
        if topic_id:
            queryset = queryset.filter(topic_id=topic_id)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class TrainingTopicViewSet(viewsets.ModelViewSet):
    queryset = TrainingTopic.objects.all()
    serializer_class = TrainingTopicSerializer
    
    def get_queryset(self):
        """Filter topics by level and station from query parameters by default"""
        queryset = TrainingTopic.objects.all()
        level_id = self.request.query_params.get("level_id")
        station_id = self.request.query_params.get("station_id")
        
        if level_id:
            queryset = queryset.filter(level_id=level_id)
        if station_id:
            queryset = queryset.filter(station_id=station_id)
            
        return queryset
    
    # Custom endpoint: filter topics by level and/or station (keeping for backward compatibility)
    @action(detail=False, methods=["get"])
    def by_level_station(self, request):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

# hanchou & shokucho 



# your_app/views.py

from rest_framework import viewsets
from .models import HanContent, HanSubtopic, HanTrainingContent
from .serializers import (
    HanContentDetailSerializer,
    HanContentListSerializer,
    HanSubtopicSerializer,
    HanTrainingContentSerializer
)

class HanContentViewSet(viewsets.ModelViewSet):
    """
    Handles listing, creating, retrieving, updating, and deleting Main Topics.
    """
    # prefetch_related is a performance optimization that prevents many small database queries.
    queryset = HanContent.objects.prefetch_related('subtopics__materials').all()

    def get_serializer_class(self):
        """
        Chooses the serializer based on the action.
        - For retrieving a single item ('retrieve'), use the detailed serializer.
        - For listing all items ('list'), use the simple serializer.
        """
        if self.action == 'retrieve':
            return HanContentDetailSerializer
        return HanContentListSerializer


# --- MODIFIED VIEWSET FOR SUBTOPICS ---
class HanSubtopicViewSet(viewsets.ModelViewSet):
    serializer_class = HanSubtopicSerializer

    def get_queryset(self):
        """
        Filters subtopics based on a query parameter from the URL.
        Example: GET /api/subtopics/?han_content_id=5
        """
        queryset = HanSubtopic.objects.all()
        han_content_id = self.request.query_params.get('han_content_id')
        if han_content_id:
            queryset = queryset.filter(han_content_id=han_content_id)
        return queryset

    def perform_create(self, serializer):
        """
        When creating, the parent ID must be sent in the request body.
        Example: POST /api/subtopics/ with body {"title": "...", "han_content": 5}
        """
        # The serializer will handle associating the parent, since the ID is in the data.
        serializer.save()

from django.http import FileResponse


class HanTrainingContentViewSet(viewsets.ModelViewSet):
    serializer_class = HanTrainingContentSerializer

    def get_queryset(self):
        """
        Filters materials based on a query parameter from the URL.
        Example: GET /api/materials/?han_subtopic_id=12
        """
        queryset = HanTrainingContent.objects.all()
        subtopic_id = self.request.query_params.get('han_subtopic_id')
        if subtopic_id:
            queryset = queryset.filter(han_subtopic_id=subtopic_id)
        return queryset

    def perform_create(self, serializer):
        """
        When creating, the parent ID must be sent in the request body.
        Example: POST /api/materials/ with body {"description": "...", "han_subtopic": 12}
        """
        serializer.save()


def serve_han_material_file(request, pk):
    """
    Serves the protected media file for a HanTrainingContent object.
    """
    material = get_object_or_404(HanTrainingContent, pk=pk)

    if not material.training_file:
        raise Http404("No file found for this material.")

    try:
        return FileResponse(material.training_file.open('rb'), as_attachment=False)
    except FileNotFoundError:
        raise Http404("File does not exist on the server.")


from rest_framework import viewsets
from django.shortcuts import get_object_or_404
from django.http import FileResponse
from .models import ShoContent, ShoSubtopic, ShoTrainingContent
from .serializers import (
    ShoContentListSerializer,
    ShoContentDetailSerializer,
    ShoSubtopicSerializer,
    ShoTrainingContentSerializer
)


# --- SHO CONTENT VIEWSET ---
class ShoContentViewSet(viewsets.ModelViewSet):
    """
    Handles listing, creating, retrieving, updating, and deleting Main Topics.
    """
    queryset = ShoContent.objects.prefetch_related('sho_subtopics__sho_materials').all()

    def get_serializer_class(self):
        """
        Chooses the serializer based on the action.
        - For retrieving a single item ('retrieve'), use the detailed serializer.
        - For listing all items ('list'), use the simple serializer.
        """
        if self.action == 'retrieve':
            return ShoContentDetailSerializer
        return ShoContentListSerializer


# --- SHO SUBTOPIC VIEWSET ---
class ShoSubtopicViewSet(viewsets.ModelViewSet):
    serializer_class = ShoSubtopicSerializer

    def get_queryset(self):
        """
        Filters subtopics based on a query parameter from the URL.
        Example: GET /api/sho-subtopics/?sho_content_id=5
        """
        queryset = ShoSubtopic.objects.all()
        sho_content_id = self.request.query_params.get('sho_content_id')
        if sho_content_id:
            queryset = queryset.filter(sho_content_id=sho_content_id)
        return queryset

    def perform_create(self, serializer):
        """
        When creating, the parent ID must be sent in the request body.
        Example: POST /api/sho-subtopics/ with body {"title": "...", "sho_content": 5}
        """
        serializer.save()


# --- SHO TRAINING CONTENT VIEWSET ---
class ShoTrainingContentViewSet(viewsets.ModelViewSet):
    serializer_class = ShoTrainingContentSerializer

    def get_queryset(self):
        """
        Filters materials based on a query parameter from the URL.
        Example: GET /api/sho-materials/?sho_subtopic_id=12
        """
        queryset = ShoTrainingContent.objects.all()
        subtopic_id = self.request.query_params.get('sho_subtopic_id')
        if subtopic_id:
            queryset = queryset.filter(sho_subtopic_id=subtopic_id)
        return queryset

    def perform_create(self, serializer):
        """
        When creating, the parent ID must be sent in the request body.
        Example: POST /api/sho-materials/ with body {"sho_description": "...", "sho_subtopic": 12}
        """
        serializer.save()


def serve_sho_material_file(request, pk):
    """
    Serves the protected media file for a ShoTrainingContent object.
    """
    material = get_object_or_404(ShoTrainingContent, pk=pk)

    if not material.training_file:
        raise Http404("No file found for this material.")

    try:
        return FileResponse(material.training_file.open('rb'), as_attachment=False)
    except FileNotFoundError:
        raise Http404("File does not exist on the server.")







from rest_framework import viewsets
from .models import HanchouExamQuestion
from .serializers import HanchouExamQuestionSerializer

class HanchouExamQuestionViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing Hanchou Exam Questions, including bulk upload.
    """
    queryset = HanchouExamQuestion.objects.all()
    serializer_class = HanchouExamQuestionSerializer
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request, *args, **kwargs):
        """
        Generates and serves an Excel template with sample data for Hanchou questions.
        """
        sample_data = [
            {
                'question': 'What is the largest mammal in the world?',
                'option_a': 'Elephant',
                'option_b': 'Blue Whale',
                'option_c': 'Giraffe',
                'option_d': 'Great White Shark',
                'correct_answer': 'Blue Whale'
            },
            {
                'question': 'Which element has the atomic number 1?',
                'option_a': 'Helium',
                'option_b': 'Oxygen',
                'option_c': 'Hydrogen',
                'option_d': 'Carbon',
                'correct_answer': 'Hydrogen'
            }
        ]
        
        df = pd.DataFrame(sample_data)
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Using a specific sheet name is good practice
            df.to_excel(writer, index=False, sheet_name='Hanchou_Questions')
            
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Hanchou_Question_Upload_Template.xlsx"'
        
        return response

    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request, *args, **kwargs):
        """
        Handles bulk creation of Hanchou questions from an Excel file.
        """
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'detail': 'No file was uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # The sheet_name must match the one used in download_template
            df = pd.read_excel(file_obj, sheet_name='Hanchou_Questions', engine='openpyxl')
            df = df.where(pd.notnull(df), None)
        except Exception as e:
            return Response({'detail': f"Error reading the Excel file. Ensure it contains a sheet named 'Hanchou_Questions'. Error: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        required_columns = {
            'question', 'option_a', 'option_b', 'option_c',
            'option_d', 'correct_answer'
        }
        
        if not required_columns.issubset(df.columns):
            missing_cols = required_columns - set(df.columns)
            return Response({'detail': f'File is missing required columns: {", ".join(missing_cols)}'}, status=status.HTTP_400_BAD_REQUEST)

        questions_to_create = []
        errors = []

        for index, row in df.iterrows():
            row_data = row.to_dict()
            serializer = self.get_serializer(data=row_data)
            
            try:
                # Validate using the model's clean() method
                instance = HanchouExamQuestion(**row_data)
                instance.clean()
                
                # Also run standard serializer validation
                serializer.is_valid(raise_exception=True)
                questions_to_create.append(serializer.validated_data)

            except (DRFValidationError, DjangoValidationError, TypeError) as e:
                error_detail = serializer.errors if hasattr(serializer, 'errors') and serializer.errors else str(e)
                errors.append({'row': index + 2, 'errors': error_detail})
        
        if questions_to_create:
            model_instances = [HanchouExamQuestion(**data) for data in questions_to_create]
            HanchouExamQuestion.objects.bulk_create(model_instances)

        response_data = {
            'status': 'Upload complete.',
            'created_count': len(questions_to_create),
            'error_count': len(errors),
            'errors': errors
        }

        return Response(response_data, status=status.HTTP_201_CREATED)
    

from rest_framework import viewsets, permissions, filters as drf_filters
from django_filters import rest_framework as filters
from .models import HanchouExamResult
from .serializers import HanchouExamResultSerializer

class HanchouExamResultFilter(filters.FilterSet):
    pay_code = filters.CharFilter(field_name="employee__pay_code", lookup_expr="iexact")
    name = filters.CharFilter(field_name="employee__name", lookup_expr="icontains")
    exam_date = filters.DateFromToRangeFilter(field_name="exam_date")
    submitted_at = filters.DateFromToRangeFilter(field_name="submitted_at")
    passed = filters.BooleanFilter()

    class Meta:
        model = HanchouExamResult
        fields = ["employee", "passed", "exam_date"]

class HanchouExamResultViewSet(viewsets.ModelViewSet):
    queryset = HanchouExamResult.objects.select_related("employee").all().order_by("-submitted_at", "-started_at")
    serializer_class = HanchouExamResultSerializer
    # permission_classes = [permissions.IsAuthenticated]  # adjust as needed

    filter_backends = [filters.DjangoFilterBackend, drf_filters.SearchFilter, drf_filters.OrderingFilter]
    filterset_class = HanchouExamResultFilter
    search_fields = ["employee__name", "employee__pay_code", "employee__card_no", "remarks"]
    ordering_fields = ["submitted_at", "exam_date", "score", "total_questions", "duration_seconds"]
    ordering = ["-submitted_at", "-started_at"]

from rest_framework import viewsets
from .models import ShokuchouExamQuestion,ShokuchouExamResult
from .serializers import ShokuchouExamQuestionSerializer,ShokuchouExamResultSerializer

# Add these imports at the top of your views.py
import pandas as pd
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django.core.exceptions import ValidationError as DjangoValidationError
from rest_framework.exceptions import ValidationError as DRFValidationError

# Your existing model and serializer
from .models import ShokuchouExamQuestion
from .serializers import ShokuchouExamQuestionSerializer


class ShokuchouExamQuestionViewSet(viewsets.ModelViewSet):
    queryset = ShokuchouExamQuestion.objects.all()
    serializer_class = ShokuchouExamQuestionSerializer
    # Add parser classes for the ViewSet to handle file uploads
    parser_classes = [JSONParser, MultiPartParser, FormParser]



    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request, *args, **kwargs):
        """
        Generates and serves an Excel template with sample data for bulk uploading questions.
        """
        # Define the sample data
        sample_data = [
            {
                'sho_question': 'What is the capital of Japan?',
                'sho_option_a': 'Seoul',
                'sho_option_b': 'Beijing',
                'sho_option_c': 'Tokyo',
                'sho_option_d': 'Bangkok',
                'sho_correct_answer': 'Tokyo'
            },
            {
                'sho_question': 'Which planet is known as the Red Planet?',
                'sho_option_a': 'Earth',
                'sho_option_b': 'Mars',
                'sho_option_c': 'Jupiter',
                'sho_option_d': 'Saturn',
                'sho_correct_answer': 'Mars'
            }
        ]
        
        # Create a pandas DataFrame
        df = pd.DataFrame(sample_data)
        
        # Use an in-memory buffer
        buffer = io.BytesIO()
        
        # Write the DataFrame to the buffer as an Excel file
        # index=False prevents pandas from writing row indices to the file
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Questions')
            
        # Set the buffer's pointer to the beginning
        buffer.seek(0)
        
        # Create the HTTP response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Question_Upload_Template.xlsx"'
        
        return response







    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request, *args, **kwargs):
        """
        Handles bulk creation of questions from an Excel file upload.
        The Excel file must have a sheet named 'Questions' and columns:
        'sho_question', 'sho_option_a', 'sho_option_b', 'sho_option_c',
        'sho_option_d', 'sho_correct_answer'.
        """
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'detail': 'No file was uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not file_obj.name.endswith(('.xlsx', '.xls')):
            return Response({'detail': 'Invalid file format. Please upload an Excel file (.xlsx, .xls).'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Explicitly read the 'Questions' sheet
            df = pd.read_excel(file_obj, sheet_name='Questions', engine='openpyxl')
            # Replace NaN values with None for proper serialization
            df = df.where(pd.notnull(df), None)
        except Exception as e:
            # Catches errors like missing sheet or unreadable file
            return Response(
                {'detail': f"Error reading the Excel file. Make sure it contains a sheet named 'Questions'. Error: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        required_columns = {
            'sho_question', 'sho_option_a', 'sho_option_b', 'sho_option_c',
            'sho_option_d', 'sho_correct_answer'
        }
        
        if not required_columns.issubset(df.columns):
            missing_cols = required_columns - set(df.columns)
            return Response(
                {'detail': f'File is missing required columns: {", ".join(missing_cols)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        questions_to_create = []
        errors = []

        for index, row in df.iterrows():
            row_data = row.to_dict()
            serializer = self.get_serializer(data=row_data)
            
            try:
                # The model's clean() method is not called by default in DRF serializers.
                # We can simulate it by creating a model instance and calling full_clean().
                # This ensures our custom validation rule is checked.
                instance = ShokuchouExamQuestion(**row_data)
                instance.clean() # This will call our custom validation
                
                # We also run serializer validation for data types, max_length, etc.
                serializer.is_valid(raise_exception=True)
                questions_to_create.append(serializer.validated_data)

            except (DRFValidationError, DjangoValidationError, TypeError) as e:
                error_detail = serializer.errors if hasattr(serializer, 'errors') and serializer.errors else str(e)
                errors.append({'row': index + 2, 'errors': error_detail}) # +2 because index is 0-based and header is row 1
        
        # Now, create the actual model instances for bulk_create
        if questions_to_create:
            model_instances = [ShokuchouExamQuestion(**data) for data in questions_to_create]
            ShokuchouExamQuestion.objects.bulk_create(model_instances)

        response_data = {
            'status': 'Upload complete.',
            'created_count': len(questions_to_create),
            'error_count': len(errors),
            'errors': errors
        }

        return Response(response_data, status=status.HTTP_201_CREATED)






class ShokuchouExamResultFilter(filters.FilterSet):
    pay_code = filters.CharFilter(field_name="employee__pay_code", lookup_expr="iexact")
    name = filters.CharFilter(field_name="employee__name", lookup_expr="icontains")
    sho_submitted_at = filters.DateFromToRangeFilter(field_name="sho_submitted_at")
    sho_passed = filters.BooleanFilter()

    class Meta:
        model = ShokuchouExamResult
        fields = ["employee", "sho_passed", "sho_submitted_at"]


# --- SHO RESULT VIEWSET ---
class ShokuchouExamResultViewSet(viewsets.ModelViewSet):
    queryset = (
        ShokuchouExamResult.objects.select_related("employee")
        .all()
        .order_by("-sho_submitted_at", "-sho_started_at")
    )
    serializer_class = ShokuchouExamResultSerializer
    # permission_classes = [permissions.IsAuthenticated]  # enable if needed

    filter_backends = [filters.DjangoFilterBackend, drf_filters.SearchFilter, drf_filters.OrderingFilter]
    filterset_class = ShokuchouExamResultFilter
    search_fields = ["employee__name", "employee__pay_code", "employee__card_no", "sho_remarks"]
    ordering_fields = [
        "sho_submitted_at",
        "sho_score",
        "sho_total_questions",
        "sho_duration_seconds",
    ]
    ordering = ["-sho_submitted_at", "-sho_started_at"]


class HanchouResultCertificatePDF(APIView):
    def get(self, request, pk):
        try:
            result = HanchouExamResult.objects.select_related('employee').get(pk=pk)
            
            employee_name = result.employee.name.strip()
            exam_name = result.exam_name.strip()

            buffer = io.BytesIO()
            p = canvas.Canvas(buffer, pagesize=landscape(letter))
            width, height = landscape(letter)

            # Define colors and margins
            main_color = (85/255, 26/255, 139/255)
            shadow_color = (200/255, 200/255, 200/255)
            border_color = (244/255, 145/255, 34/255)
            margin = 0.5 * inch

            # Draw the decorative borders
            p.setStrokeColorRGB(*border_color)
            p.setLineWidth(3)
            p.rect(margin, margin, width - 2 * margin, height - 2 * margin)
            p.setStrokeColorRGB(0.9, 0.9, 0.9)
            p.setLineWidth(1)
            p.rect(margin + 5, margin + 5, width - 2 * (margin + 5), height - 2 * (margin + 5))

            # --- DRAW CERTIFICATE CONTENT ---
            
            draw_reflected_text(p, width / 2.0, height - 1.7*inch, "KML SEATING", "Times-Bold", 36, main_color, shadow_color)
            draw_reflected_text(p, width / 2.0, height - 2.4*inch, "HANCHOUE EXAM CERTIFICATE", "Times-Bold", 28, main_color, shadow_color)
            
            p.setFont("Times-Roman", 16)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawCentredString(width / 2.0, height - 3.7*inch, "THIS IS TO CERTIFY THAT")
            
            draw_reflected_text(p, width / 2.0, height - 4.4*inch, f'“{employee_name.upper()}”', "Times-Bold", 24, main_color, shadow_color, y_offset=1.5)
            
            p.setFont("Times-Roman", 16)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawCentredString(width / 2.0, height - 5.2*inch, "HAS SUCCESSFULLY PASSED THE")
            
            draw_reflected_text(p, width / 2.0, height - 5.9*inch, exam_name.upper(), "Times-Bold", 22, main_color, shadow_color, y_offset=1.5)

            # Signature
            p.setFont("Times-Roman", 12)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawRightString(width - margin - 0.5*inch, margin + 0.8*inch, "TRAINER SIGNATURE")
            p.line(width - margin - 2.5*inch, margin + 0.7*inch, width - margin - 0.5*inch, margin + 0.7*inch)

            p.showPage()
            p.save()

            buffer.seek(0)
            return HttpResponse(buffer, content_type='application/pdf', headers={'Content-Disposition': f'attachment; filename="Hanchou_Certificate_{employee_name}.pdf"'})

        except HanchouExamResult.DoesNotExist:
            return Response({"error": "Hanchou exam result not found"}, status=status.HTTP_404_NOT_FOUND)
        except AttributeError:
            return Response({"error": "Associated employee for this result could not be found."}, status=status.HTTP_404_NOT_FOUND)



# Make sure you have all necessary imports at the top of your views.py file
import io
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch

# from .models import Score # Ensure Score model is imported

# This helper function should be defined or imported in your views.py
def draw_reflected_text(p, x, y, text, font_name, font_size, main_color, shadow_color, y_offset=2):
    """Draws centered text with a subtle shadow/reflection effect."""
    # Shadow
    p.setFont(font_name, font_size)
    p.setFillColorRGB(*shadow_color)
    p.drawCentredString(x + y_offset, y - y_offset, text)
    # Main Text
    p.setFillColorRGB(*main_color)
    p.drawCentredString(x, y, text)



# --- Place the draw_reflected_text function here ---
def draw_reflected_text(p, x, y, text, font, size, main_color, shadow_color, x_offset=1.5, y_offset=1.5):
    """Draws text with a simple shadow/reflection underneath."""
    # Draw shadow/reflection
    p.setFont(font, size)
    p.setFillColorRGB(*shadow_color)
    p.drawCentredString(x + x_offset, y - y_offset, text)
    
    # Draw main text
    p.setFillColorRGB(*main_color)
    p.drawCentredString(x, y, text)
# ----------------------------------------------------


class ShokuchouResultCertificatePDF(APIView):
    def get(self, request, pk):
        try:
            # CHANGED: Query ShokuchouExamResult instead of HanchouExamResult
            result = ShokuchouExamResult.objects.select_related('employee').get(pk=pk)
            
            employee_name = result.employee.name.strip()
            # CHANGED: Use sho_exam_name field
            exam_name = result.sho_exam_name.strip()

            buffer = io.BytesIO()
            p = canvas.Canvas(buffer, pagesize=landscape(letter))
            width, height = landscape(letter)

            # Define colors and margins (kept the same for consistent branding)
            main_color = (85/255, 26/255, 139/255)
            shadow_color = (200/255, 200/255, 200/255)
            border_color = (244/255, 145/255, 34/255)
            margin = 0.5 * inch

            # Draw the decorative borders
            p.setStrokeColorRGB(*border_color)
            p.setLineWidth(3)
            p.rect(margin, margin, width - 2 * margin, height - 2 * margin)
            p.setStrokeColorRGB(0.9, 0.9, 0.9)
            p.setLineWidth(1)
            p.rect(margin + 5, margin + 5, width - 2 * (margin + 5), height - 2 * (margin + 5))

            # --- DRAW CERTIFICATE CONTENT ---
            
            draw_reflected_text(p, width / 2.0, height - 1.7*inch, "KML SEATING", "Times-Bold", 36, main_color, shadow_color)
            # CHANGED: Updated certificate title
            draw_reflected_text(p, width / 2.0, height - 2.4*inch, "SHOKUCHOU EXAM CERTIFICATE", "Times-Bold", 28, main_color, shadow_color)
            
            p.setFont("Times-Roman", 16)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawCentredString(width / 2.0, height - 3.7*inch, "THIS IS TO CERTIFY THAT")
            
            draw_reflected_text(p, width / 2.0, height - 4.4*inch, f'“{employee_name.upper()}”', "Times-Bold", 24, main_color, shadow_color, y_offset=1.5)
            
            p.setFont("Times-Roman", 16)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawCentredString(width / 2.0, height - 5.2*inch, "HAS SUCCESSFULLY PASSED THE")
            
            draw_reflected_text(p, width / 2.0, height - 5.9*inch, exam_name.upper(), "Times-Bold", 22, main_color, shadow_color, y_offset=1.5)

            # Signature
            p.setFont("Times-Roman", 12)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawRightString(width - margin - 0.5*inch, margin + 0.8*inch, "TRAINER SIGNATURE")
            p.line(width - margin - 2.5*inch, margin + 0.7*inch, width - margin - 0.5*inch, margin + 0.7*inch)

            p.showPage()
            p.save()

            buffer.seek(0)
            # CHANGED: Updated the filename for the download
            return HttpResponse(buffer, content_type='application/pdf', headers={'Content-Disposition': f'attachment; filename="Shokuchou_Certificate_{employee_name}.pdf"'})

        # CHANGED: Catch DoesNotExist for the correct model
        except ShokuchouExamResult.DoesNotExist:
            return Response({"error": "Shokuchou exam result not found"}, status=status.HTTP_404_NOT_FOUND)
        except AttributeError:
            return Response({"error": "Associated employee for this result could not be found."}, status=status.HTTP_404_NOT_FOUND)


# 10 cycle

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from collections import defaultdict
from django.shortcuts import get_object_or_404

from .models import (
    TenCycleDayConfiguration,
    TenCycleTopics,
    TenCycleSubTopic,
    TenCyclePassingCriteria,
    OperatorPerformanceEvaluation,
    EvaluationSubTopicMarks,
    MasterTable
)
from .serializers import (
    TenCycleDayConfigurationSerializer,
    TenCycleTopicsSerializer,
    TenCycleSubTopicSerializer,
    TenCyclePassingCriteriaSerializer,
    OperatorPerformanceEvaluationSerializer,
    EvaluationSubMarksSerializer
)


class TenCycleSubTopicViewSet(viewsets.ModelViewSet):
    queryset = TenCycleSubTopic.objects.all()
    serializer_class = TenCycleSubTopicSerializer
    
    @action(detail=False, methods=['get'], url_path='by-topic')
    def by_topic(self, request):
        topic_id = request.query_params.get('topic_id')
        
        if not topic_id:
            return Response(
                {"error": "topic_id is required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
            
        queryset = self.queryset.filter(topic_id=topic_id, is_active=True)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class TenCycleDayConfigurationViewSet(viewsets.ModelViewSet):
    queryset = TenCycleDayConfiguration.objects.all()
    serializer_class = TenCycleDayConfigurationSerializer
    
    @action(detail=False, methods=['get'], url_path='get-configuration')
    def get_configuration(self, request):
        level_id = request.query_params.get('level_id')
        department_id = request.query_params.get('department_id')
        station_id = request.query_params.get('station_id')
        
        if not level_id or not department_id:
            return Response(
                {"error": "level_id and department_id are required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        filters = {
            'level_id': level_id,
            'department_id': department_id,
            'is_active': True
        }
        
        if station_id:
            filters['station_id'] = station_id
        else:
            filters['station__isnull'] = True
            
        queryset = self.queryset.filter(**filters)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class TenCycleTopicsViewSet(viewsets.ModelViewSet):
    queryset = TenCycleTopics.objects.all()
    serializer_class = TenCycleTopicsSerializer

    @action(detail=True, methods=['get'])
    def subtopics(self, request, pk=None):
        topic = self.get_object()
        subtopics = topic.subtopics.all()
        serializer = TenCycleSubTopicSerializer(subtopics, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], url_path='get-configuration')
    def get_configuration(self, request):
        level_id = request.query_params.get('level_id')
        department_id = request.query_params.get('department_id')
        station_id = request.query_params.get('station_id')
        
        if not level_id or not department_id:
            return Response(
                {"error": "level_id and department_id are required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        filters = {
            'level_id': level_id,
            'department_id': department_id,
            'is_active': True
        }
        
        if station_id:
            filters['station_id'] = station_id
        else:
            filters['station__isnull'] = True
            
        queryset = self.queryset.filter(**filters)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class TenCyclePassingCriteriaViewSet(viewsets.ModelViewSet):
    queryset = TenCyclePassingCriteria.objects.all()
    serializer_class = TenCyclePassingCriteriaSerializer
    
    @action(detail=False, methods=['get'], url_path='get-configuration')
    def get_configuration(self, request):
        level_id = request.query_params.get('level_id')
        department_id = request.query_params.get('department_id')
        station_id = request.query_params.get('station_id')
        
        if not level_id or not department_id:
            return Response(
                {"error": "level_id and department_id are required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if station_id:
            try:
                criteria = self.queryset.get(
                    level_id=level_id,
                    department_id=department_id,
                    station_id=station_id,
                    is_active=True
                )
                serializer = self.get_serializer(criteria)
                return Response(serializer.data)
            except TenCyclePassingCriteria.DoesNotExist:
                pass
        
        try:
            criteria = self.queryset.get(
                level_id=level_id,
                department_id=department_id,
                station__isnull=True,
                is_active=True
            )
            serializer = self.get_serializer(criteria)
            return Response(serializer.data)
        except TenCyclePassingCriteria.DoesNotExist:
            return Response(
                {"error": "No passing criteria found for the given configuration"}, 
                status=status.HTTP_404_NOT_FOUND
            )

class TenCycleConfigurationViewSet(viewsets.ViewSet):
    @action(detail=False, methods=['get'], url_path='complete-configuration')
    def complete_configuration(self, request):
        """Get complete configuration including days, topics, subtopics, and passing criteria"""
        level_id = request.query_params.get('level_id')
        department_id = request.query_params.get('department_id')
        station_id = request.query_params.get('station_id')
        
        if not level_id or not department_id:
            return Response(
                {"error": "level_id and department_id are required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        filters = {
            'level_id': level_id,
            'department_id': department_id,
            'is_active': True
        }
        
        if station_id:
            filters['station_id'] = station_id
        else:
            filters['station__isnull'] = True
        
        # Get days configuration
        days = TenCycleDayConfiguration.objects.filter(**filters)
        days_data = TenCycleDayConfigurationSerializer(days, many=True).data
        
        # Get topics configuration
        topics = TenCycleTopics.objects.filter(**filters)
        topics_data = TenCycleTopicsSerializer(topics, many=True).data
        
        # Get subtopics for each topic
        for topic_data in topics_data:
            topic_id = topic_data['id']
            subtopics = TenCycleSubTopic.objects.filter(topic_id=topic_id, is_active=True)
            topic_data['subtopics'] = TenCycleSubTopicSerializer(subtopics, many=True).data
        
        # Get passing criteria
        passing_criteria = None
        try:
            if station_id:
                criteria = TenCyclePassingCriteria.objects.get(
                    level_id=level_id,
                    department_id=department_id,
                    station_id=station_id,
                    is_active=True
                )
            else:
                criteria = TenCyclePassingCriteria.objects.get(
                    level_id=level_id,
                    department_id=department_id,
                    station__isnull=True,
                    is_active=True
                )
            passing_criteria = TenCyclePassingCriteriaSerializer(criteria).data
        except TenCyclePassingCriteria.DoesNotExist:
            # Set default passing criteria
            passing_criteria = {
                'passing_percentage': 60.0,
                'level': level_id,
                'department': department_id,
                'station': station_id
            }
        
        return Response({
            'days': days_data,
            'topics': topics_data,
            'passing_criteria': passing_criteria,
            'level_id': level_id,
            'department_id': department_id,
            'station_id': station_id
        })

# class OperatorPerformanceEvaluationViewSet(viewsets.ModelViewSet):
#     queryset = OperatorPerformanceEvaluation.objects.all()
#     serializer_class = OperatorPerformanceEvaluationSerializer

#     @action(detail=False, methods=['get'], url_path='by-employee-code/(?P<employee_code>[^/.]+)')
#     def by_employee_code(self, request, employee_code=None):
#         employee = get_object_or_404(MasterTable, emp_id=employee_code)
#         evaluations = OperatorPerformanceEvaluation.objects.filter(employee=employee)
#         serializer = self.get_serializer(evaluations, many=True)
#         return Response(serializer.data)


from rest_framework.decorators import action
from rest_framework.response import Response
from django.shortcuts import get_object_or_404

class OperatorPerformanceEvaluationViewSet(viewsets.ModelViewSet):
    queryset = OperatorPerformanceEvaluation.objects.all()
    serializer_class = OperatorPerformanceEvaluationSerializer

    @action(detail=False, methods=['get'], url_path='by-employee-code/(?P<employee_code>[^/.]+)')
    def by_employee_code(self, request, employee_code=None):
        employee = get_object_or_404(MasterTable, emp_id=employee_code)

        # Get filter parameters from the query string
        level_id = request.query_params.get('level_id')
        department_id = request.query_params.get('department_id')
        station_id = request.query_params.get('station_id')

        evaluations = OperatorPerformanceEvaluation.objects.filter(employee=employee)

        if level_id:
            evaluations = evaluations.filter(level_id=level_id)
        if department_id:
            evaluations = evaluations.filter(department_id=department_id)
        if station_id:
            evaluations = evaluations.filter(station_id=station_id)


        serializer = self.get_serializer(evaluations, many=True)
        return Response(serializer.data)



class EvaluationSubTopicMarksViewSet(viewsets.ModelViewSet):
    queryset = EvaluationSubTopicMarks.objects.all()
    serializer_class = EvaluationSubMarksSerializer

    @action(detail=False, methods=['get'], url_path='by-employee-code/(?P<employee_code>[^/.]+)')
    def by_employee_code(self, request, employee_code=None):
      employee = get_object_or_404(MasterTable, emp_id=employee_code)
      evaluations = OperatorPerformanceEvaluation.objects.filter(employee=employee)

      marks_qs = EvaluationSubTopicMarks.objects.filter(employee__in=evaluations)
      serializer = self.get_serializer(marks_qs, many=True)

      days_map = defaultdict(list)
      for mark_obj, mark_data in zip(marks_qs, serializer.data):
        days_map[mark_obj.day.day_name].append((mark_obj, mark_data))

      per_day_results = []
      total_score = 0
      total_possible_score = 0

      for day_name, marks_list in days_map.items():
        day_score = 0
        day_max_score = 0
        for mark_obj, mark_data in marks_list:
            subtopic_score = sum(mark_data.get(f'mark_{i}', 0) or 0 for i in range(1, 11))
            day_score += subtopic_score
            max_score = mark_obj.subtopic.score_required * 10
            day_max_score += max_score

        passing_percentage = 60.0  # or fetch dynamically 

        status_str = "Pass" if (day_max_score > 0 and (day_score / day_max_score) * 100 >= passing_percentage) else "Fail - Retraining Required"
        per_day_results.append({
            "day": day_name,
            "score": day_score,
            "max_score": day_max_score,
            "passing_percentage": passing_percentage,
            "status": status_str
        })

        total_score += day_score
        total_possible_score += day_max_score

      final_status = "Not Evaluated"
      if total_possible_score > 0:
        overall_percentage = (total_score / total_possible_score) * 100
        final_status = "Pass" if overall_percentage >= passing_percentage else "Fail - Retraining Required"

      
      return Response({
        'employee_code': employee_code,
        'evaluations': serializer.data,   # <-- this sends full detailed marks list
        'per_day_results': per_day_results,
        'total_score': total_score,
        'total_possible_score': total_possible_score,
        'final_status': final_status,
    })


    @action(detail=False, methods=['put'], url_path='update-mark/(?P<employee_code>[^/.]+)')
    def update_by_employee_code(self, request, employee_code=None):
        employee = get_object_or_404(MasterTable, emp_id=employee_code)
        employee_eval_qs = OperatorPerformanceEvaluation.objects.filter(employee=employee)
        evaluations = EvaluationSubTopicMarks.objects.filter(employee__in=employee_eval_qs)

        topic_id = request.data.get('subtopic')
        day_id = request.data.get('day')
        if not topic_id or not day_id:
            return Response({"error": "subtopic and day IDs are required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            mark_instance = evaluations.get(subtopic_id=topic_id, day_id=day_id)
        except EvaluationSubTopicMarks.DoesNotExist:
            return Response({"error": "Matching mark record not found."}, status=404)

        serializer = EvaluationSubMarksSerializer(mark_instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)
    
    @action(detail=False, methods=['get'], url_path='daily-total/(?P<employee_code>[^/.]+)')
    def daily_total(self, request, employee_code=None):
        employee = get_object_or_404(MasterTable, emp_id=employee_code)
        evaluations = OperatorPerformanceEvaluation.objects.filter(employee=employee)
        marks_qs = EvaluationSubTopicMarks.objects.filter(employee__in=evaluations)
        serializer = self.get_serializer(marks_qs, many=True)

        days_map = defaultdict(list)
        for mark_obj, mark_data in zip(marks_qs, serializer.data):
            days_map[mark_obj.day.day_name].append(mark_data)

        per_day_results = []
        grand_total = 0
        grand_max_total = 0

        for day_name, marks_list in days_map.items():
            day_score = 0
            day_max_score = 0
            for mark in marks_list:
                day_score += mark.get('total_score', 0) or 0
                day_max_score += mark.get('max_possible_score', 0) or 0
            grand_total += day_score
            grand_max_total += day_max_score
            per_day_results.append({
                'day': day_name,
                'total_score': day_score,
                'max_possible_score': day_max_score,
                'percentage': round((day_score / day_max_score)*100, 2) if day_max_score > 0 else 0
            })

        final_percentage = round((grand_total / grand_max_total)*100, 2) if grand_max_total > 0 else 0

        if evaluations.exists():
            eval_obj = evaluations.first()
            try:
                pass_crit = TenCyclePassingCriteria.objects.get(
                    level=eval_obj.level,
                    department=eval_obj.department,
                    station=eval_obj.station
                )
            except TenCyclePassingCriteria.DoesNotExist:
                pass_crit = TenCyclePassingCriteria.objects.filter(
                    level=eval_obj.level,
                    department=eval_obj.department,
                    station__isnull=True
                ).first()
            passing_percentage = pass_crit.passing_percentage if pass_crit else 60.0
        else:
            passing_percentage = 60.0  # default fallback

        final_status = "Pass" if final_percentage >= passing_percentage else "Fail - Retraining Required"

        return Response({
            'employee_code': employee_code,
            'per_day_totals': per_day_results,
            'grand_total_score': grand_total,
            'grand_max_score': grand_max_total,
            'final_percentage': final_percentage,
            'passing_percentage': passing_percentage,
            'final_status': final_status
        })
        
#=================================  10 cycle   ================================#


# ======================== Machine Allocation Approval ============ #


from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from .serializers import MachineAllocationApprovalSerializer
from .models import MachineAllocation


class MachineAllocationApprovalViewSet(viewsets.ModelViewSet):
    queryset = MachineAllocation.objects.all()
    serializer_class = MachineAllocationApprovalSerializer

    @action(detail=True, methods=['put'], url_path='set-status')
    def set_status(self, request, pk=None):
        allocation = self.get_object()
        status_value = request.data.get('approval_status')

        if status_value not in dict(MachineAllocation.APPROVAL_STATUS_CHOICES):
            return Response({'error': 'Invalid status'}, status=status.HTTP_400_BAD_REQUEST)

        allocation.approval_status = status_value
        allocation.save()
        return Response({
            'status': 'success',
            'id': allocation.id,
            'approval_status': allocation.approval_status
        })

    @action(detail=True, methods=['put'], url_path='reject')
    def reject(self, request, pk=None):
        allocation = self.get_object()
        allocation.approval_status = 'rejected'
        allocation.save()
        return Response({
            'status': 'rejected',
            'id': allocation.id,
            'approval_status': allocation.approval_status
        }, status=status.HTTP_200_OK)



from .serializers import EmployeeWithStatusSerializer

class EmployeeMachineAllocationViewSet(viewsets.ModelViewSet):
    queryset = MachineAllocation.objects.all()
    serializer_class = ...  # your main MachineAllocation serializer

    @action(detail=False, methods=['get'], url_path='eligible-employees')
    def eligible_employees(self, request):
        machine_id = request.query_params.get('machine_id')
        if not machine_id:
            return Response({'error': 'machine_id is required'}, status=400)

        try:
            machine = Machine.objects.get(id=machine_id)
        except Machine.DoesNotExist:
            return Response({'error': 'Machine not found'}, status=404)

        # matching_skills = OperatorSkill.objects.filter(station__skill=machine.process) # skilmatrix table
        # employee_ids = matching_skills.values_list('operator_id', flat=True).distinct()
        # employees = MasterTable.objects.filter(id__in=employee_ids)

        # serializer = EmployeeWithStatusSerializer(employees, many=True, context={'machine_id': machine_id})
        # return Response(serializer.data)


# =========================== Machine Allocation Approval =============================== # 

from rest_framework import viewsets
from .models import OJTTopic
from .serializers import OJTTopicSerializer

class OJTTopicViewSet(viewsets.ModelViewSet):
    queryset = OJTTopic.objects.all().order_by("sl_no")
    serializer_class = OJTTopicSerializer





from rest_framework import viewsets
from .models import OJTDay
from .serializers import OJTDaySerializer

class OJTDayViewSet(viewsets.ModelViewSet):
    queryset = OJTDay.objects.all().order_by("id")
    serializer_class = OJTDaySerializer









from rest_framework import viewsets
from .models import OJTScore
from .serializers import OJTScoreSerializer

class OJTScoreViewSet(viewsets.ModelViewSet):
    queryset = OJTScore.objects.all()
    serializer_class = OJTScoreSerializer

from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework.decorators import action
from .models import OJTScoreRange
from .serializers import OJTScoreRangeSerializer


class OJTScoreRangeViewSet(viewsets.ModelViewSet):
    queryset = OJTScoreRange.objects.all().order_by("id")
    serializer_class = OJTScoreRangeSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        department_id = self.request.query_params.get("department")
        level_id = self.request.query_params.get("level")

        if department_id:
            queryset = queryset.filter(department_id=department_id)
        if level_id:
            queryset = queryset.filter(level_id=level_id)

        return queryset

from rest_framework import viewsets
from .models import OJTPassingCriteria
from .serializers import OJTPassingCriteriaSerializer


class OJTPassingCriteriaViewSet(viewsets.ModelViewSet):
    queryset = OJTPassingCriteria.objects.all()
    serializer_class = OJTPassingCriteriaSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        department_id = self.request.query_params.get("department")
        level_id = self.request.query_params.get("level")

        if department_id:
            queryset = queryset.filter(department_id=department_id)
        if level_id:
            queryset = queryset.filter(level_id=level_id)

        return queryset

from rest_framework import viewsets, status
from rest_framework.response import Response
from .models import TraineeInfo
from .serializers import TraineeInfoSerializer


from django.core.exceptions import ObjectDoesNotExist
from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
from .models import TraineeInfo, OJTScore, OJTTopic, OJTScoreRange, OJTPassingCriteria, Station
from .serializers import TraineeInfoSerializer

from django.db import transaction
from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
from django.db.models import Sum

from .models import (
    TraineeInfo, OJTScore, MasterTable, Level,
    OJTDay, OJTScoreRange, OJTPassingCriteria, OJTTopic
)
from .serializers import TraineeInfoSerializer
from .signals import run_after_delay, update_skill_matrix


class TraineeInfoViewSet(viewsets.ModelViewSet):
    queryset = TraineeInfo.objects.all()
    serializer_class = TraineeInfoSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        emp_id = self.request.query_params.get('emp_id', '').strip()
        station_id = self.request.query_params.get('station_id')
        level_id = self.request.query_params.get('level_id')

        print(f"[GET] Params → emp_id={emp_id}, station={station_id}, level={level_id}")

        if emp_id:
            queryset = queryset.filter(emp_id__iexact=emp_id)
        if station_id:
            queryset = queryset.filter(station_id=station_id)
        if level_id:
            queryset = queryset.filter(level_id=level_id)

        print(f"[GET] → {queryset.count()} record(s)")
        return queryset

    def create(self, request, *args, **kwargs):
        print(f"[POST] Raw data → {request.data}")

        # ✅ Validate emp_id first
        emp_id = request.data.get('emp_id', '').strip()
        if not emp_id:
            print("[ERROR] emp_id missing or empty")
            return Response(
                {"emp_id": "Employee ID is required and cannot be empty."},
                status=status.HTTP_400_BAD_REQUEST
            )

        station_id = request.data.get('station')
        level_id = request.data.get('level')

        print(f"[UPSERT CHECK] emp_id={emp_id}, station_id={station_id}, level_id={level_id}")

        # ✅ Check if record exists BEFORE validation
        existing = TraineeInfo.objects.filter(
            emp_id=emp_id,
            station_id=station_id,
            level_id=level_id
        ).first()

        if existing:
            print(f"[UPSERT] Updating existing ID={existing.id}")
            # ✅ Use update logic with existing instance
            serializer = self.get_serializer(existing, data=request.data, partial=False)
            if not serializer.is_valid():
                print(f"[ERROR] Validation failed → {serializer.errors}")
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
            return self._update_and_respond(existing, serializer)
        else:
            print(f"[UPSERT] Creating new record")
            # ✅ Use create logic
            serializer = self.get_serializer(data=request.data)
            if not serializer.is_valid():
                print(f"[ERROR] Validation failed → {serializer.errors}")
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
            trainee = serializer.save()
            self._post_save(trainee)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        partial = kwargs.get('partial', False)
        instance = self.get_object()

        incoming_level = request.data.get('level')
        if incoming_level and int(incoming_level) != instance.level_id:
            return Response({
                "error": "Cannot change level",
                "detail": f"Use POST to create new level record.",
            }, status=status.HTTP_400_BAD_REQUEST)

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        return self._update_and_respond(instance, serializer)

    def _update_and_respond(self, instance, serializer):
        scores_data = serializer.validated_data.pop('scores', [])

        # Update scalar fields
        for attr, value in serializer.validated_data.items():
            setattr(instance, attr, value)
        instance.save()

        # Replace scores
        instance.scores.all().delete()
        for score_data in scores_data:
            OJTScore.objects.create(trainee=instance, **score_data)

        self._post_save(instance)
        return Response(self.get_serializer(instance).data)
    # ------------------------------------------------------------------
    # 5. POST-SAVE
    # ------------------------------------------------------------------
    def _post_save(self, trainee):
        self.perform_recalculate_status(trainee)
        self.trigger_skill_matrix_update(trainee)

    # ------------------------------------------------------------------
    # 6. RECALCULATE STATUS
    # ------------------------------------------------------------------
    def perform_recalculate_status(self, trainee):
        first_score = trainee.scores.first()
        if not first_score:
            trainee.status = "Pending"
            trainee.save(update_fields=["status"])
            return

        department = first_score.topic.department
        level = first_score.topic.level
        level_id = self._get_id(level)

        required_days = OJTDay.objects.filter(department=department, level_id=level_id).distinct()
        if not required_days.exists():
            trainee.status = "Pending"
            trainee.save(update_fields=["status"])
            return

        try:
            score_range = OJTScoreRange.objects.get(department=department, level_id=level_id)
            max_topic_score = score_range.max_score
        except OJTScoreRange.DoesNotExist:
            trainee.status = "Pending"
            trainee.save(update_fields=["status"])
            return

        all_passed = True
        for day in required_days:
            criteria = (
                OJTPassingCriteria.objects.filter(department=department, level_id=level_id, day=day).first()
                or OJTPassingCriteria.objects.filter(department=department, level_id=level_id, day__isnull=True).first()
            )
            if not criteria:
                trainee.status = "Pending"
                trainee.save(update_fields=["status"])
                return

            required_pct = criteria.percentage
            topics = OJTTopic.objects.filter(department=department, level_id=level_id)
            total_topics = topics.count()

            actual = OJTScore.objects.filter(trainee=trainee, day=day, topic__in=topics).aggregate(total=Sum("score"))["total"] or 0
            count = OJTScore.objects.filter(trainee=trainee, day=day, topic__in=topics).count()

            if count < total_topics:
                trainee.status = "Pending"
                trainee.save(update_fields=["status"])
                return

            max_possible = total_topics * max_topic_score
            pct = (actual / max_possible) * 100 if max_possible > 0 else 0

            if pct < required_pct:
                all_passed = False
                break

        trainee.status = "Pass" if all_passed else "Fail"
        trainee.save(update_fields=["status"])

    # ------------------------------------------------------------------
    # 7. SKILL MATRIX – ONLY IF VALID
    # ------------------------------------------------------------------
    def trigger_skill_matrix_update(self, trainee):
        if not trainee.emp_id or trainee.emp_id.strip() == '':
            print(f"[SKILL MATRIX] SKIPPED – emp_id missing (ID={trainee.id})")
            return

        if trainee.status != "Pass":
            print(f"[SKILL MATRIX] SKIPPED – status={trainee.status} (emp_id={trainee.emp_id})")
            return

        try:
            employee = MasterTable.objects.get(emp_id=trainee.emp_id.strip())
            run_after_delay(
                update_skill_matrix,
                5,
                employee,
                trainee.station,
                trainee.level,
                True
            )
            print(f"[SKILL MATRIX] TRIGGERED → emp_id={trainee.emp_id}")
        except MasterTable.DoesNotExist:
            print(f"[SKILL MATRIX] SKIPPED – No MasterTable for emp_id={trainee.emp_id}")
        except Exception as e:
            print(f"[SKILL MATRIX] ERROR → {e}")

    # ------------------------------------------------------------------
    # HELPER: Extract ID
    # ------------------------------------------------------------------
    def _get_id(self, obj_or_int):
        if hasattr(obj_or_int, 'pk'):
            return obj_or_int.pk
        if hasattr(obj_or_int, 'id'):
            return obj_or_int.id
        return int(obj_or_int)




# views.py
from rest_framework.generics import ListAPIView
from .models import OJTTopic
from .serializers import OJTTopicSerializer

class OJTTopicListView(ListAPIView):
    serializer_class = OJTTopicSerializer

    def get_queryset(self):
        queryset = OJTTopic.objects.all()
        department_id = self.request.query_params.get("department")
        level_id = self.request.query_params.get("level")

        if department_id:
            queryset = queryset.filter(department_id=department_id)
        if level_id:
            queryset = queryset.filter(level_id=level_id)

        return queryset
    


from rest_framework.generics import ListAPIView
from .models import OJTDay
from .serializers import OJTDaySerializer

class OJTDayListView(ListAPIView):
    serializer_class = OJTDaySerializer

    def get_queryset(self):
        queryset = OJTDay.objects.all()
        department_id = self.request.query_params.get("department")
        level_id = self.request.query_params.get("level")

        if department_id:
            queryset = queryset.filter(department_id=department_id)
        if level_id:
            queryset = queryset.filter(level_id=level_id)

        return queryset
    

from rest_framework import generics
from .models import TraineeInfo
from .serializers import TraineeInfoSerializer

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

class TraineeInfoListView(APIView):
    def get(self, request):
        emp_id = request.query_params.get('emp_id', '').strip()
        station_id = request.query_params.get('station_id')
        level_id = request.query_params.get('level_id')
        
        print(f"[LIST VIEW] Filtering: emp_id='{emp_id}', station={station_id}, level={level_id}")
        
        queryset = TraineeInfo.objects.all()
        
        if emp_id:
            queryset = queryset.filter(emp_id__iexact=emp_id)
        if station_id:
            queryset = queryset.filter(station_id=station_id)
        if level_id:
            queryset = queryset.filter(level_id=level_id)
        
        print(f"[LIST VIEW] Found {queryset.count()} records")
        
        serializer = TraineeInfoSerializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

from rest_framework import viewsets
from .models import QuantityOJTScoreRange, QuantityPassingCriteria
from .serializers import QuantityOJTScoreRangeSerializer, QuantityPassingCriteriaSerializer


class QuantityOJTScoreRangeViewSet(viewsets.ModelViewSet):
    queryset = QuantityOJTScoreRange.objects.all()
    serializer_class = QuantityOJTScoreRangeSerializer


class QuantityPassingCriteriaViewSet(viewsets.ModelViewSet):
    queryset = QuantityPassingCriteria.objects.all()
    serializer_class = QuantityPassingCriteriaSerializer

    
from rest_framework import viewsets
from .models import OJTLevel2Quantity, Level
from .serializers import OJTLevel2QuantitySerializer, LevelSerializer


class OJTLevel2QuantityViewSet(viewsets.ModelViewSet):
    """
    API endpoint for OJT Level 2 Quantity with nested evaluations.
    Supports filtering by trainee_id, level, and station.
    """
    queryset = OJTLevel2Quantity.objects.all()
    serializer_class = OJTLevel2QuantitySerializer

    def get_queryset(self):
        queryset = super().get_queryset()

        trainee_id = self.request.query_params.get("trainee_id")
        level_id = self.request.query_params.get("level")
        station_name = self.request.query_params.get("station")

        if trainee_id:
            queryset = queryset.filter(trainee_id=trainee_id)
        if level_id:
            queryset = queryset.filter(level_id=level_id)
        if station_name:
            queryset = queryset.filter(station_name__iexact=station_name)  # case-insensitive match

        return queryset


# =================== Refreshment Training ============================= #

from rest_framework import viewsets
from .models import Training_category, Curriculum, CurriculumContent, Trainer_name, Venues, Schedule, MasterTable , RescheduleLog,EmployeeAttendance
from .serializers import Training_categorySerializer, CurriculumSerializer, CurriculumContentSerializer, Trainer_nameSerializer, VenuesSerializer, ScheduleSerializer, MasterTableSerializer, EmployeeAttendanceSerializer, RescheduleLogSerializer

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from datetime import timedelta
from dateutil.relativedelta import relativedelta
from datetime import date


# ... (keep existing imports)

# class RecurrenceIntervalViewSet(viewsets.ModelViewSet):
#     """
#     ViewSet for managing recurrence intervals
#     """
#     queryset = RecurrenceInterval.objects.all()
#     serializer_class = RecurrenceIntervalSerializer

#     @action(detail=False, methods=['get'])
#     def active(self, request):
#         """Get the currently active recurrence interval"""
#         active_interval = RecurrenceInterval.objects.filter(is_active=True).first()
#         if active_interval:
#             serializer = self.get_serializer(active_interval)
#             return Response(serializer.data)
#         return Response(
#             {"detail": "No active recurrence interval found"}, 
#             status=status.HTTP_404_NOT_FOUND
#         )


class Training_categoryViewSet(viewsets.ModelViewSet):
    queryset = Training_category.objects.all()
    serializer_class = Training_categorySerializer

class CurriculumViewSet(viewsets.ModelViewSet):
    serializer_class = CurriculumSerializer

    def get_queryset(self):
        queryset = Curriculum.objects.all()
        category_id = self.request.query_params.get('category_id')
        if category_id is not None:
            queryset = queryset.filter(category_id=category_id)
        return queryset

class CurriculumContentViewSet(viewsets.ModelViewSet):
    serializer_class = CurriculumContentSerializer

    def get_queryset(self):
        queryset = CurriculumContent.objects.all()
        curriculum_id = self.request.query_params.get('curriculum')
        if curriculum_id is not None:
            queryset = queryset.filter(curriculum_id=curriculum_id)
        return queryset

class Trainer_nameViewSet(viewsets.ModelViewSet):
    queryset = Trainer_name.objects.all()
    serializer_class = Trainer_nameSerializer

class VenueViewSet(viewsets.ModelViewSet):
    queryset = Venues.objects.all()
    serializer_class = VenuesSerializer

class ScheduleViewSet(viewsets.ModelViewSet):
    queryset = Schedule.objects.all()
    serializer_class = ScheduleSerializer    

    def create(self, request, *args, **kwargs):
        """Override create to handle recurrence"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Set is_recurring based on recurrence_months
        recurrence_months = request.data.get('recurrence_months')
        if recurrence_months:
            serializer.validated_data['is_recurring'] = True
        
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(
            serializer.data, 
            status=status.HTTP_201_CREATED, 
            headers=headers
        )
    
    @action(detail=True, methods=['post'])
    def start_training(self, request, pk=None):
        """
        Start training session - Moves to Stage 1 (Pre-Test)
        Current logic: requires all employees to have attendance marked.
        (We can relax this later when fully switching to batch-based flow.)
        """
        schedule = self.get_object()
        
        # Attendance Check
        total_employees = schedule.employees.count()
        marked_attendance = EmployeeAttendance.objects.filter(schedule=schedule).count()
        
        if marked_attendance == 0:
            return Response(
                {"error": "Cannot start training. Please mark attendance for at least one employee."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Initialize Stage if not started
        if schedule.training_stage == 0:
            schedule.training_stage = 1  # Move to Pre-Test
            schedule.status = 'pending'  # Training is live but not finished
            schedule.save()
        
        return Response({
            "message": "Training started",
            "stage": schedule.training_stage,
            "status": schedule.status
        })

    @action(detail=True, methods=['post'])
    def advance_stage(self, request, pk=None):
        """
        Move to next stage:
        1 (Pre) -> 2 (Content)
        2 (Content) -> 3 (Post)
        3 (Post) -> 4 (Completed)
        """
        schedule = self.get_object()
        current = schedule.training_stage
        
        if current == 1:
            schedule.training_stage = 2
            message = "Moved to Curriculum Phase"
        elif current == 2:
            schedule.training_stage = 3
            message = "Moved to Post-Test Phase"
        elif current == 3:
            schedule.training_stage = 4
            schedule.status = 'completed'
            schedule.completed_date = date.today()
            message = "Training Completed"
        else:
            return Response({"error": "Cannot advance stage"}, status=status.HTTP_400_BAD_REQUEST)
            
        schedule.save()
        return Response({"message": message, "stage": schedule.training_stage})
    
    @action(detail=True, methods=['patch'])
    def update_next_training_date(self, request, pk=None):
        """Update the next training date manually (for holidays, etc.)"""
        schedule = self.get_object()
        
        new_date = request.data.get('next_training_date')
        if not new_date:
            return Response(
                {"error": "next_training_date is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from datetime import datetime
            parsed_date = datetime.strptime(new_date, '%Y-%m-%d').date()
            
            schedule.next_training_date_override = parsed_date
            schedule.save()
            
            serializer = self.get_serializer(schedule)
            return Response(serializer.data)
        except ValueError:
            return Response(
                {"error": "Invalid date format. Use YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['get'])
    def check_attendance_status(self, request, pk=None):
        """Check if all employees have attendance marked"""
        schedule = self.get_object()
        total_employees = schedule.employees.count()
        marked_attendance = EmployeeAttendance.objects.filter(schedule=schedule).count()
        
        return Response({
            "total_employees": total_employees,
            "marked_attendance": marked_attendance,
            "all_marked": marked_attendance >= total_employees,
            "can_start_training": marked_attendance >= total_employees
        })

    @action(detail=True, methods=['get'])
    def batches(self, request, pk=None):
        """
        List all refresher batches for this schedule.
        GET /schedules/{id}/batches/
        """
        schedule = self.get_object()
        batches = RefresherBatch.objects.filter(
            schedule=schedule
        ).prefetch_related('employees')
        serializer = RefresherBatchSerializer(batches, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def create_batch(self, request, pk=None):
        """
        Create a new batch for this schedule from employees marked as 'present'.

        Expected body:
        {
          "employee_ids": ["EMP001", "EMP002", ...]  // MasterTable.emp_id values
        }
        """
        schedule = self.get_object()
        employee_ids = request.data.get('employee_ids', [])

        if not employee_ids or not isinstance(employee_ids, list):
            return Response(
                {"error": "employee_ids (list) is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Fetch employees by emp_id
        employees = list(MasterTable.objects.filter(emp_id__in=employee_ids))
        found_ids = {str(e.emp_id) for e in employees}
        missing = [eid for eid in employee_ids if str(eid) not in found_ids]
        if missing:
            return Response(
                {"error": f"Employees not found: {', '.join(missing)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Ensure all are marked 'present' in EmployeeAttendance for this schedule
        present_attendances = EmployeeAttendance.objects.filter(
            schedule=schedule,
            employee__in=employees,
            status='present'
        )
        if present_attendances.count() != len(employees):
            return Response(
                {"error": "All employees must be marked 'present' for this schedule before creating a batch."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Ensure these employees are not already in any batch for this schedule
        already_in_batch = RefresherBatch.objects.filter(
            schedule=schedule,
            employees__in=employees
        ).distinct()
        if already_in_batch.exists():
            return Response(
                {"error": "One or more employees are already assigned to a batch for this schedule."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Determine next batch number
        existing_count = RefresherBatch.objects.filter(schedule=schedule).count()
        batch_number = existing_count + 1

        # Build a readable name: "<Category Name> - Batch X"
        category_name = schedule.training_category.name if schedule.training_category else "Batch"
        batch_name = f"{category_name} - Batch {batch_number}"

        batch = RefresherBatch.objects.create(
            schedule=schedule,
            batch_number=batch_number,
            name=batch_name,
        )
        batch.employees.add(*employees)

        serializer = RefresherBatchSerializer(batch)
        return Response(serializer.data, status=status.HTTP_201_CREATED)




    # =================== Employee Template Upload/Download ============================= #
from rest_framework.decorators import api_view
from django.utils import timezone
from datetime import date, timedelta
from dateutil.relativedelta import relativedelta
from django.views.decorators.cache import never_cache

# @api_view(['GET'])
# @never_cache
# def recurring_schedules_info(request):
#     """
#     Get information about all recurring schedules and their next recurrence dates
#     """
#     # Get all completed recurring schedules
#     recurring_schedules = Schedule.objects.filter(
#         status='completed',
#         is_recurring=True,
#         recurrence_interval__isnull=False,
#         completed_date__isnull=False
#     ).select_related(
#         'training_category',
#         'training_name',
#         'trainer',
#         'venue',
#         'recurrence_interval'
#     ).prefetch_related('employees')
    
#     result = []
    
#     for schedule in recurring_schedules:
#         # Calculate next training date
#         next_training_date = schedule.date + relativedelta(
#             months=schedule.recurrence_interval.interval_months
#         )
        
#         # Calculate creation trigger date (7 days before next training)
#         next_training_creation_date = next_training_date - timedelta(days=7)
        
#         result.append({
#             'id': schedule.id,
#             'training_category': schedule.training_category.name,
#             'training_name': schedule.training_name.topic,
#             'original_date': schedule.date,
#             'completed_date': schedule.completed_date,
#             'next_training_date': next_training_date,
#             'next_training_creation_date': next_training_creation_date,
#             'interval_months': schedule.recurrence_interval.interval_months,
#             'recurring_schedule_created': schedule.recurring_schedule_created,
#             'trainer': schedule.trainer.name if schedule.trainer else 'TBD',
#             'venue': schedule.venue.name if schedule.venue else 'TBD',
#             'employee_count': schedule.employees.count(),
#             'status': 'completed'
#         })
    
#     return Response(result, status=status.HTTP_200_OK)
# ---------------------------------------------------------------------------------------------------------
# @api_view(['GET'])
# @never_cache
# def recurring_schedules_info(request):
#     """Get information about all recurring schedules and their next recurrence dates"""
#     recurring_schedules = Schedule.objects.filter(
#         status='completed',
#         is_recurring=True,
#         recurrence_months__isnull=False,
#         completed_date__isnull=False
#     ).select_related(
#         'training_category',
#         'training_name',
#         'trainer',
#         'venue'
#     ).prefetch_related('employees')
    
#     result = []
    
#     for schedule in recurring_schedules:
#         next_training_date = schedule.get_next_training_date()
#         next_creation_date = schedule.get_next_creation_date()
        
#         result.append({
#             'id': schedule.id,
#             'training_category': schedule.training_category.name,
#             'training_name': schedule.training_name.topic,
#             'original_date': schedule.date,
#             'completed_date': schedule.completed_date,
#             'next_training_date': next_training_date,
#             'next_training_creation_date': next_creation_date,
#             'recurrence_months': schedule.recurrence_months,
#             'recurring_schedule_created': schedule.recurring_schedule_created,
#             'trainer': schedule.trainer.name if schedule.trainer else 'TBD',
#             'venue': schedule.venue.name if schedule.venue else 'TBD',
#             'employee_count': schedule.employees.count(),
#             'status': 'completed',
#             'has_override': schedule.next_training_date_override is not None,
#             'next_training_date_override': schedule.next_training_date_override
#         })
    
#     return Response(result, status=status.HTTP_200_OK)


# @api_view(['POST'])
# def test_recurring_schedule_creation(request):
#     """
#     Manual trigger to test recurring schedule creation
#     Useful for debugging Celery issues
#     """
#     from django.core.management import call_command
#     from io import StringIO
    
#     # Capture the command output
#     out = StringIO()
    
#     try:
#         call_command('create_recurring_schedules', stdout=out)
#         output = out.getvalue()
        
#         return Response({
#             'success': True,
#             'message': 'Recurring schedule creation command executed',
#             'output': output
#         }, status=status.HTTP_200_OK)
#     except Exception as e:
#         return Response({
#             'success': False,
#             'error': str(e)
#         }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['GET'])
# def celery_status(request):
#     """
#     Check Celery worker and task status
#     """
#     from celery import current_app
    
#     try:
#         # Check if Celery is running
#         inspector = current_app.control.inspect()
#         active_workers = inspector.active()
#         scheduled_tasks = inspector.scheduled()
        
#         # Get the beat schedule
#         beat_schedule = current_app.conf.beat_schedule
        
#         return Response({
#             'celery_running': active_workers is not None,
#             'active_workers': list(active_workers.keys()) if active_workers else [],
#             'scheduled_tasks_count': sum(len(tasks) for tasks in (scheduled_tasks or {}).values()),
#             'beat_schedule': {
#                 'create_recurring_schedules': {
#                     'task': beat_schedule.get('create-recurring-schedules-daily', {}).get('task'),
#                     'schedule': str(beat_schedule.get('create-recurring-schedules-daily', {}).get('schedule'))
#                 }
#             } if beat_schedule else {}
#         }, status=status.HTTP_200_OK)
#     except Exception as e:
#         return Response({
#             'celery_running': False,
#             'error': str(e),
#             'message': 'Celery might not be running. Start it with: celery -A yourproject worker -l info -B'
#         }, status=status.HTTP_200_OK)

# app1/views.py

from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from datetime import date

# ... your existing imports

@api_view(['GET'])
def scheduler_status(request):
    """
    Check APScheduler status and show scheduled tasks
    """
    from app1.scheduler import get_scheduler_status
    
    # Get scheduler info
    scheduler_info = get_scheduler_status()
    
    # Get upcoming training schedules that will be created
    recurring_schedules = Schedule.objects.filter(
        status='completed',
        is_recurring=True,
        recurrence_months__isnull=False,
        completed_date__isnull=False
    ).select_related('training_name', 'training_category')
    
    next_runs = []
    for schedule in recurring_schedules:
        next_date = schedule.get_next_training_date()
        creation_date = schedule.get_next_creation_date()
        
        if next_date and creation_date:
            next_runs.append({
                'schedule_id': schedule.id,
                'training_category': schedule.training_category.name,
                'training_name': schedule.training_name.topic,
                'original_date': str(schedule.date),
                'completed_date': str(schedule.completed_date),
                'next_training_date': str(next_date),
                'will_be_created_on': str(creation_date),
                'days_until_creation': (creation_date - date.today()).days,
                'recurrence_months': schedule.recurrence_months,
            })
    
    return Response({
        'scheduler_type': 'APScheduler (Built-in)',
        'scheduler_running': scheduler_info['running'],
        'scheduled_jobs': scheduler_info['jobs'],
        'upcoming_training_schedules': sorted(
            next_runs, 
            key=lambda x: x['will_be_created_on']
        ),
        'instructions': [
            'Scheduler runs automatically with Django',
            'No separate commands needed',
            'Task executes daily at 2:00 AM',
            'Just run: python manage.py runserver'
        ],
        'status': 'healthy' if scheduler_info['running'] else 'stopped'
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
def celery_status(request):
    """
    Show scheduler status (kept for backward compatibility)
    This endpoint exists for legacy reasons
    """
    from app1.scheduler import get_scheduler_status
    
    scheduler_info = get_scheduler_status()
    
    return Response({
        'scheduler_type': 'APScheduler (Built-in)',
        'scheduler_running': scheduler_info['running'],
        'celery_running': False,  # For backward compatibility
        'message': 'Using APScheduler - runs automatically with Django',
        'instructions': [
            'No separate commands needed',
            'Just run: python manage.py runserver',
            'Scheduler starts automatically',
            'Check detailed status: GET /api/schedules/scheduler_status/'
        ]
    }, status=status.HTTP_200_OK)


@api_view(['POST'])
def test_recurring_schedule_creation(request):
    """
    Manual trigger to test recurring schedule creation
    Useful for testing without waiting for scheduled time
    """
    from app1.scheduler import trigger_job_now
    
    try:
        success, message = trigger_job_now()
        
        if success:
            return Response({
                'success': True,
                'message': 'Recurring schedule creation job executed successfully',
                'details': message
            }, status=status.HTTP_200_OK)
        else:
            return Response({
                'success': False,
                'error': message
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def recurring_schedules_info(request):
    """
    Get information about all recurring schedules and their next recurrence dates
    This endpoint remains unchanged
    """
    recurring_schedules = Schedule.objects.filter(
        status='completed',
        is_recurring=True,
        recurrence_months__isnull=False,
        completed_date__isnull=False
    ).select_related(
        'training_category',
        'training_name',
        'trainer',
        'venue'
    ).prefetch_related('employees')
    
    result = []
    
    for schedule in recurring_schedules:
        next_training_date = schedule.get_next_training_date()
        next_creation_date = schedule.get_next_creation_date()
        
        result.append({
            'id': schedule.id,
            'training_category': schedule.training_category.name,
            'training_name': schedule.training_name.topic,
            'original_date': schedule.date,
            'completed_date': schedule.completed_date,
            'next_training_date': next_training_date,
            'next_training_creation_date': next_creation_date,
            'recurrence_months': schedule.recurrence_months,
            'recurring_schedule_created': schedule.recurring_schedule_created,
            'trainer': schedule.trainer.name if schedule.trainer else 'TBD',
            'venue': schedule.venue.name if schedule.venue else 'TBD',
            'employee_count': schedule.employees.count(),
            'status': 'completed',
            'has_override': schedule.next_training_date_override is not None,
            'next_training_date_override': schedule.next_training_date_override
        })
    
    return Response(result, status=status.HTTP_200_OK)
    
import csv
import io
from django.http import HttpResponse
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser
from django.db import transaction

@api_view(['GET'])
def download_employee_template(request):
    """
    Download CSV template with headers: Employee ID, Employee Name, Department
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="employee_template.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Employee ID', 'Employee Name', 'Department'])
    
    return response

@api_view(['POST'])
def upload_employees_csv(request):
    """
    Upload CSV file with employee data and validate against MasterTable
    """
    if 'file' not in request.FILES:
        return Response(
            {"error": "No file provided"}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    csv_file = request.FILES['file']
    
    if not csv_file.name.endswith('.csv'):
        return Response(
            {"error": "File must be a CSV"}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        decoded_file = csv_file.read().decode('utf-8')
        io_string = io.StringIO(decoded_file)
        reader = csv.DictReader(io_string)
        
        valid_employees = []
        errors = []
        
        for row_num, row in enumerate(reader, start=2):  # Start at 2 (header is row 1)
            employee_id = row.get('Employee ID', '').strip()
            employee_name = row.get('Employee Name', '').strip()
            department = row.get('Department', '').strip()
            
            if not employee_id:
                errors.append(f"Row {row_num}: Employee ID is required")
                continue
            
            # Try to find employee in MasterTable by emp_id
            try:
                employee = MasterTable.objects.get(emp_id=employee_id)
                
                # Validate name matches (optional - you can remove this if not needed)
                db_full_name = f"{employee.first_name or ''} {employee.last_name or ''}".strip()
                if employee_name and db_full_name.lower() != employee_name.lower():
                    errors.append(
                        f"Row {row_num}: Name mismatch for ID {employee_id}. "
                        f"Expected '{db_full_name}', got '{employee_name}'"
                    )
                    continue
                
                valid_employees.append({
                    'id': employee.emp_id,
                    'employee_code': employee.emp_id,
                    'full_name': db_full_name,
                    'department': department,
                    'first_name': employee.first_name,
                    'last_name': employee.last_name
                })
                
            except MasterTable.DoesNotExist:
                errors.append(f"Row {row_num}: Employee ID '{employee_id}' not found in system")
            except MasterTable.MultipleObjectsReturned:
                errors.append(f"Row {row_num}: Multiple employees found with ID '{employee_id}'")
        
        if errors and not valid_employees:
            return Response(
                {
                    "error": "No valid employees found",
                    "details": errors
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return Response({
            "valid_employees": valid_employees,
            "total_valid": len(valid_employees),
            "total_errors": len(errors),
            "errors": errors
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response(
            {"error": f"Error processing file: {str(e)}"}, 
            status=status.HTTP_400_BAD_REQUEST
        )

# =================== Employee Template Upload/Download End ============================= #

from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action

from .models import EmployeeAttendance, RescheduleLog
from .serializers import EmployeeAttendanceSerializer, RescheduleLogSerializer

class EmployeeAttendanceViewSet(viewsets.ModelViewSet):
    queryset = EmployeeAttendance.objects.all()
    serializer_class = EmployeeAttendanceSerializer

    # def partial_update is newly added 
    def partial_update(self, request, *args, **kwargs):
        """
        PATCH endpoint to update attendance record with reschedule details
        """
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        
        # If updating to rescheduled status with details, create a log
        if request.data.get('reschedule_date') and request.data.get('reschedule_reason'):
            # CHANGED: Force reset to 'pending' in save()
            serializer.save(status='pending')
        else:
            self.perform_update(serializer)
        
        return Response(serializer.data)
   
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        attendance_instance = serializer.save()

        print(f"Attendance created: status={attendance_instance.status}")

        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'])
    def bulk_save_attendance(self, request):
        """
        Bulk save/update attendance for multiple employees
        """
        schedule_id = request.data.get('schedule_id')
        attendances_data = request.data.get('attendances', [])
        
        if not schedule_id:
            return Response(
                {"error": "schedule_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not attendances_data:
            return Response(
                {"error": "attendances list is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            schedule = Schedule.objects.get(id=schedule_id)
        except Schedule.DoesNotExist:
            return Response(
                {"error": "Schedule not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        
        saved_count = 0
        errors = []
        
        with transaction.atomic():
            for attendance_data in attendances_data:
                employee_id = attendance_data.get('employee_id')
                attendance_status = attendance_data.get('status')
                
                if not employee_id or not attendance_status:
                    errors.append(f"Missing employee_id or status for entry")
                    continue
                
                try:
                    employee = MasterTable.objects.get(emp_id=employee_id)
                except MasterTable.DoesNotExist:
                    errors.append(f"Employee {employee_id} not found")
                    continue
                
                # Create or update attendance
                attendance, created = EmployeeAttendance.objects.update_or_create(
                    schedule=schedule,
                    employee=employee,
                    defaults={
                        'status': attendance_status,
                        'notes': attendance_data.get('notes', ''),
                        # Don't overwrite reschedule details if they exist
                    }
                )
                # If creating new or updating basic status, ensure reschedule fields are clean if status is present
                if attendance_status == 'present':
                     attendance.reschedule_date = None
                     attendance.reschedule_time = None
                     attendance.reschedule_reason = None
                     attendance.save()
                
                saved_count += 1
        
        return Response({
            "message": f"Successfully saved {saved_count} attendance records",
            "saved_count": saved_count,
            "errors": errors
        }, status=status.HTTP_200_OK if not errors else status.HTTP_207_MULTI_STATUS)

    @action(detail=False, methods=['get'])
    def by_schedule(self, request):
        """
        Get attendance records for a specific schedule
        FIX: Uses prefetch_related for topics instead of select_related for training_name
        """
        schedule_id = request.query_params.get('schedule_id')
        if not schedule_id:
            return Response(
                {"detail": "schedule_id query parameter is required."}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        attendances = self.queryset.filter(schedule_id=schedule_id).select_related(
            'schedule',
            'employee',
            'schedule__training_category',
            'schedule__trainer',
            'schedule__venue'
        ).prefetch_related(
            'schedule__topics'
        )
        serializer = self.get_serializer(attendances, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def all_reschedules(self, request):
        """
        Get all reschedule records (both pending and completed)
        FIX: Updated to support multiple topics
        """
        all_records = EmployeeAttendance.objects.filter(
            status='rescheduled'  # CHANGED: Exclude 'absent' per user request
        ).select_related(
            'schedule',
            'employee',
            'schedule__training_category',
            'schedule__trainer',
            'schedule__venue'
        ).prefetch_related(
            'schedule__topics'
        ).order_by('-schedule__date')

        pending = []
        completed = []

        for record in all_records:
            serialized = EmployeeAttendanceSerializer(record).data
            if not record.reschedule_date or not record.reschedule_reason:
                pending.append(serialized)
            else:
                completed.append(serialized)

        return Response({
            'pending': pending,
            'completed': completed
        })
    
    @action(detail=False, methods=['get'])
    def pending_reschedule(self, request):
        """
        Get only pending reschedule records
        FIX: Updated to support multiple topics
        """
        pending = EmployeeAttendance.objects.filter(
            status__in=['absent', 'rescheduled']
        ).filter(
            models.Q(reschedule_date__isnull=True) | 
            models.Q(reschedule_reason__isnull=True) |
            models.Q(reschedule_reason='')
        ).select_related(
            'schedule',
            'employee',
            'schedule__training_category',
            'schedule__trainer',
            'schedule__venue'
        ).prefetch_related(
            'schedule__topics'
        ).order_by('-schedule__date')
        
        serializer = self.get_serializer(pending, many=True)
        return Response(serializer.data)

    def update(self, request, pk=None, **kwargs):
        """
        Standard PUT update
        """
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=False)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)

    def partial_update(self, request, pk=None, **kwargs):
        """
        PATCH update - for updating reschedule details
        """
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        
        if instance.status in ['absent', 'rescheduled']:
            reschedule_date = request.data.get('reschedule_date', instance.reschedule_date)
            reschedule_time = request.data.get('reschedule_time', instance.reschedule_time)
            reschedule_reason = request.data.get('reschedule_reason', instance.reschedule_reason)

            if not all([reschedule_date, reschedule_time, reschedule_reason]):
                return Response(
                    {"error": "reschedule_date, reschedule_time, and reschedule_reason are all required"},
                    status=status.HTTP_400_BAD_REQUEST
                )

        self.perform_update(serializer)

        # Create RescheduleLog if all details are present
        if (instance.status in ['absent', 'rescheduled'] and
            instance.reschedule_date and 
            instance.reschedule_time and 
            instance.reschedule_reason):
            try:
                RescheduleLog.objects.get_or_create(
                    schedule=instance.schedule,
                    employee=instance.employee,
                    defaults={
                        'original_date': instance.schedule.date,
                        'original_time': instance.schedule.time,
                        'new_date': instance.reschedule_date,
                        'new_time': instance.reschedule_time,
                        'reason': instance.reschedule_reason,
                    }
                )
            except Exception as e:
                print(f"Error creating RescheduleLog: {e}")

        return Response(serializer.data)

    @action(detail=True, methods=['patch'])
    def update_reschedule(self, request, pk=None):
        """
        Update reschedule details for a specific attendance record
        """
        try:
            attendance = self.get_object()
            
            reschedule_date = request.data.get('reschedule_date')
            reschedule_time = request.data.get('reschedule_time')
            reschedule_reason = request.data.get('reschedule_reason')
            
            if not all([reschedule_date, reschedule_time, reschedule_reason]):
                return Response(
                    {"error": "reschedule_date, reschedule_time, and reschedule_reason are all required"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            attendance.reschedule_date = reschedule_date
            attendance.reschedule_time = reschedule_time
            attendance.reschedule_reason = reschedule_reason
            attendance.save()
            
            try:
                RescheduleLog.objects.get_or_create(
                    schedule=attendance.schedule,
                    employee=attendance.employee,
                    defaults={
                        'original_date': attendance.schedule.date,
                        'original_time': attendance.schedule.time,
                        'new_date': reschedule_date,
                        'new_time': reschedule_time,
                        'reason': reschedule_reason,
                    }
                )
            except Exception as e:
                print(f"Error creating RescheduleLog: {e}")
            
            serializer = self.get_serializer(attendance)
            return Response(serializer.data, status=status.HTTP_200_OK)
            
        except EmployeeAttendance.DoesNotExist:
            return Response({"error": "Attendance record not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def mark_as_present(self, request, pk=None):
        """
        Mark a rescheduled employee as present
        """
        try:
            attendance = self.get_object()
            
            if attendance.status not in ['absent', 'rescheduled']:
                return Response(
                    {"error": "Can only mark present for absent or rescheduled employees"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if not all([attendance.reschedule_date, attendance.reschedule_time, attendance.reschedule_reason]):
                return Response(
                    {"error": "Reschedule details (date, time, reason) must be filled before marking present"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            from django.utils import timezone
            from .models import RescheduleHistory
            
            history = RescheduleHistory.objects.create(
                employee=attendance.employee,
                original_schedule=attendance.schedule,
                original_date=attendance.schedule.date,
                original_time=attendance.schedule.time,
                original_venue=attendance.schedule.venue.name if attendance.schedule.venue else '',
                original_trainer=attendance.schedule.trainer.name if attendance.schedule.trainer else '',
                original_status=attendance.status,
                rescheduled_date=attendance.reschedule_date,
                rescheduled_time=attendance.reschedule_time,
                reschedule_reason=attendance.reschedule_reason,
                final_status='present',
                final_attendance_marked_at=timezone.now()
            )
            
            attendance.status = 'present'
            attendance.notes = f"Rescheduled and marked present on {timezone.now().strftime('%Y-%m-%d %H:%M')}"
            attendance.save()
            
            serializer = self.get_serializer(attendance)
            
            return Response({
                "message": "Employee marked as present successfully",
                "attendance": serializer.data,
                "history_id": history.id
            }, status=status.HTTP_200_OK)
            
        except EmployeeAttendance.DoesNotExist:
            return Response({"error": "Attendance record not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def reschedule_history(self, request):
        """
        Get complete reschedule history for all employees
        """
        from .models import RescheduleHistory
        
        employee_id = request.query_params.get('employee_id')
        schedule_id = request.query_params.get('schedule_id')
        
        queryset = RescheduleHistory.objects.select_related(
            'employee', 
            'original_schedule',
            'original_schedule__training_category',
            'original_schedule__trainer',
            'original_schedule__venue'
        ).prefetch_related(
            'original_schedule__topics'
        ).all()
        
        if employee_id:
            queryset = queryset.filter(employee__emp_id=employee_id)
        
        if schedule_id:
            queryset = queryset.filter(original_schedule_id=schedule_id)
        
        serializer = RescheduleHistorySerializer(queryset, many=True)
        return Response(serializer.data)


from rest_framework import viewsets
from rest_framework.response import Response

from .models import RescheduleLog
from .serializers import RescheduleLogSerializer


class RescheduleLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Viewset to retrieve Reschedule Logs (Read-only).
    Optional filtering by schedule or employee.
    """
    queryset = RescheduleLog.objects.all()
    serializer_class = RescheduleLogSerializer

    def create(self, request, *args, **kwargs):
        """Create a new reschedule log entry"""
        try:
            schedule = Schedule.objects.get(id=request.data.get('schedule'))
            employee = MasterTable.objects.get(emp_id=request.data.get('employee'))
            
            log = RescheduleLog.objects.create(
                schedule=schedule,
                employee=employee,
                original_date=request.data.get('original_date'),
                original_time=request.data.get('original_time'),
                new_date=request.data.get('new_date'),
                new_time=request.data.get('new_time'),
                reason=request.data.get('reason', '')
            )
            
            serializer = self.get_serializer(log)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Schedule.DoesNotExist:
            return Response({"error": "Schedule not found"}, status=status.HTTP_404_NOT_FOUND)
        except MasterTable.DoesNotExist:
            return Response({"error": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def list(self, request, *args, **kwargs):
        schedule_id = request.query_params.get('schedule_id')
        employee_id = request.query_params.get('employee_id')

        queryset = self.queryset

        if schedule_id:
            queryset = queryset.filter(schedule_id=schedule_id)

        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


# departmentlinestation filtering new 2 functions
@api_view(['GET'])
def get_employees_by_filters(request):
    """
    Get employees filtered by department, line, and/or station
    Query params: department_id, line_id, station_id
    """
    department_id = request.query_params.get('department_id')
    line_id = request.query_params.get('line_id')
    station_id = request.query_params.get('station_id')
    
    # Start with all employees
    queryset = MasterTable.objects.select_related(
        'department', 'current_line', 'current_station'
    ).all()
    
    # Apply filters if provided
    if department_id:
        queryset = queryset.filter(department_id=department_id)
    
    if line_id:
        queryset = queryset.filter(current_line_id=line_id)
    
    if station_id:
        queryset = queryset.filter(current_station_id=station_id)
    
    # Serialize and return
    serializer = MasterTableHandoverDetailSerializer(queryset, many=True)
    
    return Response({
        'count': queryset.count(),
        'employees': serializer.data
    })


@api_view(['GET'])
def get_filter_options(request):
    """
    Get all available departments, lines, and stations for filtering
    """
    from app1.models import Department, Line, Station
    
    departments = Department.objects.all().values('department_id', 'department_name')
    lines = Line.objects.all().values('line_id', 'line_name')
    stations = Station.objects.all().values('station_id', 'station_name')
    
    return Response({
        'departments': list(departments),
        'lines': list(lines),
        'stations': list(stations)
    })

# views.py
import csv
import pandas as pd
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import RefresherQuestionBank, RefresherQuestion
from django.db import transaction
import requests  # <--- Make sure to install this: pip install requests


def download_image_from_url(url):
    if not url or pd.isna(url):
        return None
    try:
        response = requests.get(str(url).strip(), timeout=5)
        if response.status_code == 200:
            # Get filename from URL or default
            filename = url.split('/')[-1].split('?')[0]
            if not filename or len(filename) > 100:
                filename = "downloaded_image.jpg"
            return ContentFile(response.content, name=filename)
    except Exception as e:
        print(f"Failed to download image from {url}: {e}")
    return None

class RefresherQuestionTemplateDownload(APIView):
    def get(self, request):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="refresher_question_template.csv"'

        writer = csv.writer(response)
        # Added Image URL columns
        writer.writerow([
            'question_text', 'option_a', 'option_b', 'option_c', 'option_d',
            'correct_answer', 'marks', 
            'question_image_url', 'option_a_image_url', 'option_b_image_url', 'option_c_image_url', 'option_d_image_url'
        ])
        
        # Example Row
        writer.writerow([
            'What does this sign mean?', 'Stop', 'Go', 'Slow', 'Yield', 'A', '1', 
            'https://example.com/stop_sign.jpg', '', '', '', ''
        ])

        return response
    



# class RefresherQuestionBulkUpload(APIView):
#     def post(self, request):
#         topic_id = request.data.get('topic_id')
#         file = request.FILES.get('file')

#         if not topic_id or not file:
#             return Response({"error": "topic_id and file are required"}, status=400)

#         try:
#             bank = RefresherQuestionBank.objects.get(topic_id=topic_id)
#         except RefresherQuestionBank.DoesNotExist:
#             return Response({"error": "Question bank not found for this topic"}, status=404)

#         df = pd.read_csv(file)
#         created = 0
#         errors = []

#         with transaction.atomic():
#             for idx, row in df.iterrows():
#                 try:
#                     RefresherQuestion.objects.create(
#                         bank=bank,
#                         question_text=row['question_text'],
#                         option_a=row['option_a'],
#                         option_b=row['option_b'],
#                         option_c=row.get('option_c', ''),
#                         option_d=row.get('option_d', ''),
#                         correct_answer=row['correct_answer'].upper(),
#                         marks=int(row.get('marks', 1)),
#                         order=idx + 1
#                     )
#                     created += 1
#                 except Exception as e:
#                     errors.append(f"Row {idx + 2}: {str(e)}")

#         return Response({
#             "message": f"{created} questions uploaded successfully",
#             "errors": errors
#         })


# refresher_training/views.py
from django.db import transaction
from django.db.models import Avg, Count, F, ExpressionWrapper, DecimalField
from django.utils import timezone
from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
import pandas as pd

from .models import (
    RefresherQuestionBank, RefresherQuestion,
    RefresherTestSession, RefresherAnswer
)
from .serializers import (
    RefresherQuestionBankSerializer, RefresherQuestionSerializer,
    RefresherTestSessionSerializer
)

# ===================== QUESTION BANK & QUESTIONS =====================
class RefresherQuestionBankDetail(APIView):
    def get(self, request, category_id): # Changed param name
        try:
            bank = RefresherQuestionBank.objects.get(category_id=category_id)
            serializer = RefresherQuestionBankSerializer(bank)
            return Response(serializer.data)
        except RefresherQuestionBank.DoesNotExist:
            # Auto-create bank if not exists
            category = get_object_or_404(Training_category, id=category_id)
            bank = RefresherQuestionBank.objects.create(category=category)
            serializer = RefresherQuestionBankSerializer(bank)
            return Response(serializer.data)
        

def process_manual_image_urls(instance, request_data):
    """
    Checks request data for specific URL fields and downloads images to the instance.
    Fields expected: 'question_image_url_input', 'option_a_image_url_input', etc.
    """
    fields = [
        ('question_image', 'question_image_url_input'),
        ('option_a_image', 'option_a_image_url_input'),
        ('option_b_image', 'option_b_image_url_input'),
        ('option_c_image', 'option_c_image_url_input'),
        ('option_d_image', 'option_d_image_url_input'),
    ]

    for model_field, input_field in fields:
        url = request_data.get(input_field)
        if url:
            img_file = download_image_from_url(url) # Reusing your helper function
            if img_file:
                getattr(instance, model_field).save(img_file.name, img_file, save=False)



class RefresherQuestionListCreate(generics.ListCreateAPIView):
    queryset = RefresherQuestion.objects.all()
    serializer_class = RefresherQuestionSerializer
    parser_classes = (MultiPartParser, FormParser)

    def get_queryset(self):
        # For filtering: GET /refresher/questions/?category_id=1
        category_id = self.request.query_params.get('category_id')
        if category_id:
            return RefresherQuestion.objects.filter(bank__category_id=category_id)
        return super().get_queryset()

    def perform_create(self, serializer):
        # 1. Get category_id from the POST body
        # Note: In FormData, even numbers come as strings, so we cast to int if present
        category_id_raw = self.request.data.get('category_id')
        
        if not category_id_raw:
            from rest_framework.exceptions import ValidationError
            raise ValidationError({"category_id": "This field is required."})

        # 2. Get the Category object first to be safe
        category = get_object_or_404(Training_category, id=category_id_raw)

        # 3. Get or Create the Bank using the Category OBJECT
        bank, created = RefresherQuestionBank.objects.get_or_create(category=category)
        
        # 4. Save instance with the bank
        instance = serializer.save(bank=bank)
        
        # 5. Handle URL downloads if any
        process_manual_image_urls(instance, self.request.data)
        
        instance.save()

class RefresherQuestionDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = RefresherQuestion.objects.all()
    serializer_class = RefresherQuestionSerializer
    parser_classes = (MultiPartParser, FormParser)

    def perform_update(self, serializer):
        # Save basic data
        instance = serializer.save()
        
        # Handle URL downloads
        process_manual_image_urls(instance, self.request.data)
        
        # Final save
        instance.save()


class RefresherQuestionBulkAction(APIView):
    def post(self, request):
        """
        Perform actions on multiple questions: 'archive', 'restore', 'delete', 'duplicate'
        """
        action = request.data.get('action')
        ids = request.data.get('ids', [])

        if not action or not ids:
            return Response({"error": "Action and IDs required"}, status=400)

        questions = RefresherQuestion.objects.filter(id__in=ids)

        if action == 'archive':
            questions.update(is_active=False)
            return Response({"message": f"Archived {len(ids)} questions"})
        
        elif action == 'restore':
            questions.update(is_active=True)
            return Response({"message": f"Restored {len(ids)} questions"})
        
        elif action == 'delete':
            count = questions.count()
            questions.delete()
            return Response({"message": f"Deleted {count} questions"})
            
        elif action == 'duplicate':
            count = 0
            with transaction.atomic():
                for q in questions:
                    # Clone the object
                    q.pk = None
                    q.question_text = f"{q.question_text} (Copy)"
                    q.save()
                    count += 1
            return Response({"message": f"Duplicated {count} questions"})

        return Response({"error": "Invalid action"}, status=400)


# ===================== BULK UPLOAD (EXCEL) =====================
import os  # <--- Make sure to import os at top of views.py if not present

class RefresherQuestionBulkUpload(APIView):
    def post(self, request):
        category_id = request.data.get('category_id')  # Changed from topic_id
        file = request.FILES.get('file')

        if not category_id or not file:
            return Response({"error": "category_id and file required"}, status=400)

        try:
            # Fetch Category Object
            category = get_object_or_404(Training_category, id=category_id)
            # Get or Create Bank using OBJECT
            bank, created = RefresherQuestionBank.objects.get_or_create(category=category)
        except Exception as e:
            return Response({"error": f"Category error: {str(e)}"}, status=404)

        # 2. Read File
        try:
            ext = os.path.splitext(file.name)[1].lower()
            if ext == '.csv':
                try:
                    df = pd.read_csv(file)
                except UnicodeDecodeError:
                    file.seek(0)
                    df = pd.read_csv(file, encoding='latin1')
            elif ext in ['.xlsx', '.xls']:
                df = pd.read_excel(file)
            else:
                return Response({"error": "Unsupported format. Use .csv or .xlsx"}, status=400)
            
            df.columns = df.columns.str.strip()
            
        except Exception as e:
            return Response({"error": f"Failed to read file: {str(e)}"}, status=400)

        created_count = 0
        errors = []

        # 3. Process Rows
        with transaction.atomic():
            for idx, row in df.iterrows():
                try:
                    def get_val(key, default=''):
                        val = row.get(key)
                        return str(val).strip() if pd.notna(val) else default

                    question = RefresherQuestion(
                        bank=bank,
                        question_text=get_val('question_text'),
                        option_a=get_val('option_a'),
                        option_b=get_val('option_b'),
                        option_c=get_val('option_c'),
                        option_d=get_val('option_d'),
                        correct_answer=get_val('correct_answer', 'A').upper(),
                        marks=int(pd.to_numeric(row.get('marks'), errors='coerce') or 1),
                        order=idx + 1
                    )

                    # Download Images
                    q_img = download_image_from_url(row.get('question_image_url'))
                    if q_img: question.question_image.save(q_img.name, q_img, save=False)

                    opt_a_img = download_image_from_url(row.get('option_a_image_url'))
                    if opt_a_img: question.option_a_image.save(opt_a_img.name, opt_a_img, save=False)

                    opt_b_img = download_image_from_url(row.get('option_b_image_url'))
                    if opt_b_img: question.option_b_image.save(opt_b_img.name, opt_b_img, save=False)

                    opt_c_img = download_image_from_url(row.get('option_c_image_url'))
                    if opt_c_img: question.option_c_image.save(opt_c_img.name, opt_c_img, save=False)

                    opt_d_img = download_image_from_url(row.get('option_d_image_url'))
                    if opt_d_img: question.option_d_image.save(opt_d_img.name, opt_d_img, save=False)

                    question.save()
                    created_count += 1

                except Exception as e:
                    errors.append(f"Row {idx+2}: {str(e)}")

        return Response({
            "message": f"{created_count} questions uploaded successfully",
            "errors": errors
        })

# ===================== TEST EXECUTION =====================

class RefresherTestStart(APIView):
    def post(self, request):
        """
        LOBBY INITIALIZATION (per schedule & optional batch)
        Body:
        {
          "schedule_id": number,
          "test_type": "pre" | "post",
          "mode": "individual" | "remote",
          "batch_id": number | null   # NEW (optional)
        }
        """
        schedule_id = request.data.get('schedule_id')
        test_type = request.data.get('test_type')
        mode = request.data.get('mode', 'individual')
        batch_id = request.data.get('batch_id')  # NEW

        if not schedule_id or not test_type:
            return Response(
                {"error": "schedule_id and test_type are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
          schedule = Schedule.objects.get(id=schedule_id)

          # 1. Question bank (by Category)
          try:
              bank = RefresherQuestionBank.objects.get(
                  category=schedule.training_category
              )
          except RefresherQuestionBank.DoesNotExist:
              return Response(
                  {"error": f"No Question Bank found for category: {schedule.training_category.name}"},
                  status=status.HTTP_404_NOT_FOUND
              )

          questions = RefresherQuestion.objects.filter(
              bank=bank, is_active=True
          ).order_by('order')
          if not questions.exists():
              return Response(
                  {"error": "Question Bank is empty."},
                  status=status.HTTP_400_BAD_REQUEST
              )

          # 2. Get Present Employees
          present_attendances = EmployeeAttendance.objects.filter(
              schedule=schedule, status='present'
          ).select_related('employee')

          # 2a. If batch_id is provided, restrict to that batch's employees
          batch = None
          if batch_id is not None:
              try:
                  batch = RefresherBatch.objects.get(
                      id=batch_id,
                      schedule=schedule
                  )
              except RefresherBatch.DoesNotExist:
                  return Response(
                      {"error": "Batch not found for this schedule"},
                      status=status.HTTP_404_NOT_FOUND
                  )
              batch_employees = batch.employees.all()
              present_attendances = present_attendances.filter(
                  employee__in=batch_employees
              )

          present_employees = [att.employee for att in present_attendances]
          if not present_employees:
              return Response(
                  {"error": "No employees marked as Present."},
                  status=status.HTTP_400_BAD_REQUEST
              )

          # 3. Session Management (per batch)
          existing_sessions = RefresherTestSession.objects.filter(
              schedule=schedule,
              test_type=test_type,
              employee__in=present_employees,
              batch=batch  # may be None (if no batch provided)
          )
          existing_map = {s.employee.pk: s for s in existing_sessions}
          new_sessions = []
          sessions_response = []

          for emp in present_employees:
              session = existing_map.get(emp.pk)
              if session:
                  # If mode changed and not submitted yet, update
                  if session.mode != mode and not session.submitted_at:
                      session.mode = mode
                      session.save(update_fields=['mode'])
                  sessions_response.append({
                      "session_id": session.id,
                      "employee_id": emp.emp_id,  # optional, for frontend
                      "employee_name": f"{emp.first_name} {emp.last_name}",
                      "test_type": test_type,
                      "is_completed": bool(session.submitted_at)
                  })
              else:
                  new_sessions.append(
                      RefresherTestSession(
                          schedule=schedule,
                          employee=emp,
                          test_type=test_type,
                          mode=mode,
                          batch=batch,              # NEW
                          started_at=timezone.now()
                      )
                  )

          if new_sessions:
              with transaction.atomic():
                  RefresherTestSession.objects.bulk_create(new_sessions)
              created = RefresherTestSession.objects.filter(
                  schedule=schedule,
                  test_type=test_type,
                  employee__in=[s.employee for s in new_sessions],
                  batch=batch
              )
              for s in created:
                  sessions_response.append({
                      "session_id": s.id,
                      "employee_id": s.employee.emp_id,
                      "employee_name": f"{s.employee.first_name} {s.employee.last_name}",
                      "test_type": test_type,
                      "is_completed": False
                  })

          total_present = len(present_employees)
          total_completed = len(
              [s for s in sessions_response if s['is_completed']]
          )

          response_data = {
              "message": "Initialized",
              "schedule_id": schedule.id,
              "topic_name": schedule.training_category.name,  # category name
              "batch_id": batch.id if batch else None,
              "sessions": sessions_response,
              "total_present": total_present,
              "total_completed": total_completed,
          }

          # For remote mode, send questions once
          if mode == 'remote':
              question_serializer = RefresherQuestionSerializer(
                  questions, many=True
              )
              response_data['questions'] = question_serializer.data

          return Response(response_data)

        except Schedule.DoesNotExist:
          return Response({"error": "Schedule not found"}, status=404)
        except Exception as e:
          return Response({"error": str(e)}, status=500)

    def get(self, request):
        """
        INDIVIDUAL TEST START (unchanged, no batch filter here
        because a single test_session already knows its batch)
        """
        session_id = request.query_params.get('session_id')
        if not session_id:
            return Response({"error": "session_id required"}, status=400)
        try:
            session = RefresherTestSession.objects.select_related(
                'schedule', 'employee'
            ).get(id=session_id)
            if session.submitted_at:
                return Response(
                    {"error": "Test already submitted", "completed": True},
                    status=403
                )
            
            # Fetch from Category Bank
            bank = RefresherQuestionBank.objects.get(
                category=session.schedule.training_category
            )
            questions = RefresherQuestion.objects.filter(
                bank=bank, is_active=True
            ).order_by('order')
            question_serializer = RefresherQuestionSerializer(
                questions, many=True
            )
            
            return Response({
                "session_id": session.id,
                "employee_name": f"{session.employee.first_name} {session.employee.last_name}",
                "time_limit": bank.time_limit_minutes,
                "questions": question_serializer.data
            })
        except Exception as e:
            return Response({"error": str(e)}, status=500)

# In views.py

class RefresherTestSubmit(APIView):
    def post(self, request):
        data = request.data
        submissions = data.get('submissions', [])
        if not submissions and 'test_session_id' in data: submissions = [data]

        results = []
        schedule_to_check = None # We will track which schedule to update

        for sub in submissions:
            session_id = sub.get('test_session_id')
            answers = sub.get('answers', {})

            try:
                session = RefresherTestSession.objects.get(id=session_id)
                schedule_to_check = session.schedule # Capture schedule for checking later

                # Fetch questions for the CATEGORY
                questions = RefresherQuestion.objects.filter(bank__category=session.schedule.training_category)
                
                correct_count = 0
                total_marks = 0
                
                with transaction.atomic():
                    RefresherAnswer.objects.filter(test_session=session).delete()
                    for q in questions:
                        total_marks += q.marks
                        user_choice = answers.get(str(q.id)) or answers.get(q.id)
                        if user_choice:
                            is_correct = (user_choice.upper() == q.correct_answer.upper())
                            if is_correct: correct_count += q.marks
                            RefresherAnswer.objects.create(test_session=session, question=q, selected_option=user_choice.upper(), is_correct=is_correct)

                total_possible = questions.aggregate(total=Sum('marks'))['total'] or 1
                percentage = (correct_count / total_possible) * 100
                
                # Get Passing Score from Bank
                bank = RefresherQuestionBank.objects.get(category=session.schedule.training_category)
                passing_score = bank.passing_score
                
                session.score = correct_count
                session.total_questions = questions.count()
                session.percentage = round(percentage, 2)
                session.passed = percentage >= passing_score
                session.submitted_at = timezone.now()
                session.save()

                results.append({"session_id": session.id, "score": correct_count})

            except Exception as e:
                results.append({"session_id": session_id, "error": str(e)})

        # --- AUTO-COMPLETE LOGIC START ---
        if schedule_to_check:
            try:
                # 1. Count Total Present Employees
                total_present = EmployeeAttendance.objects.filter(
                    schedule=schedule_to_check, 
                    status='present'
                ).count()

                # 2. Count Completed Post-Tests for this Schedule
                # We check distinct employees to avoid double counting if duplicate sessions exist
                completed_post_tests = RefresherTestSession.objects.filter(
                    schedule=schedule_to_check,
                    test_type='post',
                    submitted_at__isnull=False
                ).values('employee').distinct().count()

                # 3. If everyone present has finished Post-Test, mark Schedule as Completed
                if total_present > 0 and completed_post_tests >= total_present:
                    schedule_to_check.status = 'completed'
                    schedule_to_check.training_stage = 4
                    schedule_to_check.completed_date = timezone.now().date()
                    schedule_to_check.save()
                    print(f"Schedule {schedule_to_check.id} marked as COMPLETED")
                elif schedule_to_check.status == 'scheduled':
                     # If training has activity but isn't done, mark as pending
                     schedule_to_check.status = 'pending'
                     schedule_to_check.save()

            except Exception as e:
                print(f"Error updating schedule status: {e}")
        # --- AUTO-COMPLETE LOGIC END ---

        return Response({"message": "Submissions processed", "results": results})
    

class ScheduleContentList(APIView):
    def get(self, request, pk):
        try:
            schedule = Schedule.objects.get(id=pk)
            # Get contents for ALL topics in this schedule
            topic_ids = schedule.topics.values_list('id', flat=True)
            contents = CurriculumContent.objects.filter(curriculum_id__in=topic_ids)
            serializer = CurriculumContentSerializer(contents, many=True)
            return Response(serializer.data)
        except Schedule.DoesNotExist:
            return Response({"error": "Schedule not found"}, status=404)


from django.shortcuts import get_object_or_404
from django.db.models import Avg, Sum, Count
from django.utils import timezone
from django.db import transaction
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, generics

from .models import (
    RefresherQuestionBank, 
    RefresherQuestion,
    RefresherTestSession, 
    RefresherAnswer,
    Schedule, 
    MasterTable,
    # Assuming KeyEvent model exists in your app or you import it. 
    # If KeyEvent is in a different app, import it from there.
    # For now, I assume it's available in models or we use a placeholder.
    # from .models import KeyEvent 
)
from .serializers import (
    RefresherQuestionBankSerializer, 
    RefresherQuestionSerializer,
    RefresherTestSessionSerializer
)

# ===================== REMOTE EVENT POLLING =====================
# This drives the Remote Voting UI. It polls for hardware key presses.

class RefresherRemoteEvent(APIView):
    def get(self, request):
        """
        Poll for new key events since a specific ID.
        """
        try:
            # Get the ID sent by frontend (defaults to 0)
            since_id = int(request.query_params.get('since_id', 0))
            
            # Query the MAIN KeyEvent table where the Python script writes data
            new_events = KeyEvent.objects.filter(id__gt=since_id).order_by('id')[:50]
            
            data = []
            for evt in new_events:
                data.append({
                    'id': evt.id,
                    'key_id': evt.key_id,   
                    'info': evt.info,       
                    'timestamp': evt.timestamp
                })
            return Response(data)
            
        except Exception as e:
            # Print error to terminal so we can see if something else is wrong
            print(f"Refresher Polling Error: {e}")
            return Response([])

    def post(self, request):
        """
        Manually simulate a key press (useful for testing without hardware).
        """
        try:
            from .models import KeyEvent
            
            KeyEvent.objects.create(
                key_id=request.data.get('key_id'),
                info=request.data.get('info'),
                base_id=1,
                mode=1,
                timestamp=timezone.now(),
                client_timestamp=timezone.now(),
                event_type='manual'
            )
            return Response({"status": "received"}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)




# ===================== EFFECTIVENESS REPORT =====================
class TrainingEffectivenessReport(APIView):
    def get(self, request):
        # CHANGED: Filter by category_id instead of topic_id
        category_id = request.query_params.get('category_id')
        
        # --- DETAILED VIEW (Specific Category) ---
        if category_id:
            try:
                bank = RefresherQuestionBank.objects.get(category_id=category_id)
                
                # Fetch sessions for this Category
                sessions = RefresherTestSession.objects.filter(
                    schedule__training_category_id=category_id,
                    submitted_at__isnull=False
                ).select_related('employee')
                
                employee_map = {}
                
                for s in sessions:
                    emp_id = s.employee.pk
                    if emp_id not in employee_map:
                        employee_map[emp_id] = {
                            "id": emp_id,
                            "name": f"{s.employee.first_name} {s.employee.last_name}",
                            "code": s.employee.emp_id,
                            "pre_score": None,
                            "post_score": None,
                            "date": s.submitted_at
                        }
                    
                    if s.test_type == 'pre':
                        employee_map[emp_id]['pre_score'] = float(s.percentage)
                    elif s.test_type == 'post':
                        employee_map[emp_id]['post_score'] = float(s.percentage)
                
                results = []
                for emp_data in employee_map.values():
                    pre = emp_data['pre_score'] if emp_data['pre_score'] is not None else 0
                    post = emp_data['post_score'] if emp_data['post_score'] is not None else 0
                    gain = post - pre
                    
                    results.append({
                        **emp_data,
                        "pre_score": pre,
                        "post_score": post,
                        "gain": gain,
                        "status": "Improved" if gain > 0 else "No Change" if gain == 0 else "Declined"
                    })
                    
                return Response({
                    "topic": bank.category.name, # Return Category Name as Topic
                    "details": results
                })
                
            except RefresherQuestionBank.DoesNotExist:
                return Response({"error": "Category Bank not found"}, status=404)

        # --- SUMMARY VIEW (All Categories) ---
        data = []
        banks = RefresherQuestionBank.objects.all()
        
        for bank in banks:
            pre_tests = RefresherTestSession.objects.filter(
                schedule__training_category=bank.category,
                test_type='pre',
                submitted_at__isnull=False
            )
            
            post_tests = RefresherTestSession.objects.filter(
                schedule__training_category=bank.category,
                test_type='post',
                submitted_at__isnull=False
            )

            if not pre_tests.exists() and not post_tests.exists():
                continue

            pre_avg = pre_tests.aggregate(Avg('percentage'))['percentage__avg'] or 0
            post_avg = post_tests.aggregate(Avg('percentage'))['percentage__avg'] or 0
            
            improvement = post_avg - pre_avg
            effectiveness = (improvement / pre_avg * 100) if pre_avg > 0 else improvement 

            data.append({
                "topic_id": bank.category.id, # Use Category ID
                "topic": bank.category.name,  # Display Category Name
                "category": "General",        # Static label or group
                "pre_avg": round(pre_avg, 1),
                "post_avg": round(post_avg, 1),
                "improvement": round(improvement, 1),
                "effectiveness": round(effectiveness, 1),
                "employee_count": post_tests.count()
            })

        return Response(sorted(data, key=lambda x: x['effectiveness'], reverse=True))


# =================== Refreshment Training End ============================= #

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import Department, Station, StationSetting
from .serializers import DepartmentselectSerializer, StationselectSerializer, StationSettingSerializer

class DepartmentListView(APIView):
    def get(self, request):
        departments = Department.objects.all()
        serializer = DepartmentselectSerializer(departments, many=True)
        return Response(serializer.data)

class StationsByDepartmentView(APIView):
    def get(self, request, department_id):
        stations = Station.objects.filter(subline_linedepartment_department_id=department_id)
        serializer = StationselectSerializer(stations, many=True)
        return Response(serializer.data)


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import StationSetting, Department, Station
from .serializers import StationSettingSerializer

class StationSettingCreateView(APIView):
    def get(self, request):
        department_id = request.query_params.get('department_id')
        if not department_id:
            return Response(
                {"error": "department_id query parameter is required."},
                status=status.HTTP_400_BAD_REQUEST
            )
                 
        try:
            department = Department.objects.get(department_id=department_id)
        except Department.DoesNotExist:
            return Response(
                {"error": "Department not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Group settings by station
        settings = StationSetting.objects.filter(department=department).select_related('station', 'department')
        
        # Group by station
        station_groups = {}
        for setting in settings:
            station_key = setting.station.station_id
            if station_key not in station_groups:
                station_groups[station_key] = {
                    'department_id': setting.department.department_id,
                    'department_name': setting.department.department_name,
                    'station_id': setting.station.station_id,
                    'station_name': setting.station.station_name,
                    'all_options': []
                }
            station_groups[station_key]['all_options'].append(setting.option)
        
        # Convert to list and remove duplicates from options
        result = []
        for station_data in station_groups.values():
            station_data['all_options'] = list(set(station_data['all_options']))
            result.append(station_data)
        
        return Response(result, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = StationSettingSerializer(data=request.data)
        if serializer.is_valid():
            created_settings = serializer.save()
            return Response({"success": True, "created": len(created_settings)}, status=201)
        return Response(serializer.errors, status=400)



# =================== Retraining start ============================= #


# from rest_framework import viewsets, status
# from rest_framework.decorators import action
# from rest_framework.response import Response
# from django.db.models import Q
# from collections import defaultdict
# from django.shortcuts import get_object_or_404

# from .models import (
#     RetrainingSession, RetrainingConfig, MasterTable,
   
#     OperatorPerformanceEvaluation, TenCyclePassingCriteria,
      
#     TraineeInfo, OJTPassingCriteria, OJTScoreRange,
    
#     Score, EvaluationPassingCriteria,
#     Department, Level, Station
# )
# from .serializers import RetrainingSessionSerializer, RetrainingConfigSerializer

# class RetrainingConfigViewSet(viewsets.ModelViewSet):
#     queryset = RetrainingConfig.objects.all()
#     serializer_class = RetrainingConfigSerializer

#     def get_queryset(self):
#         queryset = super().get_queryset()
#         level_id = self.request.query_params.get('level_id')
#         evaluation_type = self.request.query_params.get('evaluation_type')
        
#         if level_id:
#             queryset = queryset.filter(level_id=level_id)
#         if evaluation_type:
#             queryset = queryset.filter(evaluation_type=evaluation_type)
            
#         return queryset
    
    
# class RetrainingSessionViewSet(viewsets.ModelViewSet):
#     queryset = RetrainingSession.objects.all().order_by("-created_at")
#     serializer_class = RetrainingSessionSerializer

#     def create(self, request, *args, **kwargs):
#      """Schedule a new retraining session"""
#      employee_id = request.data.get('employee')
#      level_id = request.data.get('level')
#      department_id = request.data.get('department')
#      station_id = request.data.get('station')
#      evaluation_type = request.data.get('evaluation_type')

#      if not (employee_id and level_id and department_id and evaluation_type):
#         return Response({
#             "error": "Missing required fields: employee, level, department, evaluation_type"
#         }, status=status.HTTP_400_BAD_REQUEST)

#      # Get max allowed retraining sessions for this level and evaluation type
#      config = RetrainingConfig.objects.filter(
#         level_id=level_id, 
#         evaluation_type=evaluation_type
#      ).first()
#      max_retraining_sessions = config.max_count if config else 2

#      # Count existing retraining sessions for this specific combination
#      filter_kwargs = {
#         'employee_id': employee_id,
#         'level_id': level_id,
#         'department_id': department_id,
#         'evaluation_type': evaluation_type
#      }
    
#      if station_id:
#         filter_kwargs['station_id'] = station_id

#      existing_sessions = RetrainingSession.objects.filter(**filter_kwargs)
#      existing_count = existing_sessions.count()

#      if existing_count >= max_retraining_sessions:
#         return Response({
#             'error': f'Maximum retraining sessions ({max_retraining_sessions}) reached for this employee and evaluation combination.'
#         }, status=status.HTTP_400_BAD_REQUEST)

#      request.data['attempt_no'] = existing_count + 2
    
     
    
#      # Create the main session with scheduling info only
#      response = super().create(request, *args, **kwargs)
     
     
    
#      return response

#     @action(detail=True, methods=['patch'], url_path='complete-session')
#     def complete_session(self, request, pk=None):
#         """Complete a retraining session with results and observations"""
#         try:
#             session = self.get_object()
            
#             if session.status != 'Pending':
#                 return Response(
#                     {'error': 'Only pending sessions can be completed'}, 
#                     status=status.HTTP_400_BAD_REQUEST
#                 )
            
#             # Update main session with completion data
#             session.status = request.data.get('status', 'Completed')
#             session.performance_percentage = request.data.get('performance_percentage')
#             session.required_percentage = request.data.get('required_percentage')
#             session.save()
            
#             # Update session detail with observations and trainer info
#             session_detail, created = RetrainingSessionDetail.objects.get_or_create(
#                 retraining_session=session
#             )
            
#             if 'observations_failure_points' in request.data:
#                 session_detail.observations_failure_points = request.data['observations_failure_points']
            
#             if 'trainer_name' in request.data:
#                 session_detail.trainer_name = request.data['trainer_name']
            
#             session_detail.save()
            
#             # Return updated session with detail
#             updated_serializer = self.get_serializer(session)
#             return Response(updated_serializer.data)
            
#         except RetrainingSession.DoesNotExist:
#             return Response(
#                 {'error': 'Session not found'}, 
#                 status=status.HTTP_404_NOT_FOUND
#             )

#     @action(detail=True, methods=['patch'], url_path='update-observations')
#     def update_observations(self, request, pk=None):
#         """Update only observations and trainer info (for partial updates)"""
#         try:
#             session = self.get_object()
            
#             # Get or create session detail
#             session_detail, created = RetrainingSessionDetail.objects.get_or_create(
#                 retraining_session=session
#             )
            
#             # Update only the fields provided
#             if 'observations_failure_points' in request.data:
#                 session_detail.observations_failure_points = request.data['observations_failure_points']
            
#             if 'trainer_name' in request.data:
#                 session_detail.trainer_name = request.data['trainer_name']
            
#             session_detail.save()
            
#             # Return updated session with detail
#             updated_serializer = self.get_serializer(session)
#             return Response(updated_serializer.data)
            
#         except RetrainingSession.DoesNotExist:
#             return Response(
#                 {'error': 'Session not found'}, 
#                 status=status.HTTP_404_NOT_FOUND
#             )

    

#     @action(detail=False, methods=['get'], url_path='employee-sessions/(?P<employee_id>[^/.]+)')
#     def get_employee_sessions(self, request, employee_id=None):
#         """Get all retraining sessions for a specific employee, ordered by attempt number"""
#         try:
#             employee = MasterTable.objects.get(emp_id=employee_id)
#             sessions = RetrainingSession.objects.filter(
#                 employee=employee
#             ).select_related('session_detail').order_by(
#                 'evaluation_type', 'level', 'department', 'station', 'attempt_no'
#             )
            
#             serializer = self.get_serializer(sessions, many=True)
#             return Response({
#                 'employee_id': employee_id,
#                 'employee_name': f"{employee.first_name or ''} {employee.last_name or ''}".strip(),
#                 'sessions': serializer.data
#             })
            
#         except MasterTable.DoesNotExist:
#             return Response(
#                 {'error': 'Employee not found'}, 
#                 status=status.HTTP_404_NOT_FOUND
#             )

#     @action(detail=False, methods=['get'], url_path='failed-employees')
#     def get_failed_employees(self, request):
#         """Get all failed employees from all evaluation types with retraining status"""
#         failed_employees = []
        
#         # Get failed employees from each evaluation system
#         ten_cycle_failed = self._get_10cycle_failed_employees()
#         failed_employees.extend(ten_cycle_failed)
        
#         ojt_failed = self._get_ojt_failed_employees()
#         failed_employees.extend(ojt_failed)
        
#         evaluation_failed = self._get_evaluation_failed_employees()
#         failed_employees.extend(evaluation_failed)
        
#         print(f"10 Cycle failures: {len(ten_cycle_failed)}")      
#         print(f"OJT failures: {len(ojt_failed)}")                  
#         print(f"Evaluation failures: {len(evaluation_failed)}")  
#         print(f"Total failures: {len(failed_employees)}") 
        
#         # Apply filters
#         department_id = request.query_params.get('department_id')
#         evaluation_type = request.query_params.get('evaluation_type')
#         status_filter = request.query_params.get('status')
        
#         if department_id:
#             failed_employees = [emp for emp in failed_employees if emp.get('department_id') == int(department_id)]
        
#         if evaluation_type:
#             failed_employees = [emp for emp in failed_employees if emp.get('evaluation_type') == evaluation_type]
            
#         if status_filter:
#             failed_employees = [emp for emp in failed_employees if emp.get('retraining_status') == status_filter]
        
#         return Response({
#             'count': len(failed_employees),
#             'results': failed_employees
#         })
        
        

#     def _get_10cycle_failed_employees(self):
#         # """Get failed employees from 10 Cycle evaluations"""
#         # failed_employees = []
        
#         # evaluations = OperatorPerformanceEvaluation.objects.filter(
#         #     final_status='Fail - Retraining Required',
#         #     is_completed=True
#         # ).select_related('employee', 'department', 'station', 'level')
    
#         failed_employees = []
    
        
#         all_evaluations = OperatorPerformanceEvaluation.objects.all()
#         print(f"Total OperatorPerformanceEvaluations: {all_evaluations.count()}")
    
        
#         unique_statuses = OperatorPerformanceEvaluation.objects.values_list('final_status', flat=True).distinct()
#         print(f"All final_status values in DB: {list(unique_statuses)}")
    
        
#         completed_values = OperatorPerformanceEvaluation.objects.values_list('is_completed', flat=True).distinct()
#         print(f"All is_completed values in DB: {list(completed_values)}")
    
        
#         fail_evaluations = OperatorPerformanceEvaluation.objects.filter(
#          final_status__icontains='Fail'
#         )
#         print(f"Evaluations containing 'Fail': {fail_evaluations.count()}")
#         for eval in fail_evaluations:
#          print(f"  - Status: '{eval.final_status}' | Completed: {eval.is_completed} | Employee: {eval.employee}")
    
        
#         evaluations = OperatorPerformanceEvaluation.objects.filter(
#           final_status='Fail - Retraining Required',
#           is_completed=True
#         ).select_related('employee', 'department', 'station', 'level')
    
#         print(f"Evaluations matching exact criteria: {evaluations.count()}")
        
#         for evaluation in evaluations:
#             # Get passing criteria
#             try:
#                 criteria = TenCyclePassingCriteria.objects.get(
#                     level=evaluation.level,
#                     department=evaluation.department,
#                     station=evaluation.station,
#                     is_active=True
#                 )
#                 required_percentage = criteria.passing_percentage
#             except TenCyclePassingCriteria.DoesNotExist:
#                 required_percentage = 60.0
            
#             # Get retraining sessions for this employee/evaluation combo
#             existing_sessions = RetrainingSession.objects.filter(
#                 employee=evaluation.employee,
#                 level=evaluation.level,
#                 department=evaluation.department,
#                 station=evaluation.station,
#                 evaluation_type='10 Cycle'
#             ).select_related('session_detail').order_by('-attempt_no')
#             # ).order_by('-attempt_no')
            
#             # Get max allowed attempts
#             config = RetrainingConfig.objects.filter(
#                 level=evaluation.level, 
#                 evaluation_type='10 Cycle'
#             ).first()
#             max_attempts = config.max_count if config else 2
            
#             # Determine retraining status
#             retraining_status = self._determine_retraining_status(existing_sessions, max_attempts)
            
#             gap = required_percentage - (evaluation.final_percentage or 0)
            
#             failed_employees.append({
#                 'employee_pk': evaluation.employee.emp_id,
#                 'employee_id': evaluation.employee.emp_id,
#                 'employee_name': f"{evaluation.employee.first_name or ''} {evaluation.employee.last_name or ''}".strip(),
#                 'department_id': evaluation.department.department_id,
#                 'department_name': evaluation.department.department_name,
#                 'station_id': evaluation.station.station_id if evaluation.station else None,
#                 'station_name': evaluation.station.station_name if evaluation.station else 'N/A',
#                 'level_id': evaluation.level.level_id,
#                 'level_name': evaluation.level.level_name,
#                 'evaluation_type': '10 Cycle',
#                 'obtained_percentage': round(evaluation.final_percentage or 0, 2),
#                 'required_percentage': required_percentage,
#                 'performance_gap': round(gap, 2),
#                 'last_evaluation_date': evaluation.date.isoformat(),
#                 'existing_sessions_count': existing_sessions.count(),
#                 'max_attempts': max_attempts,
#                 'can_schedule_retraining': existing_sessions.count() < max_attempts,
#                 'retraining_status': retraining_status,
#                 # 'retraining_records': [
#                 #     {
#                 #         'id': session.id,
#                 #         'attempt_no': session.attempt_no,
#                 #         'scheduled_date': session.scheduled_date.isoformat(),
#                 #         'scheduled_time': session.scheduled_time.strftime('%H:%M'),
#                 #         'venue': session.venue,
#                 #         'status': session.status,
#                 #         'performance_percentage': session.performance_percentage
#                 #     } for session in existing_sessions
#                 # ]
#                                 'retraining_records': [
#                        {
#                           'id': session.id,
#                            'attempt_no': session.attempt_no,
#                            'scheduled_date': session.scheduled_date.isoformat(),
#                             'scheduled_time': session.scheduled_time.strftime('%H:%M'),
#                             'venue': session.venue,
#                             'status': session.status,
#                             'performance_percentage': session.performance_percentage,
#                             'session_detail': {
#                                 'observations_failure_points': session.session_detail.observations_failure_points if hasattr(session, 'session_detail') else None,
#                                 'trainer_name': session.session_detail.trainer_name if hasattr(session, 'session_detail') else None,
#                             } if hasattr(session, 'session_detail') else None
#                         } for session in existing_sessions
#                     ]
#             })
        
#         return failed_employees


    
    
#     def _get_ojt_failed_employees(self):
#      """Get failed employees from OJT evaluations"""
#      failed_employees = []
    
#      trainees = TraineeInfo.objects.filter(status='Fail')
    
#      for trainee in trainees:
#         # Get trainee's score info to determine department/level
#         trainee_score = trainee.scores.select_related('topic_department', 'topic_level', 'day').first()
#         if not trainee_score:
#             continue

#         department = trainee_score.topic.department
#         level = trainee_score.topic.level
#         day = trainee_score.day

#         # Get passing criteria
#         criteria = (
#             OJTPassingCriteria.objects.filter(department=department, level=level, day=day).first()
#             or OJTPassingCriteria.objects.filter(department=department, level=level, day__isnull=True).first()
#         )
#         required_percentage = criteria.percentage if criteria else 60.0

#         # Calculate obtained percentage
#         from django.db.models import Sum
#         total_score = trainee.scores.aggregate(total=Sum("score"))["total"] or 0
#         total_topics = trainee.scores.count()

#         obtained_percentage = 0
#         if total_topics > 0:
#             try:
#                 score_range = OJTScoreRange.objects.filter(
#                     department=department,
#                     level=level
#                 ).first()
#                 max_score = score_range.max_score if score_range else 5
#                 obtained_percentage = (total_score / (total_topics * max_score)) * 100
#             except:
#                 obtained_percentage = 0

#         # Try to match trainee with MasterTable
#         try:
#             employee = MasterTable.objects.get(emp_id=trainee.emp_id)
#             employee_pk = employee.pk
#         except MasterTable.DoesNotExist:
#             employee = None
#             employee_pk = None

#         # Get retraining sessions (only if mapped to MasterTable)
#         if employee:
#             existing_sessions = RetrainingSession.objects.filter(
#                 employee=employee,
#                 level=level,
#                 department=department,
#                 evaluation_type='OJT'
#             ).select_related('session_detail').order_by('-attempt_no')
#             # ).order_by('-attempt_no')

#             config = RetrainingConfig.objects.filter(level=level, evaluation_type='OJT').first()
#             max_attempts = config.max_count if config else 2

#             retraining_status = self._determine_retraining_status(existing_sessions, max_attempts)
#         else:
#             existing_sessions = []
#             max_attempts = 0
#             retraining_status = "not_mapped"  # Or "N/A"

#         gap = required_percentage - obtained_percentage

#         failed_employees.append({
#             'employee_pk': employee_pk, 
#             'employee_id': trainee.emp_id,
#             'employee_name': trainee.trainee_name,
#             'department_id': department.department_id,
#             'department_name': department.department_name,
#             'station_id': None,
#             'station_name': trainee.station,
#             'level_id': level.level_id,
#             'level_name': level.level_name,
#             'evaluation_type': 'OJT',
#             'obtained_percentage': round(obtained_percentage, 2),
#             'required_percentage': required_percentage,
#             'performance_gap': round(gap, 2),
#             'last_evaluation_date': trainee.doj.isoformat() if trainee.doj else None,
#             'existing_sessions_count': len(existing_sessions),
#             'max_attempts': max_attempts,
#             'can_schedule_retraining': employee is not None and len(existing_sessions) < max_attempts,
#             'retraining_status': retraining_status,
#             # 'retraining_records': [
#             #     {
#             #         'id': session.id,
#             #         'attempt_no': session.attempt_no,
#             #         'scheduled_date': session.scheduled_date.isoformat(),
#             #         'scheduled_time': session.scheduled_time.strftime('%H:%M'),
#             #         'venue': session.venue,
#             #         'status': session.status,
#             #         'performance_percentage': session.performance_percentage
#             #     } for session in existing_sessions
#             # ]
#                             'retraining_records': [
#                        {
#                           'id': session.id,
#                            'attempt_no': session.attempt_no,
#                            'scheduled_date': session.scheduled_date.isoformat(),
#                             'scheduled_time': session.scheduled_time.strftime('%H:%M'),
#                             'venue': session.venue,
#                             'status': session.status,
#                             'performance_percentage': session.performance_percentage,
#                             'session_detail': {
#                                 'observations_failure_points': session.session_detail.observations_failure_points if hasattr(session, 'session_detail') else None,
#                                 'trainer_name': session.session_detail.trainer_name if hasattr(session, 'session_detail') else None,
#                             } if hasattr(session, 'session_detail') else None
#                         } for session in existing_sessions
#                     ]
#         })
 
#      return failed_employees


#     def _get_evaluation_failed_employees(self):
#         """Get failed employees from Evaluation tests"""
#         failed_employees = []
        
#         failed_scores = Score.objects.filter(passed=False).select_related(
#             'employee', 'level', 'skill', 'test'
#         )
        
#         for score in failed_scores:
#             if not score.employee or not score.level:
#                 continue
                
#             # Get department
#             department = None
#             if hasattr(score.employee, 'department') and score.employee.department:
#                 department = score.employee.department
#             elif score.skill and hasattr(score.skill, 'subline'):
#                 try:
#                     department = score.skill.subline.line.department
#                 except AttributeError:
#                     continue
#             else:
#                 continue
            
#             # Get passing criteria
#             criteria = EvaluationPassingCriteria.objects.filter(
#                 level=score.level,
#                 department=department
#             ).first()
#             required_percentage = float(criteria.percentage) if criteria else 80.0
            
#             # Get retraining sessions
#             existing_sessions = RetrainingSession.objects.filter(
#                 employee=score.employee,
#                 level=score.level,
#                 department=department,
#                 evaluation_type='Evaluation'
#             ).select_related('session_detail').order_by('-attempt_no')
#             # ).order_by('-attempt_no')
            
#             config = RetrainingConfig.objects.filter(level=score.level, evaluation_type='Evaluation').first()
#             max_attempts = config.max_count if config else 2
            
#             retraining_status = self._determine_retraining_status(existing_sessions, max_attempts)
#             gap = required_percentage - score.percentage
            
#             failed_employees.append({
#                 'employee_pk': score.employee.emp_id,
#                 'employee_id': score.employee.emp_id,
#                 'employee_name': f"{score.employee.first_name or ''} {score.employee.last_name or ''}".strip(),
#                 'department_id': department.department_id,
#                 'department_name': department.department_name,
#                 'station_id': score.skill.station_id if score.skill else None,
#                 'station_name': score.skill.station_name if score.skill else 'N/A',
#                 'level_id': score.level.level_id,
#                 'level_name': score.level.level_name,
#                 'evaluation_type': 'Evaluation',
#                 'obtained_percentage': round(score.percentage, 2),
#                 'required_percentage': required_percentage,
#                 'performance_gap': round(gap, 2),
#                 'last_evaluation_date': score.created_at.date().isoformat(),
#                 'existing_sessions_count': existing_sessions.count(),
#                 'max_attempts': max_attempts,
#                 'can_schedule_retraining': existing_sessions.count() < max_attempts,
#                 'retraining_status': retraining_status,
#                 # 'retraining_records': [
#                 #     {
#                 #         'id': session.id,
#                 #         'attempt_no': session.attempt_no,
#                 #         'scheduled_date': session.scheduled_date.isoformat(),
#                 #         'scheduled_time': session.scheduled_time.strftime('%H:%M'),
#                 #         'venue': session.venue,
#                 #         'status': session.status,
#                 #         'performance_percentage': session.performance_percentage
#                 #     } for session in existing_sessions
#                 # ]
#                 'retraining_records': [
#                        {
#                           'id': session.id,
#                            'attempt_no': session.attempt_no,
#                            'scheduled_date': session.scheduled_date.isoformat(),
#                             'scheduled_time': session.scheduled_time.strftime('%H:%M'),
#                             'venue': session.venue,
#                             'status': session.status,
#                             'performance_percentage': session.performance_percentage,
#                             'session_detail': {
#                                 'observations_failure_points': session.session_detail.observations_failure_points if hasattr(session, 'session_detail') else None,
#                                 'trainer_name': session.session_detail.trainer_name if hasattr(session, 'session_detail') else None,
#                             } if hasattr(session, 'session_detail') else None
#                         } for session in existing_sessions
#                     ]
#             })
        
#         return failed_employees

#     def _determine_retraining_status(self, existing_sessions, max_attempts):
#         """Determine the retraining status based on existing sessions"""
#         if not existing_sessions.exists():
#             return 'pending'
        
#         latest_session = existing_sessions.first()
        
#         if existing_sessions.count() >= max_attempts:
#             # Check if last attempt was successful
#             if latest_session.status == 'Completed' and latest_session.performance_percentage and latest_session.performance_percentage >= (latest_session.required_percentage or 60):
#                 return 'completed'
#             else:
#                 return 'failed'  # Max attempts reached without success
        
#         if latest_session.status == 'Pending':
#             return 'scheduled'
#         elif latest_session.status == 'Completed':
#             if latest_session.performance_percentage and latest_session.performance_percentage >= (latest_session.required_percentage or 60):
#                 return 'completed'
#             else:
#                 return 'pending'  # Can schedule another attempt
#         elif latest_session.status == 'Missed':
#             return 'pending'  # Can reschedule
        
#         return 'pending'

#     @action(detail=False, methods=['get'], url_path='summary')
#     def get_summary(self, request):
#         """Get summary statistics for retraining dashboard"""
#         # Get all failed employees
#         failed_response = self.get_failed_employees(request)
#         all_failed = failed_response.data['results']
        
#         # Calculate summary stats
#         total_failed = len(all_failed)
#         pending = len([emp for emp in all_failed if emp['retraining_status'] == 'pending'])
#         scheduled = len([emp for emp in all_failed if emp['retraining_status'] == 'scheduled'])
#         completed = len([emp for emp in all_failed if emp['retraining_status'] == 'completed'])
#         failed = len([emp for emp in all_failed if emp['retraining_status'] == 'failed'])
        
#         # Group by evaluation type
#         by_evaluation_type = {}
#         for emp in all_failed:
#             eval_type = emp['evaluation_type']
#             if eval_type not in by_evaluation_type:
#                 by_evaluation_type[eval_type] = {
#                     'total': 0,
#                     'pending': 0,
#                     'scheduled': 0,
#                     'completed': 0,
#                     'failed': 0
#                 }
#             by_evaluation_type[eval_type]['total'] += 1
#             by_evaluation_type[eval_type][emp['retraining_status']] += 1
        
#         # Group by department
#         by_department = {}
#         for emp in all_failed:
#             dept_name = emp['department_name']
#             if dept_name not in by_department:
#                 by_department[dept_name] = {
#                     'total': 0,
#                     'pending': 0,
#                     'scheduled': 0,
#                     'completed': 0,
#                     'failed': 0
#                 }
#             by_department[dept_name]['total'] += 1
#             by_department[dept_name][emp['retraining_status']] += 1
        
#         return Response({
#             'overall_summary': {
#                 'total_failed_employees': total_failed,
#                 'pending_retraining': pending,
#                 'scheduled_retraining': scheduled,
#                 'completed_retraining': completed,
#                 'failed_retraining': failed
#             },
#             'by_evaluation_type': by_evaluation_type,
#             'by_department': by_department
#         })


from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Q, Sum
from django.shortcuts import get_object_or_404

from .models import (
    RetrainingSession, RetrainingConfig, MasterTable,
    OperatorPerformanceEvaluation, TenCyclePassingCriteria,
    TraineeInfo, OJTPassingCriteria, OJTScoreRange,
    Score, EvaluationPassingCriteria,
    Department, Level, Station,
    OperatorObservanceSheet,  # <-- NEW
    RetrainingSessionDetail
)
from .serializers import RetrainingSessionSerializer, RetrainingConfigSerializer


class RetrainingConfigViewSet(viewsets.ModelViewSet):
    queryset = RetrainingConfig.objects.all()
    serializer_class = RetrainingConfigSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        level_id = self.request.query_params.get('level_id')
        evaluation_type = self.request.query_params.get('evaluation_type')
        if level_id:
            queryset = queryset.filter(level_id=level_id)
        if evaluation_type:
            queryset = queryset.filter(evaluation_type=evaluation_type)
        return queryset


class RetrainingSessionViewSet(viewsets.ModelViewSet):
    queryset = RetrainingSession.objects.select_related(
        'employee', 'level', 'department', 'station', 'session_detail'
    ).order_by('-created_at')
    serializer_class = RetrainingSessionSerializer

    # ------------------------------------------------------------------
    # CREATE – Schedule retraining
    # ------------------------------------------------------------------
    def create(self, request, *args, **kwargs):
        employee_id = request.data.get('employee')
        level_id = request.data.get('level')
        department_id = request.data.get('department')
        station_id = request.data.get('station')
        evaluation_type = request.data.get('evaluation_type')

        if not all([employee_id, level_id, department_id, evaluation_type]):
            return Response({
                "error": "Missing required fields: employee, level, department, evaluation_type"
            }, status=status.HTTP_400_BAD_REQUEST)

        config = RetrainingConfig.objects.filter(
            level_id=level_id, evaluation_type=evaluation_type
        ).first()
        max_retraining_sessions = config.max_count if config else 2

        filter_kwargs = {
            'employee_id': employee_id,
            'level_id': level_id,
            'department_id': department_id,
            'evaluation_type': evaluation_type
        }
        if station_id:
            filter_kwargs['station_id'] = station_id

        existing_count = RetrainingSession.objects.filter(**filter_kwargs).count()

        if existing_count >= max_retraining_sessions:
            return Response({
                'error': f'Maximum retraining sessions ({max_retraining_sessions}) reached.'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Set correct attempt number
        request.data['attempt_no'] = existing_count + 1

        return super().create(request, *args, **kwargs)

    # ------------------------------------------------------------------
    # PATCH – Complete session
    # ------------------------------------------------------------------
    @action(detail=True, methods=['patch'], url_path='complete-session')
    def complete_session(self, request, pk=None):
        session = self.get_object()
        if session.status != 'Pending':
            return Response({'error': 'Only pending sessions can be completed'}, status=status.HTTP_400_BAD_REQUEST)

        session.status = request.data.get('status', 'Completed')
        session.performance_percentage = request.data.get('performance_percentage')
        session.required_percentage = request.data.get('required_percentage')
        session.save()

        detail, _ = RetrainingSessionDetail.objects.get_or_create(retraining_session=session)
        if 'observations_failure_points' in request.data:
            detail.observations_failure_points = request.data['observations_failure_points']
        if 'trainer_name' in request.data:
            detail.trainer_name = request.data['trainer_name']
        detail.save()

        return Response(self.get_serializer(session).data)

    # ------------------------------------------------------------------
    # PATCH – Update observations only
    # ------------------------------------------------------------------
    @action(detail=True, methods=['patch'], url_path='update-observations')
    def update_observations(self, request, pk=None):
        session = self.get_object()
        detail, _ = RetrainingSessionDetail.objects.get_or_create(retraining_session=session)
        if 'observations_failure_points' in request.data:
            detail.observations_failure_points = request.data['observations_failure_points']
        if 'trainer_name' in request.data:
            detail.trainer_name = request.data['trainer_name']
        detail.save()
        return Response(self.get_serializer(session).data)

    # ------------------------------------------------------------------
    # GET – All failed employees (MAIN ENDPOINT)
    # ------------------------------------------------------------------
    @action(detail=False, methods=['get'], url_path='failed-employees')
    def get_failed_employees(self, request):
        failed_employees = []

        failed_employees.extend(self._get_10cycle_failed_employees())
        failed_employees.extend(self._get_ojt_failed_employees())
        failed_employees.extend(self._get_evaluation_failed_employees())
        failed_employees.extend(self._get_operator_observance_failed_employees())  # NEW
        failed_employees.extend(self._get_skill_level2_failed_employees())       # NEW

        # Filters
        department_id = request.query_params.get('department_id')
        evaluation_type = request.query_params.get('evaluation_type')
        status_filter = request.query_params.get('status')

        if department_id:
            failed_employees = [e for e in failed_employees if str(e.get('department_id')) == department_id]
        if evaluation_type:
            failed_employees = [e for e in failed_employees if e.get('evaluation_type') == evaluation_type]
        if status_filter:
            failed_employees = [e for e in failed_employees if e.get('retraining_status') == status_filter]

        return Response({'count': len(failed_employees), 'results': failed_employees})

    # ------------------------------------------------------------------
    # 1. 10-Cycle Failures
    # ------------------------------------------------------------------
    def _get_10cycle_failed_employees(self):
        failed = []
        evals = OperatorPerformanceEvaluation.objects.filter(
            Q(final_status__icontains='Fail') | 
            Q(final_status='Fail - Retraining Required')
        ).select_related('employee', 'department', 'station', 'level')

        for ev in evals:
            required = 60.0
            try:
                crit = TenCyclePassingCriteria.objects.get(
                    level=ev.level, department=ev.department, station=ev.station, is_active=True
                )
                required = crit.passing_percentage
            except TenCyclePassingCriteria.DoesNotExist:
                pass

            sessions = RetrainingSession.objects.filter(
                employee=ev.employee, level=ev.level, department=ev.department,
                station=ev.station, evaluation_type='10 Cycle'
            ).select_related('session_detail').order_by('-attempt_no')

            config = RetrainingConfig.objects.filter(level=ev.level, evaluation_type='10 Cycle').first()
            max_attempts = config.max_count if config else 2
            status = self._determine_retraining_status(sessions, max_attempts)

            failed.append(self._build_record(
                employee=ev.employee,
                department=ev.department,
                station=ev.station,
                level=ev.level,
                eval_type='10 Cycle',
                obtained=ev.final_percentage or 0,
                required=required,
                date=ev.date,
                sessions=sessions,
                max_attempts=max_attempts,
                status=status
            ))
        return failed

    # ------------------------------------------------------------------
    # 2. OJT Failures
    # ------------------------------------------------------------------
    def _get_ojt_failed_employees(self):
        failed = []
        for trainee in TraineeInfo.objects.filter(status='Fail').select_related('level', 'station'):
            dept = trainee.station.department if trainee.station and trainee.station.department else None
            if not dept:
                continue

            total = trainee.scores.aggregate(s=Sum('score'))['s'] or 0
            topics = trainee.scores.count()
            max_per = 5
            try:
                range_obj = OJTScoreRange.objects.get(department=dept, level=trainee.level)
                max_per = range_obj.max_score
            except OJTScoreRange.DoesNotExist:
                pass
            obtained = (total / (topics * max_per) * 100) if topics else 0

            required = 60.0
            try:
                crit = OJTPassingCriteria.objects.get(department=dept, level=trainee.level)
                required = crit.percentage
            except OJTPassingCriteria.DoesNotExist:
                pass

            employee = MasterTable.objects.filter(emp_id=trainee.emp_id).first()
            if not employee:
                continue

            sessions = RetrainingSession.objects.filter(
                employee=employee, level=trainee.level, department=dept, evaluation_type='OJT'
            ).select_related('session_detail').order_by('-attempt_no')

            config = RetrainingConfig.objects.filter(level=trainee.level, evaluation_type='OJT').first()
            max_attempts = config.max_count if config else 2
            status = self._determine_retraining_status(sessions, max_attempts)

            failed.append(self._build_record(
                employee=employee, department=dept, station=trainee.station, level=trainee.level,
                eval_type='OJT', obtained=obtained, required=required, date=trainee.doj,
                sessions=sessions, max_attempts=max_attempts, status=status
            ))
        return failed

    # ------------------------------------------------------------------
    # 3. Evaluation (Test) Failures
    # ------------------------------------------------------------------
    def _get_evaluation_failed_employees(self):
        failed = []
        for score in Score.objects.filter(passed=False).select_related('employee', 'level', 'skill', 'department'):
            dept = score.department or (score.skill.subline.line.department if score.skill and score.skill.subline else None)
            if not dept or not score.level:
                continue

            required = 80.0
            try:
                crit = EvaluationPassingCriteria.objects.get(department=dept, level=score.level)
                required = float(crit.percentage)
            except EvaluationPassingCriteria.DoesNotExist:
                pass

            sessions = RetrainingSession.objects.filter(
                employee=score.employee, level=score.level, department=dept, evaluation_type='Evaluation'
            ).select_related('session_detail').order_by('-attempt_no')

            config = RetrainingConfig.objects.filter(level=score.level, evaluation_type='Evaluation').first()
            max_attempts = config.max_count if config else 2
            status = self._determine_retraining_status(sessions, max_attempts)

            failed.append(self._build_record(
                employee=score.employee, department=dept, station=score.skill, level=score.level,
                eval_type='Evaluation', obtained=score.percentage, required=required, date=score.created_at.date(),
                sessions=sessions, max_attempts=max_attempts, status=status
            ))
        return failed

    # ------------------------------------------------------------------
    # 4. Operator Observance Failures (FIXED)
    # ------------------------------------------------------------------
    def _get_operator_observance_failed_employees(self):
        failed = []
        sheets = OperatorObservanceSheet.objects.filter(
            Q(result='Fail') | Q(result='Re-training')
        )

        for sheet in sheets:
            # Resolve Employee
            employee = None
            if sheet.operator_name.isdigit():
                employee = MasterTable.objects.filter(emp_id=sheet.operator_name).first()
            if not employee:
                parts = sheet.operator_name.strip().split()
                if len(parts) >= 2:
                    employee = MasterTable.objects.filter(
                        first_name__iexact=parts[0],
                        last_name__iexact=' '.join(parts[1:])
                    ).first()

            if not employee:
                continue

            # Resolve Dept & Level from Topics
            dept = level = None
            for topic in sheet.topics.all():
                if topic.department and not dept:
                    dept = topic.department
                if topic.level and not level:
                    level = topic.level
                if dept and level:
                    break
            if not (dept and level):
                continue

            # Obtained %
            obtained = 0.0
            if sheet.marks_obtained:
                try:
                    obtained = float(sheet.marks_obtained.split('/')[0])
                except:
                    pass
            else:
                total_o = sum(1 for day in sheet.marks.values() for v in day.values() if v == 'O')
                total = sum(len(day) for day in sheet.marks.values())
                obtained = (total_o / total * 100) if total > 0 else 0

            required = 80.0
            try:
                crit = EvaluationPassingCriteria.objects.get(department=dept, level=level)
                required = float(crit.percentage)
            except EvaluationPassingCriteria.DoesNotExist:
                pass

            sessions = RetrainingSession.objects.filter(
                employee=employee, department=dept, level=level, evaluation_type='Operator Observance'
            ).select_related('session_detail').order_by('-attempt_no')

            config = RetrainingConfig.objects.filter(level=level, evaluation_type='Operator Observance').first()
            max_attempts = config.max_count if config else 2
            status = self._determine_retraining_status(sessions, max_attempts)

            failed.append(self._build_record(
                employee=employee, department=dept, station=None, level=level,
                eval_type='Operator Observance', obtained=obtained, required=required,
                date=sheet.evaluation_end_date, sessions=sessions, max_attempts=max_attempts, status=status
            ))
        return failed

    # ------------------------------------------------------------------
    # 5. Skill Evaluation Level 2 Failures
    # ------------------------------------------------------------------
    def _get_skill_level2_failed_employees(self):
        failed = []
        evals = EvaluationLevel2.objects.filter(
            Q(status=EvaluationLevel2.STATUS_FAIL) |
            Q(status=EvaluationLevel2.STATUS_RE_EVAL_FAIL)
        ).select_related('employee', 'department', 'level')

        for ev in evals:
            required = 80.0
            try:
                crit = EvaluationPassingCriteria.objects.get(department=ev.department, level=ev.level)
                required = float(crit.percentage)
            except EvaluationPassingCriteria.DoesNotExist:
                pass

            sessions = RetrainingSession.objects.filter(
                employee=ev.employee, department=ev.department, level=ev.level,
                evaluation_type='Skill Evaluation Level 2'
            ).select_related('session_detail').order_by('-attempt_no')

            config = RetrainingConfig.objects.filter(level=ev.level, evaluation_type='Skill Evaluation Level 2').first()
            max_attempts = config.max_count if config else 2
            status = self._determine_retraining_status(sessions, max_attempts)

            failed.append(self._build_record(
                employee=ev.employee, department=ev.department, station=None, level=ev.level,
                eval_type='Skill Evaluation Level 2', obtained=ev.total_marks, required=required,
                date=ev.evaluation_date, sessions=sessions, max_attempts=max_attempts, status=status
            ))
        return failed

    # ------------------------------------------------------------------
    # Helper: Build uniform record
    # ------------------------------------------------------------------
    def _build_record(self, employee, department, station, level, eval_type,
                      obtained, required, date, sessions, max_attempts, status):
        return {
            'employee_pk': employee.pk,
            'employee_id': employee.emp_id,
            'employee_name': f"{employee.first_name} {employee.last_name}".strip(),
            'department_id': department.department_id,
            'department_name': department.department_name,
            'station_id': station.station_id if station else None,
            'station_name': station.station_name if station else 'N/A',
            'level_id': level.level_id,
            'level_name': level.level_name,
            'evaluation_type': eval_type,
            'obtained_percentage': round(obtained, 2),
            'required_percentage': required,
            'performance_gap': round(required - obtained, 2),
            'last_evaluation_date': date.isoformat(),
            'existing_sessions_count': sessions.count(),
            'max_attempts': max_attempts,
            'can_schedule_retraining': sessions.count() < max_attempts,
            'retraining_status': status,
            'retraining_records': [
                {
                    'id': s.id,
                    'attempt_no': s.attempt_no,
                    'scheduled_date': s.scheduled_date.isoformat(),
                    'scheduled_time': s.scheduled_time.strftime('%H:%M'),
                    'venue': s.venue,
                    'status': s.status,
                    'performance_percentage': s.performance_percentage,
                    'session_detail': {
                        'observations_failure_points': s.session_detail.observations_failure_points
                        if hasattr(s, 'session_detail') else None,
                        'trainer_name': s.session_detail.trainer_name
                        if hasattr(s, 'session_detail') else None,
                    } if hasattr(s, 'session_detail') else None
                } for s in sessions
            ]
        }

    # ------------------------------------------------------------------
    # Helper: Determine status
    # ------------------------------------------------------------------
    def _determine_retraining_status(self, sessions, max_attempts):
        if not sessions.exists():
            return 'pending'
        latest = sessions.first()
        if sessions.count() >= max_attempts:
            return 'completed' if (
                latest.status == 'Completed' and
                latest.performance_percentage and
                latest.performance_percentage >= (latest.required_percentage or 60)
            ) else 'failed'
        if latest.status == 'Pending':
            return 'scheduled'
        if latest.status == 'Completed':
            return 'completed' if (
                latest.performance_percentage and
                latest.performance_percentage >= (latest.required_percentage or 60)
            ) else 'pending'
        return 'pending'

    # ------------------------------------------------------------------
    # Summary
    # ------------------------------------------------------------------
    @action(detail=False, methods=['get'], url_path='summary')
    def get_summary(self, request):
        resp = self.get_failed_employees(request)
        all_failed = resp.data['results']

        total = len(all_failed)
        by_status = {'pending': 0, 'scheduled': 0, 'completed': 0, 'failed': 0}
        by_type = {}
        by_dept = {}

        for emp in all_failed:
            s = emp['retraining_status']
            by_status[s] += 1
            t = emp['evaluation_type']
            d = emp['department_name']
            by_type.setdefault(t, {'total': 0, 'pending': 0, 'scheduled': 0, 'completed': 0, 'failed': 0})
            by_type[t]['total'] += 1
            by_type[t][s] += 1
            by_dept.setdefault(d, {'total': 0, 'pending': 0, 'scheduled': 0, 'completed': 0, 'failed': 0})
            by_dept[d]['total'] += 1
            by_dept[d][s] += 1

        return Response({
            'overall_summary': {
                'total_failed_employees': total,
                'pending_retraining': by_status['pending'],
                'scheduled_retraining': by_status['scheduled'],
                'completed_retraining': by_status['completed'],
                'failed_retraining': by_status['failed']
            },
            'by_evaluation_type': by_type,
            'by_department': by_dept
        })
# =================== Retraining end ============================= #

from rest_framework import viewsets
from .models import (
    QuantityScoreSetup,
    QuantityPassingCriteria,
    OJTLevel2Quantity,
    Level2QuantityOJTEvaluation,
)
from .serializers import (
    QuantityScoreSetupSerializer,
    QuantityPassingCriteriaSerializer,
    OJTLevel2QuantitySerializer,
    Level2QuantityOJTEvaluationSerializer,
)


# ----------------------------
# Quantity Score Ranges
# ----------------------------
class QuantityOJTScoreRangeViewSet(viewsets.ModelViewSet):
    queryset = QuantityScoreSetup.objects.all()
    serializer_class = QuantityScoreSetupSerializer


# ----------------------------
# Quantity Passing Criteria
# ----------------------------
class QuantityPassingCriteriaViewSet(viewsets.ModelViewSet):
    queryset = QuantityPassingCriteria.objects.all()
    serializer_class = QuantityPassingCriteriaSerializer


# ----------------------------
# OJT Main Record (Trainee)
# ----------------------------

# ----------------------------
# Daily Evaluation
# ----------------------------
class Level2QuantityOJTEvaluationViewSet(viewsets.ModelViewSet):
    queryset = Level2QuantityOJTEvaluation.objects.all()
    serializer_class = Level2QuantityOJTEvaluationSerializer


from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import AssessmentMode
from .serializers import AssessmentModeSerializer

@api_view(['GET'])
def get_assessment_mode(request):
    """Get current assessment mode"""
    current_mode = AssessmentMode.get_current_mode()
    return Response({
        'mode': current_mode.mode,
        'updated_at': current_mode.updated_at
    })

@api_view(['POST'])
def toggle_assessment_mode(request):
    """Toggle between quality and quantity mode"""
    mode = request.data.get('mode')
    
    if mode not in ['quality', 'quantity']:
        return Response({
            'error': 'Invalid mode. Must be "quality" or "quantity"'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    current_mode = AssessmentMode.get_current_mode()
    current_mode.mode = mode
    current_mode.save()
    
    return Response({
        'mode': current_mode.mode,
        'updated_at': current_mode.updated_at,
        'message': f'Mode switched to {mode}'
    })


from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import LevelColour, Level
from .serializers import LevelColourSerializer

class LevelColourViewSet(viewsets.ModelViewSet):
    queryset = LevelColour.objects.all()
    serializer_class = LevelColourSerializer
    
    @action(detail=False, methods=['post'])
    def reset_to_defaults(self, request):
        """Reset all level colors to default values"""
        try:
            for level_id, color_code in DEFAULT_COLOURS.items():
                try:
                    level = Level.objects.get(level_id=level_id)
                    level_colour, created = LevelColour.objects.get_or_create(level=level)
                    level_colour.colour_code = color_code
                    level_colour.save()
                except Level.DoesNotExist:
                    continue
            
            # Return updated colors
            colours = LevelColour.objects.all()
            serializer = self.get_serializer(colours, many=True)
            return Response(serializer.data)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def bulk_update(self, request):
        """Update multiple level colors at once"""
        try:
            colors_data = request.data.get('colors', {})
            updated_colors = []
            
            for level_key, color_code in colors_data.items():
                # Extract level number from key like 'level1' -> 1
                level_num = int(level_key.replace('level', ''))
                
                try:
                    level = Level.objects.get(level_id=level_num)
                    level_colour, created = LevelColour.objects.get_or_create(level=level)
                    level_colour.colour_code = color_code
                    level_colour.save()
                    updated_colors.append(level_colour)
                except Level.DoesNotExist:
                    continue
            
            serializer = self.get_serializer(updated_colors, many=True)
            return Response(serializer.data)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        

from rest_framework import viewsets, status
from rest_framework.response import Response
from .models import SkillMatrixDisplaySetting
from .serializers import SkillMatrixDisplaySettingSerializer

class SkillMatrixDisplaySettingViewSet(viewsets.ViewSet):
    """
    A simple ViewSet to get/set display shape.
    """

    def list(self, request):
        # Retrieve or create singleton instance
        setting, created = SkillMatrixDisplaySetting.objects.get_or_create(id=1)
        serializer = SkillMatrixDisplaySettingSerializer(setting)
        return Response(serializer.data)

    def update(self, request, pk=None):
        setting, created = SkillMatrixDisplaySetting.objects.get_or_create(id=1)
        serializer = SkillMatrixDisplaySettingSerializer(setting, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    


import openpyxl
import re
from datetime import date
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

from django.http import HttpResponse
from django.db import transaction
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response

from .models import HierarchyStructure, SkillMatrix, Level, Department, Line, SubLine

# class SkillMatrixExcelHandlerView(APIView):
#     """
#     Handles the download and upload of a matrix-style Excel template.
#     Uses hidden IDs for robust data lookup.
#     """

#     def get(self, request, *args, **kwargs):
#         # This GET method is correct and does not need changes.
#         # It's included here so you can copy the entire class.
#         department_id = request.query_params.get('department')
#         line_id = request.query_params.get('line')
#         # ... (rest of GET method is unchanged from the last correct version) ...
#         subline_id = request.query_params.get('subline')

#         if not department_id:
#             return Response({"error": "Department ID is required."}, status=status.HTTP_400_BAD_REQUEST)

#         filters = {'department_id': department_id}
#         if line_id:
#             filters['line_id'] = line_id
#         if subline_id:
#             filters['subline_id'] = subline_id
        
#         hierarchy_nodes = HierarchyStructure.objects.filter(**filters).select_related(
#             'station', 'department', 'line'
#         ).order_by('station__station_name')

#         if not hierarchy_nodes.exists():
#             return Response({"error": "No stations found for the selected hierarchy."}, status=status.HTTP_404_NOT_FOUND)

#         first_node = hierarchy_nodes.first()
#         department_name = first_node.department.department_name
#         line_name = first_node.line.line_name if first_node.line else "All Lines"
#         actual_line_id = first_node.line.line_id if first_node.line else "" # Assuming line_id
#         stations = [node.station for node in hierarchy_nodes if node.station]

#         wb = openpyxl.Workbook()
#         ws_data = wb.active
#         ws_data.title = "Data_Upload"
#         ws_inst = wb.create_sheet("Instructions")

#         header_font = Font(bold=True, color="FFFFFF")
#         header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
#         info_font = Font(bold=True)
#         center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
#         thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

#         ws_data.merge_cells('A1:B1')
#         ws_data['A1'] = 'Department:'
#         ws_data['A1'].font = info_font
#         ws_data.merge_cells('C1:E1')
#         ws_data['C1'] = department_name

#         ws_data.merge_cells('A2:B2')
#         ws_data['A2'] = 'Line:'
#         ws_data['A2'].font = info_font
#         ws_data.merge_cells('C2:E2')
#         ws_data['C2'] = line_name

#         ws_data['A3'] = "department_id"
#         ws_data['B3'] = department_id
#         ws_data['A4'] = "line_id"
#         ws_data['B4'] = actual_line_id
#         ws_data.row_dimensions[3].hidden = True
#         ws_data.row_dimensions[4].hidden = True

#         HEADER_ROW = 5
#         main_headers = ["S.No", "Operator Name", "Employee ID", "Date of Joining (DD/MM/YYYY)"]
#         station_headers = [station.station_name for station in stations]
#         full_headers = main_headers + station_headers
        
#         ws_data.row_dimensions[HEADER_ROW].height = 40
#         for col_num, header_title in enumerate(full_headers, 1):
#             cell = ws_data.cell(row=HEADER_ROW, column=col_num, value=header_title)
#             cell.font = header_font
#             cell.fill = header_fill
#             cell.alignment = center_align
#             cell.border = thin_border
        
#         DUMMY_DATA_START_ROW = 6
#         dummy_data = [
#             (1, "Sahil", "00100022", date(2025, 1, 6)),
#             (2, "Navinder Singh", "00100035", date(2025, 1, 21)),
#         ]
        
#         current_data_row = DUMMY_DATA_START_ROW
#         for s_no, name, emp_id, doj in dummy_data:
#             ws_data.cell(row=current_data_row, column=1, value=s_no)
#             ws_data.cell(row=current_data_row, column=2, value=name)
#             ws_data.cell(row=current_data_row, column=3, value=emp_id)
#             ws_data.cell(row=current_data_row, column=4, value=doj).number_format = 'DD/MM/YYYY'
#             current_data_row += 1
            
#         if len(stations) > 1:
#             ws_data.cell(row=DUMMY_DATA_START_ROW, column=5, value="Level 4")
#             ws_data.cell(row=DUMMY_DATA_START_ROW, column=6, value="Level 2")

#         virtual_workbook = BytesIO()
#         wb.save(virtual_workbook)
#         virtual_workbook.seek(0)
#         response = HttpResponse(
#             virtual_workbook.read(),
#             content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )
#         response['Content-Disposition'] = 'attachment; filename="SkillMatrix_Template_Robust.xlsx"'
#         return response

#     @transaction.atomic
#     def post(self, request, *args, **kwargs):
#         excel_file = request.FILES.get('file')
#         if not excel_file:
#             return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
        
#         try:
#             wb = openpyxl.load_workbook(excel_file, data_only=True)
#             ws = wb["Data_Upload"]
#         except Exception:
#             return Response({"error": "Could not open the Excel file or find 'Data_Upload' sheet."}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             department_id = ws['B3'].value
#             line_id = ws['B4'].value
#             department_name = ws['C1'].value
#             HEADER_ROW = 5
#             headers = [cell.value for cell in ws[HEADER_ROW]]
#             station_headers = headers[4:]

#             if not department_id:
#                 return Response({"error": "Template is corrupt. Department ID is missing."}, status=status.HTTP_400_BAD_REQUEST)
        
#         except Exception:
#             return Response({"error": "Could not parse header information. The template might be corrupt."}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             department = Department.objects.get(department_id=department_id)
#             line = Line.objects.get(line_id=line_id) if line_id else None

#             hierarchy_filter = {'department': department}
#             if line:
#                 hierarchy_filter['line'] = line

#             hierarchy_map = {
#                 h.station.station_name: h for h in HierarchyStructure.objects.filter(
#                     **hierarchy_filter
#                 ).select_related('station')
#             }
#             # This map correctly gets the Level object for each "Level X" string
#             level_map = {level.level_name: level for level in Level.objects.all()}

#         except Department.DoesNotExist:
#             return Response({"error": f"Department with ID '{department_id}' not found in the database."}, status=status.HTTP_400_BAD_REQUEST)
#         except Line.DoesNotExist:
#              return Response({"error": f"Line with ID '{line_id}' not found for Department '{department_name}'."}, status=status.HTTP_400_BAD_REQUEST)
#         except Exception as e:
#             return Response({"error": f"An unexpected error occurred during database lookup: {e}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#         errors = []
#         successful_uploads = 0
        
#         DATA_START_ROW = 6
#         for row_idx, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
#             emp_name = row[1]
#             emp_id_from_excel = row[2]
#             doj = row[3]

#             if not emp_id_from_excel or not emp_name:
#                 continue


#             try:
#                 # Assuming your MasterTable has a field named 'emp_id' to look up against. Adjust if necessary.
#                 employee_obj = MasterTable.objects.get(emp_id=str(emp_id_from_excel).strip())
#             except MasterTable.DoesNotExist:
#                 errors.append(f"Row {row_idx}: Employee with ID '{emp_id_from_excel}' not found in the MasterTable. Please add them first.")
#                 continue # Skip all skills for this non-existent employee

#             for col_idx, skill_level_raw in enumerate(row[4:]):
#                 if not skill_level_raw:
#                     continue

#                 level_num_match = re.search(r'\d+', str(skill_level_raw))
#                 if not level_num_match:
#                     errors.append(f"Row {row_idx}: Invalid level format '{skill_level_raw}'. Please use 'Level X' or 'X'.")
#                     continue
                
#                 level_num = int(level_num_match.group(0))
#                 level_search_key = f"Level {level_num}"
                
#                 if level_search_key not in level_map:
#                     errors.append(f"Row {row_idx}: '{level_search_key}' is not a valid skill level in the database.")
#                     continue
                
#                 station_name = station_headers[col_idx]
#                 hierarchy_obj = hierarchy_map.get(station_name)
                
#                 if not hierarchy_obj:
#                     errors.append(f"Row {row_idx}: Station '{station_name}' could not be matched for the selected hierarchy.")
#                     continue
                
#                 try:
#                     SkillMatrix.objects.update_or_create(
#                         # 1. Use the correct lookup keys based on your model
#                         hierarchy=hierarchy_obj,
#                         employee=employee_obj,
                        
#                         # 2. Provide ALL other fields in the defaults dictionary
#                         defaults={
#                             'employee_name': str(emp_name).strip(),
#                             'emp_id': str(emp_id_from_excel).strip(),
#                             'doj': doj,
#                             'level': level_map[level_search_key]
#                         }
#                     )
#                     successful_uploads += 1
#                 except Exception as e:
#                     # --- FIX #2: Print the REAL exception 'e' for proper debugging ---
#                     errors.append(f"Row {row_idx}: Could not save data for Emp ID {emp_id} at Station {station_name}. Error: {e}")

#         if errors:
#             error_message = f"Processed {successful_uploads} records. However, {len(errors)} errors occurred."
#             return Response({"status": error_message, "errors": errors}, status=status.HTTP_400_BAD_REQUEST)
        
#         return Response({"status": f"Successfully uploaded/updated {successful_uploads} skill matrix records!"}, status=status.HTTP_201_CREATED)




class SkillMatrixExcelHandlerView(APIView):
    """
    Handles the download and upload of a matrix-style Excel template.
    Uses hidden IDs for robust data lookup.
    """

    def get(self, request, *args, **kwargs):
        department_id = request.query_params.get('department')
        line_id = request.query_params.get('line')
        subline_id = request.query_params.get('subline')

        if not department_id:
            return Response({"error": "Department ID is required."}, status=status.HTTP_400_BAD_REQUEST)

        filters = {'department_id': department_id}
        if line_id:
            filters['line_id'] = line_id
        if subline_id:
            filters['subline_id'] = subline_id
        
        hierarchy_nodes = HierarchyStructure.objects.filter(**filters).select_related(
            'station', 'department', 'line'
        ).order_by('station__station_name')

        if not hierarchy_nodes.exists():
            return Response({"error": "No stations found for the selected hierarchy."}, status=status.HTTP_404_NOT_FOUND)

        first_node = hierarchy_nodes.first()
        department_name = first_node.department.department_name
        line_name = first_node.line.line_name if first_node.line else "All Lines"
        actual_line_id = first_node.line.line_id if first_node.line else ""
        stations = [node.station for node in hierarchy_nodes if node.station]

        wb = openpyxl.Workbook()
        
        # --- FIX: Create Instructions first, then Data_Upload ---
        ws_inst = wb.active
        ws_inst.title = "Instructions"
        ws_data = wb.create_sheet("Data_Upload")

        # Define common styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        info_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # --- FIX: Add content to the Instructions sheet ---
        ws_inst['A1'] = "Instructions for Use"
        ws_inst['A1'].font = Font(bold=True, size=14)
        ws_inst.merge_cells('A1:C1')

        ws_inst['A3'] = "Step"
        ws_inst['A3'].font = header_font
        ws_inst['A3'].fill = header_fill
        ws_inst['B3'] = "Description"
        ws_inst['B3'].font = header_font
        ws_inst['B3'].fill = header_fill
        
        instructions_data = [
            ("1. Context", f"This template is for uploading skills for Department: '{department_name}' and Line: '{line_name}'."),
            ("2. Go to Sheet", "After reading these instructions, please switch to the 'Data_Upload' sheet to enter your data."),
            ("3. Do Not Change", "In 'Data_Upload', do not modify the headers in Row 5 or the hidden IDs in Rows 3 and 4. They are required for the upload to work correctly."),
            ("4. Fill Data", "Fill in the operator details and their skill levels for each station shown in the columns."),
            ("   - Employee ID", "This ID MUST match an existing employee in the master database. New employees cannot be added via this sheet."),
            ("   - Date of Joining", "Use the format DD/MM/YYYY."),
            ("   - Station Columns", "For each station, enter the skill level (e.g., 'Level 1', 'Level 2'). Leave the cell blank if the operator has no skill for that station."),
            ("5. Sample Data", "The first few rows in 'Data_Upload' contain sample data to show the correct format. Please DELETE these rows before entering your own data."),
            ("6. Upload", "Save the file and upload it using the 'Upload Excel' button in the application.")
        ]
        
        current_inst_row = 4
        for step, desc in instructions_data:
            cell_a = ws_inst.cell(row=current_inst_row, column=1, value=step)
            cell_b = ws_inst.cell(row=current_inst_row, column=2, value=desc)
            cell_a.font = Font(bold=True)
            cell_b.alignment = Alignment(wrap_text=True, vertical='top')
            current_inst_row += 1

        ws_inst.column_dimensions['A'].width = 20
        ws_inst.column_dimensions['B'].width = 80
        # --- END of Instructions content ---

        # --- Populate the Data_Upload sheet (ws_data) ---
        ws_data.merge_cells('A1:B1')
        ws_data['A1'] = 'Department:'
        ws_data['A1'].font = info_font
        ws_data.merge_cells('C1:E1')
        ws_data['C1'] = department_name

        ws_data.merge_cells('A2:B2')
        ws_data['A2'] = 'Line:'
        ws_data['A2'].font = info_font
        ws_data.merge_cells('C2:E2')
        ws_data['C2'] = line_name

        ws_data['A3'] = "department_id"
        ws_data['B3'] = department_id
        ws_data['A4'] = "line_id"
        ws_data['B4'] = actual_line_id
        ws_data.row_dimensions[3].hidden = True
        ws_data.row_dimensions[4].hidden = True

        HEADER_ROW = 5
        main_headers = ["S.No", "Operator Name", "Employee ID", "Date of Joining (DD/MM/YYYY)"]
        station_headers = [station.station_name for station in stations]
        full_headers = main_headers + station_headers
        
        ws_data.row_dimensions[HEADER_ROW].height = 40
        for col_num, header_title in enumerate(full_headers, 1):
            cell = ws_data.cell(row=HEADER_ROW, column=col_num, value=header_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        DUMMY_DATA_START_ROW = 6
        dummy_data = [
            (1, "Sahil", "00100022", date(2025, 1, 6)),
            (2, "Navinder Singh", "00100035", date(2025, 1, 21)),
        ]
        
        current_data_row = DUMMY_DATA_START_ROW
        for s_no, name, emp_id, doj in dummy_data:
            ws_data.cell(row=current_data_row, column=1, value=s_no)
            ws_data.cell(row=current_data_row, column=2, value=name)
            ws_data.cell(row=current_data_row, column=3, value=emp_id)
            ws_data.cell(row=current_data_row, column=4, value=doj).number_format = 'DD/MM/YYYY'
            current_data_row += 1
            
        if len(stations) > 1:
            ws_data.cell(row=DUMMY_DATA_START_ROW, column=5, value="Level 4")
            ws_data.cell(row=DUMMY_DATA_START_ROW, column=6, value="Level 2")

        # --- FIX: Set the active sheet to be Data_Upload on open ---
        wb.active = ws_data

        virtual_workbook = BytesIO()
        wb.save(virtual_workbook)
        virtual_workbook.seek(0)
        response = HttpResponse(
            virtual_workbook.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="SkillMatrix_Template_Robust.xlsx"'
        return response

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb["Data_Upload"]
        except Exception:
            return Response({"error": "Could not open the Excel file or find 'Data_Upload' sheet."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            department_id = ws['B3'].value
            line_id = ws['B4'].value
            department_name = ws['C1'].value # Used for error messages
            HEADER_ROW = 5
            headers = [cell.value for cell in ws[HEADER_ROW]]
            station_headers = headers[4:]

            if not department_id:
                return Response({"error": "Template is corrupt. Department ID is missing."}, status=status.HTTP_400_BAD_REQUEST)
        
        except Exception:
            return Response({"error": "Could not parse header information. The template might be corrupt."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            department = Department.objects.get(department_id=department_id)
            line = Line.objects.get(line_id=line_id) if line_id else None

            hierarchy_filter = {'department': department}
            if line:
                hierarchy_filter['line'] = line

            hierarchy_map = {
                h.station.station_name: h for h in HierarchyStructure.objects.filter(
                    **hierarchy_filter
                ).select_related('station')
            }
            level_map = {level.level_name: level for level in Level.objects.all()}

        except Department.DoesNotExist:
            return Response({"error": f"Department with ID '{department_id}' not found in the database."}, status=status.HTTP_400_BAD_REQUEST)
        except Line.DoesNotExist:
             return Response({"error": f"Line with ID '{line_id}' not found for Department '{department_name}'."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": f"An unexpected error occurred during database lookup: {e}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        errors = []
        successful_uploads = 0
        
        DATA_START_ROW = 6
        for row_idx, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
            emp_name = row[1]
            emp_id_from_excel = row[2]
            doj = row[3]

            if not emp_id_from_excel or not emp_name:
                continue

            try:
                employee_obj = MasterTable.objects.get(emp_id=str(emp_id_from_excel).strip())
            except MasterTable.DoesNotExist:
                errors.append(f"Row {row_idx}: Employee with ID '{emp_id_from_excel}' not found in the MasterTable. Please add them first.")
                continue

            for col_idx, skill_level_raw in enumerate(row[4:]):
                if not skill_level_raw:
                    continue

                level_num_match = re.search(r'\d+', str(skill_level_raw))
                if not level_num_match:
                    errors.append(f"Row {row_idx}: Invalid level format '{skill_level_raw}' for station '{station_headers[col_idx]}'. Please use 'Level X' or 'X'.")
                    continue
                
                level_num = int(level_num_match.group(0))
                level_search_key = f"Level {level_num}"
                
                if level_search_key not in level_map:
                    errors.append(f"Row {row_idx}: '{level_search_key}' is not a valid skill level in the database.")
                    continue
                
                station_name = station_headers[col_idx]
                hierarchy_obj = hierarchy_map.get(station_name)
                
                if not hierarchy_obj:
                    errors.append(f"Row {row_idx}: Station '{station_name}' could not be matched for the selected hierarchy.")
                    continue
                
                try:
                    SkillMatrix.objects.update_or_create(
                        hierarchy=hierarchy_obj,
                        employee=employee_obj,
                        defaults={
                            'employee_name': str(emp_name).strip(),
                            'emp_id': str(emp_id_from_excel).strip(),
                            'doj': doj,
                            'level': level_map[level_search_key]
                        }
                    )
                    successful_uploads += 1
                except Exception as e:
                    errors.append(f"Row {row_idx}: Could not save data for Emp ID {emp_id_from_excel} at Station {station_name}. Error: {e}")

        if errors:
            error_message = f"Processed {successful_uploads} records. However, {len(errors)} errors occurred."
            return Response({"status": error_message, "errors": errors}, status=status.HTTP_400_BAD_REQUEST)
        
        return Response({"status": f"Successfully uploaded/updated {successful_uploads} skill matrix records!"}, status=status.HTTP_201_CREATED)
















class DepartmentSubLineViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = SubLineSerializer
    
    def get_queryset(self):
        department_id = self.request.query_params.get('department_id')
        if department_id:
            return SubLine.objects.filter(line_department_department_id=department_id)
        return SubLine.objects.none()
    
    def list(self, request, *args, **kwargs):
        try:
            department_id = request.query_params.get('department_id')
            
            if not department_id:
                response = Response(
                    {"error": "department_id parameter is required"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
                response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
                return response
            
            # Verify department exists
            try:
                department = Department.objects.get(department_id=department_id)
            except Department.DoesNotExist:
                response = Response(
                    {"error": "Department not found"}, 
                    status=status.HTTP_404_NOT_FOUND
                )
                response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
                return response
            
            queryset = self.get_queryset()
            serializer = self.get_serializer(queryset, many=True)
            
            response_data = {
                'department_id': department.department_id,
                'department_name': department.department_name,
                'sublines': serializer.data,
                'count': queryset.count()
            }
            
            response = Response(response_data, status=status.HTTP_200_OK)
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            
            return response
            
        except Exception as e:
            response = Response(
                {"error": "Failed to fetch sublines", "details": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            return response


class DepartmentStationViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = StationSerializer
    
    def get_queryset(self):
        department_id = self.request.query_params.get('department_id')
        if department_id:
            return Station.objects.filter(subline_linedepartment_department_id=department_id)
        return Station.objects.none()
    
    def list(self, request, *args, **kwargs):
        try:
            department_id = request.query_params.get('department_id')
            
            if not department_id:
                response = Response(
                    {"error": "department_id parameter is required"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
                response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
                return response
            
            # Verify department exists
            try:
                department = Department.objects.get(department_id=department_id)
            except Department.DoesNotExist:
                response = Response(
                    {"error": "Department not found"}, 
                    status=status.HTTP_404_NOT_FOUND
                )
                response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
                return response
            
            queryset = self.get_queryset()
            serializer = self.get_serializer(queryset, many=True)
            
            response_data = {
                'department_id': department.department_id,
                'department_name': department.department_name,
                'stations': serializer.data,
                'count': queryset.count()
            }
            
            response = Response(response_data, status=status.HTTP_200_OK)
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            
            return response
            
        except Exception as e:
            response = Response(
                {"error": "Failed to fetch stations", "details": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            return response


from rest_framework import viewsets
from .models import CompanyLogo
from .serializers import CompanyLogoSerializer

class CompanyLogoViewSet(viewsets.ModelViewSet):
    queryset = CompanyLogo.objects.all()
    serializer_class = CompanyLogoSerializer


@api_view(['GET'])
def get_lines_by_department(request, department_id):
    department = get_object_or_404(Department, department_id=department_id)
    lines = Line.objects.filter(
        department=department_id
    ).select_related('department')
    
    if not lines.exists():
        return Response({
            'message': f'No lines found under department: {department.department_name}',
            'department_name': department.department_name,
            'lines': [],
            'total_count': 0
        }, status=status.HTTP_200_OK)
    
    serializer = LineReadSerializer(lines, many=True)
    
    return Response({
        'lines': serializer.data,
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
def get_sublines_by_line(request, line_id):
    # Check if line exists
    line = get_object_or_404(Line, line_id=line_id)
    
    # Get all sublines under this line
    sublines = SubLine.objects.filter(
        line=line_id
    ).select_related('line', 'line__department')
    
    if not sublines.exists():
        return Response({
            'message': f'No sublines found under line: {line.line_name}',
            'line_name': line.line_name,
            'department_name': line.department.department_name,
            'sublines': [],
            'total_count': 0
        }, status=status.HTTP_200_OK)
    
    serializer = SubLineReadSerializer(sublines, many=True)
    
    return Response({
        'line_id': line_id,
        'line_name': line.line_name,
        'sublines': serializer.data,
    }, status=status.HTTP_200_OK)

@api_view(['GET'])
def get_stations_by_subline(request, subline_id):
    
    # Check if subline exists
    subline = get_object_or_404(SubLine, subline_id=subline_id)
    
    # Get all stations under this subline
    stations = Station.objects.filter(
        subline=subline_id
    ).select_related('subline', 'subline__line', 'subline__line__department')
    
    if not stations.exists():
        return Response({
            'message': f'No stations found under subline: {subline.subline_name}',
            'subline_name': subline.subline_name,
            'line_name': subline.line.line_name,
            'department_name': subline.line.department.department_name,
            'stations': [],
            'total_count': 0
        }, status=status.HTTP_200_OK)
    
    serializer = StationReadSerializer(stations, many=True)
    
    return Response({
        'subline_id': subline_id,
        'subline_name': subline.subline_name,
        'stations': serializer.data,
    }, status=status.HTTP_200_OK)



@api_view(['GET'])
def get_stations_by_department(request, department_id):
    department = get_object_or_404(Department, department_id=department_id)
    
    # Get all stations under this department
    stations = Station.objects.filter(
        subline__line__department_id=department_id
    ).select_related('subline', 'subline__line', 'subline__line__department')
    
    if not stations.exists():
        return Response({
            'message': f'No stations found under department: {department.department_name}',
            'department_name': department.department_name,
            'stations': [],
            'total_count': 0
        }, status=status.HTTP_200_OK)
    
    serializer = StationReadSerializer(stations, many=True)
    
    return Response({
        'department_id': department_id,
        'department_name': department.department_name,
        'stations': serializer.data,
    }, status=status.HTTP_200_OK)



@api_view(['GET'])
def get_stations_by_line(request, line_id):
    
    line = get_object_or_404(Line, line_id=line_id)
    
    stations = Station.objects.filter(
        subline__line=line_id
    ).select_related('subline', 'subline__line', 'subline__line__department')
    
    if not stations.exists():
        return Response({
            'message': f'No stations found under line: {line.line_name}',
            'line_name': line.line_name,
            'department_name': line.department.department_name,
            'stations': [],
            'total_count': 0
        }, status=status.HTTP_200_OK)
    
    serializer = StationReadSerializer(stations, many=True)
    
    return Response({
        'line_id': line_id,
        'line_name': line.line_name,
        'stations': serializer.data,
    }, status=status.HTTP_200_OK)


from rest_framework.decorators import api_view

@api_view(['GET'])
def get_all_departments(request):
    departments = Department.objects.all()
    serializer = DepartmentReadSerializer(departments, many=True)
    return Response({
        'departments': serializer.data,
    }, status=status.HTTP_200_OK)





# ----------------------------------notification--------------------------#
from rest_framework import viewsets, status, serializers
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, BasePermission
from django.db.models import Q, Count
from django.utils import timezone
from datetime import timedelta
from .models import Notification
from .serializers import (
    NotificationSerializer, NotificationCreateSerializer,
    NotificationUpdateSerializer, NotificationStatsSerializer
)


class NotificationPermission(BasePermission):
    """
    Custom permission for notifications:
    - Allow read access without authentication (for testing)
    - Require authentication for write operations
    """
    def has_permission(self, request, view):
        if request.method in ['GET', 'HEAD', 'OPTIONS']:
            return True
        return request.user and request.user.is_authenticated


class NotificationViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing notifications - no authentication required for read
    """
    serializer_class = NotificationSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        return Notification.objects.all().select_related(
            'recipient', 'employee', 'level', 'training_schedule',
            'machine_allocation', 'test_session', 'retraining_session',
            'human_body_check_session'
        )

    def get_serializer_class(self):
        if self.action == 'create':
            return NotificationCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return NotificationUpdateSerializer
        return NotificationSerializer
    

    def list(self, request, *args, **kwargs):
        logger.info(f"Received request for notifications with params: {request.query_params}")
        queryset = self.get_queryset()

        is_read = request.query_params.get('is_read')
        if is_read is not None:
            queryset = queryset.filter(is_read=is_read.lower() == 'true')

        notification_types = request.query_params.getlist('notification_type')
        if notification_types:
            queryset = queryset.filter(notification_type__in=notification_types)

        notification_type = request.query_params.get('type')
        if notification_type:
            queryset = queryset.filter(notification_type=notification_type)

        priority = request.query_params.get('priority')
        if priority:
            queryset = queryset.filter(priority=priority)

        days = request.query_params.get('days')
        if days:
            try:
                days_int = int(days)
                since_date = timezone.now() - timedelta(days=days_int)
                queryset = queryset.filter(created_at__gte=since_date)
            except ValueError:
                logger.warning(f"Invalid days parameter: {days}")

        queryset = queryset.order_by('-created_at')
        logger.info(f"Returning {queryset.count()} notifications")

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    
    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        notification = self.get_object()
        notification.mark_as_read()
        serializer = self.get_serializer(notification)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def mark_unread(self, request, pk=None):
        notification = self.get_object()
        notification.mark_as_unread()
        serializer = self.get_serializer(notification)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        queryset = self.get_queryset().filter(is_read=False)
        count = queryset.count()
        for notification in queryset:
            notification.mark_as_read()
        return Response({'message': f'Marked {count} notifications as read', 'count': count})

    @action(detail=False, methods=['get'])
    def unread(self, request):
        queryset = self.get_queryset().filter(is_read=False)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        queryset = self.get_queryset()
        total_count = queryset.count()
        unread_count = queryset.filter(is_read=False).count()
        read_count = total_count - unread_count
        recent_count = queryset.filter(created_at__gte=timezone.now() - timedelta(hours=24)).count()
        by_type = dict(queryset.values('notification_type').annotate(count=Count('id')).values_list('notification_type', 'count'))
        by_priority = dict(queryset.values('priority').annotate(count=Count('id')).values_list('priority', 'count'))

        stats_data = {
            'total_count': total_count,
            'unread_count': unread_count,
            'read_count': read_count,
            'recent_count': recent_count,
            'by_type': by_type,
            'by_priority': by_priority
        }
        serializer = NotificationStatsSerializer(stats_data)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def recent(self, request):
        since_date = timezone.now() - timedelta(hours=24)
        queryset = self.get_queryset().filter(created_at__gte=since_date)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


# ==================== API FUNCTION VIEWS ====================

@api_view(['GET'])
def notification_count(request):
    if not request.user.is_authenticated:
        return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)

    count = Notification.objects.filter(
        Q(recipient=request.user) | Q(recipient_email=request.user.email),
        is_read=False
    ).count()
    return Response({'unread_count': count})


@api_view(['GET'])
def test_notifications(request):
    notifications = Notification.objects.all()[:10]
    serializer = NotificationSerializer(notifications, many=True)
    return Response({'count': notifications.count(), 'notifications': serializer.data, 'debug': 'This is a test endpoint'})


@api_view(['POST'])
def create_system_notification(request):
    if not request.user.is_authenticated:
        return Response({'error': 'Authentication required'}, status=status.HTTP_403_FORBIDDEN)

    serializer = NotificationCreateSerializer(data=request.data)
    if serializer.is_valid():
        notification = serializer.save()
        response_serializer = NotificationSerializer(notification)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
def create_test_notification(request):
    try:
        recipient = request.user if request.user.is_authenticated else None
        notification = Notification.objects.create(
            title="Test Notification",
            message="This is a test notification to verify the system is working.",
            notification_type="system_alert",
            recipient=recipient,
            priority="medium",
            metadata={"test": True}
        )
        serializer = NotificationSerializer(notification)
        return Response({'message': 'Test notification created successfully', 'notification': serializer.data}, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def trigger_employee_notification(request):
    try:
        recipient = request.user if request.user.is_authenticated else None
        latest_employee = MasterTable.objects.last()

        if not latest_employee:
            return Response({'error': 'No employees found in the system'}, status=status.HTTP_400_BAD_REQUEST)

        notification = Notification.objects.create(
            title="New Employee Registered",
            message=f"New employee {latest_employee.first_name} {latest_employee.last_name} (Employee ID: {latest_employee.emp_id}) has been registered.",
            notification_type='employee_registration',
            recipient=recipient,
            employee=latest_employee,
            priority='medium',
            metadata={
                'emp_id': latest_employee.emp_id,
                'department': latest_employee.department.department_name if latest_employee.department else None
            }
        )
        serializer = NotificationSerializer(notification)
        return Response({'message': 'Employee notification created', 'notification': serializer.data}, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def trigger_all_notification_types(request):
    try:
        recipient = request.user if request.user.is_authenticated else None
        latest_employee = MasterTable.objects.last()
        employee_name = f"{latest_employee.first_name} {latest_employee.last_name}" if latest_employee else "Test Employee"

        notification_types = [
            {'type': 'employee_registration', 'title': 'New Employee Registered', 'message': f'{employee_name} has been registered.', 'priority': 'medium'},
            {'type': 'level_exam_completed', 'title': 'Level Exam Completed', 'message': f'{employee_name} completed Level 2 evaluation.', 'priority': 'high'},
            {'type': 'training_scheduled', 'title': 'Training Scheduled', 'message': f'Training scheduled for {employee_name}.', 'priority': 'medium'},
            {'type': 'training_completed', 'title': 'Training Completed', 'message': f'{employee_name} completed training.', 'priority': 'medium'},
            {'type': 'training_reschedule', 'title': 'Training Rescheduled', 'message': f'Training for {employee_name} rescheduled.', 'priority': 'medium'},
            {'type': 'refresher_training_scheduled', 'title': 'Refresher Training Scheduled', 'message': f'Refresher training scheduled for {employee_name}.', 'priority': 'medium'},
            {'type': 'refresher_training_completed', 'title': 'Refresher Training Completed', 'message': f'{employee_name} completed refresher training.', 'priority': 'medium'},
            {'type': 'hanchou_exam_completed', 'title': 'Hanchou Exam Completed', 'message': f'{employee_name} completed Hanchou exam.', 'priority': 'high'},
            {'type': 'shokuchou_exam_completed', 'title': 'Shokuchou Exam Completed', 'message': f'{employee_name} completed Shokuchou exam.', 'priority': 'high'},
            {'type': 'ten_cycle_evaluation_completed', 'title': '10 Cycle Evaluation Completed', 'message': f'{employee_name} completed 10 Cycle evaluation.', 'priority': 'high'},
            {'type': 'ojt_completed', 'title': 'OJT Completed', 'message': f'{employee_name} completed OJT.', 'priority': 'medium'},
            {'type': 'ojt_quantity_completed', 'title': 'OJT Quantity Completed', 'message': f'{employee_name} completed OJT Quantity evaluation.', 'priority': 'medium'},
            {'type': 'machine_allocated', 'title': 'Machine Allocated', 'message': f'Machine allocated to {employee_name}.', 'priority': 'medium'},
            {'type': 'test_assigned', 'title': 'Test Assigned', 'message': f'Test assigned to {employee_name}.', 'priority': 'medium'},
            {'type': 'evaluation_completed', 'title': 'Evaluation Completed', 'message': f'{employee_name} completed an evaluation.', 'priority': 'high'},
            {'type': 'retraining_scheduled', 'title': 'Retraining Scheduled', 'message': f'Retraining scheduled for {employee_name}.', 'priority': 'medium'},
            {'type': 'retraining_completed', 'title': 'Retraining Completed', 'message': f'{employee_name} completed retraining.', 'priority': 'medium'},
            {'type': 'human_body_check_completed', 'title': 'Human Body Check Completed', 'message': f'{employee_name} completed human body check.', 'priority': 'medium'},
            {'type': 'milestone_reached', 'title': 'Milestone Reached', 'message': f'{employee_name} reached a milestone.', 'priority': 'high'},
            {'type': 'system_alert', 'title': 'System Alert', 'message': 'System maintenance at 2:00 AM.', 'priority': 'urgent'}
        ]

        created_notifications = []
        for notif_data in notification_types:
            n = Notification.objects.create(
                title=notif_data['title'],
                message=notif_data['message'],
                notification_type=notif_data['type'],
                recipient=recipient,
                employee=latest_employee if latest_employee else None,
                priority=notif_data['priority'],
                metadata={'test': True, 'employee': employee_name}
            )
            created_notifications.append(n)

        return Response({'message': f'Created {len(created_notifications)} notifications', 'count': len(created_notifications)}, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['DELETE'])
def delete_all_notifications(request):
    if not request.user.is_authenticated:
        return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)

    try:
        deleted_count = Notification.objects.filter(
            Q(recipient=request.user) | Q(recipient_email=request.user.email)
        ).delete()[0]
        return Response({'message': f'Deleted {deleted_count} notifications', 'count': deleted_count}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



from .serializers import LevelOnePassedScoreSerializer

class LevelOnePassedUsersView(generics.ListAPIView):
    """
    API view to retrieve a list of all scores for users who have 
    successfully passed a 'Level 1' assessment.

    The results are ordered by the most recent test date first.
    """
    serializer_class = LevelOnePassedScoreSerializer
    
    def get_queryset(self):
        """
        This view returns a list of all scores for
        users who have passed a test specifically at 'Level 1'.
        """
        queryset = Score.objects.filter(
            passed=True,
            level__level_name='Level 1'  # <-- CORRECTED LOOKUP HERE
        ).select_related(
            'employee', 'test', 'level', 'skill'
        ).order_by('-created_at')
        
        return queryset
    

# In views.py
from .models import HandoverSheet
from .serializers import HandoverSheetCreateSerializer


class HandoverSheetViewSet(viewsets.ModelViewSet):
    queryset = HandoverSheet.objects.all()
    serializer_class = HandoverSheetCreateSerializer

from django.shortcuts import get_object_or_404

class EmployeeHandoverView(generics.RetrieveAPIView):
    serializer_class = HandoverSheetCreateSerializer
    queryset = HandoverSheet.objects.all()

    lookup_field = "employee__emp_id"  # tell DRF to look up by emp_id
    lookup_url_kwarg = "emp_id"











from datetime import datetime, timedelta 
from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.db import transaction
from datetime import datetime
from django.db.models import Q
import calendar
from django.db.models import Sum, Avg, F
from django.db.models.functions import Coalesce

# Make sure to import your model and serializer
from .models import DailyProductionData
from .serializers import DailyProductionDataSerializer

class DailyProductionDataViewSet(viewsets.ModelViewSet):
    """
    Final, correct ViewSet for handling production data.
    All calculation logic is now handled here, not in signals.
    """
    queryset = DailyProductionData.objects.all()
    serializer_class = DailyProductionDataSerializer

    def get_queryset(self):
        """
        Overrides the default queryset to apply filters from URL parameters.
        """
        # Start with all objects
        queryset = super().get_queryset()

        # Get parameters from the URL
        factory_id = self.request.query_params.get('factory')
        department_id = self.request.query_params.get('department')

        # Apply filters if the parameters exist
        if factory_id:
            queryset = queryset.filter(factory_id=factory_id)

        if department_id:
            queryset = queryset.filter(department_id=department_id)
        
        # Return the final, correctly filtered data
        return queryset

    @action(detail=False, methods=['post'], url_path='save-plan-entry')
    @transaction.atomic
    def save_plan_entry(self, request):
        """
        Handles creating or updating a plan entry.
        It now performs all calculations itself, ignoring the pre_save signal.
        """
        data = request.data
        entry_id = data.get('id', None)

        try:
            # Step 1: Calculate all the totals into a separate dictionary
            # This is safer and doesn't modify the original request data.
            calculated_totals = {}
            calculated_totals['ctq_plan_total'] = data.get('ctq_plan_l1', 0) + data.get('ctq_plan_l2', 0) + data.get('ctq_plan_l3', 0) + data.get('ctq_plan_l4', 0)
            calculated_totals['ctq_actual_total'] = data.get('ctq_actual_l1', 0) + data.get('ctq_actual_l2', 0) + data.get('ctq_actual_l3', 0) + data.get('ctq_actual_l4', 0)
            calculated_totals['pdi_plan_total'] = data.get('pdi_plan_l1', 0) + data.get('pdi_plan_l2', 0) + data.get('pdi_plan_l3', 0) + data.get('pdi_plan_l4', 0)
            calculated_totals['pdi_actual_total'] = data.get('pdi_actual_l1', 0) + data.get('pdi_actual_l2', 0) + data.get('pdi_actual_l3', 0) + data.get('pdi_actual_l4', 0)
            calculated_totals['other_plan_total'] = data.get('other_plan_l1', 0) + data.get('other_plan_l2', 0) + data.get('other_plan_l3', 0) + data.get('other_plan_l4', 0)
            calculated_totals['other_actual_total'] = data.get('other_actual_l1', 0) + data.get('other_actual_l2', 0) + data.get('other_actual_l3', 0) + data.get('other_actual_l4', 0)
            calculated_totals['grand_total_plan'] = calculated_totals['ctq_plan_total'] + calculated_totals['pdi_plan_total'] + calculated_totals['other_plan_total']
            calculated_totals['grand_total_actual'] = calculated_totals['ctq_actual_total'] + calculated_totals['pdi_actual_total'] + calculated_totals['other_actual_total']

            # Step 2: Perform the overlap check
            start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
            end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()
            entry_mode = data['entry_mode']
            
            overlap_condition = Q(start_date__lte=end_date) & Q(end_date__gte=start_date)
            # filters = {'line_id': data.get('line'), 'station_id': data.get('station'), 'shop_floor_id': data.get('shop_floor'), 'factory_id': data.get('factory')}
            filters = {
                'Hq_id': data.get('Hq'),
                'factory_id': data.get('factory'),
                'department_id': data.get('department'),
                'line_id': data.get('line'),
                'subline_id': data.get('subline'),
                'station_id': data.get('station'),
            }
            filters = {k: v for k, v in filters.items() if v}
            overlapping_entries = DailyProductionData.objects.filter(overlap_condition, **filters)

            if entry_id:
                overlapping_entries = overlapping_entries.exclude(id=entry_id)

            if overlapping_entries.exists():
                existing_entry = overlapping_entries.first()
                if entry_mode == 'MONTHLY' and existing_entry.entry_mode != 'MONTHLY':
                    return Response({'error': "Cannot save: Weekly/Daily data already exists in this period."}, status=status.HTTP_409_CONFLICT)
                if entry_mode != 'MONTHLY' and existing_entry.entry_mode == 'MONTHLY':
                    return Response({'error': f"Cannot save: A Monthly entry already exists for this period."}, status=status.HTTP_409_CONFLICT)

            # Step 3: Serialize and Save
            if entry_id:
                instance = DailyProductionData.objects.get(id=entry_id)
                serializer = self.get_serializer(instance, data=data, partial=True)
            else:
                serializer = self.get_serializer(data=data)
            
            serializer.is_valid(raise_exception=True)
            # Pass the calculated totals as extra arguments to the save method
            saved_instance = serializer.save(**calculated_totals)
            
            # Create a new serializer from the final saved object to send back
            response_serializer = self.get_serializer(saved_instance)
            return Response(response_serializer.data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        


        # In views.py, add this method inside the DailyProductionDataViewSet class

    

    # --- ADD THIS ENTIRE NEW FUNCTION FOR THE WEEKLY VIEW ---
    @action(detail=False, methods=['get'], url_path='weekly-summary')
    def weekly_summary(self, request):
        """
        NOTE: This function aggregates all data for an entire MONTH, not a single week.
        It fetches level-based data for a given factory, year, and month.
        """
        # --- 1. Get Parameters ---
        factory_id = request.query_params.get('factory')
        month_str = request.query_params.get('month')
        year_str = request.query_params.get('year')

        line_id = request.query_params.get('line')
        station_id = request.query_params.get('station')
        # --- ADD THESE LINES ---
        Hq_id = request.query_params.get('Hq')
        department_id = request.query_params.get('department')
        subline_id = request.query_params.get('subline')

        if not all([factory_id, month_str, year_str]):
            return Response(
                {'error': 'factory, month, and year are required.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # --- 2. Build Filters ---
        try:
            filters = {
                'factory_id': factory_id,
                'start_date__year': int(year_str),
                'start_date__month': int(month_str),
            }
            # --- EDIT THIS SECTION ---
            # Apply optional filters
            if Hq_id: filters['Hq_id'] = Hq_id
            if department_id: filters['department_id'] = department_id
            if subline_id: filters['subline_id'] = subline_id
            if station_id:
                filters['station_id'] = station_id
            elif line_id:
                filters['line_id'] = line_id
                
        except (ValueError, TypeError):
            return Response({'error': 'Invalid year or month format.'}, status=status.HTTP_400_BAD_REQUEST)

        # --- 3. Define Aggregation Fields (Corrected to be a DICTIONARY) ---
        aggregation_fields = {}
        field_names_to_sum = [
            'ctq_plan_l1', 'ctq_actual_l1', 'pdi_plan_l1', 'pdi_actual_l1', 'other_plan_l1', 'other_actual_l1',
            'ctq_plan_l2', 'ctq_actual_l2', 'pdi_plan_l2', 'pdi_actual_l2', 'other_plan_l2', 'other_actual_l2',
            'ctq_plan_l3', 'ctq_actual_l3', 'pdi_plan_l3', 'pdi_actual_l3', 'other_plan_l3', 'other_actual_l3',
            'ctq_plan_l4', 'ctq_actual_l4', 'pdi_plan_l4', 'pdi_actual_l4', 'other_plan_l4', 'other_actual_l4'
        ]
        for field in field_names_to_sum:
            aggregation_fields[field] = Sum(field)

        # --- 4. Query and Aggregate ---
        summary_data = DailyProductionData.objects.filter(**filters).aggregate(**aggregation_fields)
        
        # --- 5. Clean up None values and return ---
        for key, value in summary_data.items():
            if value is None:
                summary_data[key] = 0
        
        return Response(summary_data, status=status.HTTP_200_OK)
        


    @action(detail=False, methods=['get'], url_path='trend-data')
    def trend_data(self, request):
        """
        A flexible endpoint to return trend data for different metrics.
        Accepts a 'data_key' to switch between production, manpower, etc.
        Groups data by 'monthly', 'weekly', or 'daily'.
        """
        # --- 1. Get Parameters ---
        factory_id = request.query_params.get('factory')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        group_by = request.query_params.get('group_by')
        data_key = request.query_params.get('data_key', 'production')

        if not all([factory_id, start_date_str, end_date_str, group_by]):
            return Response({'error': 'Required parameters missing'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)

        # --- 2. Determine which fields to aggregate based on data_key ---
        if data_key == 'production':
            plan_field = 'total_production_plan'
            actual_field = 'total_production_actual'
        elif data_key == 'manpower':
            plan_field = 'total_operators_required_plan'
            actual_field = 'total_operators_required_actual'
        else:
            return Response({'error': 'Invalid data_key'}, status=status.HTTP_400_BAD_REQUEST)

        # --- 3. Build Filters and get the base queryset ---
        filters = {
            'factory_id': factory_id,
            'start_date__lte': end_date,
            'end_date__gte': start_date,
        }
        # --- EDIT THIS SECTION ---
        # Add optional hierarchy filters
        Hq_id = request.query_params.get('hq')
        department_id = request.query_params.get('department')
        line_id = request.query_params.get('line')
        subline_id = request.query_params.get('subline')
        station_id = request.query_params.get('station')
        
        if Hq_id: filters['Hq_id'] = Hq_id
        if department_id: filters['department_id'] = department_id
        if line_id: filters['line_id'] = line_id
        if subline_id: filters['subline_id'] = subline_id
        if station_id: filters['station_id'] = station_id

        queryset = DailyProductionData.objects.filter(**filters)
        trend_data = []

        # --- 4. Aggregate and format data ---
        if group_by == 'monthly':
            for year in range(start_date.year, end_date.year + 1):
                for month in range(1, 13):
                    if (year == start_date.year and month < start_date.month) or \
                    (year == end_date.year and month > end_date.month):
                        continue
                    
                    aggregation = queryset.filter(start_date__year=year, start_date__month=month).aggregate(
                        plan_sum=Sum(plan_field),
                        actual_sum=Sum(actual_field)
                    )
                    trend_data.append({
                        "name": f"{calendar.month_name[month][:3]} '{str(year)[2:]}",
                        "planned": aggregation['plan_sum'] or 0,
                        "actual": aggregation['actual_sum'] or 0,
                    })

        elif group_by == 'weekly':
            # THIS IS THE NEW, CORRECTED WEEKLY LOGIC
            # Find all unique weekly records within the date range and return them.
            weekly_records = queryset.filter(entry_mode='WEEKLY').order_by('start_date')
            
            for record in weekly_records:
                trend_data.append({
                    "name": f"Week of {record.start_date.strftime('%b %d')}",
                    "planned": getattr(record, plan_field), # Use getattr to get field value by name
                    "actual": getattr(record, actual_field),
                })
        
        # (The 'daily' grouping is omitted as the frontend is no longer asking for it, but can be added back if needed)
        
        return Response(trend_data, status=status.HTTP_200_OK)

    

    @action(detail=False, methods=['get'], url_path='get-plan-data')
    def get_plan_data(self, request):
        factory_id = request.query_params.get('factory')
        Hq_id = request.query_params.get('Hq')
        department_id = request.query_params.get('department')
        subline_id = request.query_params.get('subline')
        line_id = request.query_params.get('line')
        station_id = request.query_params.get('station')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        
        if not all([factory_id, start_date_str, end_date_str]):
            return Response({'error': 'Factory, start_date, and end_date are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)
        
        filters = {
            'Hq_id': Hq_id,
            'factory_id': factory_id,
            'department_id': department_id,
            'line_id': line_id,
            'subline_id': subline_id,
            'station_id': station_id,
            'start_date': start_date,
            'end_date': end_date,
        }
        filters = {k: v for k, v in filters.items() if v and v != 'all'}
        
        try:
            plan_entry = DailyProductionData.objects.get(**filters)
            serializer = self.get_serializer(plan_entry)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except DailyProductionData.DoesNotExist:
            return Response({'message': 'No data found for this period.'}, status=status.HTTP_404_NOT_FOUND)
        except DailyProductionData.MultipleObjectsReturned:
            return Response({'error': 'Multiple entries found for this period. Please ensure unique data for the specified dates.'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': f"An error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            

    @action(detail=False, methods=['get'], url_path='get-ending-team')
    def get_ending_team(self, request):
            """
            Calculates the ending team for a given period to be used as the
            starting team for the next period.
            """
            factory_id = request.query_params.get('factory')
            Hq_id = request.query_params.get('Hq')
            department_id = request.query_params.get('department')
            subline_id = request.query_params.get('subline')

            # shop_floor_id = request.query_params.get('shop_floor')
            line_id = request.query_params.get('line')
            station_id = request.query_params.get('station')
            # We will ask for the START date of the *previous* period
            target_date_str = request.query_params.get('target_date')

            if not all([factory_id, target_date_str]):
                return Response({'error': 'Factory and target_date are required.'}, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({'error': 'Invalid date format.'}, status=status.HTTP_400_BAD_REQUEST)
                
            filters = {
                'Hq_id': Hq_id,
                'factory_id': factory_id,
                'department_id': department_id,
                'line_id': line_id,
                'subline_id': subline_id,
                'station_id': station_id,
                'start_date__lte': target_date, 'end_date__gte': target_date,
            }
            filters = {k: v for k, v in filters.items() if v and v != 'all'}
            
            try:
                previous_entry = DailyProductionData.objects.get(**filters)
                
                # Perform the 'Ending Team' calculation
                starting_team = previous_entry.total_operators_available
                attrition_rate = previous_entry.attrition_rate
                
                # Use Decimal for precision
                from decimal import Decimal
                ending_team = Decimal(starting_team) * (Decimal(1) - (Decimal(attrition_rate) / Decimal(100)))
                
                # Return the calculated value, rounded to a whole number
                return Response({'ending_team': round(ending_team)}, status=status.HTTP_200_OK)
                
            except DailyProductionData.DoesNotExist:
                # If there's no previous data, we can't calculate, so we return 0
                return Response({'ending_team': 0}, status=status.HTTP_404_NOT_FOUND)
            

    @action(detail=False, methods=['get'], url_path='get-month-lock-status')
    def get_month_lock_status(self, request):
        """
        Investigates a month to see if it has any data and what its lock mode is.
        """
        factory_id = request.query_params.get('factory')
        target_date_str = request.query_params.get('target_date')
        
        try:
            target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
            month = target_date.month
            year = target_date.year
        except:
            return Response({'error': 'Invalid date'}, status=status.HTTP_400_BAD_REQUEST)

        filters = {
            'Hq_id': request.query_params.get('Hq'),
            'factory_id': factory_id,
            'department_id': request.query_params.get('department'),
            'line_id': request.query_params.get('line'),
            'subline_id': request.query_params.get('subline'),
            'station_id': request.query_params.get('station'),
            'start_date__year': year,
            'start_date__month': month,
        }
        filters = {k: v for k, v in filters.items() if v and v != 'all'}
        
        first_entry = DailyProductionData.objects.filter(**filters).first()
        
        if first_entry:
            return Response({'lock_mode': first_entry.entry_mode}, status=status.HTTP_200_OK)
        else:
            # Important to send 200 OK with null, so frontend knows the check was successful
            return Response({'lock_mode': None}, status=status.HTTP_200_OK)
        

    @action(detail=False, methods=['get'], url_path='get-period-summary')
    def get_period_summary(self, request):
        """
        This is our "Summary Chef".
        It aggregates all data for a specific period (month or week)
        to power the summary cards and stats components.
        """
        # --- 1. Get all the filter parameters from the control panel ---
        factory_id = request.query_params.get('factory')
        # shop_floor_id = request.query_params.get('shop_floor')
        Hq_id = request.query_params.get('Hq')
        department_id = request.query_params.get('department')
        subline_id = request.query_params.get('subline')
        line_id = request.query_params.get('line')
        station_id = request.query_params.get('station')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        if not all([factory_id, start_date_str, end_date_str]):
            return Response(
                {'error': 'Factory, start_date, and end_date are required.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response(
                {'error': 'Invalid date format. Use YYYY-MM-DD.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # --- 2. Build the filter dictionary dynamically ---
        # This correctly handles the "All" selections from the frontend.
        filters = {
            'factory_id': factory_id,
            # Find any record that overlaps with the selected date range
            'start_date__lte': end_date,
            'end_date__gte': start_date,
        }
        if Hq_id and Hq_id != 'all':
            filters['Hq_id'] = Hq_id
        if department_id and department_id != 'all':
            filters['department_id'] = department_id
        if subline_id and subline_id != 'all':
            filters['subline_id'] = subline_id
        if line_id and line_id != 'all':
            filters['line_id'] = line_id
        if station_id and station_id != 'all':
            filters['station_id'] = station_id
        # --- 3. Define all the fields we need to aggregate (Sum or Average) ---
        from django.db.models import Sum, Avg

        aggregation_fields = {
            # Main Totals
            'total_production_plan': Sum('total_production_plan'),
            'total_production_actual': Sum('total_production_actual'),
            'total_operators_available': Sum('total_operators_available'),
            'total_operators_required_plan': Sum('total_operators_required_plan'),
            'total_operators_required_actual': Sum('total_operators_required_actual'),
            # Rates should be averaged
            'attrition_rate': Avg('attrition_rate'),
            'absenteeism_rate': Avg('absenteeism_rate'),
            # Bifurcation fields
            'bifurcation_plan_l1': Sum('bifurcation_plan_l1'),
            'bifurcation_actual_l1': Sum('bifurcation_actual_l1'),
            'bifurcation_plan_l2': Sum('bifurcation_plan_l2'),
            'bifurcation_actual_l2': Sum('bifurcation_actual_l2'),
            'bifurcation_plan_l3': Sum('bifurcation_plan_l3'),
            'bifurcation_actual_l3': Sum('bifurcation_actual_l3'),
            'bifurcation_plan_l4': Sum('bifurcation_plan_l4'),
            'bifurcation_actual_l4': Sum('bifurcation_actual_l4'),
            # You can add all your CTQ, PDI, Other fields here if needed for stats
        }

        # --- 4. Perform the database query ---
        summary_data = DailyProductionData.objects.filter(**filters).aggregate(**aggregation_fields)

        # --- 5. Clean up the data and send the response ---
        # If the query finds nothing, aggregate returns None for all fields.
        # We need to replace None with 0 so the frontend doesn't crash.
        for key, value in summary_data.items():
            if value is None:
                summary_data[key] = 0
        
        return Response(summary_data, status=status.HTTP_200_OK)
    

    # Add this code inside your DailyProductionDataViewSet class in views.py

    @action(detail=False, methods=['get'], url_path='monthly-summary')
    def monthly_summary(self, request):
        """
        Aggregates bifurcation and other key data for a specific calendar month.
        Used by the main dashboard cards.
        """
        # --- 1. Get query parameters ---
        factory_id = request.query_params.get('factory')
        Hq_id = request.query_params.get('Hq')
        department_id = request.query_params.get('department')
        subline_id = request.query_params.get('subline')
        # shop_floor_id = request.query_params.get('shop_floor')
        line_id = request.query_params.get('line')
        station_id = request.query_params.get('station')
        month = request.query_params.get('month')
        year = request.query_params.get('year')

        if not all([factory_id, month, year]):
            return Response(
                {'error': 'Factory, month, and year are required.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # --- 2. Build the filter dictionary ---
        try:
            filters = {
                'factory_id': factory_id,
                'start_date__year': int(year),
                'start_date__month': int(month),
            }
            if Hq_id: filters['Hq_id'] = Hq_id
            if department_id: filters['department_id'] = department_id
            if subline_id: filters['subline_id'] = subline_id
            if line_id: filters['line_id'] = line_id
            if station_id: filters['station_id'] = station_id
        except (ValueError, TypeError):
            return Response({'error': 'Invalid year or month format.'}, status=status.HTTP_400_BAD_REQUEST)

        # --- 3. Define aggregation fields ---
        from django.db.models import Sum

        aggregation_fields = {
            'bifurcation_plan_l1': Sum('bifurcation_plan_l1'),
            'bifurcation_actual_l1': Sum('bifurcation_actual_l1'),
            'bifurcation_plan_l2': Sum('bifurcation_plan_l2'),
            'bifurcation_actual_l2': Sum('bifurcation_actual_l2'),
            'bifurcation_plan_l3': Sum('bifurcation_plan_l3'),
            'bifurcation_actual_l3': Sum('bifurcation_actual_l3'),
            'bifurcation_plan_l4': Sum('bifurcation_plan_l4'),
            'bifurcation_actual_l4': Sum('bifurcation_actual_l4'),
        }

        # --- 4. Query and Aggregate ---
        summary_data = DailyProductionData.objects.filter(**filters).aggregate(**aggregation_fields)
        
        # --- 5. Clean up None values and return ---
        for key, value in summary_data.items():
            if value is None:
                summary_data[key] = 0
        
        return Response(summary_data, status=status.HTTP_200_OK)


    @action(detail=False, methods=['get'], url_path='aggregated-weekly-data')
    def aggregated_weekly_data(self, request):
        """
        Aggregates bifurcation and other key data for a 7-day period (a week).
        Used by the main dashboard cards in weekly view.
        """
        # --- 1. Get query parameters ---
        factory_id = request.query_params.get('factory')
        # shop_floor_id = request.query_params.get('shop_floor')
        Hq_id = request.query_params.get('Hq')
        department_id = request.query_params.get('department')
        subline_id = request.query_params.get('subline')
        line_id = request.query_params.get('line')
        station_id = request.query_params.get('station')
        start_date_str = request.query_params.get('start_date')

        if not all([factory_id, start_date_str]):
            return Response(
                {'error': 'Factory and start_date are required.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # --- 2. Calculate date range and build filters ---
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = start_date + timedelta(days=6)

            filters = {
                'factory_id': factory_id,
                'start_date__lte': end_date,
                'end_date__gte': start_date,
            }
            if Hq_id and Hq_id != 'all':
                filters['hq_id'] = Hq_id
            if department_id and department_id != 'all':
                filters['department_id'] = department_id
            if subline_id and subline_id != 'all':
                filters['subline_id'] = subline_id
            if line_id and line_id != 'all':
                filters['line_id'] = line_id
            if station_id and station_id != 'all':
                filters['station_id'] = station_id
        except ValueError:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

        # --- 3. Define aggregation fields (same as monthly) ---
        from django.db.models import Sum

        aggregation_fields = {
            'bifurcation_plan_l1': Sum('bifurcation_plan_l1'),
            'bifurcation_actual_l1': Sum('bifurcation_actual_l1'),
            'bifurcation_plan_l2': Sum('bifurcation_plan_l2'),
            'bifurcation_actual_l2': Sum('bifurcation_actual_l2'),
            'bifurcation_plan_l3': Sum('bifurcation_plan_l3'),
            'bifurcation_actual_l3': Sum('bifurcation_actual_l3'),
            'bifurcation_plan_l4': Sum('bifurcation_plan_l4'),
            'bifurcation_actual_l4': Sum('bifurcation_actual_l4'),
        }

        # --- 4. Query and Aggregate ---
        summary_data = DailyProductionData.objects.filter(**filters).aggregate(**aggregation_fields)
        
        # --- 5. Clean up None values and return ---
        for key, value in summary_data.items():
            if value is None:
                summary_data[key] = 0
        
        return Response(summary_data, status=status.HTTP_200_OK)
    


    @action(detail=False, methods=['get'], url_path='gap-analysis')
    def gap_analysis(self, request):
        """
        Calculates the gap between plan and actual values for a given period.
        """
        # --- 1. Get and Validate Parameters ---
        factory_id = request.query_params.get('factory')
        hq_id = request.query_params.get('hq')
        department_id = request.query_params.get('department')
        line_id = request.query_params.get('line')
        subline_id = request.query_params.get('subline')
        station_id = request.query_params.get('station')
        month_name = request.query_params.get('month')
        year_str = request.query_params.get('year')

        if not all([factory_id, month_name, year_str]):
            return Response(
                {'error': 'Factory, month, and year are required parameters.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # --- 2. Convert Month Name to Number ---
        try:
            # Create a mapping from month name to month number
            month_map = {name: num for num, name in enumerate(calendar.month_name) if num}
            month_num = month_map.get(month_name.capitalize())
            if not month_num:
                raise ValueError("Invalid month name")
            
            year = int(year_str)
        except (ValueError, TypeError, AttributeError):
            return Response({'error': 'Invalid month or year format.'}, status=status.HTTP_400_BAD_REQUEST)

        # --- 3. Build Database Filters ---
        filters = {
            'factory_id': factory_id,
            'start_date__year': year,
            'start_date__month': month_num,
        }
        # Add optional filters only if they are provided
        if hq_id: filters['Hq_id'] = hq_id
        if department_id: filters['department_id'] = department_id
        if line_id: filters['line_id'] = line_id
        if subline_id: filters['subline_id'] = subline_id
        if station_id: filters['station_id'] = station_id
        # You can add shop_floor and line here in the same way if needed
        
        # --- 4. Define All Aggregation Fields ---
        aggregation_fields = {
            'plan_production': Sum('total_production_plan'),
            'actual_production': Sum('total_production_actual'),
            'plan_operators': Sum('total_operators_required_plan'),
            'actual_operators': Sum('total_operators_required_actual'),
            'plan_ctq_l1': Sum('ctq_plan_l1'), 'actual_ctq_l1': Sum('ctq_actual_l1'),
            'plan_ctq_l2': Sum('ctq_plan_l2'), 'actual_ctq_l2': Sum('ctq_actual_l2'),
            'plan_ctq_l3': Sum('ctq_plan_l3'), 'actual_ctq_l3': Sum('ctq_actual_l3'),
            'plan_ctq_l4': Sum('ctq_plan_l4'), 'actual_ctq_l4': Sum('ctq_actual_l4'),
            # Add PDI, Other, etc. in the same pattern if needed by the frontend
        }

        # --- 5. Query the Database ---
        summary_data = DailyProductionData.objects.filter(**filters).aggregate(**aggregation_fields)

        # --- 6. Calculate Gaps and Structure the Response ---
        # Helper function to create the plan/actual/gap structure
        def create_metric(plan_key, actual_key):
            plan = summary_data.get(plan_key) or 0
            actual = summary_data.get(actual_key) or 0
            return {
                "plan": plan,
                "actual": actual,
                "gap": plan - actual
            }

        response_payload = {
            "production": create_metric('plan_production', 'actual_production'),
            "operators": create_metric('plan_operators', 'actual_operators'),
            "ctq_l1": create_metric('plan_ctq_l1', 'actual_ctq_l1'),
            "ctq_l2": create_metric('plan_ctq_l2', 'actual_ctq_l2'),
            "ctq_l3": create_metric('plan_ctq_l3', 'actual_ctq_l3'),
            "ctq_l4": create_metric('plan_ctq_l4', 'actual_ctq_l4'),
        }

        return Response(response_payload, status=status.HTTP_200_OK)
    


   
    @action(detail=False, methods=['get'], url_path='buffer-analysis')
    def buffer_analysis(self, request):
        """
        Performs a two-part analysis for a target period:
        1. Calculates the TOTAL production volume lost due to the overall operator gap.
        2. Calculates the NET AVAILABLE manpower and GAP for each of the 4 levels,
           factoring in attrition and absenteeism.
        """
        # --- 1. Get Simplified Parameters (Unchanged) ---
        factory_id = request.query_params.get('factory')
        Hq_id = request.query_params.get('Hq')
        department_id = request.query_params.get('department')
        line_id = request.query_params.get('line')
        subline_id = request.query_params.get('subline')
        station_id = request.query_params.get('station')
        target_date_str = request.query_params.get('start_date')

        if not all([factory_id, target_date_str]):
            return Response({'error': 'Factory and start_date are required.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Invalid date format.'}, status=status.HTTP_400_BAD_REQUEST)

        # --- 2. Find the Single Target Record (Unchanged) ---
        filters = {
            'factory_id': factory_id,
            'start_date__lte': target_date,
            'end_date__gte': target_date,
        }
        if Hq_id: filters['hq_id'] = Hq_id
        if department_id: filters['department_id'] = department_id
        if line_id: filters['line_id'] = line_id
        if subline_id: filters['subline_id'] = subline_id
        if station_id: filters['station_id'] = station_id

        matching_records = DailyProductionData.objects.filter(**filters)

        # From the matches, find the one with the shortest duration (most specific)
        # We annotate the query with the duration and order by it, ascending.
        target_record = matching_records.annotate(
            duration=F('end_date') - F('start_date')
        ).order_by('duration').first()

        # If no data exists, return a default structure for consistency
        if not target_record:
            return Response({
                "lost_production_due_to_operator_gap": 0,
                "manpower_analysis_by_level": []
            }, status=status.HTTP_200_OK)

        # --- 3. NEW: Calculate Total Lost Production Volume ---
        lost_production_volume = Decimal('0.0')

        # Prevent division by zero
        if target_record.total_operators_available > 0:
            # Calculate the average production per available operator ("$B$4" from the formula)
            avg_prod_per_operator = (
                Decimal(target_record.total_production_actual) /
                Decimal(target_record.total_operators_available)
            )

            # Check if there was an operator shortfall
            if target_record.total_operators_required_actual < target_record.total_operators_required_plan:
                # Calculate the number of missing operators
                operator_gap = (
                    target_record.total_operators_required_plan -
                    target_record.total_operators_required_actual
                )
                # Calculate the final lost volume
                lost_production_volume = operator_gap * avg_prod_per_operator

        # --- 4. Per-Level Manpower Gap Calculation (Your original logic) ---
        level_analysis_data = []
        attrition_rate = Decimal(target_record.attrition_rate or 0) / Decimal(100)
        absenteeism_rate = Decimal(target_record.absenteeism_rate or 0) / Decimal(100)

        for level in range(1, 5):
            available_field = f'bifurcation_actual_l{level}'
            required_field = f'bifurcation_plan_l{level}'

            operators_available = Decimal(getattr(target_record, available_field, 0))
            operators_required = Decimal(getattr(target_record, required_field, 0))

            net_available = operators_available - (operators_available * absenteeism_rate) - (operators_available * attrition_rate)
            net_available_rounded = int(net_available)
            gap = max(0, int(operators_required - Decimal(net_available_rounded)))

            level_analysis_data.append({
                "name": f"Level {level}",
                "operators_required": int(operators_required),
                "net_available": net_available_rounded,
                "gap": gap,
            })

        # --- 5. Combine results into a single structured response ---
        response_data = {
            "lost_production_due_to_operator_gap": round(lost_production_volume, 2),
            "manpower_analysis_by_level": level_analysis_data
        }

        return Response(response_data, status=status.HTTP_200_OK)
    

    
    # In views.py, replace the gap_volume_analysis function

    @action(detail=False, methods=['get'], url_path='gap-volume-analysis')
    def gap_volume_analysis(self, request):
        factory_id = request.query_params.get('factory')
        Hq_id = request.query_params.get('Hq')
        department_id = request.query_params.get('department')
        line_id = request.query_params.get('line')
        subline_id = request.query_params.get('subline')
        target_date_str = request.query_params.get('start_date') # Only need one date

        if not all([factory_id, target_date_str]):
            return Response({'error': 'Factory and start_date are required.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Invalid date format.'}, status=status.HTTP_400_BAD_REQUEST)

        filters = {'factory_id': factory_id, 'start_date__lte': target_date, 'end_date__gte': target_date}
        station_id = request.query_params.get('station')
        if Hq_id: filters['hq_id'] = Hq_id
        if department_id: filters['department_id'] = department_id
        if line_id: filters['line_id'] = line_id
        if subline_id: filters['subline_id'] = subline_id
        if station_id: filters['station_id'] = station_id

        # Find the single record for the target period
        target_record = DailyProductionData.objects.filter(**filters).first()

        if not target_record:
            return Response({'L1': 0, 'L2': 0, 'L3': 0, 'L4': 0}, status=status.HTTP_200_OK)

        gap_data = {}
        attrition_rate = Decimal(target_record.attrition_rate or 0) / Decimal(100)
        absenteeism_rate = Decimal(target_record.absenteeism_rate or 0) / Decimal(100)

        for level in range(1, 5):
            # Use bifurcation fields which are the sums of departments
            operators_available = Decimal(getattr(target_record, f'bifurcation_actual_l{level}', 0))
            operators_required = Decimal(getattr(target_record, f'bifurcation_plan_l{level}', 0))
            net_available = operators_available - (operators_available * absenteeism_rate) - (operators_available * attrition_rate)
            gap_volume = int(operators_required - net_available)
            gap_data[f'L{level}'] = max(0, gap_volume)
        
        return Response(gap_data, status=status.HTTP_200_OK)

    
def get_planning_data(request):
    # --- 1. Get and Validate Parameters ---
    Hq_id = request.GET.get('Hq')
    department_id = request.GET.get('department')
    line_id = request.GET.get('line')
    subline_id = request.GET.get('subline')
    station_id = request.GET.get('station')
    factory_id = request.GET.get('factory')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    group_by = request.GET.get('group_by') # No default, we need it to be explicit

    if not all([factory_id, start_date_str, end_date_str, group_by]):
        return JsonResponse({'success': False, 'error': 'Factory, start_date, end_date, and group_by are required.'}, status=400)

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

    # --- 2. Build Base Filters ---
    filters = {
        'factory_id': factory_id,
        'start_date__lte': end_date,
        'end_date__gte': start_date,
    }
    # Add optional filters here if needed (station, line, etc.)
    # station_id = request.GET.get('station')
    if station_id:
        filters['station_id'] = station_id
    if Hq_id: filters['hq_id'] = Hq_id
    if department_id: filters['department_id'] = department_id
    if line_id: filters['line_id'] = line_id
    if subline_id: filters['subline_id'] = subline_id
    if station_id: filters['station_id'] = station_id
    
    queryset = DailyProductionData.objects.filter(**filters)
    graph_data = []

    # --- 3. Aggregate data based on the group_by parameter ---
    if group_by == 'month':
        # THIS IS THE NEW, SIMPLER, AND CORRECT LOOP FOR MONTHS
        current_year = start_date.year
        current_month = start_date.month
        while (current_year < end_date.year) or (current_year == end_date.year and current_month <= end_date.month):
            month_name = calendar.month_name[current_month]
            aggregation = queryset.filter(
                start_date__year=current_year,
                start_date__month=current_month
            ).aggregate(
                total_plan_prod=Sum('total_production_plan'),
                total_actual_manpower=Sum('total_operators_required_actual'),
                total_operators_plan=Sum('total_operators_required_plan'),
                total_operators_actual=Sum('total_operators_required_actual')
            )
            graph_data.append({
                "label": month_name,
                "planned_production": aggregation['total_plan_prod'] or 0,
                "actual_manpower": aggregation['total_actual_manpower'] or 0,
                "total_operators_required_plan": aggregation['total_operators_plan'] or 0,
                "total_operators_required_actual": aggregation['total_operators_actual'] or 0,
            })
            # Advance to the next month
            current_month += 1
            if current_month > 12:
                current_month = 1
                current_year += 1

    elif group_by == 'week':
        current_week_start = start_date
        while current_week_start <= end_date:
            current_week_end = current_week_start + timedelta(days=6)
            week_label = f"Week of {current_week_start.strftime('%b %d')}"
            aggregation = queryset.filter(
                start_date__lte=current_week_end,
                end_date__gte=current_week_start
            ).aggregate(
                total_plan_prod=Sum('total_production_plan'),
                total_actual_manpower=Sum('total_operators_required_actual'),
                total_operators_plan=Sum('total_operators_required_plan'),
                total_operators_actual=Sum('total_operators_required_actual')
            )
            graph_data.append({
                "label": week_label,
                "planned_production": aggregation['total_plan_prod'] or 0,
                "actual_manpower": aggregation['total_actual_manpower'] or 0,
                "total_operators_required_plan": aggregation['total_operators_plan'] or 0,
                "total_operators_required_actual": aggregation['total_operators_actual'] or 0,
            })
            current_week_start += timedelta(days=7)

    response_payload = { "success": True, "data": graph_data }
    return JsonResponse(response_payload, status=200)





from django.http import JsonResponse
from django.db.models import Sum
from django.utils import timezone
import calendar
from datetime import datetime
from dateutil import parser

def get_operators_required_trend(request):
    factory_id = request.GET.get('factory')
    time_view = request.GET.get('time_view', 'Monthly')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if not factory_id or not start_date_str or not end_date_str:
        return JsonResponse({'error': 'Factory, start_date, and end_date are required.'}, status=400)

    try:
        start_date = parser.parse(start_date_str).date()
        end_date = parser.parse(end_date_str).date()
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid date format.'}, status=400)

    if start_date > end_date:
        return JsonResponse({'error': 'start_date cannot be later than end_date.'}, status=400)

    filters = {'factory_id': factory_id}
    # --- EDIT THIS SECTION ---
    if request.GET.get('Hq'):
        filters['Hq_id'] = request.GET.get('Hq')
    if request.GET.get('department'):
        filters['department_id'] = request.GET.get('department')
    if request.GET.get('line'):
        filters['line_id'] = request.GET.get('line')
    if request.GET.get('subline'):
        filters['subline_id'] = request.GET.get('subline')
    if request.GET.get('station'):
        filters['station_id'] = request.GET.get('station')
    # if request.GET.get('shop_floor'): # Keep if still in use
    #     filters['shop_floor_id'] = request.GET.get('shop_floor')

    trend_data = []
    data_type = time_view

    if time_view == 'Weekly':
        # Existing weekly logic (working fine)
        from dateutil.rrule import rrule, WEEKLY
        weeks = list(rrule(WEEKLY, dtstart=start_date, until=end_date))
        for week_start in weeks:
            week_end = min(week_start.date() + timedelta(days=6), end_date)
            weekly_filters = filters.copy()
            weekly_filters['start_date__range'] = (week_start.date(), week_end)
            weekly_aggregation = DailyProductionData.objects.filter(**weekly_filters).aggregate(
                plan_sum=Sum('total_operators_required_plan'),
                actual_sum=Sum('total_operators_required_actual')
            )
            trend_data.append({
                'period': week_start.date().isoformat(),
                'total_operators_required_plan': weekly_aggregation['plan_sum'] or 0,
                'total_operators_required_actual': weekly_aggregation['actual_sum'] or 0,
            })
    else:  # Monthly
        year = start_date.year
        for month in range(1, 13):
            # Create filters for this specific month and year
            monthly_filters = filters.copy()
            monthly_filters['start_date__year'] = year
            monthly_filters['start_date__month'] = month

            # Aggregate data for just this month
            monthly_aggregation = DailyProductionData.objects.filter(**monthly_filters).aggregate(
                plan_sum=Sum('total_operators_required_plan'),
                actual_sum=Sum('total_operators_required_actual')
            )
            
            # Append this month's data to the list
            trend_data.append({
                'period': calendar.month_name[month],
                'year': year,
                'total_operators_required_plan': monthly_aggregation['plan_sum'] or 0,
                'total_operators_required_actual': monthly_aggregation['actual_sum'] or 0,
            })

    return JsonResponse({
        'data_type': data_type,
        'data': trend_data
    }, safe=False, status=200)





from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.db.models import Sum, Avg
from .models import DailyProductionData
from datetime import datetime
from dateutil import parser
import calendar
from math import floor, ceil

@require_GET
def monthly_availability_analysis(request):
    """
    Monthly gap analysis endpoint for /production-data/gap-analysis/.
    Computes GAP, gap_message, and Lost Production for multiple production plan scenarios.
    """
    # Get query parameters
    factory_id = request.GET.get('factory')
    month = request.GET.get('month')  # e.g., "March"
    year = request.GET.get('year')  # e.g., "2025"
    # shop_floor_id = request.GET.get('shop_floor')
    line_id = request.GET.get('line')
    station_id = request.GET.get('station')
    Hq_id = request.GET.get('Hq')
    department_id = request.GET.get('department')
    subline_id = request.GET.get('subline')

    # Validate required parameters
    if not (factory_id and month and year):
        return JsonResponse(
            {"success": False, "error": "Missing required parameters: factory, month, year"},
            status=400
        )

    try:
        year = int(year)
        month_num = list(calendar.month_name).index(month.capitalize())
    except (ValueError, IndexError):
        return JsonResponse(
            {"success": False, "error": "Invalid month or year format"},
            status=400
        )

    # Determine date range for the month
    start_date = datetime(year, month_num, 1)
    _, last_day = calendar.monthrange(year, month_num)
    end_date = datetime(year, month_num, last_day)

    # Calculate dynamic productivity rate from previous month's data
    prev_month = start_date.replace(day=1, month=month_num - 1 if month_num > 1 else 12)
    prev_year = year if month_num > 1 else year - 1
    prev_data = DailyProductionData.objects.filter(
        factory_id=factory_id,
        start_date__year=prev_year,
        start_date__month=prev_month.month,
        entry_mode='MONTHLY'
    ).aggregate(
        total_production_actual=Sum('total_production_actual'),
        total_operators_actual=Sum('total_operators_required_actual')
    )
    productivity_rate = (
        prev_data['total_production_actual'] / prev_data['total_operators_actual']
        if prev_data['total_production_actual'] and prev_data['total_operators_actual']
        else 12  # Fallback to Excel's default
    )

    # Build queryset with filters
    queryset = DailyProductionData.objects.filter(
        factory_id=factory_id,
        start_date__gte=start_date,
        end_date__lte=end_date,
        entry_mode='MONTHLY'
    )
    if Hq_id:
        queryset = queryset.filter(Hq_id=Hq_id)
    if department_id:
        queryset = queryset.filter(department_id=department_id)
    if subline_id:
        queryset = queryset.filter(subline_id=subline_id)
    # if shop_floor_id:
    #     queryset = queryset.filter(shop_floor_id=shop_floor_id)
    if line_id:
        queryset = queryset.filter(line_id=line_id)
    if station_id:
        queryset = queryset.filter(station_id=station_id)

    # Process multiple scenarios
    data_points = []
    for record in queryset:
        # Calculate Excel-derived values
        ending_team = record.total_operators_available * (1 - record.attrition_rate / 100)
        operators_available = floor(ending_team * (1 - record.absenteeism_rate / 100))  # ROUNDDOWN
        gap = operators_available - record.total_operators_required_plan
        lost_production = abs(gap) * productivity_rate if gap < 0 else 0
        gap_message = (
            f"Shortage: Hire {abs(gap)}" if gap < 0 else
            f"Surplus of {gap}" if gap > 0 else
            "Balanced"
        )

        # Previous month's actual operators for context
        prev_data = DailyProductionData.objects.filter(
            factory_id=factory_id,
            start_date__year=prev_year,
            start_date__month=prev_month.month,
            entry_mode='MONTHLY'
        ).aggregate(total_actual=Sum('total_operators_required_actual'))

        data_points.append({
            "month": month,
            "year": year,
            "production_plan": {"value": record.total_production_plan},
            "operators_required": {"value": record.total_operators_required_plan},
            "operators_available": {"value": operators_available},
            "gap": {"value": gap},
            "gap_message": gap_message,
            "lost_production": {"value": round(lost_production)},
            "attrition_rate": float(record.attrition_rate),
            "previous_month_data": {
                "month": calendar.month_name[prev_month.month],
                "actual_operators": prev_data['total_actual'] or 0
            } if prev_data['total_actual'] else None
        })

    if not data_points:
        return JsonResponse(
            {"success": False, "error": "No data found for the selected filters"},
            status=404
        )

    return JsonResponse({
        "success": True,
        "operators_availability_graph": data_points[0] if len(data_points) == 1 else data_points
    })

@require_GET
def weekly_availability_summary(request):
    """
    Weekly summary endpoint for /production-data/date-range-summary/.
    Aggregates data for a week and computes GAP, gap_message, and Lost Production.
    """
    # Get query parameters
    factory_id = request.GET.get('factory')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    # shop_floor_id = request.GET.get('shop_floor')
    line_id = request.GET.get('line')
    station_id = request.GET.get('station')
    Hq_id = request.GET.get('Hq')
    department_id = request.GET.get('department')
    subline_id = request.GET.get('subline')

    # Validate required parameters
    if not (factory_id and start_date_str and end_date_str):
        return JsonResponse(
            {"success": False, "error": "Missing required parameters: factory, start_date, end_date"},
            status=400
        )

    try:
        start_date = parser.parse(start_date_str).date()
        end_date = parser.parse(end_date_str).date()
    except ValueError:
        return JsonResponse(
            {"success": False, "error": "Invalid date format"},
            status=400
        )

    # Calculate dynamic productivity rate from previous week's data
    prev_week_end = start_date - timedelta(days=1)
    prev_week_start = prev_week_end - timedelta(days=6)
    prev_data = DailyProductionData.objects.filter(
        factory_id=factory_id,
        start_date__gte=prev_week_start,
        end_date__lte=prev_week_end,
        entry_mode='WEEKLY'
    ).aggregate(
        total_production_actual=Sum('total_production_actual'),
        total_operators_actual=Sum('total_operators_required_actual')
    )
    productivity_rate = (
        prev_data['total_production_actual'] / prev_data['total_operators_actual']
        if prev_data['total_production_actual'] and prev_data['total_operators_actual']
        else 12  # Fallback to Excel's default
    )

    # Build queryset with filters
    queryset = DailyProductionData.objects.filter(
        factory_id=factory_id,
        start_date__gte=start_date,
        end_date__lte=end_date,
        entry_mode='WEEKLY'
    )
    if Hq_id:
        queryset = queryset.filter(Hq_id=Hq_id)
    if department_id:
        queryset = queryset.filter(department_id=department_id)
    if subline_id:
        queryset = queryset.filter(subline_id=subline_id)
    # if shop_floor_id:
    #     queryset = queryset.filter(shop_floor_id=shop_floor_id)
    if line_id:
        queryset = queryset.filter(line_id=line_id)
    if station_id:
        queryset = queryset.filter(station_id=station_id)

    # Aggregate data for the week
    aggregates = queryset.aggregate(
        total_production_plan=Sum('total_production_plan'),
        total_operators_required_plan=Sum('total_operators_required_plan'),
        total_operators_available=Sum('total_operators_available'),
        avg_attrition_rate=Avg('attrition_rate'),
        avg_absenteeism_rate=Avg('absenteeism_rate')
    )

    if not aggregates['total_production_plan']:
        return JsonResponse(
            {"success": False, "error": "No data found for the selected week"},
            status=404
        )

    # Calculate Excel-derived values
    ending_team = aggregates['total_operators_available'] * (1 - aggregates['avg_attrition_rate'] / 100)
    operators_available = floor(ending_team * (1 - aggregates['avg_absenteeism_rate'] / 100))  # ROUNDDOWN
    gap = operators_available - aggregates['total_operators_required_plan']
    lost_production = abs(gap) * productivity_rate if gap < 0 else 0
    gap_message = (
        f"Shortage: Hire {abs(gap)}" if gap < 0 else
        f"Surplus of {gap}" if gap > 0 else
        "Balanced"
    )

    return JsonResponse({
        "success": True,
        "total_production_plan": aggregates['total_production_plan'] or 0,
        "total_operators_required_plan": aggregates['total_operators_required_plan'] or 0,
        "total_operators_required_actual": operators_available,
        "attrition_rate": float(aggregates['avg_attrition_rate'] or 0),
        "gap": gap,
        "gap_message": gap_message,
        "lost_production": round(lost_production)
    })



from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import SkillMatrix, MasterTable
from .serializers import SkillMatrixSerializer
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
# ⭐ ADDED: Import Score model
from .models import SkillMatrix, MasterTable, Score 
from .serializers import SkillMatrixSerializer

class SkillMatrixViewSet(viewsets.ModelViewSet):
    queryset = SkillMatrix.objects.select_related(
        'employee',
        'hierarchy',
        'hierarchy__station',
        'level',).all()
    serializer_class = SkillMatrixSerializer

    # Custom endpoint: filter by station_id
    @action(detail=False, methods=["get"])
    def by_station(self, request):
        station_id = request.query_params.get("station_id")
        if not station_id:
            return Response({"error": "station_id parameter is required"}, status=400)

        queryset = self.queryset.filter(hierarchy__station_id=station_id)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    # ---------------------------------------------------------
    # NEW: Get ONLY Eligible Employees (Includes Exam Pass Check)
    # ---------------------------------------------------------
    @action(detail=False, methods=['get'], url_path='eligible-employees')
    def eligible_employees(self, request):
        station_id = request.query_params.get("station_id")
        target_level = request.query_params.get("target_level")

        if not station_id or not target_level:
            return Response({"error": "station_id and target_level required"}, status=400)

        try:
            target_level = int(target_level)
        except ValueError:
            return Response({"error": "target_level must be int"}, status=400)

        # --- EXCLUSION LOGIC ---
        
        # 1. Fully Certified (Exam + OJT)
        # These people are in SkillMatrix. They are finished.
        certified_pks = SkillMatrix.objects.filter(
            hierarchy__station__station_id=station_id,
            level__level_id=target_level
        ).values_list('employee_id', flat=True)

        # 2. Passed Written Exam (But maybe not OJT yet)
        # These people should NOT write the exam again.
        # We check the Score table for passed=True.
        exam_passed_pks = Score.objects.filter(
            skill__station_id=station_id,
            level__level_id=target_level,
            passed=True  # ⭐ KEY CHANGE: Only block if they passed
        ).values_list('employee_id', flat=True)

        # Combine both lists to get everyone who shouldn't take the test
        excluded_pks = list(set(list(certified_pks) + list(exam_passed_pks)))

        candidates = []

        # --- ELIGIBILITY LOGIC ---

        if target_level == 1:
            # Level 1: Everyone eligible EXCEPT excluded_pks
            candidates = MasterTable.objects.exclude(pk__in=excluded_pks)

        elif target_level == 2:
            # Level 2: 
            # 1. Must have passed Level 1 (Anywhere)
            # 2. Must NOT have passed Level 2 Exam or OJT (excluded_pks)
            
            l1_holders_pks = SkillMatrix.objects.filter(
                level__level_id=1
            ).values_list('employee_id', flat=True)
            
            candidates = MasterTable.objects.filter(
                pk__in=l1_holders_pks
            ).exclude(
                pk__in=excluded_pks
            )

        else:
            # Level 3+: 
            # 1. Must have passed (Level - 1) at THIS STATION (Full Cert: Exam + OJT)
            # 2. Must NOT have passed Current Level Exam or OJT
            
            prev_level = target_level - 1
            
            # Use SkillMatrix here because you said Level 3 only opens 
            # if they pass OJT (which adds them to SkillMatrix)
            prev_level_holders_pks = SkillMatrix.objects.filter(
                hierarchy__station__station_id=station_id,
                level__level_id=prev_level
            ).values_list('employee_id', flat=True)

            candidates = MasterTable.objects.filter(
                pk__in=prev_level_holders_pks
            ).exclude(
                pk__in=excluded_pks
            )

        # --- SERIALIZATION ---
        data = []
        for emp in candidates:
            # Safe attribute access
            name = getattr(emp, 'name', None) or f"{getattr(emp, 'first_name', '')} {getattr(emp, 'last_name', '')}".strip()
            
            section = getattr(emp, 'section', None)
            if not section and hasattr(emp, 'department') and emp.department:
                section = emp.department.department_name
            if not section:
                section = 'N/A'

            pay_code = getattr(emp, 'pay_code', None) or getattr(emp, 'emp_id', 'N/A')
            emp_id_str = str(emp.pk)

            data.append({
                "id": emp_id_str,
                "name": name,
                "pay_code": pay_code,
                "section": section,
                "eligible": True
            })

        return Response(data)

    # Legacy endpoint
    @action(detail=False, methods=['get'], url_path='check-eligibility')
    def check_eligibility(self, request):
        return Response({"eligible": True})




from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import HierarchyStructure

class HierarchyAllDepartmentsView(APIView):
    """
    Fetch hierarchy for all departments with flexible nesting:
    department → line → subline → station
    department → line → station
    department → subline → station
    department → station
    """

    def get(self, request, *args, **kwargs):
        try:
            structures = HierarchyStructure.objects.select_related(
                "department", "line", "subline", "station"
            )

            if not structures.exists():
                return Response(
                    {"error": "No hierarchy data found"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            departments_data = {}

            for structure in structures:
                department = structure.department
                if not department:
                    continue  # skip if no department linked

                dept_id = department.department_id
                if dept_id not in departments_data:
                    departments_data[dept_id] = {
                        "department_id": dept_id,
                        "department_name": department.department_name,
                        "lines": {},
                        "sublines": {},
                        "stations": {},
                    }

                department_data = departments_data[dept_id]
                line = structure.line
                subline = structure.subline
                station = structure.station

                if line:
                    if line.line_id not in department_data["lines"]:
                        department_data["lines"][line.line_id] = {
                            "line_id": line.line_id,
                            "line_name": line.line_name,
                            "sublines": {},
                            "stations": {},
                        }

                    if subline:
                        if subline.subline_id not in department_data["lines"][line.line_id]["sublines"]:
                            department_data["lines"][line.line_id]["sublines"][subline.subline_id] = {
                                "subline_id": subline.subline_id,
                                "subline_name": subline.subline_name,
                                "stations": {},
                            }

                        if station:
                            department_data["lines"][line.line_id]["sublines"][subline.subline_id]["stations"][
                                station.station_id
                            ] = {
                                "station_id": station.station_id,
                                "station_name": station.station_name,
                            }
                    else:
                        if station:
                            department_data["lines"][line.line_id]["stations"][station.station_id] = {
                                "station_id": station.station_id,
                                "station_name": station.station_name,
                            }

                elif subline:
                    if subline.subline_id not in department_data["sublines"]:
                        department_data["sublines"][subline.subline_id] = {
                            "subline_id": subline.subline_id,
                            "subline_name": subline.subline_name,
                            "stations": {},
                        }

                    if station:
                        department_data["sublines"][subline.subline_id]["stations"][station.station_id] = {
                            "station_id": station.station_id,
                            "station_name": station.station_name,
                        }

                elif station:
                    department_data["stations"][station.station_id] = {
                        "station_id": station.station_id,
                        "station_name": station.station_name,
                    }

            # convert dicts to lists for clean JSON
            for dept_id, department_data in departments_data.items():
                department_data["lines"] = list(department_data["lines"].values())
                for line in department_data["lines"]:
                    line["sublines"] = list(line["sublines"].values())
                    line["stations"] = list(line["stations"].values())
                    for subline in line["sublines"]:
                        subline["stations"] = list(subline["stations"].values())

                department_data["sublines"] = list(department_data["sublines"].values())
                for subline in department_data["sublines"]:
                    subline["stations"] = list(subline["stations"].values())

                department_data["stations"] = list(department_data["stations"].values())

            return Response(list(departments_data.values()), status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import HttpResponse, Http404
from django.shortcuts import get_object_or_404
from django.conf import settings
import os
import mimetypes
from .models import UserManualdocs
from .serializers import UserManualdocsSerializer

class UserManualdocsListCreateView(generics.ListCreateAPIView):
    queryset = UserManualdocs.objects.all()
    serializer_class = UserManualdocsSerializer
    parser_classes = [MultiPartParser, FormParser]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    'message': 'Document uploaded successfully!',
                    'data': serializer.data
                },
                status=status.HTTP_201_CREATED
            )
        return Response(
            {
                'message': 'Failed to upload document',
                'errors': serializer.errors
            },
            status=status.HTTP_400_BAD_REQUEST
        )

class UserManualdocsDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = UserManualdocs.objects.all()
    serializer_class = UserManualdocsSerializer
    parser_classes = [MultiPartParser, FormParser]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response(
            {'message': 'Document deleted successfully!'},
            status=status.HTTP_200_OK
        )

@api_view(['GET'])
def view_file(request, doc_id):
    try:
        doc = get_object_or_404(UserManualdocs, id=doc_id)
        
        if not doc.file:
            return Response(
                {'error': 'No file associated with this document'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        file_path = doc.file.path
        
        if not os.path.exists(file_path):
            return Response(
                {'error': 'File not found on server'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get the file's MIME type
        mime_type, _ = mimetypes.guess_type(file_path)
        if mime_type is None:
            mime_type = 'application/octet-stream'
        
        # Read the file
        try:
            with open(file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type=mime_type)
                
                # Set headers for inline viewing (opens in browser)
                filename = os.path.basename(file_path)
                response['Content-Disposition'] = f'inline; filename="{filename}"'
                
                return response
                
        except Exception as e:
            return Response(
                {'error': f'Error reading file: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            
    except UserManualdocs.DoesNotExist:
        return Response(
            {'error': 'Document not found'},
            status=status.HTTP_404_NOT_FOUND
        )

@api_view(['GET'])
def download_file(request, doc_id):
    try:
        doc = get_object_or_404(UserManualdocs, id=doc_id)
        
        if not doc.file:
            return Response(
                {'error': 'No file associated with this document'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        file_path = doc.file.path
        
        if not os.path.exists(file_path):
            return Response(
                {'error': 'File not found on server'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get the file's MIME type
        mime_type, _ = mimetypes.guess_type(file_path)
        if mime_type is None:
            mime_type = 'application/octet-stream'
        
        # Read the file
        try:
            with open(file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type=mime_type)
                
                # Set headers for forced download
                filename = os.path.basename(file_path)
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                
                return response
                
        except Exception as e:
            return Response(
                {'error': f'Error reading file: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            
    except UserManualdocs.DoesNotExist:
        return Response(
            {'error': 'Document not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
# 


# STANDARD_A0002(trainingattendance)
# =================== TrainingAttendance ============================= #



from django.utils import timezone
from django.db import transaction
from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response

# Make sure all your models and serializers are imported
from .models import TrainingBatch, UserRegistration, TrainingAttendance, Days
from .serializers import TrainingBatchSerializer, BatchAttendanceDetailSerializer, TrainingAttendanceSerializer

# --- Active/Past Batch Views (Unchanged, but included for context) ---
class ActiveTrainingBatchListView(generics.ListAPIView):
    queryset = TrainingBatch.objects.filter(is_active=True)
    serializer_class = TrainingBatchSerializer


from django.utils import timezone # Make sure timezone is imported

class BatchAttendanceDetailView(APIView):
    """
    GET /api/attendance-detail/{batch_id}/
    Returns all data needed for the attendance page.
    The concept of 'next_training_day_to_mark' is removed from the logic,
    as all days are now editable. It is kept in the response for compatibility but is null.
    """
    def get(self, request, batch_id, *args, **kwargs):
        try:
            batch = TrainingBatch.objects.get(batch_id=batch_id)
        except TrainingBatch.DoesNotExist:
            return Response({"error": "Batch not found."}, status=status.HTTP_404_NOT_FOUND)

        users = UserRegistration.objects.filter(batch_id=batch_id)
        if not users.exists():
            return Response({"error": "No users found for this batch."}, status=status.HTTP_404_NOT_FOUND)

        data = {
            'batch_id': batch_id,
            'next_training_day_to_mark': None, # This is no longer used to lock UI
            'is_completed': not batch.is_active,
            'users': users
        }
        
        serializer = BatchAttendanceDetailSerializer(data, context={'batch_id': batch_id})
        return Response(serializer.data, status=status.HTTP_200_OK)







from django.utils import timezone
from django.db import transaction
from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response

# Make sure all your models and serializers are imported
from .models import TrainingBatch, UserRegistration, TrainingAttendance, Days
from .serializers import TrainingBatchSerializer, BatchAttendanceDetailSerializer, TrainingAttendanceSerializer


class PastTrainingBatchListView(generics.ListAPIView):
    queryset = TrainingBatch.objects.filter(is_active=False)
    serializer_class = TrainingBatchSerializer




from django.db.models import Max, Count
from django.utils import timezone
from django.db import transaction
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from datetime import datetime as dt



from .models import TrainingBatch, UserRegistration, TrainingAttendance, Days
from .serializers import TrainingAttendanceSerializer
from datetime import datetime as dt

class BulkAttendanceUpdateView(APIView):
    """
    POST /api/attendances/
    Creates or updates a list of attendance records. This view is designed to be
    flexible, accepting records for multiple users across multiple days in a single request.
    It uses a transaction to ensure all updates succeed or none do.
    
    Auto-completion logic:
    - Dynamically calculates final day from Days model
    - Only completes batch when ALL days are marked for ALL users
    - No hardcoded day numbers
    """
    
    def post(self, request, *args, **kwargs):
        attendance_data = request.data
        
        # Validate payload structure
        if not isinstance(attendance_data, list):
            return Response(
                {"error": "Expected a list of attendance objects."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not attendance_data:
            return Response(
                {"message": "No attendance data provided to update."},
                status=status.HTTP_200_OK
            )

        # Extract batch ID from first record
        target_batch_id = attendance_data[0].get('batch')
        if not target_batch_id:
            return Response(
                {"error": "Batch ID is missing in payload."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verify batch exists
        try:
            batch = TrainingBatch.objects.get(batch_id=target_batch_id)
        except TrainingBatch.DoesNotExist:
            return Response(
                {"error": f"Batch '{target_batch_id}' not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        created_records = []
        updated_records = []
        errors = []

        try:
            # Use a transaction to make the entire operation atomic
            with transaction.atomic():
                for idx, item in enumerate(attendance_data):
                    user_id = item.get('user')
                    day_id = item.get('day_number')
                    status_val = item.get('status')
                    attendance_date_str = item.get('attendance_date')

                    if not attendance_date_str:
                        raise ValueError(f"attendance_date missing for user {user_id}")

                    try:
                        # attendance_date = datetime.strptime(attendance_date_str, "%Y-%m-%d").date()
                        attendance_date = dt.strptime(attendance_date_str, "%Y-%m-%d").date()
                    except ValueError:
                        raise ValueError(f"Invalid attendance_date format: {attendance_date_str}")

                    # Validate required fields
                    if not all([user_id, day_id, status_val]):
                        errors.append({
                            "index": idx,
                            "error": "Missing required fields (user, day_number, or status)",
                            "item": item
                        })
                        continue

                    # Validate and parse attendance_date
                    if not attendance_date_str:
                        errors.append({
                            "index": idx,
                            "error": "attendance_date is required",
                            "item": item
                        })
                        continue

                    try:
                        # attendance_date = datetime.strptime(attendance_date_str, "%Y-%m-%d").date()
                        attendance_date = dt.strptime(attendance_date_str, "%Y-%m-%d").date()
                    except ValueError:
                        errors.append({
                            "index": idx,
                            "error": f"Invalid attendance_date format: {attendance_date_str}. Expected YYYY-MM-DD",
                            "item": item
                        })
                        continue

                    # Validate status value
                    if status_val not in ['present', 'absent']:
                        errors.append({
                            "index": idx,
                            "error": f"Invalid status value: {status_val}. Must be 'present' or 'absent'",
                            "item": item
                        })
                        continue

                    # Validate user exists
                    try:
                        user = UserRegistration.objects.get(id=user_id)
                    except UserRegistration.DoesNotExist:
                        errors.append({
                            "index": idx,
                            "error": f"User with ID {user_id} not found",
                            "item": item
                        })
                        continue

                    # Validate day exists
                    try:
                        day = Days.objects.get(days_id=day_id)
                    except Days.DoesNotExist:
                        errors.append({
                            "index": idx,
                            "error": f"Day with ID {day_id} not found",
                            "item": item
                        })
                        continue

                    # Use update_or_create to handle both new and existing records
                    try:
                        obj, created = TrainingAttendance.objects.update_or_create(
                            user=user,
                            batch=batch,
                            day_number=day,
                            defaults={
                                'status': status_val,
                                'attendance_date': attendance_date
                            }
                        )
                        
                        serialized_data = TrainingAttendanceSerializer(obj).data
                        
                        if created:
                            created_records.append(serialized_data)
                        else:
                            updated_records.append(serialized_data)
                            
                    except Exception as e:
                        errors.append({
                            "index": idx,
                            "error": f"Database error: {str(e)}",
                            "item": item
                        })
                        continue

                # If there were any errors, rollback the transaction
                if errors:
                    raise ValueError("Validation errors occurred")

            # --- AUTO-COMPLETION LOGIC (DYNAMIC) ---
            # Only runs if the transaction was successful
            self._check_and_complete_batch(target_batch_id)

        except ValueError as e:
            # Validation errors - transaction was rolled back
            return Response({
                "error": "Attendance save failed due to validation errors",
                "validation_errors": errors,
                "created": 0,
                "updated": 0
            }, status=status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            # Unexpected errors
            return Response({
                "error": f"An unexpected error occurred: {str(e)}",
                "created": len(created_records),
                "updated": len(updated_records)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Success response
        total_processed = len(created_records) + len(updated_records)
        
        return Response({
            "message": f"Attendance saved successfully. {len(created_records)} created, {len(updated_records)} updated.",
            "summary": {
                "total_processed": total_processed,
                "created": len(created_records),
                "updated": len(updated_records),
                "batch_id": target_batch_id
            },
            "created": created_records,
            "updated": updated_records
        }, status=status.HTTP_201_CREATED)

    def _check_and_complete_batch(self, batch_id):
        """
        Internal method to check if a batch should be auto-completed.
        
        Completion criteria:
        - ALL days must be marked for ALL users
        - No hardcoded day numbers
        - Dynamically calculated from Days model
        
        Args:
            batch_id: The batch ID to check
        """
        try:
            batch = TrainingBatch.objects.get(batch_id=batch_id)
            
            # Skip if already completed
            if not batch.is_active:
                print(f"ℹ️  Batch {batch_id} is already marked as completed")
                return

            # Get total number of training days
            total_days = Days.objects.count()
            
            if total_days == 0:
                print(f"⚠️  No training days found in Days model")
                return

            # Get total number of users in this batch
            users_in_batch_count = UserRegistration.objects.filter(batch_id=batch_id).count()
            
            if users_in_batch_count == 0:
                print(f"⚠️  No users found in batch {batch_id}")
                return

            # Calculate total required attendance records
            # (each user must have attendance for each day)
            total_required_records = total_days * users_in_batch_count

            # Count actual attendance records for this batch
            actual_attendance_count = TrainingAttendance.objects.filter(
                batch_id=batch_id
            ).count()

            # Calculate completion percentage
            completion_percentage = (actual_attendance_count / total_required_records * 100) if total_required_records > 0 else 0

            print(f"📊 Batch {batch_id} completion status:")
            print(f"   - Total days: {total_days}")
            print(f"   - Total users: {users_in_batch_count}")
            print(f"   - Required records: {total_required_records}")
            print(f"   - Actual records: {actual_attendance_count}")
            print(f"   - Completion: {completion_percentage:.1f}%")

            # Complete the batch only if ALL records are present
            if actual_attendance_count >= total_required_records:
                batch.is_active = False
                batch.save()
                
                print(f"✅ Batch {batch_id} automatically completed!")
                print(f"   All {total_days} days marked for all {users_in_batch_count} users")
            else:
                missing_records = total_required_records - actual_attendance_count
                print(f"⏳ Batch {batch_id} not yet complete ({missing_records} records remaining)")

        except TrainingBatch.DoesNotExist:
            print(f"❌ Batch {batch_id} not found during completion check")
        except Exception as e:
            print(f"❌ Error checking batch completion: {str(e)}")


# Optional: Helper method to manually verify completion status
    def _get_batch_completion_details(self, batch_id):
        """
        Get detailed completion information for a batch.
        Useful for debugging or displaying progress.
        
        Returns:
            dict: Completion details including per-day breakdown
        """
        try:
            batch = TrainingBatch.objects.get(batch_id=batch_id)
            total_days = Days.objects.count()
            users_count = UserRegistration.objects.filter(batch_id=batch_id).count()
            
            # Get per-day statistics
            day_breakdown = []
            for day in Days.objects.all().order_by('days_id'):
                marked_count = TrainingAttendance.objects.filter(
                    batch_id=batch_id,
                    day_number=day
                ).count()
                
                day_breakdown.append({
                    "day_id": day.days_id,
                    "day_name": day.day,
                    "marked": marked_count,
                    "total_users": users_count,
                    "is_complete": marked_count == users_count,
                    "completion_percentage": (marked_count / users_count * 100) if users_count > 0 else 0
                })
            
            total_required = total_days * users_count
            total_marked = TrainingAttendance.objects.filter(batch_id=batch_id).count()
            
            return {
                "batch_id": batch_id,
                "is_active": batch.is_active,
                "total_days": total_days,
                "total_users": users_count,
                "total_required_records": total_required,
                "total_marked_records": total_marked,
                "overall_completion_percentage": (total_marked / total_required * 100) if total_required > 0 else 0,
                "all_days_complete": all(day['is_complete'] for day in day_breakdown),
                "day_by_day": day_breakdown
            }
        except Exception as e:
            return {"error": str(e)}


class BatchCompletionStatusView(APIView):
    """
    GET /api/batch/{batch_id}/completion-status/
    Returns detailed completion information for a batch
    """
    def get(self, request, batch_id, *args, **kwargs):
        try:
            batch = TrainingBatch.objects.get(batch_id=batch_id)
            
            total_days = Days.objects.count()
            users_count = UserRegistration.objects.filter(batch_id=batch_id).count()
            
            # Get attendance stats per day
            day_stats = []
            for day in Days.objects.all().order_by('days_id'):
                marked_count = TrainingAttendance.objects.filter(
                    batch_id=batch_id,
                    day_number=day
                ).count()
                
                day_stats.append({
                    "day_id": day.days_id,
                    "day_name": day.day,
                    "marked": marked_count,
                    "total": users_count,
                    "is_complete": marked_count == users_count,
                    "percentage": round((marked_count / users_count * 100) if users_count > 0 else 0, 1)
                })
            
            total_required = total_days * users_count
            total_marked = TrainingAttendance.objects.filter(
                batch_id=batch_id
            ).count()
            
            all_days_complete = all(day['is_complete'] for day in day_stats)
            
            return Response({
                "batch_id": batch_id,
                "is_active": batch.is_active,
                "status": "Active" if batch.is_active else "Completed",
                "total_days": total_days,
                "total_users": users_count,
                "total_required_records": total_required,
                "total_marked_records": total_marked,
                "completion_percentage": round((total_marked / total_required * 100) if total_required > 0 else 0, 2),
                "all_days_complete": all_days_complete,
                "day_by_day_status": day_stats,
                "can_auto_complete": all_days_complete and batch.is_active
            }, status=status.HTTP_200_OK)
            
        except TrainingBatch.DoesNotExist:
            return Response(
                {"error": "Batch not found"},
                status=status.HTTP_404_NOT_FOUND
            )


class ToggleBatchCompletionView(APIView):
    """
    POST /api/batch/{batch_id}/toggle-completion/
    Manually toggle batch completion status (for admin override)
    """
    def post(self, request, batch_id, *args, **kwargs):
        try:
            batch = TrainingBatch.objects.get(batch_id=batch_id)
            
            # Toggle the status
            old_status = "completed" if not batch.is_active else "active"
            batch.is_active = not batch.is_active
            batch.save()
            
            new_status = "completed" if not batch.is_active else "active"
            
            return Response({
                "message": f"Batch status changed from {old_status} to {new_status}",
                "batch_id": batch_id,
                "is_active": batch.is_active,
                "status": new_status
            }, status=status.HTTP_200_OK)
            
        except TrainingBatch.DoesNotExist:
            return Response(
                {"error": "Batch not found"},
                status=status.HTTP_404_NOT_FOUND
            )


class ReopenBatchView(APIView):
    """
    POST /api/batch/{batch_id}/reopen/
    Reopen a completed batch (sets is_active = True)
    """
    def post(self, request, batch_id, *args, **kwargs):
        try:
            batch = TrainingBatch.objects.get(batch_id=batch_id)
            
            if batch.is_active:
                return Response({
                    "message": "Batch is already active",
                    "batch_id": batch_id,
                    "is_active": True
                }, status=status.HTTP_200_OK)
            
            batch.is_active = True
            batch.save()
            
            return Response({
                "message": f"Batch {batch_id} has been reopened for editing",
                "batch_id": batch_id,
                "is_active": True
            }, status=status.HTTP_200_OK)
            
        except TrainingBatch.DoesNotExist:
            return Response(
                {"error": "Batch not found"},
                status=status.HTTP_404_NOT_FOUND
            )

class CompleteTrainingBatchView(APIView):
    def post(self, request, batch_id, *args, **kwargs):
        try:
            batch = TrainingBatch.objects.get(batch_id=batch_id)
            if not batch.is_active:
                return Response({"message": "Batch is already completed."}, status=status.HTTP_200_OK)
            batch.is_active = False
            batch.save()
            return Response(TrainingBatchSerializer(batch).data, status=status.HTTP_200_OK)
        except TrainingBatch.DoesNotExist:
            return Response({"error": "Batch not found."}, status=status.HTTP_404_NOT_FOUND)
        




from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from django.db.models import Q, Count, Exists, OuterRef

class AbsentUsersListView(APIView):
    """
    GET /api/batch/{batch_id}/absentees/
    Returns only absences that are NOT yet rescheduled.
    After rescheduling, they disappear from this list permanently.
    """
    def get(self, request, batch_id, *args, **kwargs):
        batch = get_object_or_404(TrainingBatch, batch_id=batch_id)
        
        day_id = request.query_params.get('day')
        overall = request.query_params.get('overall', '').lower() == 'true'

        users_in_batch = UserRegistration.objects.filter(batch_id=batch_id)
        if not users_in_batch.exists():
            return Response({"detail": "No users in this batch."}, status=status.HTTP_404_NOT_FOUND)

        # Subquery: Does a rescheduled session exist for this exact absence?
        rescheduled_subquery = RescheduledSession.objects.filter(
            employee=OuterRef('user'),
            batch=OuterRef('batch'),
            original_day=OuterRef('day_number'),
        )

        # Base query: all absences that are NOT rescheduled
        absent_attendances = TrainingAttendance.objects.filter(
            batch_id=batch_id,
            status='absent'
        ).exclude(
            Exists(rescheduled_subquery)
        ).select_related('user', 'day_number')

        if day_id:
            # Specific day
            day_id = int(day_id)
            absent_attendances = absent_attendances.filter(day_number__days_id=day_id)
            
            # ✅ Build list with MasterTable lookup
            absent_list = []
            for att in absent_attendances.select_related('user'):
                user = att.user
                
                # Check MasterTable for emp_id
                emp_id = user.emp_id
                try:
                    master_emp = MasterTable.objects.get(
                        first_name__iexact=user.first_name,
                        last_name__iexact=user.last_name or ''
                    )
                    emp_id = master_emp.emp_id
                except (MasterTable.DoesNotExist, MasterTable.MultipleObjectsReturned):
                    pass  # Use original emp_id
                
                absent_list.append({
                    "id": user.id,
                    "temp_id": user.temp_id,
                    "emp_id": emp_id,
                    "full_name": f"{user.first_name} {user.last_name or ''}".strip(),
                    "attendance_date": att.attendance_date,
                })
            
            return Response({
                "batch_id": batch_id,
                "on_day": day_id,
                "total_absent": len(absent_list),
                "absent_users": absent_list
            }, status=status.HTTP_200_OK)

        elif overall:
            # Overall unique users
            absent_user_ids = absent_attendances.values_list('user_id', flat=True).distinct()
            users = users_in_batch.filter(id__in=absent_user_ids)
            
            # ✅ Build list with MasterTable lookup
            absent_list = []
            for user in users:
                emp_id = user.emp_id
                try:
                    master_emp = MasterTable.objects.get(
                        first_name__iexact=user.first_name,
                        last_name__iexact=user.last_name or ''
                    )
                    emp_id = master_emp.emp_id
                except (MasterTable.DoesNotExist, MasterTable.MultipleObjectsReturned):
                    pass
                
                absent_list.append({
                    "id": user.id,
                    "temp_id": user.temp_id,
                    "emp_id": emp_id,
                    "full_name": f"{user.first_name} {user.last_name or ''}".strip(),
                })
            
            return Response({
                "batch_id": batch_id,
                "scope": "overall",
                "total_absent": len(absent_list),
                "absent_users": absent_list
            }, status=status.HTTP_200_OK)

        else:
            # Grouped by day
            absences = absent_attendances.values(
                'day_number_id', 'day_number__days_id'
            ).annotate(count=Count('id')).order_by('day_number__days_id')

            result = []
            for item in absences:
                day_absent_attendances = absent_attendances.filter(
                    day_number_id=item['day_number_id']
                )
                
                # ✅ Build user list with MasterTable lookup
                absent_users = []
                for att in day_absent_attendances.select_related('user'):
                    user = att.user
                    emp_id = user.emp_id
                    
                    try:
                        master_emp = MasterTable.objects.get(
                            first_name__iexact=user.first_name,
                            last_name__iexact=user.last_name or ''
                        )
                        emp_id = master_emp.emp_id
                    except (MasterTable.DoesNotExist, MasterTable.MultipleObjectsReturned):
                        pass
                    
                    absent_users.append({
                        "id": user.id,
                        "temp_id": user.temp_id,
                        "emp_id": emp_id,
                        "full_name": f"{user.first_name} {user.last_name or ''}".strip(),
                        "attendance_date": att.attendance_date.isoformat() if att.attendance_date else None,
                    })

                result.append({
                    "day_number": item['day_number__days_id'],
                    "absent_count": item['count'],
                    "absent_users": absent_users
                })

            return Response({"absentees_by_day": result}, status=status.HTTP_200_OK)


class BatchAllAttendanceView(APIView):
    """
    GET /api/batch/{batch_id}/all-attendance/
    Returns all attendance records for a batch to get saved dates
    """
    def get(self, request, batch_id, *args, **kwargs):
        try:
            batch = TrainingBatch.objects.get(batch_id=batch_id)
            attendances = TrainingAttendance.objects.filter(batch=batch).values(
                'day_number', 'attendance_date'
            ).distinct()
            
            return Response(list(attendances), status=status.HTTP_200_OK)
        except TrainingBatch.DoesNotExist:
            return Response({"error": "Batch not found."}, status=status.HTTP_404_NOT_FOUND)



from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from django.db.models import Q
from django.utils import timezone

from .models import RescheduledSession, SubTopic, Days, TrainingBatch, UserRegistration
from .serializers import (
    RescheduledSessionSerializer,
    RescheduledSessionCreateSerializer,
    MarkRescheduledAttendanceSerializer,
    SubTopicSimpleSerializer
)


class SubTopicsByDayView(generics.ListAPIView):
    """
    GET /api/subtopics/by-day/{day_id}/
    Returns all subtopics for a specific training day.
    """
    serializer_class = SubTopicSimpleSerializer
    
    def get_queryset(self):
        day_id = self.kwargs.get('day_id')
        return SubTopic.objects.filter(days__days_id=day_id).select_related('days')


class RescheduledSessionListView(generics.ListAPIView):
    def get(self, request):
        sessions = RescheduledSession.objects.all().order_by('-rescheduled_date', '-rescheduled_time')
        serializer = RescheduledSessionSerializer(sessions, many=True, context={'request': request})
        return Response(serializer.data)
    """
    GET /api/rescheduled-sessions/
    Optional query params:
        ?batch_id=xxx     → Filter by batch
        ?status=scheduled → Filter by status (scheduled/completed/cancelled)
        ?employee_id=123  → Filter by employee
    """
    serializer_class = RescheduledSessionSerializer
    
    def get_queryset(self):
        queryset = RescheduledSession.objects.all().select_related(
            'employee',
            'batch',
            'original_day',
            'training_subtopic'
        )
    
        
        # Apply filters
        batch_id = self.request.query_params.get('batch_id')
        status_filter = self.request.query_params.get('status')
        employee_id = self.request.query_params.get('employee_id')
        
        if batch_id:
            queryset = queryset.filter(batch__batch_id=batch_id)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        
        return queryset


class RescheduledSessionCreateView(generics.CreateAPIView):
    """
    POST /api/rescheduled-sessions/create/
    Body: {
        "employee": 1,
        "batch": "BATCH-21-11-2025-1",
        "original_day": 3,
        "original_date": "2025-11-15",
        "rescheduled_date": "2025-11-25",
        "rescheduled_time": "10:30",
        "training_subtopic": 5,
        "training_name": "React Advanced Patterns",
        "notes": "Make-up session for Day 3"
    }
    """
    serializer_class = RescheduledSessionCreateSerializer
    
    def perform_create(self, serializer):
        serializer.save()
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        
        # Return full session data
        session = RescheduledSession.objects.get(id=serializer.instance.id)
        response_serializer = RescheduledSessionSerializer(session)
        
        return Response(
            response_serializer.data,
            status=status.HTTP_201_CREATED
        )


class RescheduledSessionDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    GET/PUT/PATCH/DELETE /api/rescheduled-sessions/{id}/
    Retrieve, update, or delete a specific rescheduled session.
    """
    queryset = RescheduledSession.objects.all()
    serializer_class = RescheduledSessionSerializer
    
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        
        # Prevent updating if attendance already marked
        if instance.attendance_marked and not partial:
            return Response(
                {"error": "Cannot fully update a session with marked attendance."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        
        return Response(serializer.data)


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.utils import timezone
from .models import RescheduledSession
from .serializers import MarkRescheduledAttendanceSerializer


class MarkRescheduledAttendanceView(APIView):
    def post(self, request):
        serializer = MarkRescheduledAttendanceSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        session_id = serializer.validated_data['session_id']
        attendance_status = serializer.validated_data['attendance_status']
        marked_by = serializer.validated_data.get('marked_by', 'Admin')
        photo = serializer.validated_data.get('photo')

        try:
            session = RescheduledSession.objects.get(id=session_id)
        except RescheduledSession.DoesNotExist:
            return Response({'error': 'Session not found'}, status=404)

        if photo:
            session.attendance_photo = photo

        session.attendance_marked = True
        session.attendance_status = attendance_status
        session.status = 'completed'
        session.attendance_marked_at = timezone.now()
        session.marked_by = marked_by
        session.save()

        # THE ONLY CHANGE YOU NEED
        photo_url = None
        if session.attendance_photo:
            photo_url = request.build_absolute_uri(session.attendance_photo.url)

        return Response({
            'message': 'Attendance marked successfully',
            'session': {
                'id': session.id,
                'attendance_photo_url': photo_url,
                'attendance_marked': True,
                'attendance_status': 'present',
                'status': 'completed'
            }
        })




class BatchRescheduledSessionsView(APIView):
    """
    GET /api/batch/{batch_id}/rescheduled-sessions/
    Returns all rescheduled sessions for a specific batch with statistics.
    """
    
    def get(self, request, batch_id, *args, **kwargs):
        batch = get_object_or_404(TrainingBatch, batch_id=batch_id)
        
        sessions = RescheduledSession.objects.filter(
            batch=batch
        ).select_related(
            'employee',
            'original_day',
            'training_subtopic'
        ).order_by('-rescheduled_date', '-rescheduled_time')
        
        serializer = RescheduledSessionSerializer(sessions, many=True)
        
        # Calculate statistics
        total_sessions = sessions.count()
        scheduled = sessions.filter(status='scheduled').count()
        completed = sessions.filter(status='completed').count()
        cancelled = sessions.filter(status='cancelled').count()
        
        return Response({
            "batch_id": batch_id,
            "statistics": {
                "total": total_sessions,
                "scheduled": scheduled,
                "completed": completed,
                "cancelled": cancelled
            },
            "sessions": serializer.data
        }, status=status.HTTP_200_OK)


class RescheduleFromAbsentView(APIView):
    """
    POST /api/reschedule-from-absent/
    Convenience endpoint to create a rescheduled session directly from absent record.
    
    Body: {
        "employee_id": 1,
        "batch_id": "BATCH-21-11-2025-1",
        "absent_day": 3,
        "rescheduled_date": "2025-11-25",
        "rescheduled_time": "10:30",
        "training_subtopic_id": 5,
        "training_name": "React Advanced Patterns",
        "notes": "Optional notes"
    }
    """
    
    def post(self, request, *args, **kwargs):
        data = request.data
        
        # Validate required fields
        required_fields = [
            'employee_id', 'batch_id', 'absent_day',
            'rescheduled_date', 'rescheduled_time', 'training_name'
        ]
        missing = [f for f in required_fields if f not in data]
        if missing:
            return Response(
                {"error": f"Missing required fields: {', '.join(missing)}"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            employee = UserRegistration.objects.get(id=data['employee_id'])
            batch = TrainingBatch.objects.get(batch_id=data['batch_id'])
            original_day = Days.objects.get(days_id=data['absent_day'])
            
            # Get training subtopic if provided
            training_subtopic = None
            if data.get('training_subtopic_id'):
                training_subtopic = SubTopic.objects.get(
                    subtopic_id=data['training_subtopic_id']
                )
            
            # Find the original absence date
            from .models import TrainingAttendance
            absent_record = TrainingAttendance.objects.filter(
                user=employee,
                batch=batch,
                day_number=original_day,
                status='absent'
            ).first()
            
            original_date = absent_record.attendance_date if absent_record else timezone.now().date()
            
            # Create rescheduled session
            session = RescheduledSession.objects.create(
                employee=employee,
                batch=batch,
                original_day=original_day,
                original_date=original_date,
                rescheduled_date=data['rescheduled_date'],
                rescheduled_time=data['rescheduled_time'],
                training_subtopic=training_subtopic,
                training_name=data['training_name'],
                notes=data.get('notes', ''),
                status='scheduled'
            )
            
            serializer = RescheduledSessionSerializer(session)
            
            return Response({
                "message": "Session rescheduled successfully",
                "session": serializer.data
            }, status=status.HTTP_201_CREATED)
            
        except UserRegistration.DoesNotExist:
            return Response(
                {"error": "Employee not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except TrainingBatch.DoesNotExist:
            return Response(
                {"error": "Batch not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Days.DoesNotExist:
            return Response(
                {"error": "Invalid day"},
                status=status.HTTP_404_NOT_FOUND
            )
        except SubTopic.DoesNotExist:
            return Response(
                {"error": "Training subtopic not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# # =================== TrainingAttendance ============================= #



  
# views.py (COMPLETE WITH FINANCIAL YEAR SUPPORT)

from django.utils import timezone
from django.db.models import Sum, Avg, Q
from rest_framework import viewsets, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from django.http import HttpResponse
from django.db import transaction
import pandas as pd
import io

from .models import ManagementReview, Hq, Factory, Department, Line, SubLine, Station
from .serializers import ManagementReviewSerializer

# =================== VIEWSET ===================
class ManagementReviewViewSet(viewsets.ModelViewSet):
    queryset = ManagementReview.objects.all().select_related('hq', 'factory', 'department','line','subline','station')
    serializer_class = ManagementReviewSerializer
    pagination_class = None

# =================== EXCEL DOWNLOAD TEMPLATE ===================
class ManagementDownloadTemplateView(APIView):
    def get(self, request):
        columns = [
            'HQ Name', 'Factory Name', 'Department Name', 'Line Name',
            'Subline Name', 'Station Name', 'Month (1-12)', 'Year (YYYY)',
            'New Operators Joined', 'New Operators Trained',
            'Total Training Plans', 'Total Trainings Actual',
            'Total Defects MSIL', 'CTQ Defects MSIL',
            'Total Defects Tier1', 'CTQ Defects Tier1',
            'Total Internal Rejection', 'CTQ Internal Rejection',
            'Manpower Available', 'Manpower Required', 'GCA Defects'
        ]
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_template = pd.DataFrame(columns=columns)
            df_template.to_excel(writer, sheet_name='Fill_Data_Here', index=False)
            
            instructions = [
                ["Column", "Instruction"],
                ["Hierarchy", "Copy names exactly from 'Reference_Data' sheet."],
                ["Month/Year", "Numeric only (e.g., Month: 10, Year: 2025)."],
                ["Metrics", "Numeric integers only. Use 0 for empty values."],
                ["GCA Defects", "Decimal numbers allowed (e.g., 12.5)."],
            ]
            pd.DataFrame(instructions[1:], columns=instructions[0]).to_excel(writer, sheet_name='Instructions', index=False)
            
            hqs = list(Hq.objects.values_list('hq_name', flat=True))
            factories = list(Factory.objects.values_list('factory_name', flat=True))
            departments = list(Department.objects.values_list('department_name', flat=True))
           
            max_len = max(len(hqs), len(factories), len(departments))
            ref_data = {
                'Valid HQs': hqs + [''] * (max_len - len(hqs)),
                'Valid Factories': factories + [''] * (max_len - len(factories)),
                'Valid Departments': departments + [''] * (max_len - len(departments)),
            }
            pd.DataFrame(ref_data).to_excel(writer, sheet_name='Reference_Data', index=False)
        
        output.seek(0)
        filename = "Management_Review_Template.xlsx"
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response

# =================== EXCEL UPLOAD ===================
class ManagementUploadExcelView(APIView):
    parser_classes = [MultiPartParser]
    
    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)
        
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response({"error": "File must be Excel (.xlsx, .xls)"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            try:
                df = pd.read_excel(file, sheet_name='Fill_Data_Here')
            except:
                df = pd.read_excel(file, sheet_name=0)
            
            df = df.where(pd.notnull(df), None)
            success_count = 0
            errors = []
            
            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        hq_obj = Hq.objects.filter(hq_name__iexact=row.get('HQ Name')).first()
                        factory_obj = Factory.objects.filter(factory_name__iexact=row.get('Factory Name')).first()
                       
                        if not factory_obj:
                            errors.append(f"Row {index+2}: Factory not found.")
                            continue
                        
                        dept_obj = Department.objects.filter(department_name__iexact=row.get('Department Name')).first()
                        line_obj = Line.objects.filter(line_name__iexact=row.get('Line Name')).first()
                        subline_obj = SubLine.objects.filter(subline_name__iexact=row.get('Subline Name')).first()
                        station_obj = Station.objects.filter(station_name__iexact=row.get('Station Name')).first()
                        
                        def get_val(val):
                            try:
                                return int(val) if val is not None else 0
                            except:
                                return 0
                        
                        def get_decimal(val):
                            try:
                                return float(val) if val is not None else None
                            except:
                                return None
                        
                        ManagementReview.objects.update_or_create(
                            hq=hq_obj,
                            factory=factory_obj,
                            department=dept_obj,
                            month=get_val(row.get('Month (1-12)')),
                            year=get_val(row.get('Year (YYYY)')),
                            defaults={
                                'line': line_obj,
                                'subline': subline_obj,
                                'station': station_obj,
                                'new_operators_joined': get_val(row.get('New Operators Joined')),
                                'new_operators_trained': get_val(row.get('New Operators Trained')),
                                'total_training_plans': get_val(row.get('Total Training Plans')),
                                'total_trainings_actual': get_val(row.get('Total Trainings Actual')),
                                'total_defects_msil': get_val(row.get('Total Defects MSIL')),
                                'ctq_defects_msil': get_val(row.get('CTQ Defects MSIL')),
                                'total_defects_tier1': get_val(row.get('Total Defects Tier1')),
                                'ctq_defects_tier1': get_val(row.get('CTQ Defects Tier1')),
                                'total_internal_rejection': get_val(row.get('Total Internal Rejection')),
                                'ctq_internal_rejection': get_val(row.get('CTQ Internal Rejection')),
                                'manpower_available': get_val(row.get('Manpower Available')),
                                'manpower_required': get_val(row.get('Manpower Required')),
                                'gca_defects': get_decimal(row.get('GCA Defects')),
                            }
                        )
                        success_count += 1
                    except Exception as e:
                        errors.append(f"Row {index+2}: Error - {str(e)}")
            
            return Response({
                "message": f"Processed {success_count} records successfully.",
                "errors": errors
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": f"Failed to parse: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# =================== HELPER FUNCTIONS ===================
def apply_hierarchy_filters(request, queryset):
    """Apply hierarchy filters to queryset"""
    hq_id = request.query_params.get('hq')
    factory_id = request.query_params.get('factory')
    department_id = request.query_params.get('department')
    line_id = request.query_params.get('line')
    subline_id = request.query_params.get('subline')
    station_id = request.query_params.get('station')
    
    if hq_id and hq_id != 'null':
        queryset = queryset.filter(hq_id=hq_id)
    if factory_id and factory_id != 'null':
        queryset = queryset.filter(factory_id=factory_id)
    if department_id and department_id != 'null':
        queryset = queryset.filter(department_id=department_id)
    if line_id and line_id != 'null':
        queryset = queryset.filter(line_id=line_id)
    if subline_id and subline_id != 'null':
        queryset = queryset.filter(subline_id=subline_id)
    if station_id and station_id != 'null':
        queryset = queryset.filter(station_id=station_id)
    
    return queryset

def get_financial_year_filter(request):
    """
    Returns Q object for filtering financial year (Apr-Mar)
    If year param = 2025, returns: Apr 2025 - Mar 2026
    """
    year_param = request.query_params.get('year')
    
    if year_param:
        try:
            fy_start = int(year_param)
        except ValueError:
            fy_start = timezone.now().year
    else:
        now = timezone.now()
        fy_start = now.year if now.month >= 4 else now.year - 1
    
    fy_end = fy_start + 1
    
    # Return Q filter: (year=2025 AND month>=4) OR (year=2026 AND month<=3)
    return Q(year=fy_start, month__gte=4) | Q(year=fy_end, month__lte=3)

# =================== CURRENT MONTH SUMMARY CARDS ===================
class CurrentMonthTrainingDataView(APIView):
    """Summary card for training data (current month only)"""
    def get(self, request):
        current_time = timezone.now()
        queryset = ManagementReview.objects.filter(year=current_time.year, month=current_time.month)
        queryset = apply_hierarchy_filters(request, queryset)
        
        data = queryset.aggregate(
            new_operators_joined=Sum('new_operators_joined'),
            new_operators_trained=Sum('new_operators_trained'),
            total_training_plans=Sum('total_training_plans'),
            total_trainings_actual=Sum('total_trainings_actual')
        )
        
        response_data = {key: value or 0 for key, value in data.items()}
        return Response(response_data, status=status.HTTP_200_OK)

class CurrentMonthDefectsDataView(APIView):
    """Summary card for defects data (current month only)"""
    def get(self, request):
        current_time = timezone.now()
        queryset = ManagementReview.objects.filter(year=current_time.year, month=current_time.month)
        queryset = apply_hierarchy_filters(request, queryset)
        
        data = queryset.aggregate(
            total_defects_msil=Sum('total_defects_msil'),
            ctq_defects_msil=Sum('ctq_defects_msil'),
            total_defects_tier1=Sum('total_defects_tier1'),
            ctq_defects_tier1=Sum('ctq_defects_tier1'),
            total_internal_rejection=Sum('total_internal_rejection'),
            ctq_internal_rejection=Sum('ctq_internal_rejection')
        )
        
        response_data = {key: value or 0 for key, value in data.items()}
        return Response(response_data, status=status.HTTP_200_OK)

# =================== CHART VIEWS (FINANCIAL YEAR) ===================
class OperatorsChartView(APIView):
    """Operators Joined vs Trained - Financial Year"""
    def get(self, request):
        queryset = ManagementReview.objects.filter(get_financial_year_filter(request))
        queryset = apply_hierarchy_filters(request, queryset)
        
        data = queryset.values('year', 'month').annotate(
            operators_joined=Sum('new_operators_joined'),
            operators_trained=Sum('new_operators_trained')
        ).order_by('year', 'month')
        
        formatted_data = []
        for item in data:
            formatted_data.append({
                'year': item['year'],
                'month': item['month'],
                'month_year': f"{item['year']}-{str(item['month']).zfill(2)}",
                'operators_joined': item['operators_joined'] or 0,
                'operators_trained': item['operators_trained'] or 0
            })
        
        return Response(formatted_data, status=status.HTTP_200_OK)

class TrainingPlansChartView(APIView):
    """Training Plans vs Actual - Financial Year"""
    def get(self, request):
        queryset = ManagementReview.objects.filter(get_financial_year_filter(request))
        queryset = apply_hierarchy_filters(request, queryset)
        
        data = queryset.values('year', 'month').annotate(
            training_plans=Sum('total_training_plans'),
            trainings_actual=Sum('total_trainings_actual')
        ).order_by('year', 'month')
        
        formatted_data = []
        for item in data:
            formatted_data.append({
                'year': item['year'],
                'month_year': f"{item['year']}-{str(item['month']).zfill(2)}",
                'training_plans': item['training_plans'] or 0,
                'trainings_actual': item['trainings_actual'] or 0
            })
        
        return Response(formatted_data, status=status.HTTP_200_OK)

class DefectsChartView(APIView):
    """Man Related Defects Trend - Financial Year"""
    def get(self, request):
        queryset = ManagementReview.objects.filter(get_financial_year_filter(request))
        queryset = apply_hierarchy_filters(request, queryset)
        
        data = queryset.values('year', 'month').annotate(
            defects_msil=Sum('total_defects_msil'),
            ctq_defects_msil=Sum('ctq_defects_msil')
        ).order_by('year', 'month')
        
        formatted_data = []
        for item in data:
            formatted_data.append({
                'year': item['year'],
                'month_year': f"{item['year']}-{str(item['month']).zfill(2)}",
                'defects_msil': item['defects_msil'] or 0,
                'ctq_defects_msil': item['ctq_defects_msil'] or 0
            })
        
        return Response(formatted_data, status=status.HTTP_200_OK)

class Tier1DefectsChartView(APIView):
    """Tier 1 Defects - Financial Year"""
    def get(self, request):
        queryset = ManagementReview.objects.filter(get_financial_year_filter(request))
        queryset = apply_hierarchy_filters(request, queryset)
        
        data = queryset.values('year', 'month').annotate(
            total_tier1=Sum('total_defects_tier1'),
            ctq_tier1=Sum('ctq_defects_tier1')
        ).order_by('year', 'month')
        
        formatted_data = []
        for item in data:
            formatted_data.append({
                'year': item['year'],
                'month_year': f"{item['year']}-{str(item['month']).zfill(2)}",
                'total_defects_tier1': item['total_tier1'] or 0,
                'ctq_defects_tier1': item['ctq_tier1'] or 0
            })
        
        return Response(formatted_data, status=status.HTTP_200_OK)

class InternalRejectionChartView(APIView):
    """Internal Rejection Chart - Financial Year"""
    def get(self, request):
        queryset = ManagementReview.objects.filter(get_financial_year_filter(request))
        queryset = apply_hierarchy_filters(request, queryset)
        
        data = queryset.values('year', 'month').annotate(
            internal_rej=Sum('total_internal_rejection'),
            ctq_rej=Sum('ctq_internal_rejection')
        ).order_by('year', 'month')
        
        formatted_data = []
        for item in data:
            formatted_data.append({
                'year': item['year'],
                'month_year': f"{item['year']}-{str(item['month']).zfill(2)}",
                'total_internal_rejection': item['internal_rej'] or 0,
                'ctq_internal_rejection': item['ctq_rej'] or 0
            })
        
        return Response(formatted_data, status=status.HTTP_200_OK)

class MonthPlanningChartView(APIView):
    """Manpower Planning - Financial Year"""
    def get(self, request):
        queryset = ManagementReview.objects.filter(get_financial_year_filter(request))
        queryset = apply_hierarchy_filters(request, queryset)
        
        data = queryset.values('year', 'month').annotate(
            total_req=Sum('manpower_required'),
            total_avail=Sum('manpower_available')
        ).order_by('year', 'month')
        
        formatted_data = []
        for item in data:
            formatted_data.append({
                'year': item['year'],
                'month_year': f"{item['year']}-{str(item['month']).zfill(2)}",
                'manpower_required': item['total_req'] or 0,
                'manpower_available': item['total_avail'] or 0
            })
        
        return Response(formatted_data, status=status.HTTP_200_OK)

class GcaDefectsChartView(APIView):
    """GCA Defects Chart - Financial Year (Average)"""
    def get(self, request):
        queryset = ManagementReview.objects.filter(get_financial_year_filter(request))
        queryset = apply_hierarchy_filters(request, queryset)
        
        data = queryset.values('year', 'month').annotate(
            avg_gca=Avg('gca_defects')
        ).order_by('year', 'month')
        
        formatted_data = []
        for item in data:
            formatted_data.append({
                'year': item['year'],
                'month_year': f"{item['year']}-{str(item['month']).zfill(2)}",
                'gca_defects': round(float(item['avg_gca'] or 0), 2)
            })
        
        return Response(formatted_data, status=status.HTTP_200_OK)












# # views.py (CORRECTED) ====== management review ==========================================




# # ✅ CORRECTED: Attrition Chart View
# class AttritionChartView(APIView):
#     def get(self, request):
#         req_year = request.query_params.get('year')
        
#         # Determine FY Start Year
#         if req_year:
#             fy_start_year = int(req_year)
#         else:
#             now = timezone.now()
#             fy_start_year = now.year - 1 if now.month <= 3 else now.year
        
#         fy_end_year = fy_start_year + 1

#         # ✅ Query for Financial Year: Apr-Dec of start year + Jan-Mar of end year
#         queryset = AdvanceManpowerDashboard.objects.filter(
#             Q(year=fy_start_year, month__gte=4) |  # Apr to Dec of start year
#             Q(year=fy_end_year, month__lte=3)      # Jan to Mar of end year
#         )

#         # Apply hierarchy filters
#         queryset = apply_hierarchy_filters(request, queryset)

#         # Aggregate Data (AVERAGE for Rates)
#         data = queryset.values('year', 'month').annotate(
#             avg_attrition=Avg('attrition_rate')
#         ).order_by('year', 'month')

#         formatted_data = []
#         for item in data:
#             formatted_data.append({
#                 'month': item['month'],
#                 'year': item['year'],
#                 'attrition_rate': round(float(item['avg_attrition'] or 0), 2)
#             })

#         return Response(formatted_data, status=status.HTTP_200_OK)


#         return Response(formatted_data, status=status.HTTP_200_OK)


# # ✅ CORRECTED: Absenteeism Chart View
# class AbsenteeismChartView(APIView):
#     def get(self, request):
#         req_year = request.query_params.get('year')
        
#         # Determine FY Start Year
#         if req_year:
#             fy_start_year = int(req_year)
#         else:
#             now = timezone.now()
#             fy_start_year = now.year - 1 if now.month <= 3 else now.year
        
#         fy_end_year = fy_start_year + 1

#         # ✅ Query for Financial Year: Apr-Dec of start year + Jan-Mar of end year
#         queryset = AdvanceManpowerDashboard.objects.filter(
#             Q(year=fy_start_year, month__gte=4) |  # Apr to Dec of start year
#             Q(year=fy_end_year, month__lte=3)      # Jan to Mar of end year
#         )

#         # Apply hierarchy filters
#         queryset = apply_hierarchy_filters(request, queryset)

#         # Aggregate (Avg)
#         data = queryset.values('year', 'month').annotate(
#             avg_absenteeism=Avg('absenteeism_rate')
#         ).order_by('year', 'month')

#         formatted_data = []
#         for item in data:
#             formatted_data.append({
#                 'month': item['month'],
#                 'year': item['year'],
#                 'absenteeism_rate': round(float(item['avg_absenteeism'] or 0), 2)
#             })

#         return Response(formatted_data, status=status.HTTP_200_OK)




# from django.utils import timezone
# from rest_framework import viewsets, status
# from rest_framework.views import APIView
# from rest_framework.response import Response

# from .models import ManagementReview
# from .serializers import (
#     ManagementReviewSerializer, TrainingDataSerializer, DefectsDataSerializer,
#     OperatorsChartSerializer, TrainingPlansChartSerializer, DefectsChartSerializer
# )

# class ManagementReviewViewSet(viewsets.ModelViewSet):
#     # This viewset is fine, no changes needed here.
#     queryset = ManagementReview.objects.all().select_related('hq', 'factory', 'department','line','subline','station')
#     serializer_class = ManagementReviewSerializer
#     pagination_class = None




# import pandas as pd
# import io
# from django.http import HttpResponse
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
# from rest_framework.parsers import MultiPartParser
# from django.db import transaction
# from .models import ManagementReview, Hq, Factory, Department, Line, SubLine, Station


# class ManagementDownloadTemplateView(APIView):
#     """
#     Generates an Excel template specifically for Management Review Data.
#     """
#     def get(self, request):
#         columns = [
#             'HQ Name', 'Factory Name', 'Department Name', 'Line Name', 
#             'Subline Name', 'Station Name', 'Month (1-12)', 'Year (YYYY)',
#             'New Operators Joined', 'New Operators Trained', 
#             'Total Training Plans', 'Total Trainings Actual',
#             'Total Defects MSIL', 'CTQ Defects MSIL',
#             'Total Defects Tier1', 'CTQ Defects Tier1',
#             'Total Internal Rejection', 'CTQ Internal Rejection'
#         ]

#         output = io.BytesIO()
#         with pd.ExcelWriter(output, engine='openpyxl') as writer:
#             # Sheet 1: Data Entry
#             df_template = pd.DataFrame(columns=columns)
#             df_template.to_excel(writer, sheet_name='Fill_Data_Here', index=False)

#             # Sheet 2: Instructions
#             instructions = [
#                 ["Column", "Instruction"],
#                 ["Hierarchy", "Copy names exactly from 'Reference_Data' sheet."],
#                 ["Month/Year", "Numeric only (e.g., Month: 10, Year: 2025)."],
#                 ["Metrics", "Numeric integers only. Use 0 for empty values."]
#             ]
#             pd.DataFrame(instructions[1:], columns=instructions[0]).to_excel(writer, sheet_name='Instructions', index=False)

#             # Sheet 3: Reference Data
#             # --- FIX IS HERE: Changed 'name' to 'hq_name' ---
#             hqs = list(Hq.objects.values_list('hq_name', flat=True))
#             factories = list(Factory.objects.values_list('factory_name', flat=True))
#             departments = list(Department.objects.values_list('department_name', flat=True))
            
#             max_len = max(len(hqs), len(factories), len(departments))
#             ref_data = {
#                 'Valid HQs': hqs + [''] * (max_len - len(hqs)),
#                 'Valid Factories': factories + [''] * (max_len - len(factories)),
#                 'Valid Departments': departments + [''] * (max_len - len(departments)),
#             }
#             pd.DataFrame(ref_data).to_excel(writer, sheet_name='Reference_Data', index=False)

#         output.seek(0)
#         filename = "Management_Review_Template.xlsx"
#         response = HttpResponse(
#             output,
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = f'attachment; filename={filename}'
#         return response

# class ManagementUploadExcelView(APIView):
#     """
#     Handles bulk upload for Management Review.
#     """
#     parser_classes = [MultiPartParser]

#     def post(self, request):
#         file = request.FILES.get('file')
#         if not file:
#             return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

#         if not file.name.endswith(('.xlsx', '.xls')):
#             return Response({"error": "File must be Excel (.xlsx, .xls)"}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             try:
#                 df = pd.read_excel(file, sheet_name='Fill_Data_Here')
#             except:
#                 df = pd.read_excel(file, sheet_name=0)

#             # Convert NaN to None (standard Python null)
#             df = df.where(pd.notnull(df), None)
            
#             success_count = 0
#             errors = []

#             with transaction.atomic():
#                 for index, row in df.iterrows():
#                     try:
#                         # 1. Fetch Foreign Keys
#                         hq_name = row.get('HQ Name')
#                         factory_name = row.get('Factory Name')
#                         dept_name = row.get('Department Name')
#                         line_name = row.get('Line Name')
#                         subline_name = row.get('Subline Name')
#                         station_name = row.get('Station Name')

#                         hq_obj = Hq.objects.filter(hq_name__iexact=hq_name).first() if hq_name else None
                        
#                         factory_obj = Factory.objects.filter(factory_name__iexact=factory_name).first() if factory_name else None
#                         if not factory_obj:
#                             errors.append(f"Row {index+2}: Factory '{factory_name}' not found.")
#                             continue

#                         # Note: If dept_name exists in Excel but not in DB, dept_obj becomes None.
#                         # This creates "Unknown" entries. Ensure DB has these departments.
#                         dept_obj = Department.objects.filter(department_name__iexact=dept_name).first() if dept_name else None
                        
#                         line_obj = Line.objects.filter(line_name__iexact=line_name).first() if line_name else None
#                         subline_obj = SubLine.objects.filter(subline_name__iexact=subline_name).first() if subline_name else None
#                         station_obj = Station.objects.filter(station_name__iexact=station_name).first() if station_name else None

#                         def get_val(val): 
#                             return int(val) if val is not None else 0

#                         # 2. UPDATE OR CREATE
#                         # CRITICAL FIX: 'line', 'subline', 'station' moved OUT of defaults
#                         # so they act as filters (Unique identifiers)
#                         ManagementReview.objects.update_or_create(
#                             hq=hq_obj,
#                             factory=factory_obj,
#                             department=dept_obj,
#                             line=line_obj,          # Moved here
#                             subline=subline_obj,    # Moved here
#                             station=station_obj,    # Moved here
#                             month=get_val(row.get('Month (1-12)')),
#                             year=get_val(row.get('Year (YYYY)')),
#                             defaults={
#                                 # Only the numerical data acts as the "Update" payload
#                                 'new_operators_joined': get_val(row.get('New Operators Joined')),
#                                 'new_operators_trained': get_val(row.get('New Operators Trained')),
#                                 'total_training_plans': get_val(row.get('Total Training Plans')),
#                                 'total_trainings_actual': get_val(row.get('Total Trainings Actual')),
#                                 'total_defects_msil': get_val(row.get('Total Defects MSIL')),
#                                 'ctq_defects_msil': get_val(row.get('CTQ Defects MSIL')),
#                                 'total_defects_tier1': get_val(row.get('Total Defects Tier1')),
#                                 'ctq_defects_tier1': get_val(row.get('CTQ Defects Tier1')),
#                                 'total_internal_rejection': get_val(row.get('Total Internal Rejection')),
#                                 'ctq_internal_rejection': get_val(row.get('CTQ Internal Rejection')),
#                             }
#                         )
#                         success_count += 1
#                     except Exception as e:
#                         errors.append(f"Row {index+2}: Error - {str(e)}")

#             return Response({
#                 "message": f"Processed {success_count} records successfully.",
#                 "errors": errors
#             }, status=status.HTTP_200_OK)

#         except Exception as e:
#             return Response({"error": f"Failed to parse: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

# # class ManagementUploadExcelView(APIView):
# #     """
# #     Handles bulk upload for Management Review.
# #     """
# #     parser_classes = [MultiPartParser]

# #     def post(self, request):
# #         file = request.FILES.get('file')
# #         if not file:
# #             return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

# #         if not file.name.endswith(('.xlsx', '.xls')):
# #             return Response({"error": "File must be Excel (.xlsx, .xls)"}, status=status.HTTP_400_BAD_REQUEST)

# #         try:
# #             try:
# #                 df = pd.read_excel(file, sheet_name='Fill_Data_Here')
# #             except:
# #                 df = pd.read_excel(file, sheet_name=0)

# #             df = df.where(pd.notnull(df), None)
# #             success_count = 0
# #             errors = []

# #             with transaction.atomic():
# #                 for index, row in df.iterrows():
# #                     try:
# #                         # Hierarchy Lookup
# #                         # --- FIX IS HERE: Changed name__iexact to hq_name__iexact ---
# #                         hq_obj = Hq.objects.filter(hq_name__iexact=row.get('HQ Name')).first()
                        
# #                         factory_obj = Factory.objects.filter(factory_name__iexact=row.get('Factory Name')).first()
                        
# #                         if not factory_obj:
# #                             errors.append(f"Row {index+2}: Factory not found.")
# #                             continue

# #                         dept_obj = Department.objects.filter(department_name__iexact=row.get('Department Name')).first()
# #                         line_obj = Line.objects.filter(line_name__iexact=row.get('Line Name')).first()
# #                         subline_obj = SubLine.objects.filter(subline_name__iexact=row.get('Subline Name')).first()
# #                         station_obj = Station.objects.filter(station_name__iexact=row.get('Station Name')).first()

# #                         def get_val(val): return int(val) if val is not None else 0

# #                         ManagementReview.objects.update_or_create(
# #                             hq=hq_obj,
# #                             factory=factory_obj,
# #                             department=dept_obj,
# #                             month=get_val(row.get('Month (1-12)')),
# #                             year=get_val(row.get('Year (YYYY)')),
# #                             defaults={
# #                                 'line': line_obj,
# #                                 'subline': subline_obj,
# #                                 'station': station_obj,
# #                                 'new_operators_joined': get_val(row.get('New Operators Joined')),
# #                                 'new_operators_trained': get_val(row.get('New Operators Trained')),
# #                                 'total_training_plans': get_val(row.get('Total Training Plans')),
# #                                 'total_trainings_actual': get_val(row.get('Total Trainings Actual')),
# #                                 'total_defects_msil': get_val(row.get('Total Defects MSIL')),
# #                                 'ctq_defects_msil': get_val(row.get('CTQ Defects MSIL')),
# #                                 'total_defects_tier1': get_val(row.get('Total Defects Tier1')),
# #                                 'ctq_defects_tier1': get_val(row.get('CTQ Defects Tier1')),
# #                                 'total_internal_rejection': get_val(row.get('Total Internal Rejection')),
# #                                 'ctq_internal_rejection': get_val(row.get('CTQ Internal Rejection')),
# #                             }
# #                         )
# #                         success_count += 1
# #                     except Exception as e:
# #                         errors.append(f"Row {index+2}: Error - {str(e)}")

# #             return Response({
# #                 "message": f"Processed {success_count} records successfully.",
# #                 "errors": errors
# #             }, status=status.HTTP_200_OK)

# #         except Exception as e:
# #             return Response({"error": f"Failed to parse: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# from django.db.models import Sum
# from django.utils import timezone
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
# from .models import ManagementReview

# # --- Helper function to apply filters ---
# def apply_hierarchy_filters(request, queryset):
#     # Get parameters
#     hq_id = request.query_params.get('hq')
#     factory_id = request.query_params.get('factory')
#     department_id = request.query_params.get('department')
#     line_id = request.query_params.get('line')
#     subline_id = request.query_params.get('subline')
#     station_id = request.query_params.get('station')

#     # Apply filters ONLY if value exists and is not 'null'
#     if hq_id and hq_id != 'null':
#         queryset = queryset.filter(hq_id=hq_id)
        
#     if factory_id and factory_id != 'null':
#         queryset = queryset.filter(factory_id=factory_id)
        
#     if department_id and department_id != 'null':
#         queryset = queryset.filter(department_id=department_id)
        
#     if line_id and line_id != 'null':
#         queryset = queryset.filter(line_id=line_id)
        
#     if subline_id and subline_id != 'null':
#         queryset = queryset.filter(subline_id=subline_id)
        
#     if station_id and station_id != 'null':
#         queryset = queryset.filter(station_id=station_id)
    
#     return queryset

# # 1. View for the Training Summary Card
# class CurrentMonthTrainingDataView(APIView):
#     def get(self, request):
#         current_time = timezone.now()
        
#         # Start with current month filter
#         queryset = ManagementReview.objects.filter(
#             year=current_time.year,
#             month=current_time.month
#         )

#         # Apply hierarchy
#         queryset = apply_hierarchy_filters(request, queryset)

#         # Aggregate (Sum)
#         data = queryset.aggregate(
#             new_operators_joined=Sum('new_operators_joined'),
#             new_operators_trained=Sum('new_operators_trained'),
#             total_training_plans=Sum('total_training_plans'),
#             total_trainings_actual=Sum('total_trainings_actual')
#         )

#         # Clean up None values (convert to 0)
#         response_data = {
#             key: value or 0 for key, value in data.items()
#         }

#         return Response(response_data, status=status.HTTP_200_OK)


# # 2. View for the Defects Summary Card
# class CurrentMonthDefectsDataView(APIView):
#     def get(self, request):
#         current_time = timezone.now()
        
#         queryset = ManagementReview.objects.filter(
#             year=current_time.year,
#             month=current_time.month
#         )

#         queryset = apply_hierarchy_filters(request, queryset)

#         data = queryset.aggregate(
#             total_defects_msil=Sum('total_defects_msil'),
#             ctq_defects_msil=Sum('ctq_defects_msil'),
#             total_defects_tier1=Sum('total_defects_tier1'),
#             ctq_defects_tier1=Sum('ctq_defects_tier1'),
#             total_internal_rejection=Sum('total_internal_rejection'),
#             ctq_internal_rejection=Sum('ctq_internal_rejection')
#         )

#         response_data = {
#             key: value or 0 for key, value in data.items()
#         }

#         return Response(response_data, status=status.HTTP_200_OK)




# from django.db.models import Sum
# from django.utils import timezone
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
# from .models import ManagementReview

# class InternalRejectionChartView(APIView):
#     def get(self, request):
#         # 1. Determine Year
#         req_year = request.query_params.get('year')
#         if req_year:
#             target_year = int(req_year)
#         else:
#             target_year = timezone.now().year
            
#         queryset = ManagementReview.objects.filter(year=target_year)

#         # 2. Apply Hierarchy Filters
#         queryset = apply_hierarchy_filters(request, queryset)

#         # 3. Aggregate Data (Sum)
#         data = queryset.values('year', 'month').annotate(
#             internal_rej=Sum('total_internal_rejection'),
#             ctq_rej=Sum('ctq_internal_rejection')
#         ).order_by('year', 'month')

#         # 4. Format Response
#         formatted_data = []
#         for item in data:
#             formatted_data.append({
#                 'year': item['year'],
#                 'month_year': f"{item['year']}-{str(item['month']).zfill(2)}",
#                 'total_internal_rejection': item['internal_rej'] or 0,
#                 'ctq_internal_rejection': item['ctq_rej'] or 0
#             })

#         return Response(formatted_data, status=status.HTTP_200_OK)





# from django.db.models import Sum
# from django.utils import timezone
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
# from .models import ManagementReview

# # --- 2. UPDATED VIEW ---
# class OperatorsChartView(APIView):
#     def get(self, request):
#         # 1. Year Logic (Default to current year if not provided)
#         req_year = request.query_params.get('year')
#         if req_year:
#             target_year = int(req_year)
#         else:
#             target_year = timezone.now().year
        
#         queryset = ManagementReview.objects.filter(year=target_year)

#         # 2. Apply Dynamic Filters using the Helper
#         # This handles the "All" logic automatically
#         queryset = apply_hierarchy_filters(request, queryset)

#         # 3. Aggregation
#         # Because we filtered the queryset above, this SUM will calculate:
#         # - Specific Station data (if Station ID was provided)
#         # - OR Total Line data (if Line ID provided, but Station was null)
#         # - OR Total Dept data (if Dept ID provided, but Line was null)
#         data = queryset.values('year', 'month').annotate(
#             operators_joined=Sum('new_operators_joined'),
#             operators_trained=Sum('new_operators_trained')
#         ).order_by('year', 'month')

#         # 4. Format Response
#         formatted_data = []
#         for item in data:
#             formatted_data.append({
#                 'year': item['year'],
#                 'month_year': f"{item['year']}-{str(item['month']).zfill(2)}",
#                 'operators_joined': item['operators_joined'] or 0,
#                 'operators_trained': item['operators_trained'] or 0
#             })

#         return Response(formatted_data, status=status.HTTP_200_OK)
    


# # class TrainingPlansChartView(APIView):
# #     def get(self, request):
# #         current_year = timezone.now().year
# #         data = ManagementReview.objects.filter(
# #             year=current_year  # FIX: Changed from month_year__year
# #         ).order_by('year', 'month')  # FIX: Changed from 'month_year'
# #         serializer = TrainingPlansChartSerializer(data, many=True)
# #         return Response(serializer.data, status=status.HTTP_200_OK)


# from django.db.models import Sum
# from django.utils import timezone
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
# from .models import ManagementReview
# # --- UPDATED VIEW FOR PLAN VS ACTUAL ---
# class TrainingPlansChartView(APIView):
#     def get(self, request):
#         # 1. Year Logic
#         req_year = request.query_params.get('year')
#         if req_year:
#             target_year = int(req_year)
#         else:
#             target_year = timezone.now().year
            
#         queryset = ManagementReview.objects.filter(year=target_year)

#         # 2. Apply Hierarchy Filters (Uses the "All" logic)
#         # If frontend sends empty string or no param for 'line', this SKIPS the filter.
#         # This means it includes records where line is NULL + records where line is Specific.
#         queryset = apply_hierarchy_filters(request, queryset)

#         # 3. Aggregate Data (Sum)
#         data = queryset.values('year', 'month').annotate(
#             training_plans=Sum('total_training_plans'),
#             trainings_actual=Sum('total_trainings_actual')
#         ).order_by('year', 'month')

#         # 4. Format Response
#         formatted_data = []
#         for item in data:
#             formatted_data.append({
#                 'year': item['year'],
#                 'month_year': f"{item['year']}-{str(item['month']).zfill(2)}",
#                 'training_plans': item['training_plans'] or 0,
#                 'trainings_actual': item['trainings_actual'] or 0
#             })

#         return Response(formatted_data, status=status.HTTP_200_OK)


# # class DefectsChartView(APIView):
# #     def get(self, request):
# #         current_year = timezone.now().year
# #         data = ManagementReview.objects.filter(
# #             year=current_year  # FIX: Changed from month_year__year
# #         ).order_by('year', 'month')  # FIX: Changed from 'month_year'
# #         serializer = DefectsChartSerializer(data, many=True)
# #         return Response(serializer.data, status=status.HTTP_200_OK)
    
# from django.db.models import Sum
# from django.utils import timezone
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
# from .models import ManagementReview

# class DefectsChartView(APIView):
#     def get(self, request):
#         # 1. Determine Year
#         req_year = request.query_params.get('year')
#         if req_year:
#             target_year = int(req_year)
#         else:
#             target_year = timezone.now().year
            
#         queryset = ManagementReview.objects.filter(year=target_year)

#         queryset = apply_hierarchy_filters(request, queryset)

#         # 3. Aggregate Data (Sum)
#         data = queryset.values('year', 'month').annotate(
#             defects_msil=Sum('total_defects_msil'),
#             ctq_defects_msil=Sum('ctq_defects_msil')
#         ).order_by('year', 'month')

#         # 4. Format Response
#         formatted_data = []
#         for item in data:
#             formatted_data.append({
#                 'year': item['year'],
#                 'month_year': f"{item['year']}-{str(item['month']).zfill(2)}",
#                 'defects_msil': item['defects_msil'] or 0,
#                 'ctq_defects_msil': item['ctq_defects_msil'] or 0
#             })

#         return Response(formatted_data, status=status.HTTP_200_OK)    





# from django.db.models import Sum
# from django.utils import timezone
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
# from .models import ManagementReview

# class Tier1DefectsChartView(APIView):
#     def get(self, request):
#         # 1. Determine Year (Default to System Year)
#         req_year = request.query_params.get('year')
#         if req_year:
#             target_year = int(req_year)
#         else:
#             target_year = timezone.now().year
            
#         queryset = ManagementReview.objects.filter(year=target_year)

#         queryset = apply_hierarchy_filters(request, queryset)

#         # 3. Aggregate Data (Sum Tier 1 fields)
#         data = queryset.values('year', 'month').annotate(
#             total_tier1=Sum('total_defects_tier1'),
#             ctq_tier1=Sum('ctq_defects_tier1')
#         ).order_by('year', 'month')

#         # 4. Format Response
#         formatted_data = []
#         for item in data:
#             formatted_data.append({
#                 'year': item['year'],
#                 'month_year': f"{item['year']}-{str(item['month']).zfill(2)}",
#                 'total_defects_tier1': item['total_tier1'] or 0,
#                 'ctq_defects_tier1': item['ctq_tier1'] or 0
#             })

#         return Response(formatted_data, status=status.HTTP_200_OK)


# =============================end of management reviw graph ===============================










# =================================== Advance manpower =======================================


import pandas as pd
import io
from django.http import HttpResponse
from django.db import transaction
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Sum
from rest_framework import viewsets, status
from rest_framework.response import Response
from .models import AdvanceManpowerDashboard, HierarchyStructure, Station
from .serializers import AdvanceManpowerDashboardSerializer

class AdvanceManpowerDashboardViewSet(viewsets.ModelViewSet):
    queryset = AdvanceManpowerDashboard.objects.all()
    serializer_class = AdvanceManpowerDashboardSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        
        hq_id = self.request.query_params.get('hq')
        factory_id = self.request.query_params.get('factory')
        department_id = self.request.query_params.get('department')
        line_id = self.request.query_params.get('line')
        subline_id = self.request.query_params.get('subline')
        station_id = self.request.query_params.get('station')
        month = self.request.query_params.get('month')
        year = self.request.query_params.get('year')

        if hq_id and hq_id != 'null': queryset = queryset.filter(hq_id=hq_id)
        if factory_id and factory_id != 'null': queryset = queryset.filter(factory_id=factory_id)
        if department_id and department_id != 'null': queryset = queryset.filter(department_id=department_id)
        if line_id and line_id != 'null': queryset = queryset.filter(line_id=line_id)
        if subline_id and subline_id != 'null': queryset = queryset.filter(subline_id=subline_id)
        if station_id and station_id != 'null': queryset = queryset.filter(station_id=station_id)
        
        if month and month != 'null': queryset = queryset.filter(month=month)
        if year and year != 'null': queryset = queryset.filter(year=year)
        
        return queryset
    def create(self, request, *args, **kwargs):
        data = request.data.copy()

        # Extract key fields
        hq_id = data.get('hq')
        factory_id = data.get('factory')
        department_id = data.get('department')
        line_id = data.get('line')
        month = data.get('month')
        year = data.get('year')

        # --- NEW REQUIREMENT: Line is REQUIRED for saving ---
        if not line_id or line_id in ('null', '', None):
            return Response(
                {"error": "Line selection is required. Data is saved at Line level."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Remove station from payload (we don't use it anymore)
        data.pop('station', None)
        data.pop('subline', None)  # Optional: also ignore subline if you want pure Line-level

        # Use update_or_create at Line level (ignore station/subline in unique constraint if needed)
        # But your model has unique_together including station → we need to allow null station

        # Option: Set station = None explicitly
        data['station'] = None
        data['subline'] = None  # or keep if you want to support subline grouping later

        # Ensure month/year are integers
        try:
            data['month'] = int(month)
            data['year'] = int(year)
        except:
            return Response({"error": "Invalid month or year"}, status=status.HTTP_400_BAD_REQUEST)

        # Update or create single record at Line level
        obj, created = AdvanceManpowerDashboard.objects.update_or_create(
            hq_id=hq_id if hq_id and hq_id != 'null' else None,
            factory_id=factory_id,
            department_id=department_id if department_id and department_id != 'null' else None,
            line_id=line_id,
            subline_id=None,        # Force null
            station_id=None,        # Force null → Line-level record
            month=data['month'],
            year=data['year'],
            defaults={
                'total_stations': data.get('total_stations', 0),
                'operators_required': data.get('operators_required', 0),
                'operators_available': data.get('operators_available', 0),
                'buffer_manpower_required': data.get('buffer_manpower_required', 0),
                'buffer_manpower_available': data.get('buffer_manpower_available', 0),
                'l1_required': data.get('l1_required', 0),
                'l1_available': data.get('l1_available', 0),
                'l2_required': data.get('l2_required', 0),
                'l2_available': data.get('l2_available', 0),
                'l3_required': data.get('l3_required', 0),
                'l3_available': data.get('l3_available', 0),
                'l4_required': data.get('l4_required', 0),
                'l4_available': data.get('l4_available', 0),
                'attrition_rate': data.get('attrition_rate', '0.00'),
                'absenteeism_rate': data.get('absenteeism_rate', '0.00'),
                'oet_attrition':data.get('oet_attrition', '0.00'),
                'associate_attrition':data.get('associate_attrition', '0.00'),
            }
        )

        serializer = self.get_serializer(obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)

    # def create(self, request, *args, **kwargs):
    #     data = request.data.copy()
        
    #     # Extract IDs
    #     hq_id = data.get('hq')
    #     factory_id = data.get('factory')
    #     department_id = data.get('department')
    #     line_id = data.get('line')
    #     subline_id_input = data.get('subline')
    #     station_id = data.get('station')
        
    #     month = data.get('month')
    #     year = data.get('year')

    #     # LOGIC: If Line is selected, but Station is "All" (null)
    #     if line_id and (not station_id or station_id == 'null'):
            
    #         # 1. Query the HIERARCHY table, not the Station table.
    #         # The Hierarchy table contains the mapping: Line -> Stations
    #         hierarchy_qs = HierarchyStructure.objects.filter(line_id=line_id)
            
    #         # Filter by subline only if the user specifically selected one
    #         if subline_id_input and subline_id_input != 'null':
    #             hierarchy_qs = hierarchy_qs.filter(subline_id=subline_id_input)

    #         # Ensure we actually found stations
    #         if not hierarchy_qs.exists():
    #             return Response(
    #                 {"error": "No stations found in Hierarchy for this Line."}, 
    #                 status=status.HTTP_400_BAD_REQUEST
    #             )

    #         created_instances = []

    #         # 2. Loop through the Hierarchy records found
    #         for h_entry in hierarchy_qs:
    #             # h_entry is a Hierarchy object containing .station, .subline, etc.
                
    #             target_station = h_entry.station
    #             if not target_station:
    #                 continue # Skip if hierarchy entry has no station

    #             # Prepare data for this specific station
    #             station_data = data.copy()
    #             station_data['station'] = target_station.station_id
                
    #             # CRITICAL: Use the Subline from the Hierarchy entry.
    #             # This ensures that if the station belongs to a subline, it gets saved correctly
    #             # even if the user selected "All" sublines.
    #             station_data['subline'] = h_entry.subline_id if h_entry.subline else None

    #             # 3. Update or Create Logic
    #             existing_record = AdvanceManpowerDashboard.objects.filter(
    #                 hq_id=hq_id,
    #                 factory_id=factory_id,
    #                 department_id=department_id,
    #                 line_id=line_id,
    #                 # Match the subline strictly to the hierarchy's definition
    #                 subline_id=station_data['subline'], 
    #                 station_id=target_station.station_id,
    #                 month=month,
    #                 year=year
    #             ).first()

    #             if existing_record:
    #                 # Update existing record
    #                 serializer = self.get_serializer(existing_record, data=station_data, partial=True)
    #             else:
    #                 # Create new record
    #                 serializer = self.get_serializer(data=station_data)

    #             if serializer.is_valid():
    #                 serializer.save()
    #                 created_instances.append(serializer.data)
    #             else:
    #                 print(f"Error saving station {target_station.id}: {serializer.errors}")

    #         return Response(created_instances, status=status.HTTP_201_CREATED)

    #     # Normal Create (Single specific station selected)
    #     return super().create(request, *args, **kwargs)

    # --- 1. EXCEL DOWNLOAD TEMPLATE (MULTI-SHEET) ---
    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        """
        Generates an Excel template with 3 sheets: Data, Instructions, Reference Data.
        """
        # --- SHEET 1: DATA HEADERS ---
        headers = [
            'hq_name', 'factory_name', 'department_name', 
            'line_name', 'subline_name', 'station_name',
            'month', 'year',
            'total_stations', 'operators_required', 'operators_available',
            'buffer_manpower_required', 'buffer_manpower_available',
            'l1_required', 'l1_available',
            'l2_required', 'l2_available',
            'l3_required', 'l3_available',
            'l4_required', 'l4_available',
            'attrition_rate', 'absenteeism_rate'
        ]
        
        # Example Data
        dummy_data = [
            {
                'hq_name': 'Corporate HQ', 'factory_name': 'Main Plant', 'department_name': 'Assembly',
                'line_name': 'Line A', 'subline_name': 'Sub A1', 'station_name': 'Station 101',
                'month': 1, 'year': 2025, 
                'total_stations': 1, 'operators_required': 10, 'operators_available': 8,
                'buffer_manpower_required': 2, 'buffer_manpower_available': 1,
                'l1_required': 5, 'l1_available': 4, 'l2_required': 3, 'l2_available': 3,
                'l3_required': 1, 'l3_available': 1, 'l4_required': 1, 'l4_available': 0,
                'attrition_rate': 0.05, 'absenteeism_rate': 0.02, 
            },
            {
                'hq_name': 'Corporate HQ', 'factory_name': 'Main Plant', 'department_name': 'Logistics',
                'line_name': '', 'subline_name': '', 'station_name': '', # Empty means "All"
                'month': 1, 'year': 2025, 
                'total_stations': 50, 'operators_required': 50, 'operators_available': 45,
                'buffer_manpower_required': 5, 'buffer_manpower_available': 2,
                'l1_required': 20, 'l1_available': 18, 'l2_required': 10, 'l2_available': 10,
                'l3_required': 10, 'l3_available': 9, 'l4_required': 10, 'l4_available': 8,
                'attrition_rate': 0.10, 'absenteeism_rate': 0.05,
            },
        ]

        # --- SHEET 2: INSTRUCTIONS ---
        instructions = [
            ["Column", "Instruction"],
            ["Hierarchy Names", "Copy names EXACTLY from the 'Reference_Data' sheet to avoid errors."],
            ["Hierarchy Logic", "To select 'All' for a level (e.g., All Lines), leave that column EMPTY."],
            ["Mandatory", "Factory Name, Month, and Year are mandatory."],
            ["Numeric Fields", "Enter integers for counts. Enter decimals (0.0 - 1.0) or percentages for rates."],
            ["Duplicates", "If a record exists for the same Factory/Dept/Line/Date, it will be updated."]
        ]

        # --- SHEET 3: REFERENCE DATA ---
        # Fetch all valid names from DB
        hqs = list(Hq.objects.values_list('hq_name', flat=True))
        factories = list(Factory.objects.values_list('factory_name', flat=True))
        departments = list(Department.objects.values_list('department_name', flat=True))
        lines = list(Line.objects.values_list('line_name', flat=True))
        sublines = list(SubLine.objects.values_list('subline_name', flat=True))
        stations = list(Station.objects.values_list('station_name', flat=True))

        # Find max length to frame the DataFrame
        max_len = max(len(hqs), len(factories), len(departments), len(lines), len(sublines), len(stations))

        ref_data = {
            'Valid HQs': hqs + [''] * (max_len - len(hqs)),
            'Valid Factories': factories + [''] * (max_len - len(factories)),
            'Valid Departments': departments + [''] * (max_len - len(departments)),
            'Valid Lines': lines + [''] * (max_len - len(lines)),
            'Valid Sublines': sublines + [''] * (max_len - len(sublines)),
            'Valid Stations': stations + [''] * (max_len - len(stations)),
        }

        # --- BUILD EXCEL FILE ---
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            pd.DataFrame(dummy_data, columns=headers).to_excel(writer, sheet_name='Fill_Data_Here', index=False)
            pd.DataFrame(instructions[1:], columns=instructions[0]).to_excel(writer, sheet_name='Instructions', index=False)
            pd.DataFrame(ref_data).to_excel(writer, sheet_name='Reference_Data', index=False)
        
        output.seek(0)
        
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="manpower_planning_template.xlsx"'
        return response

        # --- EXCEL UPLOAD ---
        # --- EXCEL UPLOAD ---
    @action(detail=False, methods=['post'], url_path='upload-data')
    @transaction.atomic
    def upload_data(self, request):
        """
        Accepts Excel. 
        Logic:
        1. If Station Name is provided -> Update/Create that specific station.
        2. If Station Name is EMPTY but Line Name is provided -> Find ALL stations in Hierarchy for that line and update/create them.
        """
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(file_obj, na_filter=False) # na_filter=False makes empty cells empty strings ''
        except Exception as e:
            return Response({"error": f"Error reading Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        # Normalize headers to lowercase/strip
        df.columns = [str(c).strip().lower().replace(' ', '_') for c in df.columns]

        required_columns = {'factory_name', 'month', 'year'}
        if not required_columns.issubset(df.columns):
            return Response({"error": f"Missing mandatory columns: {required_columns - set(df.columns)}"}, status=status.HTTP_400_BAD_REQUEST)

        errors = []
        created_count = 0
        updated_count = 0

        # Helper to safely get numbers
        def get_val(row, key, is_float=False):
            val = row.get(key, 0)
            if val == '': return 0.0 if is_float else 0
            try:
                return float(val) if is_float else int(val)
            except:
                return 0.0 if is_float else 0

        for index, row in df.iterrows():
            row_num = index + 2
            try:
                # 1. Clean Inputs
                hq_name = str(row.get('hq_name', '')).strip()
                factory_name = str(row.get('factory_name', '')).strip()
                department_name = str(row.get('department_name', '')).strip()
                line_name = str(row.get('line_name', '')).strip()
                subline_name = str(row.get('subline_name', '')).strip()
                station_name = str(row.get('station_name', '')).strip()

                month = int(row['month'])
                year = int(row['year'])

                # 2. Resolve Hierarchy Objects (By Name)
                # Note: We find them independently because models are decoupled
                factory = Factory.objects.filter(factory_name__iexact=factory_name).first()
                if not factory:
                    errors.append(f"Row {row_num}: Factory '{factory_name}' not found.")
                    continue

                hq = None
                if hq_name:
                    hq = Hq.objects.filter(hq_name__iexact=hq_name).first()

                department = None
                if department_name:
                    department = Department.objects.filter(department_name__iexact=department_name).first()

                line = None
                if line_name:
                    line = Line.objects.filter(line_name__iexact=line_name).first()
                    if not line:
                        errors.append(f"Row {row_num}: Line '{line_name}' not found.")
                        continue

                subline = None
                if subline_name:
                    subline = SubLine.objects.filter(subline_name__iexact=subline_name).first()

                # 3. Prepare Metric Defaults
                defaults = {
                    'total_stations': get_val(row, 'total_stations'),
                    'operators_required': get_val(row, 'operators_required'),
                    'operators_available': get_val(row, 'operators_available'),
                    'buffer_manpower_required': get_val(row, 'buffer_manpower_required'),
                    'buffer_manpower_available': get_val(row, 'buffer_manpower_available'),
                    'l1_required': get_val(row, 'l1_required'),
                    'l1_available': get_val(row, 'l1_available'),
                    'l2_required': get_val(row, 'l2_required'),
                    'l2_available': get_val(row, 'l2_available'),
                    'l3_required': get_val(row, 'l3_required'),
                    'l3_available': get_val(row, 'l3_available'),
                    'l4_required': get_val(row, 'l4_required'),
                    'l4_available': get_val(row, 'l4_available'),
                    'attrition_rate': get_val(row, 'attrition_rate', True),
                    'absenteeism_rate': get_val(row, 'absenteeism_rate', True),
                }

                # ============================================================
                # SCENARIO A: Station Name PROVIDED (Single Record)
                # ============================================================
                if station_name:
                    station_obj = Station.objects.filter(station_name__iexact=station_name).first()
                    if not station_obj:
                        errors.append(f"Row {row_num}: Station '{station_name}' not found.")
                        continue
                    
                    # Use the .pk logic here to fix the previous AttributeError
                    obj, created = AdvanceManpowerDashboard.objects.update_or_create(
                        hq=hq,
                        factory=factory,
                        department=department,
                        line=line,
                        subline=subline,
                        station_id=station_obj.pk, # Use .pk
                        month=month,
                        year=year,
                        defaults=defaults
                    )
                    if created: created_count += 1
                    else: updated_count += 1

                # ============================================================
                # SCENARIO B: Station Name EMPTY but Line Name PROVIDED (Bulk/Explode)
                # ============================================================
                elif line:
                    # User wants to apply this data to ALL stations in this Line
                    # 1. Query Hierarchy to find stations
                    hierarchy_qs = HierarchyStructure.objects.filter(line=line)
                    
                    if subline:
                        hierarchy_qs = hierarchy_qs.filter(subline=subline)

                    if not hierarchy_qs.exists():
                        errors.append(f"Row {row_num}: No stations found in Hierarchy for Line '{line_name}'.")
                        continue
                    
                    # Helper to avoid duplicates in loop
                    processed_station_ids = set()

                    for h_entry in hierarchy_qs:
                        target_station = h_entry.station
                        if not target_station: continue
                        if target_station.pk in processed_station_ids: continue
                        
                        processed_station_ids.add(target_station.pk)

                        # Force the correct subline from hierarchy if not explicitly provided in Excel
                        actual_subline = h_entry.subline if h_entry.subline else subline

                        obj, created = AdvanceManpowerDashboard.objects.update_or_create(
                            hq=hq,
                            factory=factory,
                            department=department,
                            line=line,
                            subline=actual_subline,
                            station_id=target_station.pk, # Use .pk
                            month=month,
                            year=year,
                            defaults=defaults
                        )
                        if created: created_count += 1
                        else: updated_count += 1
                else:
                    # Neither Station nor Line provided?
                    errors.append(f"Row {row_num}: Must provide at least a Line Name.")

            except Exception as e:
                errors.append(f"Row {row_num}: Unexpected error: {str(e)}")

        if errors:
            transaction.set_rollback(True)
            return Response({
                "status": "Error",
                "message": "Upload failed. No data saved.",
                "errors": errors
            }, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            "status": "Success",
            "message": "Data uploaded successfully.",
            "records_created": created_count,
            "records_updated": updated_count,
        }, status=status.HTTP_201_CREATED)

        
     # Existing helper actions...
    @action(detail=False, methods=["get"])
    def department_wise_stations(self, request):
        data = (
            AdvanceManpowerDashboard.objects
            .values("department__id", "department__department_name")
            .annotate(total_stations=Sum("total_stations"))
        )
        return Response(data)

    @action(detail=False, methods=["get"])
    def department_month_year(self, request):
        # You might want to update this to filter by line/subline too if needed
        qs = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)



from django.utils import timezone
from django.http import JsonResponse
from .models import BiometricAttendance, SkillMatrix, AdvanceManpowerDashboard

# def Buffer_manpower_live(request):
#     now = timezone.now()
    
#     # 1. Year Handling: Defaults to current system year, but allows ?year=2023 override
#     try:
#         target_year = int(request.GET.get('year', now.year))
#     except ValueError:
#         target_year = now.year

#     # Filters
#     hq_id = request.GET.get('hq')
#     factory_id = request.GET.get('factory')
#     department_id = request.GET.get('department')
#     line_id = request.GET.get('line')
#     subline_id = request.GET.get('subline')
#     station_id = request.GET.get('station')

#     monthly_data = []

#     # LOOP THROUGH ALL 12 MONTHS
#     for m in range(1, 13):
        
#         # ====================================================
#         # STEP 1: CALCULATE TOTAL AVAILABLE (Internal use only)
#         # ====================================================
#         biometric_qs = BiometricAttendance.objects.filter(
#             attendance_date__month=m,
#             attendance_date__year=target_year
#         ).exclude(status__istartswith='A') 

#         present_pay_codes = biometric_qs.values_list('pay_code', flat=True).distinct()

#         available_qs = SkillMatrix.objects.filter(
#             employee__emp_id__in=present_pay_codes
#         )

#         # Apply Filters to SkillMatrix
#         if hq_id and hq_id != 'null': available_qs = available_qs.filter(hierarchy__hq_id=hq_id)
#         if factory_id and factory_id != 'null': available_qs = available_qs.filter(hierarchy__factory_id=factory_id)
#         if department_id and department_id != 'null': available_qs = available_qs.filter(hierarchy__department_id=department_id)
#         if line_id and line_id != 'null': available_qs = available_qs.filter(hierarchy__line_id=line_id)
#         if subline_id and subline_id != 'null': available_qs = available_qs.filter(hierarchy__subline_id=subline_id)
#         if station_id and station_id != 'null': available_qs = available_qs.filter(hierarchy__station_id=station_id)

#         total_available = available_qs.values('employee__emp_id').distinct().count()


#         # ====================================================
#         # STEP 2: CALCULATE REQUIRED & BUFFER
#         # ====================================================
#         dashboard_qs = AdvanceManpowerDashboard.objects.filter(month=m, year=target_year)

#         if hq_id and hq_id != 'null': dashboard_qs = dashboard_qs.filter(hq_id=hq_id)
#         if factory_id and factory_id != 'null': dashboard_qs = dashboard_qs.filter(factory_id=factory_id)
#         if department_id and department_id != 'null': dashboard_qs = dashboard_qs.filter(department_id=department_id)
#         if line_id and line_id != 'null': dashboard_qs = dashboard_qs.filter(line_id=line_id)
#         if subline_id and subline_id != 'null': dashboard_qs = dashboard_qs.filter(subline_id=subline_id)
#         if station_id and station_id != 'null': dashboard_qs = dashboard_qs.filter(station_id=station_id)

#         # Deduplication Logic (Same as before)
#         rows = dashboard_qs.order_by('created_at')
#         unique_records = {}
#         for row in rows:
#             key = (row.station_id, row.subline_id, row.line_id, row.department_id, row.factory_id)
#             unique_records[key] = row

#         total_required = 0
#         total_buffer_required = 0 # New variable accumulator

#         for row in unique_records.values():
#             # Sum Operators Required
#             total_required += (row.operators_required or 0)
#             # Sum Buffer Manpower Required
#             total_buffer_required += (row.buffer_manpower_required or 0)

#         # ====================================================
#         # STEP 3: PREPARE FINAL OUTPUT
#         # ====================================================
        
#         # Calculate Difference (Available - Required)
#         difference = total_available - total_required

#         monthly_data.append({
#             "month": m,
#             "year": target_year,
#             "difference": difference,
#             "buffer_manpower_required": total_buffer_required
#         })

#     return JsonResponse(monthly_data, safe=False)
def Buffer_manpower_live(request):
    now = timezone.now()
    
    # 1. Year Handling: Defaults to current system year, but allows ?year=2023 override
    try:
        target_year = int(request.GET.get('year', now.year))
    except ValueError:
        target_year = now.year

    # Filters
    hq_id = request.GET.get('hq')
    factory_id = request.GET.get('factory')
    department_id = request.GET.get('department')
    line_id = request.GET.get('line')
    subline_id = request.GET.get('subline')
    station_id = request.GET.get('station')

    monthly_data = []

    # LOOP THROUGH ALL 12 MONTHS
    for m in range(1, 13):
        
        # ====================================================
        # STEP 1: CALCULATE TOTAL AVAILABLE (Internal use only)
        # ====================================================
        biometric_qs = BiometricAttendance.objects.filter(
            attendance_date__month=m,
            attendance_date__year=target_year
        ).exclude(status__istartswith='A') 

        present_pay_codes = biometric_qs.values_list('pay_code', flat=True).distinct()

        available_qs = SkillMatrix.objects.filter(
            employee__emp_id__in=present_pay_codes
        )

        # Apply Filters to SkillMatrix
        if hq_id and hq_id != 'null': available_qs = available_qs.filter(hierarchy__hq_id=hq_id)
        if factory_id and factory_id != 'null': available_qs = available_qs.filter(hierarchy__factory_id=factory_id)
        if department_id and department_id != 'null': available_qs = available_qs.filter(hierarchy__department_id=department_id)
        if line_id and line_id != 'null': available_qs = available_qs.filter(hierarchy__line_id=line_id)
        if subline_id and subline_id != 'null': available_qs = available_qs.filter(hierarchy__subline_id=subline_id)
        if station_id and station_id != 'null': available_qs = available_qs.filter(hierarchy__station_id=station_id)

        total_available = available_qs.values('employee__emp_id').distinct().count()


        # ====================================================
        # STEP 2: CALCULATE REQUIRED & BUFFER
        # ====================================================
        dashboard_qs = AdvanceManpowerDashboard.objects.filter(month=m, year=target_year)

        if hq_id and hq_id != 'null': dashboard_qs = dashboard_qs.filter(hq_id=hq_id)
        if factory_id and factory_id != 'null': dashboard_qs = dashboard_qs.filter(factory_id=factory_id)
        if department_id and department_id != 'null': dashboard_qs = dashboard_qs.filter(department_id=department_id)
        if line_id and line_id != 'null': dashboard_qs = dashboard_qs.filter(line_id=line_id)
        if subline_id and subline_id != 'null': dashboard_qs = dashboard_qs.filter(subline_id=subline_id)
        if station_id and station_id != 'null': dashboard_qs = dashboard_qs.filter(station_id=station_id)

        # Deduplication Logic
        rows = dashboard_qs.order_by('created_at')
        unique_records = {}
        for row in rows:
            key = (row.station_id, row.subline_id, row.line_id, row.department_id, row.factory_id)
            unique_records[key] = row

        total_required = 0
        total_buffer_required = 0

        for row in unique_records.values():
            total_required += (row.operators_required or 0)
            total_buffer_required += (row.buffer_manpower_required or 0)

        # ====================================================
        # STEP 3: FINAL CALCULATION - NEVER SHOW NEGATIVE DIFFERENCE
        # ====================================================
        
        difference = total_available - total_required
        difference = max(0, difference)   # ← This is what you asked for: negative values become 0

        monthly_data.append({
            "month": m,
            "year": target_year,
            "difference": difference,                    # Surplus manpower (0 if shortage)
            "buffer_manpower_required": total_buffer_required
        })

    return JsonResponse(monthly_data, safe=False)




# # --- NEW CHART VIEW ---
class AdvancedManpowerTrendChartView(APIView):
    def get(self, request):
        # 1. Determine Year (Default to current)
        req_year = request.query_params.get('year')
        target_year = int(req_year) if req_year else timezone.now().year
        
        # 2. Base Query
        queryset = AdvanceManpowerDashboard.objects.filter(year=target_year)

        # 3. Apply Hierarchy Filters (Handles "All" automatically)
        queryset = apply_hierarchy_filters(request, queryset)

        # 4. Aggregate Data by Month
        data = queryset.values('year', 'month').annotate(
            total_required=Sum('operators_required'),
            total_available=Sum('operators_available')
        ).order_by('year', 'month')

        # 5. Format Response
        formatted_data = []
        for item in data:
            formatted_data.append({
                'month': item['month'],
                'year': item['year'],
                'operators_required': item['total_required'] or 0,
                'operators_available': item['total_available'] or 0
            })

        return Response(formatted_data, status=status.HTTP_200_OK)
    




from django.db.models import Avg



# from django.http import JsonResponse
# from django.utils import timezone
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
# from django.db.models import Q
# from .models import BiometricAttendance, SkillMatrix, AdvanceManpowerDashboard

# class BufferManpowerChartView(APIView):

#     def get(self, request):
#         now = timezone.now()
#         # 1. Determine Financial Year Start
#         target_fy_start = int(request.GET.get("year", now.year - 1 if now.month <= 3 else now.year))
        
#         # 2. Setup Hierarchy Filters
#         # Note: SkillMatrix uses 'hierarchy__', Dashboard uses direct fields
#         filters_skill = {}
#         filters_dashboard = {}
#         for f in ['hq', 'factory', 'department', 'line', 'subline', 'station']:
#             val = request.GET.get(f)
#             if val and val != "null" and val.strip():
#                 filters_skill[f"hierarchy__{f}_id"] = val
#                 filters_dashboard[f"{f}_id"] = val

#         monthly_data = []

#         # 3. Loop through 12 months (April to March)
#         # Financial year sequence: 4,5,6,7,8,9,10,11,12, 1,2,3
#         months_sequence = [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3]

#         for month in months_sequence:
#             year = target_fy_start if month >= 4 else target_fy_start + 1

#             # --- A. Get Available Manpower (From Attendance + Skill Matrix) ---
#             biometric_qs = BiometricAttendance.objects.filter(
#                 attendance_date__year=year,
#                 attendance_date__month=month,
#             ).exclude(status__in=['A', 'Absent', 'Leave', 'WO'])

#             present_emp_codes = biometric_qs.values_list('pay_code', flat=True).distinct()

#             available_qs = SkillMatrix.objects.filter(
#                 employee__emp_id__in=present_emp_codes
#             )
#             if filters_skill:
#                 available_qs = available_qs.filter(**filters_skill)

#             total_available = available_qs.values('employee__emp_id').distinct().count()

#             # --- B. Get Required Manpower & Target Buffer (From Dashboard) ---
#             dashboard_qs = AdvanceManpowerDashboard.objects.filter(
#                 year=year,
#                 month=month,
#                 **filters_dashboard
#             )

#             # Deduplicate entries per hierarchy
#             unique_dashboard = {}
#             for rec in dashboard_qs:
#                 key = (rec.station_id, rec.subline_id, rec.line_id, rec.department_id, rec.factory_id)
#                 unique_dashboard[key] = rec

#             total_required = sum(r.operators_required or 0 for r in unique_dashboard.values())
#             # We also get the 'Buffer Required' target to show on the chart
#             total_buffer_required = sum(r.buffer_manpower_required or 0 for r in unique_dashboard.values())

#             # --- C. THE CONCEPT: Buffer Available = Required - Available ---
#             # If Required=37 and Available=35, gap = 2
#             gap = total_required - total_available
#             buffer_manpower_available = max(0, gap) 

#             monthly_data.append({
#                 "month": month,
#                 "year": year,
#                 "buffer_manpower_required": total_buffer_required, # Target line
#                 "buffer_manpower_available": buffer_manpower_available, # The Area (Gap)
#             })

#         return Response(monthly_data, status=status.HTTP_200_OK)



class AdvanceCardStatsView(APIView):
    def get(self, request):
        req_year = request.query_params.get('year')
        target_year = int(req_year) if req_year else timezone.now().year
        
        queryset = AdvanceManpowerDashboard.objects.filter(year=target_year)
        queryset = apply_hierarchy_filters(request, queryset)

        # 1. Aggregate the raw DB fields
        data = queryset.aggregate(
            stations=Sum('total_stations'),
            req=Sum('operators_required'),
            avail=Sum('operators_available'),
            buf_req=Sum('buffer_manpower_required'),
            buf_avail=Sum('buffer_manpower_available'),
            
            # Raw DB Fields
            db_l1_req=Sum('l1_required'), 
            db_l1_avail=Sum('l1_available'),
            db_l2_req=Sum('l2_required'), 
            db_l2_avail=Sum('l2_available'),
            db_l3_req=Sum('l3_required'), 
            db_l3_avail=Sum('l3_available'),
            db_l4_req=Sum('l4_required'), 
            db_l4_avail=Sum('l4_available')
        )

        # 2. Map them to the keys your Frontend expects
        response_data = {
            # Main Card
            "total_stations": data['stations'] or 0,
            "operators_required": data['req'] or 0,
            "operators_available": data['avail'] or 0,
            "buffer_manpower_required": data['buf_req'] or 0,
            "buffer_manpower_available": data['buf_avail'] or 0,

            # MAPPING LOGIC:
            # Frontend Key           <--   Database Sum
            "bifurcation_plan_l1":        data['db_l1_req'] or 0,
            "bifurcation_actual_l1":      data['db_l1_avail'] or 0,
            
            "bifurcation_plan_l2":        data['db_l2_req'] or 0,
            "bifurcation_actual_l2":      data['db_l2_avail'] or 0,
            
            "bifurcation_plan_l3":        data['db_l3_req'] or 0,
            "bifurcation_actual_l3":      data['db_l3_avail'] or 0,
            
            "bifurcation_plan_l4":        data['db_l4_req'] or 0,
            "bifurcation_actual_l4":      data['db_l4_avail'] or 0,
        }

        # Return list (so frontend data[0] works)
        return Response([response_data], status=status.HTTP_200_OK)




from django.db.models import Sum, Avg, Q
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.core.cache import cache
from django.db import connection
import hashlib
import json

# Helper function for hierarchy filters
def apply_hierarchy_filters(request, queryset):
    """Apply hierarchy filters to queryset based on request params"""
    hq_id = request.query_params.get('hq')
    factory_id = request.query_params.get('factory')
    department_id = request.query_params.get('department')
    line_id = request.query_params.get('line')
    subline_id = request.query_params.get('subline')
    station_id = request.query_params.get('station')

    if hq_id and hq_id != 'null':
        queryset = queryset.filter(hq_id=hq_id)
    if factory_id and factory_id != 'null':
        queryset = queryset.filter(factory_id=factory_id)
    if department_id and department_id != 'null':
        queryset = queryset.filter(department_id=department_id)
    if line_id and line_id != 'null':
        queryset = queryset.filter(line_id=line_id)
    if subline_id and subline_id != 'null':
        queryset = queryset.filter(subline_id=subline_id)
    if station_id and station_id != 'null':
        queryset = queryset.filter(station_id=station_id)
    
    return queryset

def generate_cache_key(request, view_name):
    """Generate a unique cache key based on request parameters"""
    params = dict(request.query_params)
    params_str = json.dumps(params, sort_keys=True)
    key_hash = hashlib.md5(params_str.encode()).hexdigest()
    return f"{view_name}_{key_hash}"



from django.db.models import Avg, Q
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.core.cache import cache
from django.utils import timezone
from .models import AdvanceManpowerDashboard
# assuming you have this helper


class AttritionChartView(APIView):
    def get(self, request):
        cache_key = generate_cache_key(request, 'attrition_trend_multi')
        cached = cache.get(cache_key)
        if cached:
            return Response(cached, status=status.HTTP_200_OK)

        req_year = request.query_params.get('year')
        
        # Determine financial year start
        if req_year:
            fy_start_year = int(req_year)
        else:
            now = timezone.now()
            fy_start_year = now.year - 1 if now.month <= 3 else now.year
        
        fy_end_year = fy_start_year + 1

        # Base queryset with hierarchy filters
        queryset = AdvanceManpowerDashboard.objects.filter(
            Q(year=fy_start_year, month__gte=4) |
            Q(year=fy_end_year, month__lte=3)
        ).select_related('hq', 'factory', 'department', 'line', 'subline', 'station')

        queryset = apply_hierarchy_filters(request, queryset)

        # Aggregate all three metrics
        aggregated = queryset.values('year', 'month').annotate(
            avg_attrition=Avg('attrition_rate'),
            avg_oet=Avg('oet_attrition'),
            avg_associate=Avg('associate_attrition'),
        ).order_by('year', 'month')

        # Format response
        result = []
        for row in aggregated:
            # Convert Decimal → float safely
            def safe_float(val):
                return float(val) if val is not None else 0.0

            result.append({
                "month": row["month"],
                "year": row["year"],
                "attrition_rate":    round(safe_float(row["avg_attrition"]), 2),
                "oet_attrition":     round(safe_float(row["avg_oet"]),       2),
                "associate_attrition": round(safe_float(row["avg_associate"]), 2),
            })

        # Cache 5 minutes
        cache.set(cache_key, result, 300)

        return Response(result, status=status.HTTP_200_OK)





# ✅ CORRECTED: Absenteeism Chart View
class AbsenteeismChartView(APIView):
    def get(self, request):
        # Check cache first
        cache_key = generate_cache_key(request, 'absenteeism_trend')
        cached_data = cache.get(cache_key)
        if cached_data:
            return Response(cached_data, status=status.HTTP_200_OK)

        req_year = request.query_params.get('year')
        
        # Determine FY Start Year
        if req_year:
            fy_start_year = int(req_year)
        else:
            now = timezone.now()
            fy_start_year = now.year - 1 if now.month <= 3 else now.year
        
        fy_end_year = fy_start_year + 1

        # ✅ Query for Financial Year
        queryset = AdvanceManpowerDashboard.objects.filter(
            Q(year=fy_start_year, month__gte=4) |  # Apr-Dec of start year
            Q(year=fy_end_year, month__lte=3)      # Jan-Mar of end year
        ).select_related('hq', 'factory', 'department', 'line', 'subline', 'station')

        # Apply hierarchy filters
        queryset = apply_hierarchy_filters(request, queryset)

        # ✅ CRITICAL FIX: Aggregate by BOTH year AND month
        data = queryset.values('year', 'month').annotate(
            avg_absenteeism=Avg('absenteeism_rate')
        ).order_by('year', 'month')

        # ✅ Format response - Include year in response
        formatted_data = []
        for item in data:
            absenteeism_value = item['avg_absenteeism']
            
            # Convert Decimal to float, handle None
            if absenteeism_value is not None:
                try:
                    absenteeism_value = float(absenteeism_value)
                except (ValueError, TypeError):
                    absenteeism_value = 0.0
            else:
                absenteeism_value = 0.0
                
            formatted_data.append({
                'month': item['month'],
                'year': item['year'],  # ✅ INCLUDE YEAR
                'absenteeism_rate': round(absenteeism_value, 2)
            })

        # Cache for 5 minutes
        cache.set(cache_key, formatted_data, 300)
        
        return Response(formatted_data, status=status.HTTP_200_OK)

# ✅ OPTIMIZED: Bifurcation Stats View
class BifurcationStatsView(APIView):
    def get(self, request):
        # Check cache first
        cache_key = generate_cache_key(request, 'bifurcation_stats')
        cached_data = cache.get(cache_key)
        if cached_data:
            return Response(cached_data, status=status.HTTP_200_OK)

        req_year = request.query_params.get('year')
        
        # Determine FY Start Year
        if req_year:
            fy_start_year = int(req_year)
        else:
            now = timezone.now()
            fy_start_year = now.year - 1 if now.month <= 3 else now.year
        
        fy_end_year = fy_start_year + 1

        # ✅ Optimized Query
        queryset = AdvanceManpowerDashboard.objects.filter(
            Q(year=fy_start_year, month__gte=4) |
            Q(year=fy_end_year, month__lte=3)
        ).select_related('hq', 'factory', 'department', 'line', 'subline', 'station')

        # Apply hierarchy filters
        queryset = apply_hierarchy_filters(request, queryset)

        # Aggregate Sums
        data = queryset.aggregate(
            l1_req=Sum('l1_required'),
            l1_avail=Sum('l1_available'),
            l2_req=Sum('l2_required'),
            l2_avail=Sum('l2_available'),
            l3_req=Sum('l3_required'),
            l3_avail=Sum('l3_available'),
            l4_req=Sum('l4_required'),
            l4_avail=Sum('l4_available')
        )

        response_data = {
            'l1_required': data['l1_req'] or 0,
            'l1_available': data['l1_avail'] or 0,
            'l2_required': data['l2_req'] or 0,
            'l2_available': data['l2_avail'] or 0,
            'l3_required': data['l3_req'] or 0,
            'l3_available': data['l3_avail'] or 0,
            'l4_required': data['l4_req'] or 0,
            'l4_available': data['l4_avail'] or 0,
        }

        # Cache for 5 minutes
        cache.set(cache_key, response_data, 300)

        return Response(response_data, status=status.HTTP_200_OK)

# # ============================== end ===================================================================




























# from django.http import HttpResponse, JsonResponse
# from django.views import View
# from django.utils.decorators import method_decorator
# from django.views.decorators.csrf import csrf_exempt
# from reportlab.lib.pagesizes import letter
# from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
# from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
# from reportlab.lib import colors
# from reportlab.lib.units import inch

# import os
# import json
# import traceback
# from io import BytesIO


# # Import the new models and serializers MultiSkilling
# from datetime import datetime 
# from .models import MasterTable, SkillMatrix, Score, Schedule, CompanyLogo
# from .serializers import (
#     CardEmployeeMasterSerializer,
#     OperatorCardSkillSerializer,
#     CardScoreSerializer,
#     CardScheduleSerializer
# )

# @method_decorator(csrf_exempt, name='dispatch')
# class EmployeeReportPDFView(View):
#     def post(self, request, *args, **kwargs):
#         """
#         Handle PDF generation requests
#         Accepts both form data and JSON input
#         """
#         try:
#             print("\n=== Received PDF generation request ===")
            
#             # 1. Parse input data
#             card_no = self._get_card_number(request)
#             if not card_no:
#                 return JsonResponse({'error': 'card_no is required'}, status=400)
#             print(f"Processing card_no: {card_no}")

#             # 2. Get employee record
#             try:
#                 # Use the new MasterTable model
#                 employee = MasterTable.objects.get(emp_id=card_no)
#                 print(f"Found employee: {employee.first_name}")
#             except MasterTable.DoesNotExist:
#                 print(f"Employee not found for card_no: {card_no}")
#                 return JsonResponse({'error': 'Employee not found'}, status=404)

#             # 3. Fetch and serialize all data using the new serializers
#             print("Fetching and serializing data...")
#             employee_data = CardEmployeeMasterSerializer(employee).data
#             operator_skills = OperatorCardSkillSerializer(
#                 SkillMatrix.objects.filter(employee=employee), many=True).data
#             scores = CardScoreSerializer(
#                 Score.objects.filter(employee=employee), many=True).data
#             # Note: You need to create a serializer for MultiSkilling if you haven't already
#             # multi_skilling = MultiSkilling.objects.filter(employee=employee) 
#             scheduled_trainings = CardScheduleSerializer(
#                 Schedule.objects.filter(employees=employee), many=True).data
            
#             # 4. Generate PDF content
#             print("Generating PDF content...")
#             buffer = BytesIO()
#             doc = SimpleDocTemplate(
#                 buffer, 
#                 pagesize=letter,
#                 rightMargin=72,
#                 leftMargin=72,
#                 topMargin=72,
#                 bottomMargin=72
#             )
#             story = self.create_pdf_content(
#                 employee_data,
#                 operator_skills,
#                 scores,
#                 # multi_skilling, # Pass the queryset for now, as no serializer is provided
#                 scheduled_trainings
#             )
            
#             # 5. Build PDF document
#             print("Building PDF document...")
#             doc.build(story)
#             buffer.seek(0)
#             print("PDF generation completed successfully")

#             # 6. Return PDF response
#             response = HttpResponse(
#                 buffer.getvalue(), 
#                 content_type='application/pdf'
#             )
#             response['Content-Disposition'] = (
#                 f'attachment; filename="employee_report_{card_no}.pdf"'
#             )
#             return response
            
#         except Exception as e:
#             print("\n!!! PDF generation failed !!!")
#             traceback.print_exc()
#             return JsonResponse(
#                 {
#                     'error': 'Internal server error',
#                     'detail': str(e),
#                     'traceback': traceback.format_exc()
#                 }, 
#                 status=500
#             )

#     def _get_card_number(self, request):
#         if request.content_type == 'application/json':
#             try:
#                 data = json.loads(request.body)
#                 return data.get('card_no')
#             except json.JSONDecodeError:
#                 return None
#         return request.POST.get('card_no')

#     def _format_date(self, date_obj):
#         """Format date object to string"""
#         from datetime import datetime as dt
        
#         if isinstance(date_obj, str):
#             try:
#                 date_obj = dt.strptime(date_obj, '%Y-%m-%d').date()
#             except ValueError:
#                 return 'N/A'
#         return date_obj.strftime('%d-%b-%Y') if date_obj else 'N/A'

#     def _get_company_logo(self):
#         try:
#             logo = CompanyLogo.objects.first()
#             if logo and logo.logo:
#                 return logo.logo.path
#         except Exception as e:
#             print(f"Error getting company logo: {e}")
#         return None

#     def _add_header(self, story, styles):
#         # """Add header with company logo and title"""
#         logo_path = self._get_company_logo()
        
#         if logo_path and os.path.exists(logo_path):
#             try:
#                 logo = Image(logo_path, width=1*inch, height=0.7*inch)
#                 title = Paragraph("Employee Comprehensive Report", styles['Title'])
#                 header_data = [[title, logo]]
#                 header_table = Table(header_data, colWidths=[5.5*inch, 1*inch])
#                 header_table.setStyle(TableStyle([
#                     ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
#                     ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
#                     ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#                     ('LEFTPADDING', (0, 0), (-1, -1), 0),
#                     ('RIGHTPADDING', (0, 0), (-1, -1), 0),
#                     ('TOPPADDING', (0, 0), (-1, -1), 0),
#                     ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
#                     ('GRID', (0, 0), (-1, -1), 0, colors.white),
#                 ]))
#                 story.append(header_table)
#             except Exception as e:
#                 print(f"Error adding logo: {e}")
#                 story.append(Paragraph("Employee Comprehensive Report", styles['Title']))
#         else:
#             story.append(Paragraph("Employee Comprehensive Report", styles['Title']))
        
#         story.append(Spacer(1, 20))
    
#     def _get_table_style(self):
#         return TableStyle([
#             ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4682B4')),
#             ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
#             ('ALIGN', (0,0), (-1,0), 'CENTER'),
#             ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
#             ('FONTSIZE', (0,0), (-1,0), 10),
#             ('BOTTOMPADDING', (0,0), (-1,0), 8),
#             ('TOPPADDING', (0,0), (-1,0), 8),
#             ('ALIGN', (0,1), (-1,-1), 'LEFT'),
#             ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
#             ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
#             ('FONTSIZE', (0,1), (-1,-1), 9),
#             ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#E0E0E0')),
#             ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8F8F8')]),
#             ('LEFTPADDING', (0,0), (-1,-1), 8),
#             ('RIGHTPADDING', (0,0), (-1,-1), 8),
#             ('TOPPADDING', (0,0), (-1,-1), 6),
#             ('BOTTOMPADDING', (0,0), (-1,-1), 6),
#         ])

#     def _get_result_badge_style(self, result, row):
#         color = colors.green if result else colors.red
#         return [
#             ('TEXTCOLOR', (3, row), (3, row), color),
#             ('FONTNAME', (3, row), (3, row), 'Helvetica-Bold')
#         ]

#     def _get_status_badge_style(self, status, row):
#         status_colors = {
#             'active': colors.green,
#             'completed': colors.blue,
#             'inprogress': colors.orange,
#             'scheduled': colors.purple,
#             'inactive': colors.gray,
#             'rescheduled': colors.darkblue,
#         }
#         color = status_colors.get(status.lower(), colors.black)
#         return [
#             ('TEXTCOLOR', (1, row), (1, row), color),
#             ('FONTNAME', (1, row), (1, row), 'Helvetica-Bold')
#         ]

#     # New `create_pdf_content` with serialized data as arguments
#     # def create_pdf_content(self, employee_data, skills_data, scores_data, multi_skilling_queryset, training_data):
#     def create_pdf_content(self, employee_data, skills_data, scores_data, multi_skilling_queryset):
#         styles = getSampleStyleSheet()
#         story = []
        
#         # Add header
#         self._add_header(story, styles)
        
#         # Add sections using the pre-serialized data
#         self._add_basic_info(story, styles, employee_data)
#         self._add_operator_skills(story, styles, skills_data)
#         self._add_scores(story, styles, scores_data)
        
#         # You'll need to update this function to handle the serialized data
#         # self._add_multi_skills(story, styles, multi_skilling_queryset)
        
#         # Add refreshment training section
#         # self._add_refreshment_training(story, styles, training_data)
        
#         return story

#     def _add_basic_info(self, story, styles, employee_data):
#         heading_style = ParagraphStyle(
#             'CustomHeading',
#             parent=styles['Heading2'],
#             fontSize=14,
#             spaceAfter=12,
#             textColor=colors.HexColor('#4682B4'),
#             fontName='Helvetica-Bold'
#         )
#         story.append(Paragraph("Basic Information", heading_style))
#         story.append(Spacer(1, 8))
        
#         basic_data = [
#             ["Field", "Value"],
#             ["Name", f"{employee_data.get('first_name', 'N/A')} {employee_data.get('last_name', '')}"],
#             ["Card No", employee_data.get('emp_id', 'N/A')],
#             ["Department", employee_data.get('department', 'N/A')],
#             ["Joining Date", self._format_date(employee_data.get('date_of_joining'))],
#             ["Birth Date", self._format_date(employee_data.get('birth_date'))],
#             ["Sex", employee_data.get('sex', 'N/A')],
#             ["Email", employee_data.get('email', 'N/A')],
#             ["Phone", employee_data.get('phone', 'N/A')],
#         ]
        
#         basic_table = Table(basic_data, colWidths=[2*inch, 4.5*inch])
#         basic_table.setStyle(self._get_table_style())
#         story.append(basic_table)
#         story.append(Spacer(1, 24))

#     def _add_operator_skills(self, story, styles, skills_data):
#         if not skills_data:
#             return

#         heading_style = ParagraphStyle(
#             'CustomHeading',
#             parent=styles['Heading2'],
#             fontSize=14,
#             spaceAfter=12,
#             textColor=colors.HexColor('#4682B4'),
#             fontName='Helvetica-Bold'
#         )
#         story.append(Paragraph("Operator Skills", heading_style))
#         story.append(Spacer(1, 8))

#         skill_data = [["Level", "Station", "Date"]]
#         for skill in skills_data:
#             skill_data.append([
#                 skill.get('level_name', 'N/A'),
#                 skill.get('station_name', 'N/A'),
#                 self._format_date(skill.get('updated_at'))
#             ])

#         skills_table = Table(skill_data, colWidths=[1*inch, 3.5*inch, 2*inch])
#         skills_style = self._get_table_style()
#         skills_style.add('ALIGN', (0,1), (0,-1), 'CENTER')
#         skills_style.add('ALIGN', (2,1), (2,-1), 'CENTER')
#         skills_table.setStyle(skills_style)
#         story.append(skills_table)
#         story.append(Spacer(1, 24))

#     def _add_scores(self, story, styles, scores_data):
#         if not scores_data:
#             return
            
#         heading_style = ParagraphStyle(
#             'CustomHeading',
#             parent=styles['Heading2'],
#             fontSize=14,
#             spaceAfter=12,
#             textColor=colors.HexColor('#4682B4'),
#             fontName='Helvetica-Bold'
#         )
#         story.append(Paragraph("Scores and Assessments", heading_style))
#         story.append(Spacer(1, 8))
        
#         score_data = [["Test", "Marks", "Percentage", "Result", "Date"]]
#         for score in scores_data:
#             score_data.append([
#                 score.get('test_name', 'N/A'),
#                 str(score.get('marks', 'N/A')),
#                 f"{score.get('percentage', 'N/A')}%" if score.get('percentage') is not None else "N/A",
#                 "Pass" if score.get('passed') else "Fail",
#                 self._format_date(score.get('created_at'))
#             ])
        
#         scores_table = Table(score_data, colWidths=[2.5*inch, 0.8*inch, 1*inch, 0.8*inch, 1.4*inch])
#         scores_style = self._get_table_style()
#         scores_style.add('ALIGN', (1,1), (2,-1), 'CENTER')
#         scores_style.add('ALIGN', (3,1), (3,-1), 'CENTER')
#         scores_style.add('ALIGN', (4,1), (4,-1), 'CENTER')
        
#         for row in range(1, len(score_data)):
#             result = scores_data[row-1].get('passed')
#             for style_command in self._get_result_badge_style(result, row):
#                 scores_style.add(*style_command)
        
#         scores_table.setStyle(scores_style)
#         story.append(scores_table)
#         story.append(Spacer(1, 24))

from django.http import HttpResponse, JsonResponse
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER
import os
import json
import traceback
from io import BytesIO
from django.db.models import Q

from .models import (
    MasterTable, SkillMatrix, Score, Schedule, CompanyLogo,
    MultiSkilling, HanchouExamResult, ShokuchouExamResult,
    TrainingAttendance, RescheduledSession, UserRegistration
)
from .serializers import (
    CardEmployeeMasterSerializer,
    OperatorCardSkillSerializer,
    CardScoreSerializer,
    CardScheduleSerializer,
    CardMultiSkillingSerializer,
    CardHanchouExamResultSerializer,
    CardShokuchouExamResultSerializer,
    CardTrainingAttendanceSerializer,
    CardRescheduledSessionSerializer,
)

@method_decorator(csrf_exempt, name='dispatch')
class EmployeeReportPDFView(View):
    def post(self, request, *args, **kwargs):
        """
        Handle PDF generation requests
        Accepts both form data and JSON input
        """
        try:
            print("\n=== Received PDF generation request ===")
            
            # 1. Parse input data
            card_no = self._get_card_number(request)
            if not card_no:
                return JsonResponse({'error': 'card_no is required'}, status=400)
            
            print(f"Processing card_no: {card_no}")
            
            # 2. Get employee record
            try:
                employee = MasterTable.objects.get(emp_id=card_no)
                print(f"Found employee: {employee.first_name}")
            except MasterTable.DoesNotExist:
                print(f"Employee not found for card_no: {card_no}")
                return JsonResponse({'error': 'Employee not found'}, status=404)
            
            # 3. Fetch and serialize all data
            print("Fetching and serializing data...")
            employee_data = CardEmployeeMasterSerializer(employee).data
            operator_skills = OperatorCardSkillSerializer(
                SkillMatrix.objects.filter(employee=employee), many=True).data
            scores = CardScoreSerializer(
                Score.objects.filter(employee=employee), many=True).data
            multi_skilling = CardMultiSkillingSerializer(
                MultiSkilling.objects.filter(employee=employee), many=True).data
            hanchou_results = CardHanchouExamResultSerializer(
                HanchouExamResult.objects.filter(employee=employee), many=True).data
            shokuchou_results = CardShokuchouExamResultSerializer(
                ShokuchouExamResult.objects.filter(employee=employee), many=True).data
            
            # Get scheduled trainings
            all_schedules = Schedule.objects.filter(employees=employee)
            scheduled_trainings = CardScheduleSerializer(
                all_schedules, many=True, context={'employee': employee}).data
            
            # Get attendance and rescheduled sessions
            attendances_data, rescheduled_sessions_data = self._get_attendance_data(employee)
            
            # 4. Generate PDF content
            print("Generating PDF content...")
            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=letter,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72
            )
            
            story = self.create_pdf_content(
                employee_data,
                operator_skills,
                scores,
                multi_skilling,
                scheduled_trainings,
                hanchou_results,
                shokuchou_results,
                attendances_data,
                rescheduled_sessions_data
            )
            
            # 5. Build PDF document
            print("Building PDF document...")
            doc.build(story)
            buffer.seek(0)
            print("PDF generation completed successfully")
            
            # 6. Return PDF response
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/pdf'
            )
            response['Content-Disposition'] = (
                f'attachment; filename="employee_report_{card_no}.pdf"'
            )
            return response
            
        except Exception as e:
            print("\n!!! PDF generation failed !!!")
            traceback.print_exc()
            return JsonResponse(
                {
                    'error': 'Internal server error',
                    'detail': str(e),
                    'traceback': traceback.format_exc()
                },
                status=500
            )
    
    def _get_card_number(self, request):
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                return data.get('card_no')
            except json.JSONDecodeError:
                return None
        return request.POST.get('card_no')
    
    def _get_attendance_data(self, employee):
        """Get attendance and rescheduled sessions data for employee"""
        attendances_data = []
        rescheduled_sessions_data = []
        
        try:
            query = Q()
            temp_id = getattr(employee, 'temp_id', None)
            if temp_id and str(temp_id).strip():
                query |= Q(temp_id__iexact=str(temp_id).strip())
            
            first_name_clean = (employee.first_name or '').strip()
            last_name_clean = (employee.last_name or '').strip()
            full_name = f"{first_name_clean} {last_name_clean}".strip()
            
            if first_name_clean:
                if last_name_clean:
                    query |= (Q(first_name__iexact=first_name_clean) & Q(last_name__iexact=last_name_clean))
                
                if ' ' in first_name_clean:
                    name_parts = first_name_clean.split()
                    if len(name_parts) > 1:
                        possible_first = name_parts[0]
                        possible_last = ' '.join(name_parts[1:])
                        query |= (Q(first_name__iexact=possible_first) & Q(last_name__iexact=possible_last))
                
                query |= Q(first_name__icontains=first_name_clean)
                query |= Q(first_name__iexact=full_name)
            
            user_records = UserRegistration.objects.filter(query).distinct()[:5]
            
            if user_records.exists():
                attendances = TrainingAttendance.objects.filter(
                    user__in=user_records
                ).select_related('batch', 'user', 'day_number').order_by(
                    '-attendance_date', 'batch__batch_id', 'day_number'
                )
                attendances_data = CardTrainingAttendanceSerializer(attendances, many=True).data
                
                rescheduled_sessions = RescheduledSession.objects.filter(
                    employee__in=user_records
                ).select_related(
                    'employee', 'original_day', 'training_subtopic', 'batch'
                ).order_by('-rescheduled_date', '-rescheduled_time')
                
                rescheduled_sessions_data = CardRescheduledSessionSerializer(
                    rescheduled_sessions, many=True
                ).data
        except Exception as e:
            print(f"❌ Attendance fetch error: {str(e)}")
        
        return attendances_data, rescheduled_sessions_data
    
    def _format_date(self, date_obj):
        """Format date object to string"""
        from datetime import datetime as dt
        
        if isinstance(date_obj, str):
            try:
                date_obj = dt.strptime(date_obj, '%Y-%m-%d').date()
            except ValueError:
                return 'N/A'
        return date_obj.strftime('%d-%b-%Y') if date_obj else 'N/A'
    
    def _get_company_logo(self):
        try:
            logo = CompanyLogo.objects.first()
            if logo and logo.logo:
                return logo.logo.path
        except Exception as e:
            print(f"Error getting company logo: {e}")
        return None
    
    def _add_header(self, story, styles):
        """Add header with company logo and title"""
        logo_path = self._get_company_logo()
        
        if logo_path and os.path.exists(logo_path):
            try:
                logo = Image(logo_path, width=1*inch, height=0.7*inch)
                title = Paragraph("Employee Comprehensive Report", styles['Title'])
                header_data = [[title, logo]]
                header_table = Table(header_data, colWidths=[5.5*inch, 1*inch])
                header_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                    ('TOPPADDING', (0, 0), (-1, -1), 0),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                    ('GRID', (0, 0), (-1, -1), 0, colors.white),
                ]))
                story.append(header_table)
            except Exception as e:
                print(f"Error adding logo: {e}")
                story.append(Paragraph("Employee Comprehensive Report", styles['Title']))
        else:
            story.append(Paragraph("Employee Comprehensive Report", styles['Title']))
        
        story.append(Spacer(1, 20))
    
    def _get_table_style(self):
        return TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4682B4')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('TOPPADDING', (0,0), (-1,0), 8),
            ('ALIGN', (0,1), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#E0E0E0')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8F8F8')]),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
            ('RIGHTPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ])
    
    def _get_result_badge_style(self, result, row):
        color = colors.green if result else colors.red
        return [
            ('TEXTCOLOR', (3, row), (3, row), color),
            ('FONTNAME', (3, row), (3, row), 'Helvetica-Bold')
        ]
    
    def _get_status_badge_style(self, status, row, col=1):
        status_colors = {
            'active': colors.green,
            'completed': colors.blue,
            'inprogress': colors.orange,
            'scheduled': colors.purple,
            'inactive': colors.gray,
            'rescheduled': colors.darkblue,
            'present': colors.green,
            'absent': colors.red,
            'pass': colors.green,
            'fail': colors.red,
        }
        color = status_colors.get(status.lower(), colors.black)
        return [
            ('TEXTCOLOR', (col, row), (col, row), color),
            ('FONTNAME', (col, row), (col, row), 'Helvetica-Bold')
        ]
    
    def _get_heading_style(self, styles):
        return ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor('#4682B4'),
            fontName='Helvetica-Bold'
        )
    
    def create_pdf_content(self, employee_data, skills_data, scores_data, 
                          multi_skilling_data, training_data, hanchou_results,
                          shokuchou_results, attendance_data, rescheduled_data):
        styles = getSampleStyleSheet()
        story = []
        
        # Add header
        self._add_header(story, styles)
        
        # Add all sections
        self._add_basic_info(story, styles, employee_data)
        self._add_operator_skills(story, styles, skills_data)
        self._add_multi_skills(story, styles, multi_skilling_data)
        self._add_scores(story, styles, scores_data)
        self._add_hanchou_shokuchou_results(story, styles, hanchou_results, shokuchou_results)
        self._add_scheduled_training(story, styles, training_data)
        self._add_attendance(story, styles, attendance_data)
        self._add_rescheduled_sessions(story, styles, rescheduled_data)
        
        return story
    
    def _add_basic_info(self, story, styles, employee_data):
        heading_style = self._get_heading_style(styles)
        
        story.append(Paragraph("Basic Information", heading_style))
        story.append(Spacer(1, 8))
        
        basic_data = [
            ["Field", "Value"],
            ["Name", " ".join(filter(None, [
                employee_data.get('first_name'),
                employee_data.get('last_name')
            ])) or "N/A"],
            ["Card No", employee_data.get('emp_id', 'N/A')],
            ["Department", employee_data.get('department', 'N/A')],
            ["Joining Date", self._format_date(employee_data.get('date_of_joining'))],
            ["Birth Date", self._format_date(employee_data.get('birth_date'))],
            ["Gender", employee_data.get('sex', 'N/A')],
            ["Email", employee_data.get('email', 'N/A')],
            ["Phone", employee_data.get('phone', 'N/A')],
        ]
        # basic_data = [
        #     ["Field", "Value"],
        #     ["Name", f"{employee_data.get('first_name', 'N/A')} {employee_data.get('last_name', '')}"],
        #     ["Card No", employee_data.get('emp_id', 'N/A')],
        #     ["Department", employee_data.get('department', 'N/A')],
        #     ["Joining Date", self._format_date(employee_data.get('date_of_joining'))],
        #     ["Birth Date", self._format_date(employee_data.get('birth_date'))],
        #     ["Gender", employee_data.get('sex', 'N/A')],
        #     ["Email", employee_data.get('email', 'N/A')],
        #     ["Phone", employee_data.get('phone', 'N/A')],
        # ]
        
        basic_table = Table(basic_data, colWidths=[2*inch, 4.5*inch])
        basic_table.setStyle(self._get_table_style())
        story.append(basic_table)
        story.append(Spacer(1, 24))
    
    def _add_operator_skills(self, story, styles, skills_data):
        heading_style = self._get_heading_style(styles)
        
        story.append(Paragraph("Operator Skills", heading_style))
        story.append(Spacer(1, 8))
        
        if not skills_data:
            story.append(Paragraph("<i>No operator skills recorded</i>", styles['Normal']))
            story.append(Spacer(1, 24))
            return
        
        skill_data = [["Level", "Station", "Date"]]
        for skill in skills_data:
            skill_data.append([
                skill.get('level_name', 'N/A'),
                skill.get('station_name', 'N/A'),
                self._format_date(skill.get('updated_at'))
            ])
        
        skills_table = Table(skill_data, colWidths=[1.5*inch, 3.5*inch, 1.5*inch])
        skills_style = self._get_table_style()
        skills_style.add('ALIGN', (0,1), (0,-1), 'CENTER')
        skills_style.add('ALIGN', (2,1), (2,-1), 'CENTER')
        skills_table.setStyle(skills_style)
        
        story.append(skills_table)
        story.append(Spacer(1, 24))
    
    def _add_multi_skills(self, story, styles, multi_skilling_data):
        heading_style = self._get_heading_style(styles)
        
        story.append(Paragraph("Multi-Skilling Training", heading_style))
        story.append(Spacer(1, 8))
        
        if not multi_skilling_data:
            story.append(Paragraph("<i>No multi-skilling records found</i>", styles['Normal']))
            story.append(Spacer(1, 24))
            return
        
        multi_data = [["Training Name", "Status", "Date"]]
        for ms in multi_skilling_data:
            multi_data.append([
                ms.get('training_name', 'N/A'),
                ms.get('status', 'N/A').title(),
                self._format_date(ms.get('date'))
            ])
        
        multi_table = Table(multi_data, colWidths=[3*inch, 1.5*inch, 2*inch])
        multi_style = self._get_table_style()
        multi_style.add('ALIGN', (1,1), (2,-1), 'CENTER')
        
        for row in range(1, len(multi_data)):
            status = multi_skilling_data[row-1].get('status', '')
            for style_command in self._get_status_badge_style(status, row, col=1):
                multi_style.add(*style_command)
        
        multi_table.setStyle(multi_style)
        story.append(multi_table)
        story.append(Spacer(1, 24))
    
    def _add_scores(self, story, styles, scores_data):
        heading_style = self._get_heading_style(styles)
        
        story.append(Paragraph("Scores and Assessments", heading_style))
        story.append(Spacer(1, 8))
        
        if not scores_data:
            story.append(Paragraph("<i>No assessment scores available</i>", styles['Normal']))
            story.append(Spacer(1, 24))
            return
        
        score_data = [["Test", "Marks", "Percentage", "Result", "Date"]]
        for score in scores_data:
            score_data.append([
                score.get('test_name', 'N/A'),
                str(score.get('marks', 'N/A')),
                f"{score.get('percentage', 'N/A')}%" if score.get('percentage') is not None else "N/A",
                "Pass" if score.get('passed') else "Fail",
                self._format_date(score.get('created_at'))
            ])
        
        scores_table = Table(score_data, colWidths=[2.3*inch, 0.7*inch, 1*inch, 0.7*inch, 1.8*inch])
        scores_style = self._get_table_style()
        scores_style.add('ALIGN', (1,1), (2,-1), 'CENTER')
        scores_style.add('ALIGN', (3,1), (3,-1), 'CENTER')
        scores_style.add('ALIGN', (4,1), (4,-1), 'CENTER')
        
        for row in range(1, len(score_data)):
            result = scores_data[row-1].get('passed')
            for style_command in self._get_result_badge_style(result, row):
                scores_style.add(*style_command)
        
        scores_table.setStyle(scores_style)
        story.append(scores_table)
        story.append(Spacer(1, 24))
    
    def _add_hanchou_shokuchou_results(self, story, styles, hanchou_results, shokuchou_results):
        heading_style = self._get_heading_style(styles)
        
        story.append(Paragraph("Hanchou & Shokuchou Exam Results", heading_style))
        story.append(Spacer(1, 8))
        
        if not hanchou_results and not shokuchou_results:
            story.append(Paragraph("<i>No exam results found</i>", styles['Normal']))
            story.append(Spacer(1, 24))
            return
        
        # Hanchou Results
        if hanchou_results:
            story.append(Paragraph("<b>Hanchou Exams:</b>", styles['Normal']))
            story.append(Spacer(1, 4))
            
            hanchou_data = [["Exam Type", "Score", "Result", "Date"]]
            for result in hanchou_results:
                hanchou_data.append([
                    result.get('exam_type', 'N/A'),
                    str(result.get('score', 'N/A')),
                    result.get('result', 'N/A').title(),
                    self._format_date(result.get('exam_date'))
                ])
            
            hanchou_table = Table(hanchou_data, colWidths=[2*inch, 1*inch, 1.5*inch, 2*inch])
            hanchou_style = self._get_table_style()
            hanchou_style.add('ALIGN', (1,1), (3,-1), 'CENTER')
            
            for row in range(1, len(hanchou_data)):
                result = hanchou_results[row-1].get('result', '').lower()
                for style_command in self._get_status_badge_style(result, row, col=2):
                    hanchou_style.add(*style_command)
            
            hanchou_table.setStyle(hanchou_style)
            story.append(hanchou_table)
            story.append(Spacer(1, 12))
        
        # Shokuchou Results
        if shokuchou_results:
            story.append(Paragraph("<b>Shokuchou Exams:</b>", styles['Normal']))
            story.append(Spacer(1, 4))
            
            shokuchou_data = [["Exam Type", "Score", "Result", "Date"]]
            for result in shokuchou_results:
                shokuchou_data.append([
                    result.get('exam_type', 'N/A'),
                    str(result.get('score', 'N/A')),
                    result.get('result', 'N/A').title(),
                    self._format_date(result.get('exam_date'))
                ])
            
            shokuchou_table = Table(shokuchou_data, colWidths=[2*inch, 1*inch, 1.5*inch, 2*inch])
            shokuchou_style = self._get_table_style()
            shokuchou_style.add('ALIGN', (1,1), (3,-1), 'CENTER')
            
            for row in range(1, len(shokuchou_data)):
                result = shokuchou_results[row-1].get('result', '').lower()
                for style_command in self._get_status_badge_style(result, row, col=2):
                    shokuchou_style.add(*style_command)
            
            shokuchou_table.setStyle(shokuchou_style)
            story.append(shokuchou_table)
        
        story.append(Spacer(1, 24))
    
    def _add_scheduled_training(self, story, styles, training_data):
        heading_style = self._get_heading_style(styles)
        
        story.append(Paragraph("Scheduled Trainings", heading_style))
        story.append(Spacer(1, 8))
        
        if not training_data:
            story.append(Paragraph("<i>No scheduled trainings found</i>", styles['Normal']))
            story.append(Spacer(1, 24))
            return
        
        training_table_data = [["Category", "Venue", "Date", "Status"]]
        for training in training_data:
            training_table_data.append([
                training.get('category', 'N/A'),
                training.get('venue', 'N/A'),
                self._format_date(training.get('date')),
                training.get('status', 'N/A').title()
            ])
        
        training_table = Table(training_table_data, colWidths=[2*inch, 2*inch, 1.5*inch, 1*inch])
        training_style = self._get_table_style()
        training_style.add('ALIGN', (2,1), (3,-1), 'CENTER')
        
        for row in range(1, len(training_table_data)):
            status = training_data[row-1].get('status', '')
            for style_command in self._get_status_badge_style(status, row, col=3):
                training_style.add(*style_command)
        
        training_table.setStyle(training_style)
        story.append(training_table)
        story.append(Spacer(1, 24))
    
    def _add_attendance(self, story, styles, attendance_data):
        heading_style = self._get_heading_style(styles)
        
        story.append(Paragraph("Training Attendance Records", heading_style))
        story.append(Spacer(1, 8))
        
        if not attendance_data:
            story.append(Paragraph("<i>No attendance records found</i>", styles['Normal']))
            story.append(Spacer(1, 24))
            return
        
        attendance_table_data = [["Batch", "Day", "Date", "Status"]]
        for att in attendance_data:
            attendance_table_data.append([
                att.get('batch_name', 'N/A'),
                f"Day {att.get('day_number', 'N/A')}",
                self._format_date(att.get('attendance_date')),
                att.get('status', 'N/A').title()
            ])
        
        attendance_table = Table(attendance_table_data, colWidths=[2*inch, 1*inch, 1.5*inch, 1*inch])
        attendance_style = self._get_table_style()
        attendance_style.add('ALIGN', (1,1), (3,-1), 'CENTER')
        
        for row in range(1, len(attendance_table_data)):
            status = attendance_data[row-1].get('status', '')
            for style_command in self._get_status_badge_style(status, row, col=3):
                attendance_style.add(*style_command)
        
        attendance_table.setStyle(attendance_style)
        story.append(attendance_table)
        story.append(Spacer(1, 24))
    
    def _add_rescheduled_sessions(self, story, styles, rescheduled_data):
        heading_style = self._get_heading_style(styles)
        
        story.append(Paragraph("Rescheduled Training Sessions", heading_style))
        story.append(Spacer(1, 8))
        
        if not rescheduled_data:
            story.append(Paragraph("<i>No rescheduled sessions found</i>", styles['Normal']))
            story.append(Spacer(1, 24))
            return
        
        rescheduled_table_data = [["Topic", "Original Date", "Rescheduled Date", "Time"]]
        for rs in rescheduled_data:
            rescheduled_table_data.append([
                rs.get('training_subtopic', 'N/A'),
                self._format_date(rs.get('original_date')),
                self._format_date(rs.get('rescheduled_date')),
                rs.get('rescheduled_time', 'N/A')
            ])
        
        rescheduled_table = Table(rescheduled_table_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1*inch])
        rescheduled_style = self._get_table_style()
        rescheduled_style.add('ALIGN', (1,1), (3,-1), 'CENTER')
        rescheduled_table.setStyle(rescheduled_style)
        
        story.append(rescheduled_table)
        story.append(Spacer(1, 24))


# =============== Emp History Card ====================
from rest_framework.response import Response
from rest_framework import status
from rest_framework.views import APIView
from django.db.models import Q  # ← THIS WAS MISSING - CAUSING 500 ERROR
from django.core.exceptions import ObjectDoesNotExist

from .models import (
    MasterTable,
    SkillMatrix,
    Score,
    MultiSkilling,
    RescheduledSession,
    UserRegistration,
    TrainingAttendance,
    Schedule,
    HanchouExamResult,
    ShokuchouExamResult,
)

from .serializers import (
    CardEmployeeMasterSerializer,
    OperatorCardSkillSerializer,
    CardScoreSerializer,
    CardMultiSkillingSerializer,
    CardHanchouExamResultSerializer,
    CardShokuchouExamResultSerializer,
    CardTrainingAttendanceSerializer,
    CardScheduleSerializer,
    CardRescheduledSessionSerializer,
)


# class EmployeeCardDetailsView(APIView):
#     def get(self, request):
#         card_no = request.query_params.get('card_no')
#         if not card_no:
#             return Response({'error': 'card_no parameter is required'}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             employee = MasterTable.objects.get(emp_id=card_no)
#         except MasterTable.DoesNotExist:
#             return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)
#         except Exception as e:
#             return Response({'error': f'Database error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#         # === Employee Basic Info ===
#         employee_data = CardEmployeeMasterSerializer(employee).data
#         employee_data['line_name'] = (
#         employee.sub_department.line_name if employee.sub_department else None
#         )
#         employee_data['station_name'] = (
#             employee.station.station_name if employee.station else None
#         )

#         # === Skills & Scores ===
#         operator_skills = OperatorCardSkillSerializer(SkillMatrix.objects.filter(employee=employee), many=True).data
#         scores = CardScoreSerializer(Score.objects.filter(employee=employee), many=True).data
#         multi_skilling = CardMultiSkillingSerializer(MultiSkilling.objects.filter(employee=employee), many=True).data

#         # === Exam Results ===
#         hanchou_results = CardHanchouExamResultSerializer(
#             HanchouExamResult.objects.filter(employee=employee), many=True
#         ).data
#         shokuchou_results = CardShokuchouExamResultSerializer(
#             ShokuchouExamResult.objects.filter(employee=employee), many=True
#         ).data

#         # === Scheduled Trainings ===
#         all_schedules = Schedule.objects.filter(employees=employee)
#         scheduled_trainings_data = CardScheduleSerializer(
#             all_schedules, many=True, context={'employee': employee}
#         ).data

#         # === TRAINING ATTENDANCE & RESCHEDULED SESSIONS ===
#         attendances_data = []
#         rescheduled_sessions_data = []

#         try:
#             # Build flexible matching query
#             query = Q()

#             # Priority 1: temp_id match
#             temp_id = getattr(employee, 'temp_id', None)
#             if temp_id and str(temp_id).strip():
#                 query |= Q(temp_id__iexact=str(temp_id).strip())

#             # Clean names
#             first_name_clean = (employee.first_name or '').strip()
#             last_name_clean = (employee.last_name or '').strip()
#             full_name = f"{first_name_clean} {last_name_clean}".strip()

#             if first_name_clean:
#                 # Strategy 1: Exact first + last name
#                 if last_name_clean:
#                     query |= (Q(first_name__iexact=first_name_clean) & Q(last_name__iexact=last_name_clean))

#                 # Strategy 2: Split compound first_name (Sneha Gautam → first=Sneha, last=Gautam)
#                 if ' ' in first_name_clean:
#                     name_parts = first_name_clean.split()
#                     if len(name_parts) > 1:
#                         possible_first = name_parts[0]
#                         possible_last = ' '.join(name_parts[1:])
#                         query |= (Q(first_name__iexact=possible_first) & Q(last_name__iexact=possible_last))

#                 # Strategy 3: First name partial match
#                 query |= Q(first_name__icontains=first_name_clean)
                
#                 # Strategy 4: Full name in first_name field
#                 query |= Q(first_name__iexact=full_name)

#             # Execute query safely
#             user_records = UserRegistration.objects.filter(query).distinct()[:5]  # Limit to 5 to prevent overload

#             if user_records.exists():
#                 matched_users = [
#                     f"{u.first_name} {u.last_name or ''} ({u.temp_id})" 
#                     for u in user_records
#                 ]
#                 print(f"✅ Matched {len(user_records)} UserRegistration records for {employee.emp_id} ({full_name}): {matched_users}")

#                 # Get attendance
#                 attendances = TrainingAttendance.objects.filter(
#                     user__in=user_records
#                 ).select_related('batch', 'user', 'day_number').order_by(
#                     '-attendance_date', 'batch__batch_id', 'day_number'
#                 )
#                 attendances_data = CardTrainingAttendanceSerializer(attendances, many=True).data
#                 print(f"✅ Found {len(attendances_data)} attendance records")

#                 # Get rescheduled sessions
#                 rescheduled_sessions = RescheduledSession.objects.filter(
#                     employee__in=user_records
#                 ).select_related(
#                     'employee', 'original_day', 'training_subtopic', 'batch'
#                 ).order_by('-rescheduled_date', '-rescheduled_time')
                
#                 rescheduled_sessions_data = CardRescheduledSessionSerializer(
#                     rescheduled_sessions, many=True
#                 ).data
#                 print(f"✅ Found {len(rescheduled_sessions_data)} rescheduled sessions")
#             else:
#                 print(f"❌ No UserRegistration found for {employee.emp_id} ({full_name}) | temp_id: {temp_id}")

#         except Exception as e:
#             print(f"❌ Attendance fetch error for {employee.emp_id}: {str(e)}")
#             # Don't fail the whole request - just return empty lists
#             attendances_data = []
#             rescheduled_sessions_data = []

#         # === New Data Sources ===
#         from .models import ProductivityEvaluation, QualityEvaluation, TraineeInfo
#         from .serializers import CardProductivitySerializer, CardQualitySerializer, CardOJTSerializer

#         try:
#             productivity_data = CardProductivitySerializer(
#                 ProductivityEvaluation.objects.filter(employee=employee).order_by('-evaluation_date'), 
#                 many=True
#             ).data
#         except Exception as e:
#             print(f"Error fetching productivity: {e}")
#             productivity_data = []

#         try:
#             quality_data = CardQualitySerializer(
#                 QualityEvaluation.objects.filter(employee=employee).order_by('-evaluation_date'), 
#                 many=True
#             ).data
#         except Exception as e:
#             print(f"Error fetching quality: {e}")
#             quality_data = []

#         try:
#             # TraineeInfo uses emp_id string, not FK to MasterTable
#             ojt_data = CardOJTSerializer(
#                 TraineeInfo.objects.filter(emp_id=employee.emp_id).order_by('-doj'), 
#                 many=True
#             ).data
#         except Exception as e:
#             print(f"Error fetching OJT: {e}")
#             ojt_data = []

#         # === Final Response ===
#         response_data = {
#             'employee': employee_data,
#             'operator_skills': operator_skills,
#             'scores': scores,
#             'multi_skilling': multi_skilling,
#             'scheduled_trainings': scheduled_trainings_data,
#             'hanchou_results': hanchou_results,
#             'shokuchou_results': shokuchou_results,
#             'attendance': attendances_data,
#             'rescheduled_sessions': rescheduled_sessions_data,
#             'productivity_evaluations': productivity_data,
#             'quality_evaluations': quality_data,
#             'ojt_evaluations': ojt_data,
#         }

#         return Response(response_data, status=status.HTTP_200_OK)

class EmployeeCardDetailsView(APIView):
    def get(self, request):
        card_no = request.query_params.get('card_no')
        if not card_no:
            return Response({'error': 'card_no parameter is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = MasterTable.objects.get(emp_id=card_no)
        except MasterTable.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': f'Database error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # === Employee Basic Info ===
        employee_data = CardEmployeeMasterSerializer(employee).data
        employee_data['line_name'] = (
            employee.sub_department.line_name if employee.sub_department else None
        )
        employee_data['station_name'] = (
            employee.station.station_name if employee.station else None
        )

        # === Skills & Scores ===
        operator_skills = OperatorCardSkillSerializer(SkillMatrix.objects.filter(employee=employee), many=True).data
        scores = CardScoreSerializer(Score.objects.filter(employee=employee), many=True).data
        multi_skilling = CardMultiSkillingSerializer(MultiSkilling.objects.filter(employee=employee), many=True).data

        # === Exam Results ===
        hanchou_results = CardHanchouExamResultSerializer(
            HanchouExamResult.objects.filter(employee=employee), many=True
        ).data
        shokuchou_results = CardShokuchouExamResultSerializer(
            ShokuchouExamResult.objects.filter(employee=employee), many=True
        ).data

        # === Scheduled Trainings ===
        all_schedules = Schedule.objects.filter(employees=employee)
        scheduled_trainings_data = CardScheduleSerializer(
            all_schedules, many=True, context={'employee': employee}
        ).data

        # === TRAINING ATTENDANCE & RESCHEDULED SESSIONS ===
        attendances_data = []
        rescheduled_sessions_data = []

        try:
            query = Q()

            first_name_clean = (employee.first_name or '').strip()
            last_name_clean = (employee.last_name or '').strip()
            emp_id = getattr(employee, 'emp_id', None)

            # Priority 1: Match by emp_id (most reliable)
            if emp_id and str(emp_id).strip():
                query |= Q(emp_id__iexact=str(emp_id).strip())

            # Priority 2: Match by temp_id
            temp_id = getattr(employee, 'temp_id', None)
            if temp_id and str(temp_id).strip():
                query |= Q(temp_id__iexact=str(temp_id).strip())

            # Priority 3: STRICT exact first + last name (both must be present)
            if first_name_clean and last_name_clean:
                query |= (
                    Q(first_name__iexact=first_name_clean) &
                    Q(last_name__iexact=last_name_clean)
                )
            elif first_name_clean and ' ' in first_name_clean:
                # Handle compound name stored entirely in first_name e.g. "Sneha Gautam"
                name_parts = first_name_clean.split(None, 1)
                query |= (
                    Q(first_name__iexact=name_parts[0]) &
                    Q(last_name__iexact=name_parts[1])
                )
            # ❌ REMOVED: icontains partial match — was causing false matches

            if not query:
                print(f"⚠️  No valid matching criteria for {emp_id}")
            else:
                user_records = UserRegistration.objects.filter(query).distinct()[:5]

                if user_records.exists():
                    matched_users = [
                    f"{' '.join(filter(None, [u.first_name, u.last_name]))} ({u.temp_id})"
                    for u in user_records
                ]
                    # matched_users = [
                    #     f"{u.first_name} {u.last_name or ''} ({u.temp_id})"
                    #     for u in user_records
                    # ]
                    print(f"✅ Matched {user_records.count()} UserRegistration record(s) for "
                          f"{emp_id} ({first_name_clean} {last_name_clean}): {matched_users}")

                    # ── Deduplication ──────────────────────────────────────────
                    # If multiple users matched AND we have an emp_id,
                    # prefer only the ones whose emp_id matches exactly.
                    # This prevents pulling attendance from a different person
                    # who happens to share the same name.
                    if user_records.count() > 1 and emp_id:
                        emp_id_matched = user_records.filter(
                            emp_id__iexact=str(emp_id).strip()
                        )
                        if emp_id_matched.exists():
                            user_records = emp_id_matched
                            print(f"🎯 Narrowed to {user_records.count()} record(s) via emp_id match")

                    # ── Attendance ─────────────────────────────────────────────
                    attendances = TrainingAttendance.objects.filter(
                        user__in=user_records
                    ).select_related(
                        'batch', 'user', 'day_number'
                    ).order_by(
                        '-attendance_date', 'batch__batch_id', 'day_number'
                    )
                    attendances_data = CardTrainingAttendanceSerializer(attendances, many=True).data
                    print(f"✅ Found {len(attendances_data)} attendance record(s)")

                    # ── Rescheduled Sessions ───────────────────────────────────
                    rescheduled_sessions = RescheduledSession.objects.filter(
                        employee__in=user_records
                    ).select_related(
                        'employee', 'original_day', 'training_subtopic', 'batch'
                    ).order_by('-rescheduled_date', '-rescheduled_time')

                    rescheduled_sessions_data = CardRescheduledSessionSerializer(
                        rescheduled_sessions, many=True
                    ).data
                    print(f"✅ Found {len(rescheduled_sessions_data)} rescheduled session(s)")

                else:
                    print(f"❌ No UserRegistration found for {emp_id} "
                          f"({first_name_clean} {last_name_clean}) | temp_id: {temp_id}")

        except Exception as e:
            print(f"❌ Attendance fetch error for {employee.emp_id}: {str(e)}")
            # Don't fail the whole request — return empty lists
            attendances_data = []
            rescheduled_sessions_data = []

        # === Productivity / Quality / OJT ===
        from .models import ProductivityEvaluation, QualityEvaluation, TraineeInfo
        from .serializers import CardProductivitySerializer, CardQualitySerializer, CardOJTSerializer

        try:
            productivity_data = CardProductivitySerializer(
                ProductivityEvaluation.objects.filter(employee=employee).order_by('-evaluation_date'),
                many=True
            ).data
        except Exception as e:
            print(f"❌ Error fetching productivity: {e}")
            productivity_data = []

        try:
            quality_data = CardQualitySerializer(
                QualityEvaluation.objects.filter(employee=employee).order_by('-evaluation_date'),
                many=True
            ).data
        except Exception as e:
            print(f"❌ Error fetching quality: {e}")
            quality_data = []

        try:
            ojt_data = CardOJTSerializer(
                TraineeInfo.objects.filter(emp_id=employee.emp_id).order_by('-doj'),
                many=True
            ).data
        except Exception as e:
            print(f"❌ Error fetching OJT: {e}")
            ojt_data = []

        # === Final Response ===
        return Response({
            'employee':             employee_data,
            'operator_skills':      operator_skills,
            'scores':               scores,
            'multi_skilling':       multi_skilling,
            'scheduled_trainings':  scheduled_trainings_data,
            'hanchou_results':      hanchou_results,
            'shokuchou_results':    shokuchou_results,
            'attendance':           attendances_data,
            'rescheduled_sessions': rescheduled_sessions_data,
            'productivity_evaluations': productivity_data,
            'quality_evaluations':  quality_data,
            'ojt_evaluations':      ojt_data,
        }, status=status.HTTP_200_OK)

    
from datetime import datetime as dt
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from django.http import HttpResponse
from rest_framework import viewsets
from rest_framework.decorators import action

# Ensure you import your models and serializers correctly
# from .models import MasterTable
# from .serializers import MasterTableSerializer

class EmployeeExcelViewSet(viewsets.ModelViewSet):
    # Added 'sub_department' to select_related for performance
    queryset = MasterTable.objects.all().select_related('department', 'sub_department')
    serializer_class = MasterTableSerializer
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        queryset = self.get_queryset()

        wb = Workbook()
        ws = wb.active
        ws.title = "EMPLOYEE MASTER"

        # ------------------------------------------------------------------ #
        # 1. Title Section (Updated merge range to A:L for all columns)
        # ------------------------------------------------------------------ #
        # Headers now span 12 columns (A to L) instead of 9
        max_col_letter = 'L' 

        ws.merge_cells(f'A1:{max_col_letter}1')
        company_cell = ws['A1']
        company_cell.value = "Company Name: Krishna Maruti Limited Penstone"
        company_cell.font = Font(bold=True, size=10)
        company_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f'A2:{max_col_letter}2')
        run_date_cell = ws['A2']
        run_date_cell.value = f"Run Date & Time: {dt.now().strftime('%Y%m%d_%H%M%S')}"
        run_date_cell.font = Font(size=10)
        run_date_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f'A4:{max_col_letter}4')
        title_cell = ws['A4']
        title_cell.value = "EMPLOYEE MASTER"
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # ------------------------------------------------------------------ #
        # 2. Table Headers
        # ------------------------------------------------------------------ #
        headers = [
            'Srl. No.',
            'Employee ID',
            'First Name',
            'Last Name',
            'Department',
            'Sub Department', # Added
            'Designation',    # Added
            'Date of Joining',
            'Birth Date',
            'Sex',
            'Email',
            'Phone',
        ]

        header_font = Font(bold=True, size=10)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            
        ws.row_dimensions[6].height = 30
        
        # ------------------------------------------------------------------ #
        # 3. Data Rows
        # ------------------------------------------------------------------ #
        for row_num, employee in enumerate(queryset, 7):
            # Format dates
            birth_date = employee.birth_date.strftime('%d/%m/%Y') if employee.birth_date else ''
            joining_date = employee.date_of_joining.strftime('%d/%m/%Y') if employee.date_of_joining else ''
            
            # Handle Foreign Keys safely
            department_name = employee.department.department_name if employee.department else ''
            
            # Assuming 'Line' model has a __str__ method or a specific name field. 
            # If it has a specific field like 'line_name', change below to: employee.sub_department.line_name
            sub_department_name = str(employee.sub_department) if employee.sub_department else ''

            row_data = [
                row_num - 6,           # Srl No
                employee.emp_id,
                employee.first_name,
                employee.last_name,
                department_name,
                sub_department_name,    # Added
                employee.designation,   # Added
                joining_date,
                birth_date,
                employee.get_sex_display(),
                employee.email,
                employee.phone,
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                # Center align ID, Dates, Sex. Left align names/emails.
                # Updated indices: ID(2), DOJ(8), DOB(9), Sex(10)
                is_center = col_num in [1, 2, 8, 9, 10]
                cell.alignment = Alignment(horizontal="center" if is_center else "left", vertical="center")
                cell.font = Font(size=9)
        
        # ------------------------------------------------------------------ #
        # 4. Column Widths
        # ------------------------------------------------------------------ #
        column_widths = {
            'A': 6,  # Srl
            'B': 15, # Emp ID
            'C': 15, # First Name
            'D': 15, # Last Name
            'E': 20, # Dept
            'F': 20, # Sub Dept (New)
            'G': 18, # Designation (New)
            'H': 15, # DOJ
            'I': 12, # DOB
            'J': 8,  # Sex
            'K': 25, # Email
            'L': 15, # Phone
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # ------------------------------------------------------------------ #
        # 5. Response
        # ------------------------------------------------------------------ #
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"EMPLOYEE_MASTER_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        
        return response
# ==================== MultiSkilling Start ======================== #
    
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, viewsets
from django.db.models import Q
from .models import MasterTable, SkillMatrix, MultiSkilling, Level
from .serializers import MultiSkillingSerializer

class EmployeeSkillSearch(APIView):
    """
    Search employee by emp_id or name and return their current skills.
    Only returns employees who have at least one skill in SkillMatrix.
    """
    def get(self, request):
        query = request.GET.get("query", "").strip()
        if not query:
            return Response({"error": "query parameter is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        # Find all employees who match the search criteria
        employees_query = MasterTable.objects.filter(
            Q(emp_id__icontains=query) | 
            Q(first_name__icontains=query) | 
            Q(last_name__icontains=query)
        )
        
        # Get employees who have any skill matrix entries
        skilled_employee_ids = SkillMatrix.objects.filter(
            employee__in=employees_query
        ).values_list('employee__emp_id', flat=True).distinct()  
        
        employees_with_skills = employees_query.filter(emp_id__in=skilled_employee_ids) 
        
        result = []
        for emp in employees_with_skills:
            # Get all skills for this employee
            skills = SkillMatrix.objects.filter(employee=emp).select_related(
                "hierarchy__station", 
                "hierarchy__department", 
                "level"
            )
            
            skills_data = []
            for s in skills:
                skills_data.append({
                    "skill_id": s.id,
                    "station": s.hierarchy.station.station_name if s.hierarchy and s.hierarchy.station else None,
                    "department": s.hierarchy.department.department_name if s.hierarchy and s.hierarchy.department else None,
                    "level": s.level.level_name,
                    "updated_at": s.updated_at,
                })
            
            result.append({
                "emp_id": emp.emp_id,
                "first_name": emp.first_name,
                "last_name": emp.last_name,
                "department": emp.department.department_name if emp.department else None,
                "date_of_joining": emp.date_of_joining,
                "skills": skills_data,
            })
        
        return Response(result, status=status.HTTP_200_OK)




class MultiSkillingViewSet(viewsets.ModelViewSet):
    queryset = MultiSkilling.objects.all().select_related( 
       "employee",
       "department",
       "station",
       "skill_level"
    )
    serializer_class = MultiSkillingSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        emp_id = self.request.query_params.get("emp_id")
        if emp_id:
            qs = qs.filter(employee__emp_id=emp_id)
        return qs

from django.http import HttpResponse, JsonResponse
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import json
import traceback
from collections import defaultdict
from .models import (
    MasterTable, SkillMatrix, Station, SubLine, 
    HierarchyStructure, Department, Line, StationManager
)
from datetime import datetime

@method_decorator(csrf_exempt, name='dispatch')
class SkillMatrixExcelView(View):
    def post(self, request, *args, **kwargs):
        """
        Generate Skill Matrix Excel Report with actual data
        Accepts filters for department, main_line, sub_line
        """
        try:
            print("\n=== Received Skill Matrix Excel generation request ===")
            
            # 1. Parse input data
            filters = self._get_filters(request)
            print(f"Processing filters: {filters}")

            # 2. Get filtered data
            skill_matrix_data = self._get_skill_matrix_data(filters)
            
            # Check if we have data (Operators OR Stations)
            # Even if no operators, we might want to download the template with just headers
            if not skill_matrix_data['stations']:
                return JsonResponse({'error': 'No stations found for the given criteria'}, status=404)

            # 3. Generate Excel content
            print("Generating Skill Matrix Excel content...")
            buffer = BytesIO()
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Skill Matrix Report"
            
            self.create_skill_matrix_content(ws, skill_matrix_data, filters)
            
            # 4. Save to buffer
            print("Building Skill Matrix Excel document...")
            wb.save(buffer)
            buffer.seek(0)
            print("Skill Matrix Excel generation completed successfully")

            # 5. Return Excel response
            response = HttpResponse(
                buffer.getvalue(), 
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # Create filename based on filters
            filename = self._generate_filename(filters)
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            return response
            
        except Exception as e:
            print("\n!!! Skill Matrix Excel generation failed !!!")
            traceback.print_exc()
            return JsonResponse(
                {
                    'error': 'Internal server error',
                    'detail': str(e),
                    'traceback': traceback.format_exc()
                }, 
                status=500
            )

    def _get_filters(self, request):
        """Helper method to extract and validate filters from request"""
        data = {}
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                pass
        else:
            data = request.POST

        # FIX: Handle parameter mapping from Frontend (main_line_id) to Backend (line_id)
        filters = {
            'department_id': data.get('department_id'),
            'line_id': data.get('line_id') or data.get('main_line_id'), # Handle frontend specific key
            'sub_line_id': data.get('sub_line_id'),
            'station_requirements': data.get('station_requirements', []),
        }
        
        # Convert filter values to integers and validate
        for key in ['department_id', 'line_id', 'sub_line_id']:
            if filters.get(key):
                try:
                    # Handle if value is an object or string
                    value = filters[key]
                    if isinstance(value, dict):
                        # Attempt to extract ID if an object was passed erroneously
                        value = value.get('id') or value.get(key)
                    
                    if value is not None and str(value).strip() != '':
                        filters[key] = int(value)
                    else:
                        filters[key] = None
                except (ValueError, TypeError, AttributeError) as e:
                    print(f"Invalid {key}: {filters[key]}, Error: {str(e)}")
                    filters[key] = None
        
        return filters

    def _get_skill_matrix_data(self, filters):
        """Get organized skill matrix data based on filters"""
        
        # 1. Build query filters for HierarchyStructure
        # We must map input filter keys to HierarchyStructure model fields
        hierarchy_filters = {}
        
        # Note: In HierarchyStructure model, field is usually 'subline_id', not 'sub_line_id'
        if filters.get('sub_line_id'):
            hierarchy_filters['subline_id'] = filters['sub_line_id']
        elif filters.get('line_id'):
            hierarchy_filters['line_id'] = filters['line_id']
        elif filters.get('department_id'):
            hierarchy_filters['department_id'] = filters['department_id']

        print(f"Applying hierarchy_filters: {hierarchy_filters}")

        # 2. Get Hierarchies with Stations
        hierarchies = HierarchyStructure.objects.filter(
            **hierarchy_filters,
            station__isnull=False
        ).select_related(
            'station', 'department', 'line', 'subline'
        ).order_by('station__station_name')

        if not hierarchies.exists():
            return {'operators': [], 'stations': [], 'matrix': {}, 'station_managers': {}, 'hierarchy_info': {}}

        # 3. Extract Unique Stations (Columns)
        stations = []
        seen_station_ids = set()
        
        # We need to maintain order, but ensure uniqueness
        for h in hierarchies:
            if h.station and h.station_id and h.station_id not in seen_station_ids:
                stations.append(h.station)
                seen_station_ids.add(h.station_id)

        # 4. Fetch Skill Matrices (Data)
        # We only want skill matrices that belong to the filtered hierarchies
        skill_matrices = SkillMatrix.objects.filter(
            hierarchy__in=hierarchies
        ).select_related(
            'employee', 
            'hierarchy__station', 
            'level'
        ).order_by('employee__first_name')
        
        # 5. Filter Operators (Rows)
        # To match Frontend logic: Ensure we only include operators relevant to this view
        # The skill_matrices query already filters by hierarchy, so this restricts data to the view.
        operators_map = {}
        matrix = defaultdict(dict)
        
        for skill in skill_matrices:
            # OPTIONAL: If you want to strictly match Frontend "level > 0" logic
            # check skill level here. Assuming 0 or 'Level 0' exists.
            # For Excel reports, usually seeing "Level 0" is preferred over hiding the row.
            
            operators_map[skill.employee.emp_id] = skill.employee
            
            station = skill.hierarchy.station
            if station:
                # Handle Level Name safely
                level_str = "N/A"
                if skill.level:
                    if hasattr(skill.level, 'level_name'):
                        level_str = skill.level.level_name
                    elif hasattr(skill.level, 'level_number'):
                         level_str = f"Level {skill.level.level_number}"
                    else:
                        level_str = str(skill.level)

                matrix[skill.emp_id][station.station_name] = {
                    'skill_level': level_str,
                    'sequence': 0, # Add sequence logic if needed
                    'updated_at': skill.updated_at
                }

        operators = list(operators_map.values())
        operators.sort(key=lambda x: x.first_name if x.first_name else "")

        # 6. Station Managers (Requirements)
        station_ids = [s.station_id for s in stations]
        station_managers = StationManager.objects.filter(station_id__in=station_ids)
        station_manager_map = {sm.station_id: sm for sm in station_managers}

        # Merge with frontend provided requirements (if any)
        if filters.get('station_requirements'):
            for req in filters['station_requirements']:
                sid = req.get('station_id')
                if sid:
                    # Create a dummy object or update existing
                    if sid not in station_manager_map:
                        station_manager_map[sid] = type('StationManager', (), {})()
                        station_manager_map[sid].station_id = sid
                    
                    station_manager_map[sid].minimum_level_required = req.get('minimum_level_required', 'N/A')
                    station_manager_map[sid].minimum_operators = req.get('minimum_operators', 'N/A')

        return {
            'operators': operators,
            'stations': stations,
            'matrix': dict(matrix),
            'station_managers': station_manager_map,
            'hierarchy_info': self._get_hierarchy_info(filters)
        }

    def _get_hierarchy_info(self, filters):
        """Get hierarchy information for report header"""
        info = {}
        
        # Fetch names based on IDs
        if filters.get('department_id'):
            d = Department.objects.filter(department_id=filters['department_id']).first()
            if d: info['department'] = d.department_name
                
        if filters.get('line_id'):
            l = Line.objects.filter(line_id=filters['line_id']).first()
            if l: info['line'] = l.line_name
            # Fallback department fill
            if l and not info.get('department'):
                d = Department.objects.filter(department_id=l.department_id).first()
                if d: info['department'] = d.department_name
                
        if filters.get('sub_line_id'):
            sl = SubLine.objects.filter(subline_id=filters['sub_line_id']).first()
            if sl: info['sub_line'] = sl.subline_name
            # Fallback line fill
            if sl and not info.get('line'):
                l = Line.objects.filter(line_id=sl.line_id).first()
                if l: info['line'] = l.line_name
                
        return info

    def _generate_filename(self, filters):
        """Generate appropriate filename based on filters"""
        parts = ["skill_matrix"]
        
        if filters.get('sub_line_id'):
            parts.append("sub_line")
        elif filters.get('line_id'):
            parts.append("main_line")
        elif filters.get('department_id'):
            parts.append("department")
        else:
            parts.append("all")
            
        return f"{'_'.join(parts)}_report.xlsx"

    def _is_qualified(self, skill_level, min_required):
        """Determine if a skill level is considered qualified based on minimum required"""
        # Normalize inputs
        if not skill_level or not min_required or min_required == "N/A":
            return False

        # Level mapping
        level_map = {
            'Level 0': 0, 'L0': 0, 'Beginner': 0,
            'Level 1': 1, 'L1': 1, 'Learner': 1,
            'Level 2': 2, 'L2': 2, 'Practitioner': 2,
            'Level 3': 3, 'L3': 3, 'Expert': 3,
            'Level 4': 4, 'L4': 4, 'Master': 4
        }

        # Extract numeric values
        # Handle string "Level 1" vs int 1
        current_val = level_map.get(str(skill_level).strip(), 0)
        min_val = level_map.get(str(min_required).strip(), 0)
        
        # If mapping failed, try direct integer conversion
        if current_val == 0 and "Level" not in str(skill_level):
            try: current_val = int(skill_level)
            except: pass
            
        return current_val >= min_val

    def create_skill_matrix_content(self, ws, data, filters):
        """Generate the Excel content structure for skill matrix"""
        current_row = 1
        
        # Title with hierarchy information
        title_parts = ["Skill Matrix Report"]
        hierarchy = data['hierarchy_info']
        
        if hierarchy.get('department'):
            title_parts.append(f"Dept: {hierarchy['department']}")
        if hierarchy.get('line'):
            title_parts.append(f"Line: {hierarchy['line']}")
        if hierarchy.get('sub_line'):
            title_parts.append(f"Sub-Line: {hierarchy['sub_line']}")
            
        # Add title
        ws.merge_cells(f'A{current_row}:E{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = " - ".join(title_parts)
        title_cell.font = Font(size=16, bold=True, color='2E4D6B')
        title_cell.alignment = Alignment(horizontal='center')
        current_row += 2
        
        # Summary section
        current_row = self._add_summary_section(ws, data, current_row)
        current_row += 2
        
        # Main skill matrix table
        current_row = self._add_skill_matrix_table(ws, data, current_row)
        current_row += 2
        
        # Legend
        self._add_legend(ws, current_row)
        
        # Auto-adjust column widths
        self._adjust_column_widths_report(ws)

    def _add_summary_section(self, ws, data, start_row):
        """Add summary statistics section"""
        ws[f'A{start_row}'].value = "Summary Statistics"
        ws[f'A{start_row}'].font = Font(size=12, bold=True)
        start_row += 1
        
        total_operators = len(data['operators'])
        total_stations = len(data['stations'])
        
        # Calculate qualifications
        level_counts = defaultdict(int)
        for operator in data['operators']:
            operator_skills = data['matrix'].get(operator.emp_id, {})
            for station_name, skill_info in operator_skills.items():
                level = skill_info['skill_level']
                # Normalize level string for counting
                if "1" in str(level): level_counts['L1'] += 1
                elif "2" in str(level): level_counts['L2'] += 1
                elif "3" in str(level): level_counts['L3'] += 1
                elif "4" in str(level): level_counts['L4'] += 1

        summary_data = [
            ["Metric", "Count"],
            ["Total Operators", total_operators],
            ["Total Stations", total_stations],
            ["Level 1 (Learner)", level_counts['L1']],
            ["Level 2 (Practitioner)", level_counts['L2']],
            ["Level 3 (Expert)", level_counts['L3']],
            ["Level 4 (Master)", level_counts['L4']],
        ]
        
        for i, row_data in enumerate(summary_data):
            for j, value in enumerate(row_data):
                cell = ws.cell(row=start_row + i, column=j + 1, value=value)
                if i == 0:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='4682B4', end_color='4682B4', fill_type='solid')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        return start_row + len(summary_data)

    def _add_skill_matrix_table(self, ws, data, start_row):
        """Add the main skill matrix table"""
        ws[f'A{start_row}'].value = "Operator Skill Matrix"
        ws[f'A{start_row}'].font = Font(size=12, bold=True)
        start_row += 1
        
        operators = data['operators']
        stations = data['stations']
        matrix = data['matrix']
        station_managers = data['station_managers']
        
        # Headers
        headers = ["S.No", "Operator Name", "Employee ID", "DOJ"]
        for station in stations:
            headers.append(station.station_name)
        
        # Render Headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4682B4', end_color='4682B4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Minimum Requirements Rows
        req_row = start_row + 1
        ws.cell(row=req_row, column=1, value="Required Level").font = Font(bold=True)
        
        for col, station in enumerate(stations, 5):
            sm = station_managers.get(station.station_id)
            val = sm.minimum_level_required if sm and hasattr(sm, 'minimum_level_required') else "N/A"
            cell = ws.cell(row=req_row, column=col, value=val)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

        # Data Rows
        data_start_row = req_row + 1
        for idx, operator in enumerate(operators, 1):
            row_num = data_start_row + idx - 1
            
            doj = operator.date_of_joining.strftime('%d/%m/%Y') if operator.date_of_joining else "N/A"
            
            # Static Columns
            ws.cell(row=row_num, column=1, value=idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            ws.cell(row=row_num, column=2, value=f"{operator.first_name} {operator.last_name}").border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            ws.cell(row=row_num, column=3, value=operator.emp_id).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            ws.cell(row=row_num, column=4, value=doj).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Station Columns
            op_skills = matrix.get(operator.emp_id, {})
            
            for col_idx, station in enumerate(stations, 5):
                skill_data = op_skills.get(station.station_name)
                
                cell = ws.cell(row=row_num, column=col_idx)
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                if skill_data:
                    # level_text = str(skill_data['skill_level']).replace('Level ', 'L')
                    level_text = str(skill_data['skill_level']) 
                    # Check Qualification
                    sm = station_managers.get(station.station_id)
                    min_req = sm.minimum_level_required if sm and hasattr(sm, 'minimum_level_required') else "N/A"
                    is_qualified = self._is_qualified(skill_data['skill_level'], min_req)
                    
                    cell.value = level_text
                    
                    # Conditional Formatting
                    if "1" in level_text: cell.font = Font(color='FF0000', bold=True) # Red
                    elif "2" in level_text: cell.font = Font(color='FFA500', bold=True) # Orange
                    elif "3" in level_text: cell.font = Font(color='0000FF', bold=True) # Blue
                    elif "4" in level_text: cell.font = Font(color='008000', bold=True) # Green
                    
                    # Highlight if not qualified but assigned
                    if not is_qualified and min_req != "N/A":
                        cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                else:
                    cell.value = "-"
                    cell.font = Font(color='CCCCCC')

        return data_start_row + len(operators)

    def _add_legend(self, ws, start_row):
        """Add legend for skill levels"""
        ws[f'A{start_row}'].value = "Legend"
        ws[f'A{start_row}'].font = Font(size=12, bold=True)
        start_row += 1
        
        legend_data = [
            ["Code", "Description", "Color"],
            ["Level 1", "Learner", "Red"],
            ["Level 2", "Practitioner", "Orange"],
            ["Level 3", "Expert", "Blue"],
            ["Level 4", "Master", "Green"],
        ]
        
        for i, row_data in enumerate(legend_data):
            for j, value in enumerate(row_data):
                cell = ws.cell(row=start_row + i, column=j + 1, value=value)
                if i == 0:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                
                if j == 0 and i > 0:
                    if "1" in value: cell.font = Font(color='FF0000', bold=True)
                    elif "2" in value: cell.font = Font(color='FFA500', bold=True)
                    elif "3" in value: cell.font = Font(color='0000FF', bold=True)
                    elif "4" in value: cell.font = Font(color='008000', bold=True)
                
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def _adjust_column_widths_report(self, ws):
        """Auto-adjust column widths"""
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 25) # Cap width


from .models import Machine,MachineAllocation
from .serializers import MachineSerializer,MachineAllocationSerializer

class MachineViewSet(viewsets.ModelViewSet):
    queryset = Machine.objects.all().order_by("id")
    serializer_class = MachineSerializer
    @action(detail=False, methods=['get'])
    def by_department(self, request):
        """Get machines filtered by department"""
        department_id = request.query_params.get('department_id')
        if department_id:
            machines = self.queryset.filter(department_id=department_id)
            serializer = self.get_serializer(machines, many=True)
            return Response(serializer.data)
        return Response({"error": "department_id parameter required"}, 
                       status=status.HTTP_400_BAD_REQUEST)

class MachineAllocationViewSet(viewsets.ModelViewSet):
    queryset = MachineAllocation.objects.all().order_by("-allocated_at")
    serializer_class = MachineAllocationSerializer
    @action(detail=False, methods=['get'])
    def eligible_employees(self, request):
        """Get employees eligible for a specific machine based on level"""
        machine_id = request.query_params.get('machine_id')
        department_id = request.query_params.get('department_id')
        
        if not machine_id:
            return Response({"error": "machine_id parameter required"},
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            machine = Machine.objects.get(id=machine_id)
        except Machine.DoesNotExist:
            return Response({"error": "Machine not found"},
                          status=status.HTTP_404_NOT_FOUND)
        
        # Get SkillMatrix entries for employees in the target department
        # Filter by hierarchy department if department_id is provided
        target_department = department_id or machine.department.id
        
        # Get already allocated employees for this machine
        allocated_employees = MachineAllocation.objects.filter(
            machine=machine
        ).values_list('employee_id', flat=True)
        
        # Get eligible employees from SkillMatrix
        eligible_employees_queryset = SkillMatrix.objects.filter(
            hierarchy__department_id=target_department
        ).exclude(id__in=allocated_employees)
        
        # If you want to filter by specific hierarchy/station, add this:
        # .filter(hierarchy__station=machine.process)  # if machine.process links to station
        
        serializer = EligibleEmployeeSerializer(
            eligible_employees_queryset,
            many=True,
            context={'machine_level': machine.level}
        )
        
        return Response({
            'machine_level': machine.level,
            'employees': serializer.data
        })

    @action(detail=False, methods=['post'])
    def update_pending_status(self, request):
        """Manually trigger update of pending allocations"""
        MachineAllocation.update_pending_allocations()
        return Response({"message": "Pending allocations updated successfully"})

    @action(detail=False, methods=['get'])
    def by_status(self, request):
        """Get allocations filtered by approval status"""
        status_filter = request.query_params.get('status')
        if status_filter:
            allocations = self.queryset.filter(approval_status=status_filter)
            serializer = self.get_serializer(allocations, many=True)
            return Response(serializer.data)
        return Response({"error": "status parameter required"}, 
                status=status.HTTP_400_BAD_REQUEST)

    def perform_update(self, serializer):
        """Override to add any additional logic during update"""
        allocation = serializer.save()
        
        # Log the allocation update if needed
        print(f"Updated allocation: {allocation}")
        
        return allocation
    def perform_create(self, serializer):
        """Override to add any additional logic during creation"""
        # Ensure we're working with SkillMatrix instance
        employee_id = self.request.data.get('employee')
        try:
            skill_matrix_employee = SkillMatrix.objects.get(id=employee_id)
        except SkillMatrix.DoesNotExist:
            raise ValidationError("Employee not found in skill matrix")
        
        allocation = serializer.save(employee=skill_matrix_employee)
        print(f"Created allocation: {allocation}")
        return allocation







class MachineAllocationApprovalViewSet(viewsets.ModelViewSet):
    queryset = MachineAllocation.objects.all()
    serializer_class = MachineAllocationApprovalSerializer

    @action(detail=True, methods=['put'], url_path='set-status')
    def set_status(self, request, pk=None):
        allocation = self.get_object()
        status_value = request.data.get('approval_status')

        if status_value not in dict(MachineAllocation.APPROVAL_STATUS_CHOICES):
            return Response({'error': 'Invalid status'}, status=status.HTTP_400_BAD_REQUEST)

        allocation.approval_status = status_value
        allocation.save()
        return Response({
            'status': 'success',
            'id': allocation.id,
            'approval_status': allocation.approval_status
        })

    @action(detail=True, methods=['put'], url_path='reject')
    def reject(self, request, pk=None):
        allocation = self.get_object()
        allocation.approval_status = 'rejected'
        allocation.save()
        return Response({
            'status': 'rejected',
            'id': allocation.id,
            'approval_status': allocation.approval_status
        }, status=status.HTTP_200_OK)
    



from rest_framework import viewsets
from .models import EvaluationLevel2
from .serializers import EvaluationLevel2Serializer

class EvaluationLevel2ViewSet(viewsets.ModelViewSet):
    """
    ViewSet for EvaluationLevel2 with correct manual filtering.
    """
    queryset = EvaluationLevel2.objects.all().order_by('-created_at')
    serializer_class = EvaluationLevel2Serializer

    def get_queryset(self):
        """
        This method correctly filters by the provided query parameters.
        The fix is removing the unnecessary '__id' from ForeignKey lookups.
        """
        queryset = super().get_queryset() 
        
        employee_id = self.request.query_params.get('employee__emp_id', None)
        department_id = self.request.query_params.get('department', None)
        station = self.request.query_params.get('station_name', None)
        level_id = self.request.query_params.get('level', None)
        
        if employee_id is not None:
            queryset = queryset.filter(employee__emp_id=employee_id)
            
        if department_id is not None:
            # ✅ CORRECTED LINE: No '__id' needed for ForeignKey filtering by ID.
            queryset = queryset.filter(department=department_id)
            
        if station is not None:
            queryset = queryset.filter(station_name__iexact=station)
            
        if level_id is not None:
            # ✅ CORRECTED LINE: No '__id' needed here either.
            queryset = queryset.filter(level=level_id)
            
        return queryset
    
from rest_framework import viewsets, permissions
from .models import EvaluationCriterion
from .serializers import EvaluationCriterionSerializer
from rest_framework import viewsets
from .models import EvaluationCriterion
from .serializers import EvaluationCriterionSerializer

# class EvaluationCriterionViewSet(viewsets.ModelViewSet):
#     serializer_class = EvaluationCriterionSerializer
    
#     # 1. Default queryset (Order by level then display order)
#     queryset = EvaluationCriterion.objects.all().order_by('level', 'display_order')

#     def get_queryset(self):
#         """
#         Modified to support both:
#         1. General Management List (show ALL, including inactive)
#         2. Filtered List (optional, via query param)
#         """
#         queryset = super().get_queryset()

#         # Check for 'level' in query params (e.g., /criteria/?level=1)
#         # instead of the URL kwargs. This makes it more flexible.
#         level_id = self.request.query_params.get('level')
#         if level_id:
#             queryset = queryset.filter(level_id=level_id)
        
#         # NOTE: We REMOVED 'is_active=True'. 
#         # The Admin/Management panel needs to see inactive items.
#         # If you need a public endpoint that only shows active ones, 
#         # you should create a separate view or use a filter param like ?active_only=true
        
#         return queryset
# class EvaluationCriterionViewSet(viewsets.ModelViewSet):
#     """
#     API endpoint that allows criteria to be viewed.
#     Provides a list of criteria filtered by the level ID in the URL.
#     e.g., /api/levels/2/criteria/
#     """
#     serializer_class = EvaluationCriterionSerializer
#     queryset = EvaluationCriterion.objects.all().order_by('level', 'display_order')

#     def get_queryset(self):
#         """
#         This view should return a list of all the criteria
#         for the level as determined by the level_pk portion of the URL.
#         """
#         level_id = self.kwargs.get('level_pk')
#         return EvaluationCriterion.objects.filter(level_id=level_id, is_active=True)



from rest_framework import viewsets
from .models import EvaluationCriterion
from .serializers import EvaluationCriterionSerializer

class EvaluationCriterionViewSet(viewsets.ModelViewSet):
    serializer_class = EvaluationCriterionSerializer
    # Base queryset with ordering
    queryset = EvaluationCriterion.objects.all().order_by('level', 'display_order')

    def get_queryset(self):
        """
        Dual-purpose queryset:
        1. If 'level_pk' is in URL (Nested Route): Return ACTIVE items for that level.
        2. If 'level' is in Query Params (Management): Return ALL items (active & inactive).
        """
        queryset = super().get_queryset()

        # --- SCENARIO 1: Nested Route (e.g., /api/levels/2/criteria/) ---
        # Used by: The Evaluation Form (Public User)
        # Logic: Strict filter by Level ID, ONLY return Active items.
        if 'level_pk' in self.kwargs:
            level_id = self.kwargs['level_pk']
            return queryset.filter(level_id=level_id, is_active=True)

        # --- SCENARIO 2: Management Route (e.g., /api/criteria/) ---
        # Used by: Your React Management Dashboard
        # Logic: Return EVERYTHING (including inactive) so admins can edit them.
        
        # Optional: Support ?level=X filtering for the management panel
        level_param = self.request.query_params.get('level')
        if level_param:
            queryset = queryset.filter(level_id=level_param)

        # Note: We do NOT filter is_active=True here, because 
        # the manager needs to see the inactive ones to reactivate them.
        
        return queryset

# views.py   for productivity and quality sheet Level 1

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from .models import ProductivityEvaluation, ProductivitySequence
from .serializers import ProductivityEvaluationSerializer, ProductivitySequenceSerializer

class ProductivityEvaluationViewSet(viewsets.ModelViewSet):
    queryset = ProductivityEvaluation.objects.all()
    serializer_class = ProductivityEvaluationSerializer

    
    def get_queryset(self):
        """Override to allow filtering by employee"""
        queryset = ProductivityEvaluation.objects.all()
        employee = self.request.query_params.get('employee', None)
        if employee is not None:
            queryset = queryset.filter(employee=employee)
        return queryset.order_by('-evaluation_date')  # Most recent first




    @action(detail=True, methods=["post"])
    def calculate_results(self, request, pk=None):
        try:
            evaluation = self.get_object()
            sequences = evaluation.sequences.all()
            
            # Calculate individual column totals
            e1_total = 0
            e2_total = 0
            e3_total = 0
            max_total = 0
            
            for seq in sequences:
                # Skip cycle time sequence (mt = 0)
                if 'cycle time' in seq.sequence_name.lower() and seq.mt == 0:
                    continue
                
                e1_total += seq.e1
                e2_total += seq.e2
                e3_total += seq.e3
                max_total += seq.mt
            
            # Determine final result based on best column performance
            final_marks = 0
            final_percentage = 0
            final_status = "FAIL"
            
            if e1_total >= 12:
                final_marks = e1_total
                final_percentage = (e1_total / 15) * 100
                final_status = "PASS"
            elif e2_total >= 12:
                final_marks = e2_total
                final_percentage = (e2_total / 15) * 100
                final_status = "PASS"
            elif e3_total >= 12:
                final_marks = e3_total
                final_percentage = (e3_total / 15) * 100
                final_status = "PASS"
            else:
                # If none passed, use the highest score
                final_marks = max(e1_total, e2_total, e3_total)
                final_percentage = (final_marks / 15) * 100
                final_status = "FAIL"
            
            evaluation.max_marks = 15
            evaluation.obtained_marks = final_marks
            evaluation.percentage = final_percentage
            evaluation.status = final_status
            evaluation.save()
            
            return Response({
                "employee": f"{evaluation.employee.first_name} {evaluation.employee.last_name}",
                "emp_id": evaluation.employee.emp_id,
                "designation": evaluation.employee.designation,
                "department": evaluation.employee.department.department_name if evaluation.employee.department else None,
                "date_of_joining": evaluation.employee.date_of_joining,
                "obtained_marks": evaluation.obtained_marks,
                "max_marks": evaluation.max_marks,
                "percentage": evaluation.percentage,
                "status": evaluation.status,
                "e1_total": e1_total,
                "e2_total": e2_total,
                "e3_total": e3_total
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        """Override create to handle updating existing evaluations"""
        employee_id = request.data.get('employee')
        
        existing_evaluation = ProductivityEvaluation.objects.filter(employee=employee_id).first()
        
        if existing_evaluation:
            serializer = self.get_serializer(existing_evaluation, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return super().create(request, *args, **kwargs)

class ProductivitySequenceViewSet(viewsets.ModelViewSet):
    queryset = ProductivitySequence.objects.all()
    serializer_class = ProductivitySequenceSerializer

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        """Override create to handle updating existing sequences"""
        evaluation_id = request.data.get('evaluation')
        sequence_name = request.data.get('sequence_name')
        
       
        existing_sequence = ProductivitySequence.objects.filter(
            evaluation=evaluation_id, 
            sequence_name=sequence_name
        ).first()
        
        if existing_sequence:
            
            serializer = self.get_serializer(existing_sequence, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return super().create(request, *args, **kwargs)


     





from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from .models import QualityEvaluation, QualitySequence
from .serializers import QualityEvaluationSerializer, QualitySequenceSerializer

class QualityEvaluationViewSet(viewsets.ModelViewSet):
    queryset = QualityEvaluation.objects.all()
    serializer_class = QualityEvaluationSerializer

    def get_queryset(self):
        """Override to allow filtering by employee"""
        queryset = QualityEvaluation.objects.all()
        employee = self.request.query_params.get('employee', None)
        if employee is not None:
            queryset = queryset.filter(employee=employee)
        return queryset.order_by('-evaluation_date')  # Most recent first

    @action(detail=True, methods=["post"])
    def calculate_results(self, request, pk=None):
        try:
            evaluation = self.get_object()
            sequences = evaluation.qualitysequences.all()
            
           
            e1_total = 0
            e2_total = 0
            e3_total = 0
            max_total = 0
            
            for seq in sequences:
                
                if 'cycle time' in seq.sequence_name.lower() and seq.mt == 0:
                    continue
                
                e1_total += seq.e1
                e2_total += seq.e2
                e3_total += seq.e3
                max_total += seq.mt
            
            # Determine final result based on best column performance
            final_marks = 0
            final_percentage = 0
            final_status = "FAIL"
            
            if e1_total >= 12:
                final_marks = e1_total
                final_percentage = (e1_total / 15) * 100
                final_status = "PASS"
            elif e2_total >= 12:
                final_marks = e2_total
                final_percentage = (e2_total / 15) * 100
                final_status = "PASS"
            elif e3_total >= 12:
                final_marks = e3_total
                final_percentage = (e3_total / 15) * 100
                final_status = "PASS"
            else:
                # If none passed, use the highest score
                final_marks = max(e1_total, e2_total, e3_total)
                final_percentage = (final_marks / 15) * 100
                final_status = "FAIL"
            
            evaluation.max_marks = 15
            evaluation.obtained_marks = final_marks
            evaluation.percentage = final_percentage
            evaluation.status = final_status
            evaluation.save()
            
            return Response({
                "employee": f"{evaluation.employee.first_name} {evaluation.employee.last_name}",
                "emp_id": evaluation.employee.emp_id,
                "designation": evaluation.employee.designation,
                "department": evaluation.employee.department.department_name if evaluation.employee.department else None,
                "date_of_joining": evaluation.employee.date_of_joining,
                "obtained_marks": evaluation.obtained_marks,
                "max_marks": evaluation.max_marks,
                "percentage": evaluation.percentage,
                "status": evaluation.status,
                "e1_total": e1_total,
                "e2_total": e2_total,
                "e3_total": e3_total
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        """Override create to handle updating existing evaluations"""
        employee_id = request.data.get('employee')
        
        # Check if evaluation already exists for this employee
        existing_evaluation = QualityEvaluation.objects.filter(employee=employee_id).first()
        
        if existing_evaluation:
            # Update existing evaluation
            serializer = self.get_serializer(existing_evaluation, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            # Create new evaluation
            return super().create(request, *args, **kwargs)

class QualitySequenceViewSet(viewsets.ModelViewSet):
    queryset = QualitySequence.objects.all()
    serializer_class = QualitySequenceSerializer

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        """Override create to handle updating existing sequences"""
        evaluation_id = request.data.get('evaluation')
        sequence_name = request.data.get('sequence_name')
        
        
        existing_sequence = QualitySequence.objects.filter(
            evaluation=evaluation_id, 
            sequence_name=sequence_name
        ).first()
        
        if existing_sequence:
           
            serializer = self.get_serializer(existing_sequence, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
           
            return super().create(request, *args, **kwargs)


# level1revision start
from rest_framework import viewsets
from .models import Question # Import new model
from .serializers import QuestionSerializer # Import new serializer

# ... your existing views ...

class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer
    
    # This allows filtering like /api/questions/?subtopiccontent=5
    def get_queryset(self):
        queryset = super().get_queryset()
        subtopiccontent_id = self.request.query_params.get('subtopiccontent')
        if subtopiccontent_id:
            queryset = queryset.filter(subtopiccontent_id=subtopiccontent_id)
        return queryset
# end

# operatorobservance sheet 

from rest_framework import viewsets
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .models import Topic, OperatorObservanceSheet
from .serializers import TopicSerializer, OperatorObservanceSheetSerializer

class TopicViewSet(viewsets.ModelViewSet):
    queryset = Topic.objects.all().order_by('sr_no')
    serializer_class = TopicSerializer
    lookup_field = 'sr_no'
class OperatorObservanceSheetViewSet(viewsets.ModelViewSet):
    queryset = OperatorObservanceSheet.objects.all()
    serializer_class = OperatorObservanceSheetSerializer

# @api_view(['GET'])
# def get_sheet_by_operator(request, operator_name):
#     try:
#         sheet = OperatorObservanceSheet.objects.filter(operator_name=operator_name).last()
#         if not sheet:
#             return Response({"detail": "Not found"}, status=status.HTTP_404_NOT_FOUND)
#         serializer = OperatorObservanceSheetSerializer(sheet)
#         return Response(serializer.data)
#     except Exception as e:
#         return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

# @api_view(['GET'])
# def get_sheet_by_operator_level_station(request, operator_name, level,process_name):
#     try:
#         sheet = OperatorObservanceSheet.objects.filter(
#             operator_name=operator_name,
#             level=level,
#             process_name=process_name
#         ).last()
#         if not sheet:
#             return Response({"detail": "Not found"}, status=status.HTTP_404_NOT_FOUND)
#         serializer = OperatorObservanceSheetSerializer(sheet)
#         return Response(serializer.data)
#     except Exception as e:
#         return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def get_sheet_by_operator_level_station(request, operator_name, level, station_name):
    try:
        operator_name = operator_name.strip()
        station_name = station_name.strip()
        level = level.strip()
        
        sheet = OperatorObservanceSheet.objects.filter(
            operator_name__iexact=operator_name, 
            level__iexact=level,
            process_name__iexact=station_name
        ).last()
        
        if not sheet:
            return Response({"detail": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = OperatorObservanceSheetSerializer(sheet)
        return Response(serializer.data)
    except Exception as e:
        return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)


from django.shortcuts import render

def dojo_app(request):
    return render (request,'index.html')

from django.views.generic import TemplateView

class IndexView(TemplateView):
    template_name = 'index.html'




# ====================== Start Evaluvation Full Views =========================================
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from django.db import transaction



# Import all the necessary models from your models.py file
from .models import TestSession, QuestionPaper, Station, MasterTable, Level

import logging
import traceback
from datetime import datetime
from django.db import transaction
from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from collections import defaultdict

from .models import TestSession, QuestionPaper, Station, MasterTable, Level, Department, Score

logger = logging.getLogger(__name__)

class StartTestSessionView(APIView):
    def post(self, request):
        try:
            print("Incoming Request Data:", request.data)

            # --- 1. Get data from the request ---
            test_name = request.data.get("test_name")
            assignments = request.data.get("assignments", [])
            question_paper_id = request.data.get("question_paper_id")
            level_id = request.data.get("level")
            skill_id = request.data.get("skill")
            test_date = request.data.get("test_date")

            # --- 2. Basic validation ---
            if not test_name or not assignments:
                response_data = {"error": "Test name and assignments are required."}
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

            # --- 3. Fetch related objects ---
            question_paper = None
            if question_paper_id:
                question_paper = get_object_or_404(QuestionPaper, question_paper_id=question_paper_id)

            level_obj = None
            if level_id:
                level_obj = get_object_or_404(Level, level_id=level_id)

            skill_obj = None
            if skill_id:
                skill_obj = get_object_or_404(Station, station_id=skill_id)

            created_sessions = []
            updated_sessions = []

            # --- 4. Use transaction so all succeed or all rollback ---
            with transaction.atomic():
                for item in assignments:
                    raw_key_id = item.get("key_id") # e.g., "2"
                    employee_id = item.get("employee_id")

                    if not raw_key_id or not employee_id:
                        raise ValueError("key_id and employee_id are required in each assignment.")

                    # FIX 1: Prefix the ID. ResultExplorer searches for 'remote-'
                    formatted_key_id = f"remote-{raw_key_id}"

                    employee = get_object_or_404(MasterTable, emp_id=employee_id)

                    # FIX 2: Look up by BOTH test_name AND key_id.
                    # This allows "Remote 2" to exist in "Test A" and "Test B" separately.
                    session, created = TestSession.objects.update_or_create(
                        key_id=formatted_key_id,
                        test_name=test_name,  # <--- MOVED inside lookup args
                        defaults={
                            "employee": employee,
                            "level": level_obj,
                            "skill": skill_obj,
                            "question_paper": question_paper,
                            "department": employee.department,
                            "test_date": test_date
                        }
                    )

                    if created:
                        created_sessions.append(formatted_key_id)
                    else:
                        updated_sessions.append(formatted_key_id)

            response_data = {
                "status": "ok",
                "message": f"Test '{test_name}' configured successfully.",
                "created_sessions": created_sessions,
                "updated_sessions": updated_sessions,
            }
            print("Response:", response_data)
            return Response(response_data, status=status.HTTP_201_CREATED)
        except Exception as e:
            traceback.print_exc()
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class NextTestSequenceView(APIView):
    def get(self, request):
        paper_id = request.query_params.get("paper_id")
        date_str = request.query_params.get("date")
        if not paper_id or not date_str:
            return Response({"error": "paper_id and date are required"}, status=400)
        
        # Count sessions/scores for this paper and date prefix
        # This is a bit heuristic, but it works for "Paper - Date - Counter"
        # Better: Search by paper and date specifically if we have fields
        count = TestSession.objects.filter(
            question_paper_id=paper_id,
            test_date=date_str
        ).values('test_name').distinct().count()
        
        # Also check Score for historical ones where TestSession might be gone
        score_count = Score.objects.filter(
            test__question_paper_id=paper_id,
            test_date=date_str
        ).values('test__test_name').distinct().count()
        
        next_val = max(count, score_count) + 1
        return Response({"next_sequence": next_val})

import logging
import traceback
from django.db import transaction
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import TestSession, Score

# logger = logging.getLogger(__name__)
# EvaluationPassingCriteria

from decimal import Decimal
import traceback
import pandas as pd
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status


from .models import TestSession, Score, QuestionPaper, MasterTable, TraineeInfo, EvaluationPassingCriteria, Station
from .serializers import ScoreSerializer, SimpleScoreSerializer, TestSessionSerializer
from django.core.cache import cache
from decimal import Decimal
import traceback
import pandas as pd
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status


import logging
import traceback
from django.db import transaction
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import TestSession, Score

class ScoreListView(APIView):
    def get(self, request):
        # Assuming you use caching for latest test session
        session_key = cache.get("latest_test_session")
        if not session_key:
            return Response([])

        scores = Score.objects.filter(test__key_id=session_key).select_related('employee', 'level', 'skill')
        serializer = ScoreSerializer(scores, many=True)
        return Response(serializer.data)




class KeyIdToEmployeeNameMap(APIView):
    def get(self, request):
        mapping = TestSession.objects.select_related('employee').all()
        return Response({
            s.key_id: f"{s.employee.first_name} {s.employee.last_name}" 
            for s in mapping
        })


# class PastTestSessionsView(APIView):
#     def get(self, request):
#         # Get distinct test names from Score model through the test relationship
#         test_names = Score.objects.select_related('test').filter(
#             test__isnull=False
#         ).values_list('test__test_name', flat=True).distinct()
        
#         return Response(list(test_names))
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import TestSession
import logging

logger = logging.getLogger(__name__)

class PastTestSessionsView(APIView):
    def get(self, request):
        try:
            # Fetch distinct test sessions with test_name and created_at, sorted by created_at (descending)
            test_sessions = TestSession.objects.filter(
                test_name__isnull=False
            ).values('test_name', 'created_at').distinct().order_by('-created_at')
            
            # Format response with test_name and formatted date
            response_data = [
                {
                    'test_name': session['test_name'],
                    'created_at': session['created_at'].strftime('%Y-%m-%d %H:%M:%S')
                }
                for session in test_sessions
            ]
            
            logger.info("Returning %d test sessions", len(response_data))
            return Response(response_data)
        except Exception as e:
            logger.error("Error in PastTestSessionsView: %s", str(e))
            return Response({'error': str(e)}, status=500)

import traceback
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import Score
from .serializers import ScoreSerializer  # Import your serializer

class ScoresByTestView(APIView):
    def get(self, request, name):
        try:
            # Filter scores and pre-fetch related objects for efficiency
            scores = (
                Score.objects
                .filter(test__test_name=name)
                .select_related(
                    'employee', 'employee__department',
                    'department',
                    'skill', 'skill__subline', 'skill__subline__line', 'skill__subline__line__department',
                    'level', 'test', 'test__department'
                )
            )

            # Use the serializer to handle data, including get_department logic
            serializer = ScoreSerializer(scores, many=True)

            # Debug: Print the serialized data (this will trigger the serializer's prints for department sources)
            print("Final data being sent to frontend:", serializer.data)

            return Response(serializer.data)

        except Exception as e:
            print(f"Error in ScoresByTestView: {e}")
            traceback.print_exc()
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class SkillListView(APIView):
    def get(self, request):
        skills = Station.objects.values_list('station_name', flat=True).distinct()
        return Response(list(skills))


class ResultSummaryAPIView(APIView):
    def get(self, request):
        try:
            scores = Score.objects.select_related('employee', 'level', 'skill')
            data = []
            for score in scores:
                # Use the percentage from the score model instead of recalculating
                percentage = score.percentage
                result = 'Pass' if score.percentage >= 80 else 'Retraining' if score.percentage >= 50 else 'Fail'

                data.append({
                    "employee_id": score.employee.emp_id if hasattr(score.employee, 'emp_id') else score.employee.id,
                    "name": f"{score.employee.first_name} {score.employee.last_name}",
                    "marks": score.marks,
                    "percentage": percentage,
                    "section": score.employee.section if hasattr(score.employee, 'section') else '',
                    "level_name": score.level.level_name if score.level and hasattr(score.level, 'level_name') else '',
                    "skill": score.skill.station_name if score.skill and hasattr(score.skill, 'station_name') else (score.skill.skill if score.skill else ''),
                    "result": result,
                })

            serializer = SimpleScoreSerializer(data, many=True)
            return Response(serializer.data)
            
        except Exception as e:
            print(f"Error in ResultSummaryAPIView: {e}")
            traceback.print_exc()
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

class ResultSummaryAPIView(APIView):
    def get(self, request):
        scores = Score.objects.select_related('employee', 'level', 'skill')
        data = []
        for score in scores:
            percentage = round((score.marks / 10) * 100, 2)  # Adjust total marks accordingly
            result = 'Pass' if score.marks >= 8 else 'Retraining' if score.marks >= 5 else 'Fail'

            data.append({
                "employee_id": score.employee.id,
                "name": score.employee.name,
                "marks": score.marks,
                "percentage": percentage,
                "section": score.employee.section,  # assuming CharField
                "level_name": score.level.name if score.level else '',
                "skill": score.skill.skill if score.skill else '',  # Station.skill string
                "result": result,
            })

        serializer = SimpleScoreSerializer(data, many=True)
        return Response(serializer.data)


# ===== IMPORTS (at top of views.py, if not already present) =====
from django.db.models import Count, Avg, Q, Min
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import Score, Level, Department, Station, TestSession, QuestionPaper
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


# ============================================================
#  A. LEVELS WITH RESULTS (MODE-AWARE)
#  GET /api/results/levels/?mode=remote|individual
# ============================================================
class ResultsLevelsAPIView(APIView):
    """
    Returns the list of Levels that actually have scores, filtered by mode.
    mode = remote      -> only remote/group tests (test.key_id starts with "remote-")
    mode = individual  -> only web/individual tests (test.key_id starts with "W")
    mode omitted       -> all scores
    """

    def get(self, request):
        try:
            mode = request.query_params.get("mode")

            qs = Score.objects.select_related("level").exclude(level__isnull=True)

            if mode == "remote":
                qs = qs.filter(test__key_id__startswith="remote-")
            elif mode == "individual":
                qs = qs.filter(test__key_id__startswith="W")

            levels = (
                qs.values("level_id", "level__level_name")
                .distinct()
                .order_by("level_id")
            )

            data = [
                {
                    "level_id": row["level_id"],
                    "level_name": row["level__level_name"] or f"Level {row['level_id']}",
                }
                for row in levels
            ]
            return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error("Error in ResultsLevelsAPIView: %s", e, exc_info=True)
            return Response(
                {"error": "Failed to fetch levels with results"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


# ============================================================
#  B. DEPARTMENTS FOR A GIVEN LEVEL (MODE-AWARE)
#  GET /api/results/departments/?level_id=...&mode=...
# ============================================================
class ResultsDepartmentsAPIView(APIView):
    """
    Returns all departments that have scores for a given level, filtered by mode.
    """

    def get(self, request):
        level_id = request.query_params.get("level_id")
        mode = request.query_params.get("mode")

        if not level_id:
            return Response(
                {"error": "level_id query parameter is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            qs = Score.objects.select_related("department").filter(level_id=level_id).exclude(
                department__isnull=True
            )

            if mode == "remote":
                qs = qs.filter(test__key_id__startswith="remote-")
            elif mode == "individual":
                qs = qs.filter(test__key_id__startswith="W")

            departments = (
                qs.values("department_id", "department__department_name")
                .distinct()
                .order_by("department__department_name")
            )

            data = [
                {
                    "department_id": row["department_id"],
                    "department_name": row["department__department_name"] or "Unknown Department",
                }
                for row in departments
            ]
            return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error("Error in ResultsDepartmentsAPIView: %s", e, exc_info=True)
            return Response(
                {"error": "Failed to fetch departments with results"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


# ============================================================
#  C. STATIONS FOR (LEVEL, DEPARTMENT) (MODE-AWARE)
#  GET /api/results/stations/?level_id=...&department_id=...&mode=...
# ============================================================
class ResultsStationsAPIView(APIView):
    """
    Returns all stations (skills) that have scores for a given Level + Department + mode,
    with basic stats per station (participants, avg %, pass rate, batches_count).
    batches_count is only meaningful for remote mode, but we always compute it.
    """

    def get(self, request):
        level_id = request.query_params.get("level_id")
        department_id = request.query_params.get("department_id")
        mode = request.query_params.get("mode")

        if not level_id or not department_id:
            return Response(
                {"error": "level_id and department_id query parameters are required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            qs = Score.objects.select_related("skill").filter(
                level_id=level_id,
                department_id=department_id,
            ).exclude(skill__isnull=True)

            if mode == "remote":
                qs = qs.filter(test__key_id__startswith="remote-")
            elif mode == "individual":
                qs = qs.filter(test__key_id__startswith="W")

            aggregated = (
                qs.values("skill_id", "skill__station_name")
                .annotate(
                    participants=Count("id"),
                    avg_percentage=Avg("percentage"),
                    passed_count=Count("id", filter=Q(passed=True)),
                    batches_count=Count("test__test_name", distinct=True),
                )
                .order_by("skill__station_name")
            )

            data = []
            for row in aggregated:
                participants = row["participants"] or 0
                passed_count = row["passed_count"] or 0
                avg_percentage = row["avg_percentage"] or 0.0
                pass_rate = (passed_count / participants * 100.0) if participants else 0.0

                data.append(
                    {
                        "station_id": row["skill_id"],
                        "station_name": row["skill__station_name"] or "Unknown Station",
                        "participants": participants,
                        "avg_percentage": round(avg_percentage, 1),
                        "pass_rate": round(pass_rate, 1),
                        "batches_count": row["batches_count"] or 0,
                    }
                )

            return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error("Error in ResultsStationsAPIView: %s", e, exc_info=True)
            return Response(
                {"error": "Failed to fetch stations with results"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


# ============================================================
#  D. BATCHES (SESSIONS) FOR (LEVEL, DEPARTMENT, STATION) - REMOTE ONLY
#  GET /api/results/batches/?level_id=...&department_id=...&station_id=...&mode=remote
# ============================================================\

from datetime import datetime

class ResultsBatchesAPIView(APIView):
    """
    Returns distinct batches/sessions for a given Level + Department + Station,
    but only for remote/group mode.

    A "batch" is defined as a group of remote Scores sharing the same test__test_name.
    """

    def get(self, request):
        level_id = request.query_params.get("level_id")
        department_id = request.query_params.get("department_id")
        station_id = request.query_params.get("station_id")
        mode = request.query_params.get("mode")

        if not level_id or not department_id or not station_id:
            return Response(
                {
                    "error": "level_id, department_id and station_id query parameters are required"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Only meaningful for remote mode
        if mode == "individual":
            return Response([], status=status.HTTP_200_OK)

        try:
            qs = Score.objects.select_related("test", "test__question_paper").filter(
                level_id=level_id,
                department_id=department_id,
                skill_id=station_id,
                test__key_id__startswith="remote-",
            ).exclude(test__isnull=True)

            aggregated = (
                qs.values("test__test_name")
                .annotate(
                    participants=Count("id"),
                    avg_percentage=Avg("percentage"),
                    passed_count=Count("id", filter=Q(passed=True)),
                    created_at=Min("test__created_at"),
                    test_date=Min("test_date"),
                    paper_name=Min("test__question_paper__question_paper_name"),
                )
                .order_by("-test_date", "-created_at")
            )

            data = []
            for row in aggregated:
                session_key = row["test__test_name"]
                participants = row["participants"] or 0
                passed_count = row["passed_count"] or 0
                avg_percentage = row["avg_percentage"] or 0.0
                pass_rate = (passed_count / participants * 100.0) if participants else 0.0
                created_at = row["created_at"]
                test_date = row.get("test_date")
                paper_name = row["paper_name"] or "Paper"

                if test_date:
                    label_date = test_date.strftime("%d %b %Y")
                    created_at_iso = test_date.isoformat()
                elif created_at:
                    label_date = created_at.strftime("%d %b %H:%M")  # "10 Dec 09:30"
                    created_at_iso = created_at.isoformat()
                else:
                    label_date = "Unknown time"
                    created_at_iso = None


                display_label = f"{label_date} • {paper_name} • {participants} ppl"

                data.append(
                    {
                        "session_key": session_key,
                        "display_label": display_label,
                        "question_paper_name": paper_name,
                        "created_at": created_at_iso,
                        "participants": participants,
                        "avg_percentage": round(avg_percentage, 1),
                        "pass_rate": round(pass_rate, 1),
                    }
                )

            return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error("Error in ResultsBatchesAPIView: %s", e, exc_info=True)
            return Response(
                {"error": "Failed to fetch batches/sessions"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


# ============================================================
#  E. INDIVIDUAL SCORES FOR (LEVEL, DEPARTMENT, STATION)
#  GET /api/results/individual-scores/?level_id=...&department_id=...&station_id=...
# ============================================================
class ResultsIndividualScoresAPIView(APIView):
    """
    Returns all individual (web/tablet) scores for a given Level + Department + Station.
    Uses ScoreSerializer (same shape as ScoresByTestView).
    """

    def get(self, request):
        level_id = request.query_params.get("level_id")
        department_id = request.query_params.get("department_id")
        station_id = request.query_params.get("station_id")

        if not level_id or not department_id or not station_id:
            return Response(
                {
                    "error": "level_id, department_id and station_id query parameters are required"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            qs = (
                Score.objects.select_related(
                    "employee",
                    "employee__department",
                    "department",
                    "skill",
                    "level",
                    "test",
                    "test__department",
                )
                .filter(
                    level_id=level_id,
                    department_id=department_id,
                    skill_id=station_id,
                    test__key_id__startswith="W",  # Individual mode
                )
                .order_by("-created_at")
            )

            serializer = ScoreSerializer(qs, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error("Error in ResultsIndividualScoresAPIView: %s", e, exc_info=True)
            return Response(
                {"error": "Failed to fetch individual scores"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
    


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import *
from .serializers import *
from rest_framework import viewsets


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
from .serializers import KeyEventSerializer

class KeyEventCreateView(APIView):
    def post(self, request):
        serializer = KeyEventSerializer(data=request.data)
        if serializer.is_valid():
            # 1. Save to DB
            serializer.save()
            
            # 2. Broadcast to Frontend instantly
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                "quiz_session",
                {
                    "type": "broadcast_key_event", # Matches method name in Consumer
                    "payload": serializer.data
                }
            )
            return Response({'message': 'Key event saved & broadcasted'}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    
class LatestKeyEventView(APIView):
    def get(self, request):
        try:
            latest_event = KeyEvent.objects.latest('timestamp')
            serializer = KeyEventSerializer(latest_event)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except KeyEvent.DoesNotExist:
            return Response({"message": "No key events yet."}, status=status.HTTP_404_NOT_FOUND)

from django.core.cache import cache # Import Cache
from rest_framework.decorators import api_view
from rest_framework.response import Response
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
from .serializers import ConnectEventSerializer

# 1. RECEIVE STATUS from Client.py
@api_view(['POST'])
def connect_event_create(request):
    serializer = ConnectEventSerializer(data=request.data)
    if serializer.is_valid():
        # A. Save to DB (Optional: You can comment this out if you don't need history logs)
        serializer.save()
        
        # B. Save to Cache (Fast, No DB Lock)
        # "1" = Connected, "0" = Disconnected
        status_info = serializer.data.get('info')
        cache.set('receiver_connection_status', status_info, timeout=None) 

        # C. Broadcast to Frontend via WebSocket
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            "quiz_session", 
            {
                "type": "broadcast_connect_event",
                "payload": serializer.data
            }
        )
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)

# 2. CHECK STATUS (Frontend Initial Check)
class ReceiverStatusView(APIView):
    def get(self, request):
        # Read from RAM (Cache) instead of DB to prevent locking
        cached_status = cache.get('receiver_connection_status')
        
        # If cache is empty, fallback to DB or assume False
        if cached_status is None:
            try:
                latest = ConnectEvent.objects.latest('created_at')
                is_connected = (latest.info == "1")
            except ConnectEvent.DoesNotExist:
                is_connected = False
        else:
            is_connected = (cached_status == "1")

        return Response({"connected": is_connected}, status=200)


@api_view(['POST'])
def vote_event_create(request):
    serializer = VoteEventSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)


from .models import EvaluationPassingCriteria, Level, Department
from .serializers import EvaluationPassingCriteriaSerializer, LevelSerializer, DepartmentSerializer

class EvaluationPassingCriteriaViewSet(viewsets.ModelViewSet):
    queryset = EvaluationPassingCriteria.objects.select_related('level', 'department').all()
    serializer_class = EvaluationPassingCriteriaSerializer
    
    def create(self, request, *args, **kwargs):
        try:
            # Check if criteria already exists for this level-department combination
            level_id = request.data.get('level')
            department_id = request.data.get('department')
            
            if not level_id or not department_id:
                return Response(
                    {'detail': 'Level and Department are required.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            existing = EvaluationPassingCriteria.objects.filter(
                level_id=level_id,
                department_id=department_id
            ).first()
            
            if existing:
                return Response(
                    {'detail': 'Passing criteria already exists for this Level and Department combination.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            return super().create(request, *args, **kwargs)
        except IntegrityError:
            return Response(
                {'detail': 'Passing criteria already exists for this Level and Department combination.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'detail': f'Error creating criteria: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def update(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            level_id = request.data.get('level')
            department_id = request.data.get('department')
            
            if not level_id or not department_id:
                return Response(
                    {'detail': 'Level and Department are required.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Check if another criteria exists for this level-department combination
            existing = EvaluationPassingCriteria.objects.filter(
                level_id=level_id,
                department_id=department_id
            ).exclude(id=instance.id).first()
            
            if existing:
                return Response(
                    {'detail': 'Another passing criteria already exists for this Level and Department combination.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            return super().update(request, *args, **kwargs)
        except IntegrityError:
            return Response(
                {'detail': 'Another passing criteria already exists for this Level and Department combination.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'detail': f'Error updating criteria: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except Exception as e:
            return Response(
                {'detail': f'Error deleting criteria: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            ) 
        


from .models import EvaluationPassingCriteria, Level, Department
from .serializers import EvaluationPassingCriteriaSerializer, LevelSerializer, DepartmentSerializer

class EvaluationPassingCriteriaViewSet(viewsets.ModelViewSet):
    queryset = EvaluationPassingCriteria.objects.select_related('level', 'department').all()
    serializer_class = EvaluationPassingCriteriaSerializer
    
    def create(self, request, *args, **kwargs):
        try:
            # Check if criteria already exists for this level-department combination
            level_id = request.data.get('level')
            department_id = request.data.get('department')
            
            if not level_id or not department_id:
                return Response(
                    {'detail': 'Level and Department are required.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            existing = EvaluationPassingCriteria.objects.filter(
                level_id=level_id,
                department_id=department_id
            ).first()
            
            if existing:
                return Response(
                    {'detail': 'Passing criteria already exists for this Level and Department combination.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            return super().create(request, *args, **kwargs)
        except IntegrityError:
            return Response(
                {'detail': 'Passing criteria already exists for this Level and Department combination.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'detail': f'Error creating criteria: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def update(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            level_id = request.data.get('level')
            department_id = request.data.get('department')
            
            if not level_id or not department_id:
                return Response(
                    {'detail': 'Level and Department are required.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Check if another criteria exists for this level-department combination
            existing = EvaluationPassingCriteria.objects.filter(
                level_id=level_id,
                department_id=department_id
            ).exclude(id=instance.id).first()
            
            if existing:
                return Response(
                    {'detail': 'Another passing criteria already exists for this Level and Department combination.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            return super().update(request, *args, **kwargs)
        except IntegrityError:
            return Response(
                {'detail': 'Another passing criteria already exists for this Level and Department combination.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'detail': f'Error updating criteria: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except Exception as e:
            return Response(
                {'detail': f'Error deleting criteria: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            ) 

from .models import EvaluationPassingCriteria, Level, Department
from .serializers import EvaluationPassingCriteriaSerializer, LevelSerializer, DepartmentSerializer

class EvaluationPassingCriteriaViewSet(viewsets.ModelViewSet):
    queryset = EvaluationPassingCriteria.objects.select_related('level', 'department').all()
    serializer_class = EvaluationPassingCriteriaSerializer
    
    def create(self, request, *args, **kwargs):
        try:
            # Check if criteria already exists for this level-department combination
            level_id = request.data.get('level')
            department_id = request.data.get('department')
            
            if not level_id or not department_id:
                return Response(
                    {'detail': 'Level and Department are required.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            existing = EvaluationPassingCriteria.objects.filter(
                level_id=level_id,
                department_id=department_id
            ).first()
            
            if existing:
                return Response(
                    {'detail': 'Passing criteria already exists for this Level and Department combination.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            return super().create(request, *args, **kwargs)
        except IntegrityError:
            return Response(
                {'detail': 'Passing criteria already exists for this Level and Department combination.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'detail': f'Error creating criteria: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def update(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            level_id = request.data.get('level')
            department_id = request.data.get('department')
            
            if not level_id or not department_id:
                return Response(
                    {'detail': 'Level and Department are required.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Check if another criteria exists for this level-department combination
            existing = EvaluationPassingCriteria.objects.filter(
                level_id=level_id,
                department_id=department_id
            ).exclude(id=instance.id).first()
            
            if existing:
                return Response(
                    {'detail': 'Another passing criteria already exists for this Level and Department combination.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            return super().update(request, *args, **kwargs)
        except IntegrityError:
            return Response(
                {'detail': 'Another passing criteria already exists for this Level and Department combination.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'detail': f'Error updating criteria: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except Exception as e:
            return Response(
                {'detail': f'Error deleting criteria: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            ) 

#---------------------------- Eveluetion-------------------------------------
#STANDARD_JR0001

import logging
import traceback

from django.db import transaction
from django.shortcuts import get_object_or_404

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import TestSession, QuestionPaper, Score

logger = logging.getLogger(__name__)


class EndTestSessionView(APIView):
    """
    Remote grading endpoint.

    Expects payload like:
    {
      "paper_id": <question_paper_id>,
      "answers": {
         "<remote_id>": [0,1,-1,2,...],
         ...
      }
    }

    Saves:
    - Score.question_ids: snapshot of question ids used in this grading run
    - Score.raw_answers: dict { "<question_id>": selected_index }
    Uses TemplateQuestion.correct_index for scoring (stable).
    """

    def post(self, request):
        try:
            print(f"\n========== STARTING GRADING DEBUG ==========")
            submitted_data = request.data

            if not isinstance(submitted_data, dict):
                return Response({"error": "Invalid payload format"}, status=400)

            paper_id = submitted_data.get("paper_id")
            answers_data = submitted_data.get("answers", {})
            test_date_val = submitted_data.get("test_date")

            if not paper_id:
                return Response({"error": "paper_id is required"}, status=400)

            # --- Load Question Paper + Questions (stable ordering) ---
            try:
                question_paper = QuestionPaper.objects.get(question_paper_id=paper_id)
                questions = list(question_paper.template_questions.order_by("order", "id"))
                total_questions = len(questions)
                question_ids = [q.id for q in questions]

                print(f"DEBUG: Loaded {total_questions} questions for Paper ID {paper_id}")
            except QuestionPaper.DoesNotExist:
                return Response({"error": f"Question Paper {paper_id} not found"}, status=400)

            successful_results = []
            error_results = []
            ignored_remotes = []

            if not isinstance(answers_data, dict):
                return Response({"error": "answers must be an object/dict"}, status=400)

            for raw_key_id, answers in answers_data.items():
                if not isinstance(answers, list):
                    continue

                formatted_key_id = f"remote-{raw_key_id}"
                print(f"DEBUG: Processing Remote {raw_key_id} (Format: {formatted_key_id})")

                try:
                    with transaction.atomic():
                        # Find latest Session
                        session = (
                            TestSession.objects.filter(
                                key_id=formatted_key_id,
                                question_paper=question_paper
                            )
                            .order_by("-created_at")
                            .first()
                        )

                        if not session:
                            print(f"DEBUG: ❌ Session NOT FOUND for {formatted_key_id}")
                            ignored_remotes.append(raw_key_id)
                            continue

                        print(f"DEBUG: ✅ Found Session {session.id} for Employee {session.employee.emp_id}")

                        # --- Convert answers list -> dict keyed by question id ---
                        answers_by_qid = {}
                        for i, qid in enumerate(question_ids):
                            if i < len(answers) and isinstance(answers[i], int):
                                answers_by_qid[str(qid)] = int(answers[i])
                            else:
                                answers_by_qid[str(qid)] = -1

                        # --- GRADING using correct_index only ---
                        correct_count = 0
                        for q in questions:
                            submitted_ans = answers_by_qid.get(str(q.id), -1)

                            if not (isinstance(submitted_ans, int) and 0 <= submitted_ans <= 3):
                                # skipped/invalid
                                continue

                            db_correct = int(q.correct_index) if q.correct_index is not None else -1
                            if submitted_ans == db_correct:
                                correct_count += 1

                        percentage = round((correct_count / total_questions) * 100, 2) if total_questions > 0 else 0.0
                        passed = percentage >= 80  # keep your existing rule in DB

                        print(f"DEBUG: 🏁 Final Score: {correct_count}/{total_questions} ({percentage}%) Passed={passed}")

                        # --- SAVE SCORE ---
                        existing_score = Score.objects.filter(test=session).first()

                        if existing_score:
                            print(f"DEBUG: Updating existing score ID {existing_score.id}")
                            existing_score.marks = correct_count
                            existing_score.percentage = percentage
                            existing_score.passed = passed
                            existing_score.test_date = test_date_val or session.test_date

                            # snapshot + mapping
                            existing_score.question_ids = question_ids
                            existing_score.raw_answers = answers_by_qid

                            existing_score.save()
                        else:
                            Score.objects.create(
                                employee=session.employee,
                                test=session,
                                marks=correct_count,
                                percentage=percentage,
                                passed=passed,
                                level=session.level,
                                skill=session.skill,
                                department=session.department,
                                test_date=test_date_val or session.test_date,

                                # snapshot + mapping
                                question_ids=question_ids,
                                raw_answers=answers_by_qid,
                            )
                            print("DEBUG: Created new Score record")

                        emp_name = f"{session.employee.first_name} {session.employee.last_name}".strip()
                        successful_results.append({
                            "key_id": raw_key_id,
                            "employee_name": emp_name,
                            "marks": correct_count,
                            "percentage": percentage,
                            "passed": passed,
                            "test_name": session.test_name
                        })

                except Exception as e:
                    traceback.print_exc()
                    error_results.append({"key_id": raw_key_id, "error": str(e)})

            print(f"========== GRADING FINISHED ==========\n")

            real_test_name = successful_results[0]["test_name"] if successful_results else "Unknown Test"

            response_data = {
                "message": "Test processed",
                "grouped_results": [{
                    "test_name": real_test_name,
                    "employees": [r["employee_name"] for r in successful_results]
                }],
                "ignored_remotes": ignored_remotes,
                "errors": error_results
            }

            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            traceback.print_exc()
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db.models import OuterRef, Subquery
from .models import MasterTable, Score, RetrainingSession, RetrainingSessionDetail, RetrainingConfig, SkillMatrix


class EvaluationEligibleEmployeeListView(APIView):
    """
    API to list employees eligible to attempt evaluation test or show reason otherwise.
    Accepts query params: level_id, department_id, station_id
    """
    def get(self, request):
        level_id = request.query_params.get('level_id')
        department_id = request.query_params.get('department_id')
        station_id = request.query_params.get('station_id')
        if not level_id or not department_id:
            return Response({"error": "level_id and department_id are required"}, status=status.HTTP_400_BAD_REQUEST)

        employees = MasterTable.objects

        # Latest score for employee with filters
        latest_score_qs = Score.objects.filter(
            employee=OuterRef('pk'),
            level_id=level_id,
            department_id=department_id,
            skill_id=station_id
        ).order_by('-attempt_no')

        employees = employees.annotate(
            last_attempt_no=Subquery(latest_score_qs.values('attempt_no')[:1]),
            last_passed=Subquery(latest_score_qs.values('passed')[:1])
        )

        config = RetrainingConfig.objects.filter(evaluation_type='Evaluation').first()
        max_attempts = config.max_count if config else 2
        
        
        response_list = []
        for emp in employees:
            messages = []
            is_eligible = True

            department_name = ""
            if hasattr(emp, 'department') and emp.department:
                department_name = emp.department.department_name

            # ----------------------------
            # 1️⃣ UPDATED CHECK (MATCHES check_eligibility)
            # ----------------------------
            completed_levels = list(
                SkillMatrix.objects.filter(emp_id=emp.emp_id)
                .values_list("level__level_id", flat=True)
            )

            target_level_int = int(level_id)

            if not completed_levels:
                # No records → only Level 1 allowed
                if target_level_int != 1:
                    is_eligible = False
                    messages.append("Employee has not completed Level 1")
            else:
                highest_level = max(completed_levels)

                # If employee already has equal or higher level → allow
                if highest_level >= target_level_int:
                    pass  # eligible, no block
                else:
                    # Otherwise require immediate previous level
                    required_prev_level = target_level_int - 1

                    if required_prev_level not in completed_levels:
                        is_eligible = False
                        messages.append(
                            f"Employee has not completed Level {required_prev_level}"
                        )

            # ----------------------------
            # 2️⃣ EXISTING RETRAINING CHECK LOGIC
            # ----------------------------
            last_attempt_no = emp.last_attempt_no

            if last_attempt_no is None:
                # Never attended → eligible unless previous-level failed
                if is_eligible:
                    messages.append("Never attended test")
                response_list.append({
                    "id": emp.emp_id,
                    "emp_id": emp.emp_id,
                    "name": f"{emp.first_name or ''} {emp.last_name or ''}".strip(),
                    "department": department_name,
                    "eligible": is_eligible,
                    "reason": " & ".join(messages) if messages else "Eligible"
                })
                continue

            # Passed last attempt → not allowed to attempt again
            if emp.last_passed:
                continue

            # Failed + max attempts reached
            if last_attempt_no >= max_attempts:
                is_eligible = False
                messages.append(f"Max attempts ({max_attempts}) reached")

                response_list.append({
                    "id": emp.emp_id,
                    "emp_id": emp.emp_id,
                    "name": f"{emp.first_name or ''} {emp.last_name or ''}".strip(),
                    "department": department_name,
                    "eligible": is_eligible,
                    "reason": " & ".join(messages)
                })
                continue

            # Retraining check
            retraining_qs = RetrainingSession.objects.filter(
                employee=emp,
                level_id=level_id,
                department_id=department_id,
                station_id=station_id,
                evaluation_type='Evaluation',
                attempt_no=last_attempt_no,
                status='Scheduled'
            )

            has_retraining = retraining_qs.exists()

            has_retraining_detail = RetrainingSessionDetail.objects.filter(
                retraining_session__in=retraining_qs
            ).exists()

            if has_retraining and has_retraining_detail:
                messages.append("Retraining scheduled and session detail exists")
            else:
                is_eligible = False
                messages.append("Retraining not scheduled or session detail missing")

            response_list.append({
                "id": emp.emp_id,
                "emp_id": emp.emp_id,
                "name": f"{emp.first_name or ''} {emp.last_name or ''}".strip(),
                "department": department_name,
                "eligible": is_eligible,
                "reason": " & ".join(messages)
            })

        return Response({"employees": response_list}, status=status.HTTP_200_OK)



class ScoreListView(APIView):
    def get(self, request):
        # Assuming you use caching for latest test session
        session_key = cache.get("latest_test_session")
        if not session_key:
            return Response([])

        scores = Score.objects.filter(test__key_id=session_key).select_related('employee', 'level', 'skill')
        serializer = ScoreSerializer(scores, many=True)
        return Response(serializer.data)


class KeyIdToEmployeeNameMap(APIView):
    def get(self, request):
        mapping = TestSession.objects.select_related('employee').all()
        return Response({
            s.key_id: f"{s.employee.first_name} {s.employee.last_name}" 
            for s in mapping
        })


class PastTestSessionsView(APIView):
    def get(self, request):
        # Get distinct test names from Score model through the test relationship
        test_names = Score.objects.select_related('test').filter(
            test__isnull=False
        ).values_list('test__test_name', flat=True).distinct()
        
        return Response(list(test_names))

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from app1.models import Score
from app1.serializers import ScoreSerializer
import traceback

class ScoresByTestView(APIView):
    def get(self, request, name):
        try:
            scores = (
                Score.objects
                .filter(test__test_name=name)
                .select_related(
                    'employee', 'employee__department',
                    'department',
                    'skill', 'skill__subline', 'skill__subline__line', 'skill__subline__line__department',
                    'level', 'test', 'test__department'
                )
            )
            serializer = ScoreSerializer(scores, many=True)
            print("Final data being sent to frontend:", serializer.data)
            return Response(serializer.data)
        except Exception as e:
            print(f"Error in ScoresByTestView: {e}")
            traceback.print_exc()
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class SkillListView(APIView):
    def get(self, request):
        skills = Station.objects.values_list('station_name', flat=True).distinct()
        return Response(list(skills))


class ResultSummaryAPIView(APIView):
    def get(self, request):
        try:
            scores = Score.objects.select_related('employee', 'level', 'skill')
            data = []
            for score in scores:
                # Use the percentage from the score model instead of recalculating
                percentage = score.percentage
                result = 'Pass' if score.percentage >= 80 else 'Retraining' if score.percentage >= 50 else 'Fail'

                data.append({
                    "employee_id": score.employee.emp_id if hasattr(score.employee, 'emp_id') else score.employee.id,
                    "name": f"{score.employee.first_name} {score.employee.last_name}",
                    "marks": score.marks,
                    "percentage": percentage,
                    "section": score.employee.section if hasattr(score.employee, 'section') else '',
                    "level_name": score.level.level_name if score.level and hasattr(score.level, 'level_name') else '',
                    "skill": score.skill.station_name if score.skill and hasattr(score.skill, 'station_name') else (score.skill.skill if score.skill else ''),
                    "result": result,
                })

            serializer = SimpleScoreSerializer(data, many=True)
            return Response(serializer.data)
            
        except Exception as e:
            print(f"Error in ResultSummaryAPIView: {e}")
            traceback.print_exc()
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


        
# ===== imports needed for this view =====
import logging
import traceback
from decimal import Decimal

from django.db import transaction
from django.http import Http404
from django.shortcuts import get_object_or_404
from django.utils import timezone

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import (
    QuestionPaper, Score, Station, TestSession, MasterTable,
    Level, TemplateQuestion, Department, EvaluationPassingCriteria
)

logger = logging.getLogger(__name__)

# ===== imports needed for this view =====
import logging
import traceback
import uuid
from decimal import Decimal

from django.db import transaction
from django.http import Http404
from django.shortcuts import get_object_or_404
from django.utils import timezone

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import (
    QuestionPaper, Score, Station, TestSession, MasterTable,
    Level, TemplateQuestion, Department, EvaluationPassingCriteria
)

logger = logging.getLogger(__name__)


class SubmitWebTestAPIView(APIView):
    """
    Submit test answers from web/tablet exam.

    Key improvements:
    - Freeze attempt question set into Score.question_ids
    - Store raw_answers as dict keyed by question id: { "<qid>": selected_index }
    - Score using TemplateQuestion.correct_index (reliable)
    - Keeps percentage/passed in DB (you can choose not to display on paper)
    """

    def post(self, request):
        employee_id = request.data.get("employee_id")
        question_paper_id = request.data.get("question_paper_id")
        level_id = request.data.get("level_id")
        station_id = request.data.get("skill_id")

        try:
            logger.info("SubmitWebTestAPIView called with data: %s", request.data)

            test_name = request.data.get("test_name")
            answers = request.data.get("answers", [])
            department_name_signal = request.data.get("department_name")
            department_id = request.data.get("departmentId")

            if not all([employee_id, test_name, question_paper_id, level_id]) or not isinstance(answers, list):
                logger.warning("Validation failed: missing core required fields.")
                return Response(
                    {"error": "employee_id, test_name, question_paper_id, and level_id are required."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            employee = get_object_or_404(MasterTable, emp_id=employee_id)
            question_paper = get_object_or_404(QuestionPaper, question_paper_id=question_paper_id)
            level_object = get_object_or_404(Level, level_id=level_id)

            is_level_1_global = (
                (level_object.level_name or "").lower() == "level 1"
                or str(level_object.pk) == "1"
            )

            # --- Resolve department ---
            department_object = None

            if department_id:
                department_object = Department.objects.filter(pk=department_id).first()
                if not department_object:
                    logger.warning("Department id=%s not found; falling back.", department_id)

            if not department_object and department_name_signal:
                department_object = Department.objects.filter(
                    department_name=department_name_signal
                ).first()
                if not department_object:
                    logger.warning("Department name='%s' not found; falling back.", department_name_signal)

            if not department_object:
                department_object = employee.department

            if is_level_1_global and not department_object:
                raise Http404(
                    "No valid department found (neither request nor employee.department). "
                    "Department is required for Level 1 history."
                )

            # --- Resolve station (skill) ---
            skill_object = None
            if station_id:
                skill_object = get_object_or_404(Station, pk=station_id)

            # --- Load questions in stable ordering and freeze snapshot ---
            questions = list(question_paper.template_questions.all().order_by("order", "id"))
            total_questions = len(questions)
            logger.info("Total questions found: %s", total_questions)

            if total_questions == 0:
                return Response({"error": "No questions found for this paper."},
                                status=status.HTTP_400_BAD_REQUEST)

            question_ids = [q.id for q in questions]

            # --- Normalize answers list -> dict keyed by question id ---
            answers_by_qid = {}
            for i, qid in enumerate(question_ids):
                if i < len(answers) and isinstance(answers[i], int):
                    answers_by_qid[str(qid)] = int(answers[i])
                else:
                    answers_by_qid[str(qid)] = -1

            # --- Scoring using correct_index (reliable) ---
            correct_count = 0
            for q in questions:
                submitted_ans = answers_by_qid.get(str(q.id), -1)
                if not (isinstance(submitted_ans, int) and 0 <= submitted_ans <= 3):
                    continue
                db_correct = int(q.correct_index) if q.correct_index is not None else -1
                if submitted_ans == db_correct:
                    correct_count += 1

            # --- Keep percentage/passed in DB (not required to display) ---
            required_percentage = Decimal("80.00")
            if total_questions > 0:
                percentage_decimal = (Decimal(correct_count) / Decimal(total_questions) * Decimal("100")).quantize(
                    Decimal("0.01")
                )
            else:
                percentage_decimal = Decimal("0.00")

            percentage = float(percentage_decimal)

            try:
                criteria = EvaluationPassingCriteria.objects.filter(level=level_object).first()
                if criteria and criteria.percentage is not None:
                    required_percentage = criteria.percentage
            except Exception as e:
                logger.warning("Error fetching EvaluationPassingCriteria: %s", e)

            passed = percentage_decimal >= required_percentage

            logger.info(
                "Scoring: %s/%s correct (%.2f%%), required=%.2f%%, passed=%s",
                correct_count, total_questions, percentage, float(required_percentage), passed
            )

            now = timezone.now()
            timestamp_display = now.strftime("%Y-%m-%d %H:%M")

            with transaction.atomic():
                if not skill_object:
                    return Response(
                        {"error": "Submission failed. Station ID (skill_id) is required."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                target_station = skill_object
                score_department = employee.department if employee.department else department_object

                unique_key = "W" + uuid.uuid4().hex[:9]

                level_label = level_object.level_name if level_object else "N/A"
                dept_label = score_department.department_name if score_department else "N/A"

                unique_test_name = test_name or (
                    f"{question_paper.question_paper_name} - "
                    f"{target_station.station_name} - "
                    f"{level_label} - {dept_label} - {timestamp_display}"
                )
                
                # Fetch test_date from incoming request for web mode or from TestSession for remote
                test_date_val = request.data.get("test_date")

                test_session = TestSession.objects.create(
                    employee=employee,
                    question_paper=question_paper,
                    skill=target_station,
                    key_id=unique_key,
                    test_name=unique_test_name,
                    level=level_object,
                    department=score_department,
                    test_date=test_date_val
                )

                Score.objects.create(
                    employee=employee,
                    test=test_session,
                    skill=target_station,
                    level=level_object,
                    department=score_department,
                    marks=correct_count,
                    percentage=percentage,
                    passed=passed,
                    test_date=test_date_val or (test_session.test_date if test_session else None),

                    # NEW: snapshot + stable answer mapping
                    question_ids=question_ids,
                    raw_answers=answers_by_qid,
                )

            return Response(
                {
                    "employee": f"{employee.first_name} {employee.last_name}".strip(),
                    "level_received": level_object.level_name,
                    "department": department_object.department_name if department_object else "N/A",
                    "total_stations_updated": 1,
                    "stations_list": [skill_object.station_name] if skill_object else [],
                    "marks": correct_count,
                    "percentage": percentage,
                    "passed": passed,
                    "message": "Score successfully saved.",
                },
                status=status.HTTP_200_OK,
            )

        except MasterTable.DoesNotExist:
            return Response({"error": f"Employee with ID {employee_id} not found."},
                            status=status.HTTP_404_NOT_FOUND)
        except QuestionPaper.DoesNotExist:
            return Response({"error": f"Question Paper with ID {question_paper_id} not found."},
                            status=status.HTTP_404_NOT_FOUND)
        except Level.DoesNotExist:
            return Response({"error": f"Level with ID {level_id} not found."},
                            status=status.HTTP_404_NOT_FOUND)
        except Station.DoesNotExist:
            return Response({"error": f"Station with ID {station_id} not found."},
                            status=status.HTTP_404_NOT_FOUND)
        except Http404 as e:
            logger.warning("Submission lookup error: %s", str(e))
            return Response({"error": str(e)}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            error_trace = traceback.format_exc()
            logger.error("Error in SubmitWebTestAPIView: %s\n%s", str(e), error_trace)
            return Response(
                {"error": "Internal server error: " + str(e), "traceback": error_trace},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


#STANDARD_JR0001
#---------------------------- Eveluetion -------------------------------------

import logging

from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import Score, TemplateQuestion

logger = logging.getLogger(__name__)


class AnswerSheetView(APIView):
    """
    Answer sheet for a Score attempt.

    Supports:
    - New format:
        Score.question_ids = [qid1, qid2, ...]  # snapshot in order
        Score.raw_answers = {"<qid>": selected_index, ...}
    - Legacy format:
        Score.raw_answers = [0,1,-1,...] (mapped by question_ids snapshot if present,
        otherwise by current paper ordering)

    Uses TemplateQuestion.correct_index (0..3) for correctness.
    Includes language 2 fields.
    """

    def get(self, request, score_id):
        try:
            score = get_object_or_404(
                Score.objects.select_related(
                    "test__question_paper",
                    "employee",
                    "department",
                    "skill",
                    "level",
                ),
                pk=score_id,
            )

            test_session = score.test
            question_paper = test_session.question_paper if test_session else None
            if not question_paper:
                return Response(
                    {"error": "Test session is not linked to a question paper."},
                    status=status.HTTP_404_NOT_FOUND,
                )

            # --- Header info ---
            employee = score.employee
            employee_name = f"{employee.first_name} {employee.last_name}".strip() if employee else "N/A"
            current_employee_id = employee.emp_id if employee else "N/A"

            qp_name = getattr(question_paper, "question_paper_name", None) or "N/A"
            dept_name = score.department.department_name if score.department else "N/A"
            station_name = score.skill.station_name if score.skill else "N/A"
            level_name = score.level.level_name if score.level else "N/A"

            # --- Attempt snapshot question ids (preferred) ---
            attempt_qids = score.question_ids if isinstance(score.question_ids, list) else []
            attempt_qids = [int(x) for x in attempt_qids if str(x).isdigit()]

            # If no snapshot exists (legacy), fallback to current question paper ordering
            if not attempt_qids:
                current_questions = list(
                    TemplateQuestion.objects.filter(question_paper=question_paper).order_by("order", "id")
                )
                attempt_qids = [q.id for q in current_questions]

            # --- Fetch questions (only those in attempt_qids) ---
            # NOTE: questions might have been deleted later; we handle missing gracefully.
            qs = TemplateQuestion.objects.filter(question_paper=question_paper, id__in=attempt_qids)
            qs_by_id = {q.id: q for q in qs}

            ordered_questions = [qs_by_id[qid] for qid in attempt_qids if qid in qs_by_id]

            # --- Build answers_by_qid (int keys) ---
            submitted = score.raw_answers
            answers_by_qid = {}

            if isinstance(submitted, dict):
                # New format: {"1": 0, "8": -1, ...}
                for k, v in submitted.items():
                    try:
                        qid = int(k)
                        answers_by_qid[qid] = int(v)
                    except Exception:
                        continue

            elif isinstance(submitted, list):
                # Legacy format: [0,1,-1,...]
                # Map by attempt_qids order
                for idx, qid in enumerate(attempt_qids):
                    if idx < len(submitted):
                        try:
                            answers_by_qid[qid] = int(submitted[idx])
                        except Exception:
                            answers_by_qid[qid] = -1
                    else:
                        answers_by_qid[qid] = -1
            else:
                # Unknown/missing: treat all as skipped
                answers_by_qid = {qid: -1 for qid in attempt_qids}

            # --- Create answer sheet payload ---
            answer_sheet = []
            wrong_count = 0
            skipped_count = 0
            correct_count_display = 0  # correctness computed from answers_by_qid (for display only)

            for q in ordered_questions:
                employee_ans_index = answers_by_qid.get(q.id, -1)

                correct_index = int(q.correct_index) if q.correct_index is not None else -2
                is_skipped = employee_ans_index == -1
                is_correct = (employee_ans_index == correct_index and not is_skipped)

                if is_skipped:
                    skipped_count += 1
                elif is_correct:
                    correct_count_display += 1
                else:
                    wrong_count += 1

                question_image_url = (
                    request.build_absolute_uri(q.question_image.url) if q.question_image else None
                )

                options_with_images = [
                    {
                        "text": q.option_a,
                        "text_lang2": q.option_a_lang2,
                        "image_url": request.build_absolute_uri(q.option_a_image.url) if q.option_a_image else None,
                    },
                    {
                        "text": q.option_b,
                        "text_lang2": q.option_b_lang2,
                        "image_url": request.build_absolute_uri(q.option_b_image.url) if q.option_b_image else None,
                    },
                    {
                        "text": q.option_c,
                        "text_lang2": q.option_c_lang2,
                        "image_url": request.build_absolute_uri(q.option_c_image.url) if q.option_c_image else None,
                    },
                    {
                        "text": q.option_d,
                        "text_lang2": q.option_d_lang2,
                        "image_url": request.build_absolute_uri(q.option_d_image.url) if q.option_d_image else None,
                    },
                ]

                answer_sheet.append(
                    {
                        "id": q.id,
                        "question_text": q.question,
                        "question_text_lang2": q.question_lang2,
                        "question_image_url": question_image_url,
                        "options": options_with_images,
                        "correct_index": correct_index,
                        "employee_answer_index": employee_ans_index,
                        "is_correct": is_correct,
                    }
                )

            # IMPORTANT:
            # total_questions should be the snapshot count, not "ordered_questions length",
            # because questions might be deleted later and you still want stable fraction.
            total_questions_snapshot = len(attempt_qids)

            response_data = {
                "test_name": test_session.test_name if test_session else "N/A",
                "employee_name": employee_name,
                "employee_id": current_employee_id,
                "question_paper_name": qp_name,
                "department_name": dept_name,
                "station_name": station_name,
                "level_name": level_name,
                "test_date": score.test_date.strftime("%d %b %Y") if score.test_date else None,
                "questions": answer_sheet,

                # You said: take score & percentage from Score model (source of truth)
                "score_summary": {
                    "total_questions": total_questions_snapshot,
                    "correct_answers": score.marks,
                    "score": score.marks,
                    "percentage": score.percentage,
                    "passed": score.passed,

                    # Optional debug/UX helpers (safe to keep)
                    "wrong_answers_display": wrong_count,
                    "skipped_answers_display": skipped_count,
                    "correct_answers_display": correct_count_display,
                    "missing_questions_after_snapshot": total_questions_snapshot - len(ordered_questions),
                },
            }

            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(
                "Error in AnswerSheetView (ID: %s): %s", score_id, str(e), exc_info=True
            )
            return Response(
                {"error": "An unexpected error occurred while generating the answer sheet."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


            

# ====================== End answer matrix full Views =========================================









class LevelStationMatrixView(APIView):
    """
    Returns a matrix view of all employee results for a specific Level and Station (Skill).
    URL: /api/results/matrix/<int:level_id>/<int:station_id>/
    """
    def get(self, request, level_id, station_id):
        try:
            logger.info("LevelStationMatrixView called for Level %d, Station %d", level_id, station_id)

            # 1. Find the Question Paper corresponding to the Level and Station
            # Assuming QuestionPaper model links directly to Level and Station/Skill
            qp = QuestionPaper.objects.filter(
                level_id=level_id, 
                station_id=station_id
            ).first()

            if not qp:
                return Response({"error": "No Question Paper found for this Level and Station."}, status=status.HTTP_404_NOT_FOUND)

            # 2. Get all questions for the paper (to form the columns)
            questions_qs = TemplateQuestion.objects.filter(
                question_paper=qp
            ).order_by("id")
            
            # Map questions to column headers
            question_headers = [{
                "id": q.id,
                "text_preview": q.question[:50] + "...",
                "correct_index": options.index(q.correct_answer) if q.correct_answer in (options := [q.option_a, q.option_b, q.option_c, q.option_d]) else -1
            } for q in questions_qs]

            # 3. Get all Score records linked to this Question Paper
            # Filter scores using the TestSession which links to the QP
            all_scores = Score.objects.select_related(
                'employee', 'test'
            ).filter(
                test__question_paper=qp
            )
            
            # 4. Compile the Matrix Rows (Employee Results)
            employee_results = []
            for score in all_scores:
                employee = score.employee
                raw_answers = score.raw_answers
                
                # Normalize raw answers list (must match the number of questions)
                submitted_answers = raw_answers if isinstance(raw_answers, list) else []
                
                # Check performance per question
                q_results = []
                for i, q_header in enumerate(question_headers):
                    # Handle both dictionary (new) and list (legacy) answer formats
                    if isinstance(raw_answers, dict):
                        # New format: {"question_id": answer_index}
                        qid_str = str(q_header['id'])
                        employee_ans_index = raw_answers.get(qid_str, -1)
                        # Ensure it's an int
                        try:
                            employee_ans_index = int(employee_ans_index)
                        except (ValueError, TypeError):
                            employee_ans_index = -1
                    elif isinstance(raw_answers, list):
                        # Legacy format: [ans_index, ans_index, ...] matched by array index
                        employee_ans_index = raw_answers[i] if i < len(submitted_answers) else -1
                    else:
                        employee_ans_index = -1

                    correct_index = q_header['correct_index']
                    
                    is_correct = (employee_ans_index == correct_index and employee_ans_index != -1)
                    
                    q_results.append({
                        "is_correct": is_correct,
                        "submitted_ans_index": employee_ans_index,
                    })

                employee_results.append({
                    "employee_id": employee.emp_id,
                    "employee_name": f"{employee.first_name} {employee.last_name}",
                    "score_id": score.pk, # Link back to individual answer sheet
                    "overall_marks": score.marks,
                    "overall_percentage": score.percentage,
                    "question_results": q_results
                })

            # 5. Final Response
            response_data = {
                "level_id": level_id,
                "station_id": station_id,
                "question_paper_name": qp.question_paper_name,
                "question_headers": question_headers,
                "employee_results": employee_results
            }

            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error("Error in LevelStationMatrixView: %s", str(e))
            return Response({"error": f"An unexpected error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ====================== End Evaluvation Full Views =========================================



# =================== Station Manager =================================

from rest_framework import viewsets, status
from rest_framework.response import Response
from .models import StationManager
from .serializers import StationManagerSerializer
import logging

logger = logging.getLogger(__name__)

class StationManagerViewSet(viewsets.ModelViewSet):
    queryset = StationManager.objects.all()
    serializer_class = StationManagerSerializer

    def create(self, request, *args, **kwargs):
        # Standard create override to catch errors gracefully
        try:
            return super().create(request, *args, **kwargs)
        except Exception as e:
            logger.error(f"Error creating StationManager: {e}")
            return Response(
                {"error": str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )

    def update(self, request, *args, **kwargs):
        # Standard update override for debugging
        logger.info(f"Updating StationManager {kwargs.get('pk')} with data: {request.data}")
        try:
            return super().update(request, *args, **kwargs)
        except Exception as e:
            logger.error(f"Error updating StationManager: {e}")
            return Response(
                {"error": str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )


# from rest_framework import viewsets, status
# from rest_framework.response import Response
# from django.db import transaction
# from .models import StationManager  # Import from models.py
# from .serializers import StationManagerSerializer  # or SimpleStationManagerSerializer
# import logging

# logger = logging.getLogger(__name__)

# class StationManagerViewSet(viewsets.ModelViewSet):
#     queryset = StationManager.objects.all()
#     serializer_class = StationManagerSerializer  # Use debug serializer first

#     def create(self, request, *args, **kwargs):
#         """Override create to add better error handling and logging"""
#         logger.info(f"StationManager create request data: {request.data}")
#         print(f"DEBUG: Received data: {request.data}")
        
#         # Debug: Check what HierarchyStructure objects exist
#         print("DEBUG: Checking HierarchyStructure objects...")
#         try:
#             all_hierarchy = HierarchyStructure.objects.all()
#             for h in all_hierarchy[:10]:  # Show first 10
#                 print(f"  - ID: {h.pk}, Name: {getattr(h, 'structure_name', 'No name')}")
#         except Exception as e:
#             print(f"DEBUG: Error checking HierarchyStructure: {e}")
        
#         try:
#             serializer = self.get_serializer(data=request.data)
#             if serializer.is_valid():
#                 with transaction.atomic():
#                     self.perform_create(serializer)
#                 headers = self.get_success_headers(serializer.data)
#                 logger.info(f"StationManager created successfully: {serializer.data}")
#                 return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
#             else:
#                 logger.error(f"StationManager validation errors: {serializer.errors}")
#                 print(f"DEBUG: Validation errors: {serializer.errors}")
#                 return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
#         except Exception as e:
#             logger.error(f"StationManager creation error: {str(e)}")
#             print(f"DEBUG: Exception: {str(e)}")
#             return Response(
#                 {"detail": f"An error occurred: {str(e)}"}, 
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR
#             )

#     def perform_create(self, serializer):
#         serializer.save()
        
# views.py# views.py# views.py
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import StationManager, HierarchyStructure
from .serializers import StationManagerSerializer
import logging

logger = logging.getLogger(__name__)


@api_view(['GET'])
def get_station_requirements(request):
    """Get minimum skill requirements for all stations"""
    try:
        logger.debug("Entering get_station_requirements view")
        # Ensure pre-fetching related data for performance
        requirements = StationManager.objects.all().select_related('department', 'station')

        data = []
        level_mapping = {
            'Level 1': 1,
            'Level 2': 2,
            'Level 3': 3,
            'Level 4': 4
        }

        for req in requirements:
            logger.debug(f"Processing StationManager ID: {req.id}")

            # 1. Skip records with no HierarchyStructure link
            if not req.station:
                logger.warning(f"Skipping StationManager ID {req.id} due to missing station FK")
                continue

            # 🛑 CRITICAL FIX: Get the PK of the actual Station Model 🛑
            # req.station is the HierarchyStructure object.
            # req.station.station is the linked Station model object.
            # We need the PK of the actual Station, not the HierarchyStructure.
            current_station_pk = req.station.station.pk if req.station.station else None
            
            if not current_station_pk:
                logger.warning(f"StationManager ID {req.id} lacks link to actual Station model, skipping.")
                continue

            # Use serializer to get display names
            serializer = StationManagerSerializer(req)
            serialized_data = serializer.data
            logger.debug(f"Serialized data for StationManager ID {req.id}: {serialized_data}")

            # Calculate numerical level
            level_name = level_mapping.get(req.minimum_level_required, 0)
            logger.debug(f"Mapping level '{req.minimum_level_required}' to number {level_name}")


            data.append({
                'id': req.id,
                # Use the PK of the actual Station model (e.g., 1, 2, 3) for the frontend lookup
                'station_id': current_station_pk, 
                
                'station_name': serialized_data.get('station_name', 'Unknown Station'),
                
                # Use the PK of the Department model (req.department is the HierarchyStructure)
                'department_id': req.department.department.pk if req.department and req.department.department else None, 
                'department_name': serialized_data.get('department_name', 'Unknown Department'), 
                
                'minimum_operators': req.minimum_operators,
                'minimum_level_required': req.minimum_level_required,
                'minimum_level_name': level_name
            })

        logger.debug(f"Returning data: {data}")
        return Response(data)

    except Exception as e:
        logger.error(f"Error in get_station_requirements: {str(e)}", exc_info=True)
        return Response({'error': str(e)}, status=500)



@api_view(['POST'])
def create_station_manager(request):
    """Create a new StationManager record"""
    try:
        logger.info(f"StationManager create request data: {request.data}")
        serializer = StationManagerSerializer(data=request.data)
        if serializer.is_valid():
            station_manager = serializer.save()
            logger.info(f"StationManager created successfully: {serializer.data}")
            return Response(serializer.data, status=201)
        logger.error(f"StationManager creation failed: {serializer.errors}")
        return Response(serializer.errors, status=400)
    except Exception as e:
        logger.error(f"Error in create_station_manager: {str(e)}", exc_info=True)
        return Response({'error': str(e)}, status=500)


# ====================== Station Mnager End =========================================


#================== BiometricAttendance ==================#

#================== BiometricAttendance ==================#

from rest_framework import viewsets
from .models import BiometricAttendance
from .serializers import BiometricAttendanceSerializer

class BiometricAttendanceViewSet(viewsets.ModelViewSet):
    queryset = BiometricAttendance.objects.all()
    serializer_class = BiometricAttendanceSerializer


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status
from tablib import Dataset
from django.db.models import Count, Sum, Case, When, IntegerField, FloatField
from .models import BiometricAttendance
# --- FIXED IMPORTS ---
from datetime import datetime, time 

class ExcelUploadView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, format=None):
        file_obj = request.FILES.get('file')
        date_str = request.data.get('date')

        if not file_obj or not date_str:
            return Response({'error': 'File and Date are required'}, status=400)

        file_format = 'xls' if file_obj.name.endswith('.xls') else 'xlsx'
        dataset = Dataset()
        
        try:
            imported_data = dataset.load(file_obj.read(), format=file_format)
        except Exception as e:
            return Response({'error': f'Failed to read Excel: {str(e)}'}, status=400)

        # 1. CLEAN HEADERS
        if imported_data.headers:
            clean_headers = [str(h).strip() for h in imported_data.headers]
            imported_data.headers = clean_headers
            print(f"DEBUG: Found Headers: {clean_headers}")

        saved_count = 0
        failed_rows = []
        
        # 2. ITERATE ROWS
        for index, row in enumerate(imported_data.dict):
            try:
                card_no = row.get('Card No')
                if not card_no:
                    failed_rows.append(f"Row {index+2}: Missing 'Card No'")
                    continue
                
                card_no = str(card_no).strip()

                BiometricAttendance.objects.update_or_create(
                    card_no=card_no,
                    attendance_date=date_str,
                    defaults={
                        'sr_no': self.safe_int(row.get('Sr.No.')),
                        'pay_code': row.get('PayCode'),
                        'employee_name': row.get('Employee Name'),
                        'department': row.get('Department'),
                        'designation': row.get('Designation'),
                        'shift': row.get('Shift'),
                        
                        # FIXED TIME PARSING
                        'start': self.parse_time(row.get('Start')),
                        'in_time': self.parse_time(row.get('In')),
                        'out_time': self.parse_time(row.get('Out')),
                        
                        'hrs_works': str(row.get('Hrs Works') or ''),
                        'status': row.get('Status'),
                        'early_arrival': str(row.get('Early Arriv.') or ''),
                        'late_arrival': str(row.get('Late Arriv.') or ''),
                        'shift_early': str(row.get('Shift Early') or ''),
                        'excess_lunch': str(row.get('Excess Lunch') or ''),
                        'ot': str(row.get('Ot') or ''),
                        'ot_amount': str(row.get('Ot Amount') or ''),
                        'os': str(row.get('Os') or ''),
                        'manual': row.get('Manual'),
                    }
                )
                saved_count += 1
            except Exception as e:
                error_msg = f"Row {index+2} (Card {row.get('Card No')}): {str(e)}"
                print(error_msg)
                failed_rows.append(error_msg)

        response_data = {
            'success': True, 
            'message': f'Uploaded {saved_count} records.',
        }

        if failed_rows:
            response_data['message'] += f" Failed {len(failed_rows)} rows."
            response_data['details'] = failed_rows[:5]

        return Response(response_data, status=201)

    # --- HELPER FUNCTIONS ---

    def safe_int(self, val):
        try:
            if not val: return None
            return int(float(val))
        except:
            return None

    
    def parse_time(self, val):
        # Import the specific types needed for comparison
        from datetime import datetime as dt_class, time as time_class

        if not val or str(val).lower() == 'nan': 
            return None
        
        # 1. Check if it is already a datetime object (e.g., 2023-01-01 09:00:00)
        if isinstance(val, dt_class):
            return val.time()
            
        # 2. Check if it is already a time object (e.g., 09:00:00)
        # This was likely causing the crash if 'time' referred to the module
        if isinstance(val, time_class):
            return val
            
        # 3. If it's a string, parse it
        val_str = str(val).strip()
        
        # List of formats to try
        formats = [
            '%H:%M:%S', 
            '%H:%M', 
            '%I:%M %p', 
            '%H:%M:%S.%f'
        ]
        
        for fmt in formats:
            try:
                return dt_class.strptime(val_str, fmt).time()
            except ValueError:
                continue
        
        return None


import os
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from tablib import Dataset
from .resources import BiometricAttendanceResource

class ExcelUploadFromPathView(APIView):
    # Set the path to your Excel file here
    EXCEL_FILE_PATH = r"E:\attendance.xlsx"  # Change this to your actual path

    def post(self, request, format=None):
        if not os.path.exists(self.EXCEL_FILE_PATH):
            return Response({'error': 'File not found on server path'}, status=status.HTTP_400_BAD_REQUEST)

        file_format = 'xls' if self.EXCEL_FILE_PATH.endswith('.xls') else 'xlsx'

        dataset = Dataset()
        try:
            with open(self.EXCEL_FILE_PATH, 'rb') as file_obj:
                imported_data = dataset.load(file_obj.read(), format=file_format)
        except Exception as e:
            return Response({'error': f'Failed to read Excel file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        resource = BiometricAttendanceResource()
        result = resource.import_data(imported_data, dry_run=True)

        if result.has_errors():
            errors = []
            for row_number, row_errors in result.row_errors():
                for error in row_errors:
                    errors.append(f"Row {row_number}: {str(error.error)}")
            return Response({'error': 'Import failed', 'details': errors}, status=status.HTTP_400_BAD_REQUEST)

        # Perform actual import
        resource.import_data(imported_data, dry_run=False)
        return Response({'success': 'Data imported successfully'}, status=status.HTTP_201_CREATED)



def schedule_task(task_name, task_path, time_str):
  

    try:
        dt = datetime.strptime(time_str.strip(), "%I:%M %p")
        hour, minute = dt.hour, dt.minute
    except ValueError:
        return None, "Invalid time format. Use HH:MM AM/PM."

    # Clean up existing task
    existing_task = PeriodicTask.objects.filter(name=task_name).first()
    if existing_task:
        old_cron = existing_task.crontab
        existing_task.delete()
        if not PeriodicTask.objects.filter(crontab=old_cron).exists():
            old_cron.delete()

    # Always ensure valid cron values
    cron, _ = CrontabSchedule.objects.get_or_create(
        minute=str(minute),
        hour=str(hour),
        day_of_week='*',
        day_of_month='*',
        month_of_year='*',
        timezone='Asia/Kolkata'
    )

    task = PeriodicTask.objects.create(
        name=task_name,
        crontab=cron,
        task=task_path,
        args=json.dumps([]),
        enabled=True
    )

    kolkata = timezone.now().astimezone(pytz.timezone('Asia/Kolkata'))
    return {
        "message": f"{task_name} scheduled at {time_str} (24-hour: {hour:02}:{minute:02})",
        "current_kolkata_time": kolkata.strftime('%Y-%m-%d %H:%M:%S'),
        "task_enabled": task.enabled
    }, None


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django_celery_beat.models import PeriodicTask, CrontabSchedule
import json
import uuid 

class SetAttendanceTaskTimeView(APIView):
    
    # --- GET: Show existing alarms ---
    def get(self, request):
        tasks = PeriodicTask.objects.filter(task="app1.tasks.import_attendance_from_excel")
        data = []
        for t in tasks:
            try:
                # Parse Time
                h = int(t.crontab.hour)
                m = int(t.crontab.minute)
                ampm = "AM" if h < 12 else "PM"
                h_12 = h if 0 < h <= 12 else h - 12
                if h == 0: h_12 = 12
                time_str = f"{h_12:02}:{m:02} {ampm}"

                # Parse Days (Crucial for your requirement)
                cron_days = t.crontab.day_of_week # e.g. "1,2,3" or "*"
                if cron_days == '*':
                    days_display = "Daily"
                    selected_days = [0,1,2,3,4,5,6]
                else:
                    # Convert "1,2,5" string to list [1,2,5]
                    selected_days = [int(d) for d in cron_days.split(',')]
                    # Helper to make nice text like "Mon, Wed"
                    day_map = {0: 'Sun', 1: 'Mon', 2: 'Tue', 3: 'Wed', 4: 'Thu', 5: 'Fri', 6: 'Sat'}
                    if len(selected_days) == 7:
                        days_display = "Daily"
                    elif len(selected_days) == 2 and 0 in selected_days and 6 in selected_days:
                        days_display = "Weekends"
                    elif len(selected_days) == 5 and 0 not in selected_days and 6 not in selected_days:
                        days_display = "Weekdays"
                    else:
                        days_display = ", ".join([day_map[d] for d in sorted(selected_days)])

                data.append({
                    "id": t.id,
                    "time": time_str,
                    "label": t.description or "Scheduled Import",
                    "repeat_text": days_display,
                    "selected_days": selected_days,
                    "enabled": t.enabled
                })
            except Exception as e:
                print(e)
                continue 
                
        return Response(data)

    # --- POST: Create New Alarm ---
    def post(self, request):
        time_str = request.data.get("time") 
        label = request.data.get("label", "Auto Sync")
        days = request.data.get("days", []) # Expects list: [1, 2, 3] (Mon,Tue,Wed)
        
        try:
            from datetime import datetime
            dt = datetime.strptime(time_str.strip(), "%I:%M %p")
            hour, minute = dt.hour, dt.minute
        except ValueError:
            return Response({"error": "Invalid time"}, status=400)

        # Logic for Days
        # Celery uses 0-6 (Sunday=0, Saturday=6)
        if len(days) == 7 or len(days) == 0:
            cron_days = '*' # Daily
        else:
            cron_days = ','.join(str(d) for d in days)

        # Create Schedule
        cron, _ = CrontabSchedule.objects.get_or_create(
            minute=str(minute),
            hour=str(hour),
            day_of_week=cron_days, 
            timezone='Asia/Kolkata'
        )

        unique_id = str(uuid.uuid4())[:8]
        task_name = f"excel_import_{unique_id}"

        task = PeriodicTask.objects.create(
            name=task_name,
            crontab=cron,
            task="app1.tasks.import_attendance_from_excel", 
            args=json.dumps([]),
            description=label,
            enabled=True
        )

        return Response({"success": True})

    # --- DELETE ---
    def delete(self, request):
        task_id = request.query_params.get("id")
        try:
            PeriodicTask.objects.get(id=task_id).delete()
            return Response({"success": True})
        except:
            return Response({"error": "Not found"}, status=404)

    # --- PATCH (Toggle) ---
    def patch(self, request):
        task_id = request.data.get("id")
        try:
            task = PeriodicTask.objects.get(id=task_id)
            task.enabled = request.data.get("enabled")
            task.save()
            return Response({"success": True})
        except:
            return Response({"error": "Not found"}, status=404)



from django.db.models import Count, Sum, Case, When, IntegerField, FloatField
from django.db.models.functions import Cast # <--- NEW IMPORT
# ... keep other imports ...

class MonthlySummaryView(APIView):
    def get(self, request):
        month = request.query_params.get('month')
        year = request.query_params.get('year')

        if not month or not year:
            return Response({'error': 'Month and Year required'}, status=400)

        summary = BiometricAttendance.objects.filter(
            attendance_date__year=year,
            attendance_date__month=month
        ).values('card_no', 'employee_name', 'department', 'designation').annotate(
            # 1. Counts
            total_present=Count(Case(When(status__istartswith='P', then=1), output_field=IntegerField())),
            total_absent=Count(Case(When(status__istartswith='A', then=1), output_field=IntegerField())),
            
            # 2. SUM of Hours (Converted from Text to Number)
            # Coalesce/Cast is used to handle empty strings safely
            total_early=Sum(Cast('early_arrival', FloatField())),
            total_late=Sum(Cast('late_arrival', FloatField())),
            total_ot_hours=Sum(Cast('ot', FloatField())) # <--- Sum of 'Ot' column, NOT 'Ot Amount'
            
        ).order_by('card_no')

        return Response(summary)
    



class EmployeeMonthlyDetailView(APIView):
    def get(self, request):
        card_no = request.query_params.get('card_no')
        month = request.query_params.get('month')
        year = request.query_params.get('year')

        if not card_no or not month or not year:
            return Response({'error': 'Missing parameters'}, status=400)

        # Fetch all logs for this person in this month
        logs = BiometricAttendance.objects.filter(
            card_no=card_no,
            attendance_date__year=year,
            attendance_date__month=month
        ).values('attendance_date', 'status', 'in_time', 'out_time', 'late_arrival', 'ot')

        return Response(logs)
    


import os
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import SystemSettings
from .serializers import SystemSettingsSerializer


class SystemSettingsView(APIView):
    """
    GET: Retrieve current system settings
    POST: Create settings if not exists
    PATCH: Update the excel source path with validation
    """

    def get(self, request):
        settings = SystemSettings.objects.first()
        if not settings:
            # Create default settings if none exist
            settings = SystemSettings.objects.create(excel_source_path="")
        
        serializer = SystemSettingsSerializer(settings)
        return Response(serializer.data)

    def patch(self, request):
        settings = SystemSettings.objects.first()
        if not settings:
            settings = SystemSettings.objects.create(excel_source_path="")

        new_path = request.data.get('excel_source_path', '').strip()
        
        # Validation 1: Check if path is provided
        if not new_path:
            return Response({
                'success': False,
                'error': 'Path cannot be empty'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Validation 2: Check if path exists on the system
        if not os.path.exists(new_path):
            return Response({
                'success': False,
                'error': f'Path does not exist on server: {new_path}',
                'suggestion': 'Please create the folder first, then try again.'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Validation 3: Check if it's a directory (not a file)
        if not os.path.isdir(new_path):
            return Response({
                'success': False,
                'error': 'The path must be a folder, not a file.'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Try to create Processed folder
        processed_path = os.path.join(new_path, "Processed")
        processed_created = False
        
        if not os.path.exists(processed_path):
            try:
                os.makedirs(processed_path)
                processed_created = True
            except Exception as e:
                return Response({
                    'success': False,
                    'error': f'Could not create Processed folder: {str(e)}',
                    'suggestion': 'Please create a "Processed" folder manually inside your source folder.'
                }, status=status.HTTP_400_BAD_REQUEST)

        # Save the path
        settings.excel_source_path = new_path
        settings.save()

        serializer = SystemSettingsSerializer(settings)
        return Response({
            'success': True,
            'message': 'Path updated successfully!',
            'processed_folder_created': processed_created,
            'data': serializer.data
        })


class ValidatePathView(APIView):
    """
    POST: Validate a path without saving it
    Used for real-time validation in frontend
    """

    def post(self, request):
        path = request.data.get('path', '').strip()
        
        if not path:
            return Response({
                'valid': False,
                'error': 'Path is empty'
            })

        path_exists = os.path.exists(path)
        is_directory = os.path.isdir(path) if path_exists else False
        processed_path = os.path.join(path, "Processed")
        processed_exists = os.path.exists(processed_path)

        # Count excel files in folder
        excel_count = 0
        if path_exists and is_directory:
            try:
                files = os.listdir(path)
                excel_count = len([f for f in files if f.endswith(('.xlsx', '.xls'))])
            except:
                pass

        return Response({
            'valid': path_exists and is_directory,
            'path_exists': path_exists,
            'is_directory': is_directory,
            'processed_folder_exists': processed_exists,
            'excel_files_count': excel_count,
            'processed_path': processed_path
        })




    

#================== BiometricAttendance End ==================#

# -----------------------ESSL Biometric Man-Machine interlinkage End----------------------------
# ======================= Biometric Realtime ===================================

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import BioUser, BiometricDevice
from .serializers import BioUserSerializer
# Import the NEW functions
from .utils import enroll_user_face_in_device, enroll_user_fp_in_device,  add_employee_to_specific_device

class BioUserViewSet(viewsets.ModelViewSet):
    queryset = BioUser.objects.all()
    serializer_class = BioUserSerializer

    # ---------------------------------------------------------
    # ACTION: Trigger Face Enrollment
    # ---------------------------------------------------------
    @action(detail=True, methods=['post'])
    def enroll_face(self, request, pk=None):
        biouser = self.get_object()
        
        # 1. Get the Target Device ID from the Request
        device_id = request.data.get('device_id')
        is_overwrite = request.data.get('is_overwrite', False)

        if not device_id:
            return Response(
                {'error': 'device_id is required. You must select a machine to enroll on.'}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # 2. Find the Device
            device_instance = BiometricDevice.objects.get(id=device_id)
            
            # 3. Call the Utility
            result = enroll_user_face_in_device(biouser, device_instance, is_overwrite=is_overwrite)
            
            # 4. Format Result
            if hasattr(result, 'EnrollUserFaceResult') and hasattr(result, 'CommandId'):
                result_dict = {
                    'EnrollUserFaceResult': result.EnrollUserFaceResult,
                    'CommandId': result.CommandId,
                    'Message': f"Please stand in front of {device_instance.name} now."
                }
            else:
                result_dict = str(result)

            return Response({'result': result_dict}, status=status.HTTP_200_OK)

        except BiometricDevice.DoesNotExist:
            return Response({'error': 'Device not found'}, status=status.HTTP_404_NOT_FOUND)

    # ---------------------------------------------------------
    # ACTION: Trigger Fingerprint Enrollment
    # ---------------------------------------------------------
    @action(detail=True, methods=['post'])
    def enroll_fingerprint(self, request, pk=None):
        biouser = self.get_object()
        
        # 1. Get Inputs
        device_id = request.data.get('device_id')
        finger_index = int(request.data.get('finger_index', 1))
        is_overwrite = request.data.get('is_overwrite', False)

        if not device_id:
            return Response(
                {'error': 'device_id is required.'}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # 2. Find Device
            device_instance = BiometricDevice.objects.get(id=device_id)

            # 3. Call Utility
            result = enroll_user_fp_in_device(
                biouser, 
                device_instance, 
                finger_index_number=finger_index, 
                is_overwrite=is_overwrite
            )

            # 4. Format Result
            if hasattr(result, 'EnrollUserFPResult') and hasattr(result, 'CommandId'):
                result_dict = {
                    'EnrollUserFPResult': result.EnrollUserFPResult,
                    'CommandId': result.CommandId,
                    'Message': f"Please place finger on {device_instance.name} scanner."
                }
            else:
                result_dict = str(result)
                
            return Response({'result': result_dict}, status=status.HTTP_200_OK)

        except BiometricDevice.DoesNotExist:
            return Response({'error': 'Device not found'}, status=status.HTTP_404_NOT_FOUND)


    @action(detail=True, methods=['post'])
    def sync_to_device(self, request, pk=None):
        """
        Manually push a user to a specific biometric device.
        """
        biouser = self.get_object()
        device_id = request.data.get('device_id')

        if not device_id:
            return Response({'error': 'device_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            device_instance = BiometricDevice.objects.get(id=device_id)
            
            # Call the utility function (Make sure APIKey is INCLUDED in this function in utils.py)
            success = add_employee_to_specific_device(biouser, device_instance)
            
            if success:
                # Log the enrollment to prevent duplicates later
                from .models import BiometricEnrollment
                BiometricEnrollment.objects.get_or_create(bio_user=biouser, device=device_instance)
                return Response({'status': 'success', 'message': f'User added to {device_instance.name}'}, status=200)
            else:
                return Response({'status': 'error', 'message': 'Failed to connect to device'}, status=500)

        except BiometricDevice.DoesNotExist:
            return Response({'error': 'Device not found'}, status=404)
        


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
# from datetime import datetime
import datetime
from .models import BiometricDevice
from .utils import get_transactions_log_from_device, save_log_entry

class AttendanceLogView(APIView):
    def get(self, request):
        """
        Fetch logs.
         Save to DB (History)
        Params:
        - date (optional): YYYY-MM-DD (Defaults to today)
        - device_id (optional): ID of specific device (Defaults to ALL devices)
        """
        # 1. Date Setup
        # date_str = request.query_params.get('date', datetime.now().strftime("%Y-%m-%d"))
        date_str = request.query_params.get('date', datetime.datetime.now().strftime("%Y-%m-%d"))
        from_dt = f"{date_str} 00:00:00"
        to_dt = f"{date_str} 23:59:59"

        # 2. Determine which devices to check
        device_id = request.query_params.get('device_id')
        
        if device_id:
            # Check only one specific device
            devices = BiometricDevice.objects.filter(id=device_id)
        else:
            # Check ALL devices
            devices = BiometricDevice.objects.all()

        all_logs = []

        # 3. Loop through devices and fetch logs
        for device in devices:
            print(f"Fetching logs from: {device.name}")
            result = get_transactions_log_from_device(device, from_dt, to_dt)
            
            # 4. Parse the messy string data from eSSL
            # Result usually comes as an object with 'strDataList'
            str_data = result.strDataList if hasattr(result, 'strDataList') else ""
            
            if str_data:
                log_lines = str_data.strip().split('\n')
                for line in log_lines:
                    if line.strip():
                        parts = line.split('\t')
                        # eSSL usually returns: EmpID, DateTime, StatusCode, etc.
                        if len(parts) >= 2:
                            emp_code = parts[0]
                            time_str = parts[1]

                            # --- A. ADD TO LIST (For Frontend Live View) ---
                            all_logs.append({
                                "device_name": device.name,  # Helpful to know which machine
                                "device_sn": device.serial_number,
                                "employee_code": emp_code,
                                "datetime": time_str,
                                # "status": parts[2] if len(parts) > 2 else "Unknown"
                            })

                            # --- B. SAVE TO DATABASE (For History) ---
                            # This saves the record or updates 'last_punch' if it exists
                            save_log_entry(device, emp_code, time_str)

        # 5. SORT: Latest First (Descending)
        all_logs.sort(key=lambda x: x['datetime'], reverse=True)

        return Response({
            "date": date_str,
            "device_count": len(devices),
            "total_logs": len(all_logs),
            "logs": all_logs
        }, status=status.HTTP_200_OK)


class AttendanceDatabaseView(APIView):
    """
    View for History Page.
    Returns: Daily Summary (default) or Detailed Timeline (type=detailed)
    """
    def get(self, request):
        date_str = request.query_params.get('date', datetime.datetime.now().strftime("%Y-%m-%d"))
        view_type = request.query_params.get('type', 'summary') # 'summary' or 'detailed'

        if view_type == 'detailed':
            # --- DETAILED TIMELINE (Every Punch) ---
            records = (BiometricPunch.objects.filter(punch_time__date=date_str)
                .select_related('bio_user', 'device')
                .prefetch_related('device__machines')
                .order_by('punch_time')) # Global Timeline sorted by time

            data = []
            for record in records:
                display_name = "Unknown Device"
                if record.device:
                    device_real_name = record.device.name
                    linked_machine = record.device.machines.first()
                    display_name = f"{linked_machine.name} ({device_real_name})" if linked_machine else device_real_name

                data.append({
                    "id": record.id,
                    "employee_code": record.bio_user.employeeid,
                    "employee_name": f"{record.bio_user.first_name} {record.bio_user.last_name}",
                    "device_name": display_name,
                    "time": record.punch_time.strftime("%H:%M:%S"),
                    "date": record.punch_time.strftime("%Y-%m-%d")
                })
        else:
            # --- DAILY SUMMARY (First/Last Punch) ---
            records = (AttendanceLog.objects.filter(date=date_str)
                .select_related('bio_user', 'device')
                .prefetch_related('device__machines') 
                .order_by('bio_user__employeeid'))

            data = []
            for record in records:
                last_punch_display = record.last_punch if record.last_punch != record.first_punch else "-"
                
                display_name = "Unknown Device"
                if record.device:
                    device_real_name = record.device.name
                    linked_machine = record.device.machines.first()
                    display_name = f"{linked_machine.name} ({device_real_name})" if linked_machine else device_real_name

                data.append({
                    "id": record.id,
                    "employee_code": record.bio_user.employeeid,
                    "employee_name": f"{record.bio_user.first_name} {record.bio_user.last_name}",
                    "device_name": display_name, 
                    "first_punch": record.first_punch.strftime("%H:%M:%S"),
                    "last_punch": last_punch_display.strftime("%H:%M:%S") if last_punch_display != "-" else "-",
                    "date": record.date
                })

        return Response({
            "date": date_str,
            "type": view_type,
            "count": len(data),
            "logs": data
        }, status=status.HTTP_200_OK)

    def post(self, request):
        """
        ACTION: Smart Cleanup (Prune records older than X days)
        """
        action = request.data.get('action')
        if action == 'cleanup':
            # 1. Permission Check (Developer/Admin only)
            # user = request.user
            # if not user.role or user.role.name != 'Developer':
            #     return Response({"error": "Unauthorized. Only Developers can prune logs."}, status=status.HTTP_403_FORBIDDEN)

            # 2. Cutoff Logic (Keep last 3 months / 90 days)
            days_to_keep = 90
            cutoff_date = datetime.datetime.now() - datetime.timedelta(days=days_to_keep)
            
            # 3. Delete
            deleted_count, _ = BiometricPunch.objects.filter(punch_time__lt=cutoff_date).delete()
            
            return Response({
                "status": "success",
                "message": f"Pruned {deleted_count} logs older than {days_to_keep} days.",
                "cutoff_date": cutoff_date.strftime("%Y-%m-%d")
            }, status=status.HTTP_200_OK)
        
        return Response({"error": "Invalid action"}, status=status.HTTP_400_BAD_REQUEST)
    


from .models import BiometricDevice, BiometricEnrollment
from .tasks import manual_sync_all_devices

class BiometricDeviceViewSet(viewsets.ModelViewSet):
    queryset = BiometricDevice.objects.all()
    serializer_class = BiometricDeviceSerializer

    @action(detail=False, methods=['post'])
    def sync_all(self, request):
        """
        Triggers a manual sync for all devices for a given date range.
        """
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')

        if not start_date or not end_date:
            return Response({"error": "start_date and end_date are required (YYYY-MM-DD)"}, status=status.HTTP_400_BAD_REQUEST)

        # Trigger Celery Task
        manual_sync_all_devices.delay(start_date, end_date)

        return Response({
            "message": f"Sync started in background for range: {start_date} to {end_date}. This may take a few minutes."
        }, status=status.HTTP_202_ACCEPTED)

class BiometricEnrollmentViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only view to see logs of who is on which machine.
    """
    queryset = BiometricEnrollment.objects.all().order_by('-synced_at')
    serializer_class = BiometricEnrollmentSerializer
    filterset_fields = ['bio_user__employeeid', 'device__id']



# ======================= Biometric Realtime End ===================================# your_app/views.py
# -----------------------ESSL Biometric Man-Machine interlinkage End----------------------------



# ==========================================ojt status start ================================#

# views.py
from rest_framework import viewsets
from django.db.models import Prefetch, Q
from .models import TraineeInfo, OJTScore, OJTDay
from .serializers import OJTStatusListSerializer


class OJTStatusDashboardViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = OJTStatusListSerializer

    def get_queryset(self):
        qs = TraineeInfo.objects.select_related('station', 'level').prefetch_related(
            Prefetch(
                'scores',
                queryset=OJTScore.objects.select_related('topic__department', 'topic__level', 'day')
            )
        )

        # Filter by level
        level_id = self.request.query_params.get('level_id')
        if level_id:
            try:
                qs = qs.filter(level_id=int(level_id))
            except ValueError:
                return TraineeInfo.objects.none()

        # Optional search
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(emp_id__iexact=search) | Q(trainee_name__icontains=search)
            )

        return qs
    
    
# ==========================================ojt status end ================================#

# ==========================================operator status start ================================#

# views.py  (add to the existing file)

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticatedOrReadOnly
from rest_framework.response import Response
from django.db.models import Q
from .models import OperatorObservanceSheet
from .serializers import OperatorObservanceSheetSerializer

# -------------------------------------------------
# 1. Compact list for the UI
# -------------------------------------------------
class SheetListSerializer(serializers.ModelSerializer):
    topics_count = serializers.IntegerField(read_only=True)

    class Meta:
        model = OperatorObservanceSheet
        fields = [
            'id', 'operator_name', 'operator_category', 'process_name',
            'supervisor_name', 'level', 'score', 'result',
            'evaluation_start_date', 'evaluation_end_date',
            'topics_count',
        ]

@api_view(['GET'])
@permission_classes([IsAuthenticatedOrReadOnly])
def sheet_list_by_level(request, level):
    """
    GET /observancesheets/list/<level>/
    level = 'Level 2' | 'Level 3' | 'Level 4'
    Returns a lightweight list of sheets for that level.
    """
    if level not in ['Level 2', 'Level 3', 'Level 4']:
        return Response({"detail": "Invalid level"}, status=400)

    qs = OperatorObservanceSheet.objects.filter(level=level)\
        .annotate(topics_count=Count('topics'))\
        .order_by('-evaluation_end_date')

    serializer = SheetListSerializer(qs, many=True)
    return Response(serializer.data)


# ==========================================operator status end ================================#







#=======================auto fetch biomatric to dashboard ===========================================


# from django.http import JsonResponse
# from django.utils import timezone
# from .models import BiometricAttendance, SkillMatrix, AdvanceManpowerDashboard

# def bifurcation_stats_view(request):
#     # ==============================================================================
#     # 1. Setup Dates & Parameters
#     # ==============================================================================
#     now = timezone.now()
#     today_date = now.date()
#     current_month = now.month
#     current_year = now.year

#     # Get Query Parameters
#     hq_id = request.GET.get('hq')
#     factory_id = request.GET.get('factory')
#     department_id = request.GET.get('department')
#     line_id = request.GET.get('line')
#     subline_id = request.GET.get('subline')
#     station_id = request.GET.get('station')

#     # ==============================================================================
#     # SECTION A: AVAILABLE (Live Data)
#     # ==============================================================================
    
#     # 1. Get IDs of employees present today
#     present_emp_ids = BiometricAttendance.objects.filter(
#         created_at__date=today_date
#     ).values_list('card_no', flat=True).distinct()

#     # 2. Base QuerySet for SkillMatrix based on attendance
#     available_qs = SkillMatrix.objects.filter(employee__emp_id__in=present_emp_ids)

#     # 3. Apply Hierarchy Filters to Available QS
#     if hq_id and hq_id != 'null': 
#         available_qs = available_qs.filter(hierarchy__hq_id=hq_id)
#     if factory_id and factory_id != 'null': 
#         available_qs = available_qs.filter(hierarchy__factory_id=factory_id)
#     if department_id and department_id != 'null': 
#         available_qs = available_qs.filter(hierarchy__department_id=department_id)
#     if line_id and line_id != 'null': 
#         available_qs = available_qs.filter(hierarchy__line_id=line_id)
#     if subline_id and subline_id != 'null': 
#         available_qs = available_qs.filter(hierarchy__subline_id=subline_id)
#     if station_id and station_id != 'null': 
#         available_qs = available_qs.filter(hierarchy__station_id=station_id)

#     # 4. Calculate Counts (Using DISTINCT to avoid duplicate employee counts)
#     # This ensures 1 person = 1 count, even if they have multiple skills 
#     # relevant to the current filter context.
#     l1_available = available_qs.filter(level__level_name__icontains="1").values('employee').distinct().count()
#     l2_available = available_qs.filter(level__level_name__icontains="2").values('employee').distinct().count()
#     l3_available = available_qs.filter(level__level_name__icontains="3").values('employee').distinct().count()
#     l4_available = available_qs.filter(level__level_name__icontains="4").values('employee').distinct().count()

#     # ==============================================================================
#     # SECTION B: REQUIRED (Budget Data with Overwrite Logic)
#     # ==============================================================================
    
#     # 1. Fetch ALL records for this Month/Year
#     dashboard_qs = AdvanceManpowerDashboard.objects.filter(
#         month=current_month,
#         year=current_year
#     )

#     # 2. Apply Hierarchy Filters to Required QS
#     if hq_id and hq_id != 'null': 
#         dashboard_qs = dashboard_qs.filter(hq_id=hq_id)
#     if factory_id and factory_id != 'null': 
#         dashboard_qs = dashboard_qs.filter(factory_id=factory_id)
#     if department_id and department_id != 'null': 
#         dashboard_qs = dashboard_qs.filter(department_id=department_id)
#     if line_id and line_id != 'null': 
#         dashboard_qs = dashboard_qs.filter(line_id=line_id)
#     if subline_id and subline_id != 'null': 
#         dashboard_qs = dashboard_qs.filter(subline_id=subline_id)
#     if station_id and station_id != 'null': 
#         dashboard_qs = dashboard_qs.filter(station_id=station_id)

#     # 3. Order by created_at ASC (Oldest -> Newest)
#     rows = dashboard_qs.order_by('created_at')

#     # 4. Deduplicate Logic
#     # Key: (Station, Subline, Line, Dept)
#     unique_records = {}

#     for row in rows:
#         key = (row.station_id, row.subline_id, row.line_id, row.department_id)
#         # Because we sort by created_at ASC, the last iteration for a specific key
#         # will be the newest record, effectively overwriting older uploads.
#         unique_records[key] = row

#     # 5. Calculate Sums from unique records
#     l1_required = 0
#     l2_required = 0
#     l3_required = 0
#     l4_required = 0

#     for row in unique_records.values():
#         l1_required += (row.l1_required or 0) # Handle potential NoneType safety
#         l2_required += (row.l2_required or 0)
#         l3_required += (row.l3_required or 0)
#         l4_required += (row.l4_required or 0)

#     # ==============================================================================
#     # SECTION C: RETURN DATA
#     # ==============================================================================
#     data = {
#         "l1_required": l1_required,
#         "l1_available": l1_available,
#         "l2_required": l2_required,
#         "l2_available": l2_available,
#         "l3_required": l3_required,
#         "l3_available": l3_available,
#         "l4_required": l4_required,
#         "l4_available": l4_available,
#     }

#     return JsonResponse(data)


# from django.http import JsonResponse
# from django.utils import timezone
# from django.db.models import Count
# from .models import BiometricAttendance, SkillMatrix, AdvanceManpowerDashboard

# def total_manpower_stats_view(request):
#     now = timezone.now()
#     target_fy_start = int(request.GET.get("year", now.year - 1 if now.month <= 3 else now.year))

#     filters = {}
#     for f in ['hq', 'factory', 'department', 'line', 'subline', 'station']:
#         val = request.GET.get(f)
#         if val and val != "null" and val.strip():
#             filters[f"hierarchy__{f}_id"] = val   # ← important: hierarchy relation

#     monthly_data = []

#     for month in range(1, 13):
#         year = target_fy_start if month >= 4 else target_fy_start + 1

#         # ── 1. Get present employees this month ───────────────────────────────
#         biometric_qs = BiometricAttendance.objects.filter(
#             attendance_date__year=year,
#             attendance_date__month=month,
#         ).exclude(status__in=['A', 'Absent', 'Leave', 'WO'])  # ← adjust exclusion

#         present_emp_codes = biometric_qs.values_list('pay_code', flat=True).distinct()

#         # ── 2. Match with SkillMatrix + hierarchy filters ─────────────────────
#         available_qs = SkillMatrix.objects.filter(
#             employee__emp_id__in=present_emp_codes
#         )

#         if filters:
#             available_qs = available_qs.filter(**filters)

#         total_available = available_qs.values('employee__emp_id').distinct().count()

#         # ── 3. Get required from dashboard (same filters) ─────────────────────
#         dashboard_qs = AdvanceManpowerDashboard.objects.filter(
#             year=year,
#             month=month,
#             **{k.replace("hierarchy__", ""): v for k, v in filters.items()}
#         )

#         # Deduplicate if multiple entries per hierarchy level
#         unique_dashboard = {}
#         for rec in dashboard_qs:
#             key = (rec.station_id, rec.subline_id, rec.line_id, rec.department_id, rec.factory_id)
#             unique_dashboard[key] = rec

#         total_required = sum(r.operators_required or 0 for r in unique_dashboard.values())

#         monthly_data.append({
#             "month": month,
#             "year": year,
#             "total_required": total_required,
#             "total_available": total_available,
#             "gap": total_required - total_available,
#         })

#     return JsonResponse(monthly_data, safe=False)



# ==============================================================================
# UNIFIED DASHBOARD API (Advance Manpower)
# ==============================================================================
class AdvanceDashboardUnifiedView(APIView):
    """
    Consolidated API for the Advance Manpower Dashboard.
    Returns:
      - card_stats: { total_stations, operators_required, ... }
      - manpower_trend: List of monthly { month, total_required, total_available }
      - attrition_trend: List of monthly { month, attrition_rate }
      - buffer_trend: List of monthly { month, buffer_req, buffer_avail }
      - absenteeism_trend: List of monthly { month, absenteeism_rate }
    """
    def get(self, request):
        now = timezone.now()
        req_year = request.query_params.get('year')
        try:
            target_year = int(req_year) if req_year else now.year
        except ValueError:
            target_year = now.year

        # --- 1. FILTER QUERYSETS ---
        # Base Dashboard Queryset
        dashboard_qs = AdvanceManpowerDashboard.objects.filter(year=target_year)
        dashboard_qs = apply_hierarchy_filters(request, dashboard_qs)

        # --- 2. CARD STATS (Aggregates) ---
        # Same logic as AdvanceCardStatsView
        card_aggs = dashboard_qs.aggregate(
            stations=Sum('total_stations'),
            req=Sum('operators_required'),
            avail=Sum('operators_available'),
            buf_req=Sum('buffer_manpower_required'),
            buf_avail=Sum('buffer_manpower_available'),
            # Bifurcation
            l1_req=Sum('l1_required'), l1_avail=Sum('l1_available'),
            l2_req=Sum('l2_required'), l2_avail=Sum('l2_available'),
            l3_req=Sum('l3_required'), l3_avail=Sum('l3_available'),
            l4_req=Sum('l4_required'), l4_avail=Sum('l4_available')
        )

        card_stats = {
            "total_stations": card_aggs['stations'] or 0,
            "operators_required": card_aggs['req'] or 0,
            "operators_available": card_aggs['avail'] or 0,
            "buffer_manpower_required": card_aggs['buf_req'] or 0,
            "buffer_manpower_available": card_aggs['buf_avail'] or 0,
            "bifurcation_plan_l1": card_aggs['l1_req'] or 0,
            "bifurcation_actual_l1": card_aggs['l1_avail'] or 0,
            "bifurcation_plan_l2": card_aggs['l2_req'] or 0,
            "bifurcation_actual_l2": card_aggs['l2_avail'] or 0,
            "bifurcation_plan_l3": card_aggs['l3_req'] or 0,
            "bifurcation_actual_l3": card_aggs['l3_avail'] or 0,
            "bifurcation_plan_l4": card_aggs['l4_req'] or 0,
            "bifurcation_actual_l4": card_aggs['l4_avail'] or 0,
        }

     
        hq_id = request.GET.get("hq")
        factory_id = request.GET.get("factory")
        department_id = request.GET.get("department")
        line_id = request.GET.get("line")
        subline_id = request.GET.get("subline")
        station_id = request.GET.get("station")

        manpower_trend = []
        
        # Pre-fetch dashboard rows for the whole year to avoid DB hits inside loop
        # We group them by month
        
        # Note: dashboard_qs is already filtered by hierarchy and year
        # We just need to group by month.
        rows_by_month = {}
        # Fetch only needed fields
        all_rows = dashboard_qs.values(
            'month', 'operators_required', 'buffer_manpower_required', 'buffer_manpower_available', 'attrition_rate', 'absenteeism_rate'
        )
        
        # Prepare data structures for other charts too
        buffer_trend_map = {} # month -> {req, avail}
        attrition_trend_map = {} # month -> rate (avg? or sum? Views used diff logic)
        absenteeism_trend_map = {} # month -> rate

        # Aggregate simple stats by month FIRST (Frequency Map)
        # Since we have multiple rows per month (one per station), we sum them up.
        
        # Django 'values().annotate()' is cleaner for this:
        monthly_aggs = dashboard_qs.values('month').annotate(
            req_ops=Sum('operators_required'),
            req_buf=Sum('buffer_manpower_required'),
            avail_buf=Sum('buffer_manpower_available'),
            avg_attr=Avg('attrition_rate'), # Using Avg for rates as per original implementation
            avg_abs=Avg('absenteeism_rate')
        )
        
        agg_map = {item['month']: item for item in monthly_aggs}

        for month in range(1, 13):
            # FINANCIAL YEAR LOGIC for Biometric Year
            if month >= 4:
                year_for_month = target_year
            else:
                year_for_month = target_year + 1
            
            # --- Biometric Availability (The Heavy Part) ---
            # There is no easy way to optimize this without avoiding the loop 
            # because BiometricAttendance table is massive and partitioned by date ideally.
            # We keep specific logic but ensure filters are applied efficiently.
            
            # Optimization: Try to use Count() directly instead of fetching IDs if possible,
            # but we need distinct employee_ids from SkillMatrix. 
            
            biometric_qs = BiometricAttendance.objects.filter(
                attendance_date__month=month,
                attendance_date__year=year_for_month
            ).exclude(status__istartswith="A")
            
            # Present pay codes
            present_pay_codes = biometric_qs.values_list("pay_code", flat=True).distinct()
            
            available_qs = SkillMatrix.objects.filter(employee__emp_id__in=present_pay_codes)
            
            if hq_id and hq_id != "null": available_qs = available_qs.filter(hierarchy__hq_id=hq_id)
            if factory_id and factory_id != "null": available_qs = available_qs.filter(hierarchy__factory_id=factory_id)
            if department_id and department_id != "null": available_qs = available_qs.filter(hierarchy__department_id=department_id)
            if line_id and line_id != "null": available_qs = available_qs.filter(hierarchy__line_id=line_id)
            if subline_id and subline_id != "null": available_qs = available_qs.filter(hierarchy__subline_id=subline_id)
            if station_id and station_id != "null": available_qs = available_qs.filter(hierarchy__station_id=station_id)

            total_available_bio = available_qs.values("employee__emp_id").distinct().count()
            
            # Get Aggregated Data for this month
            month_data = agg_map.get(month, {})
            
            # Manpower Trend Entry
            manpower_trend.append({
                "month": month,
                "total_required": month_data.get('req_ops', 0),
                "total_available": total_available_bio 
            })

        # --- 4. OTHER TRENDS (From Aggregations) ---
        attrition_trend = []
        buffer_trend = []
        absenteeism_trend = []

        # We can just iterate the agg_map which contains months present in DB
        # But front-end expects all months probably? The previous APIs returned ordered lists.
        # Let's populate the full list for chart consistency.
        
        # Note: The original chart logic returned data for the requested financial year duration?
        # ManpowerTrendChart uses a loop 1-12.
        # AttritionTrendChart uses `queryset.values('year', 'month')` so it returns only what exists.
        # To be safe, we return what exists in the aggs, sorted by month.
        
        # We need to sort months properly if we want them ordered.
        # Actually the frontend maps them to 'Apr', 'May' etc based on ID.
        
        for m in range(1, 13):
             # If no data for month m, we can send 0 or skip. 
             # Existing usage implies mapping by config.id.
             d = agg_map.get(m, {})
             
             attrition_trend.append({
                 "month": m,
                 "attrition_rate": round(float(d.get('avg_attr', 0) or 0), 2)
             })
             
             buffer_trend.append({
                 "month": m,
                 "buffer_manpower_required": d.get('req_buf', 0),
                 "buffer_manpower_available": d.get('avail_buf', 0)
             })
             
             absenteeism_trend.append({
                 "month": m,
                 "absenteeism_rate": round(float(d.get('avg_abs', 0) or 0), 2)
             })

        return Response({
            "card_stats": card_stats,
            "manpower_trend": manpower_trend,
            "attrition_trend": attrition_trend,
            "buffer_trend": buffer_trend,
            "absenteeism_trend": absenteeism_trend
        }, status=status.HTTP_200_OK)





# from django.db.models import Count, Case, When, IntegerField, Q
# from django.db.models.functions import ExtractMonth, ExtractYear
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
# import datetime
# from .models import BiometricAttendance # Ensure this is imported

# class AbsenteeismTrendView(APIView):
#     def get(self, request):
#         # 1. Get Financial Year Start (e.g., 2024)
#         req_year = request.query_params.get('year')
#         if req_year and req_year.isdigit():
#             fy_start = int(req_year)
#         else:
#             now = datetime.datetime.now()
#             fy_start = now.year - 1 if now.month <= 3 else now.year
        
#         fy_end = fy_start + 1

#         # 2. Filter for the full Financial Year (Apr FY_Start to Mar FY_End)
#         # This is the "Correct Financial Year" logic
#         queryset = BiometricAttendance.objects.filter(
#             (Q(attendance_date__year=fy_start, attendance_date__month__gte=4)) |
#             (Q(attendance_date__year=fy_end, attendance_date__month__lte=3))
#         )

#         # 3. Apply Hierarchy Filters (Mapping IDs to Department Names if necessary)
#         # Example: if you pass department ID, filter the string column
#         dept_name = request.query_params.get('department_name') 
#         if dept_name:
#             queryset = queryset.filter(department__icontains=dept_name)

#         # 4. Aggregate Data by Year and Month
#         summary = queryset.annotate(
#             month=ExtractMonth('attendance_date'),
#             year=ExtractYear('attendance_date')
#         ).values('year', 'month').annotate(
#             absent_count=Count(
#                 Case(When(status__istartswith='A', then=1), output_field=IntegerField())
#             ),
#             present_count=Count(
#                 Case(When(status__istartswith='P', then=1), output_field=IntegerField())
#             )
#         ).order_by('year', 'month')

#         # 5. Calculate Rates
#         formatted_data = []
#         for item in summary:
#             p = item['present_count'] or 0
#             a = item['absent_count'] or 0
#             total = p + a
            
#             rate = round((a / total * 100), 2) if total > 0 else 0.0

#             formatted_data.append({
#                 'month': item['month'],
#                 'year': item['year'],
#                 'absenteeism_rate': rate
#             })

#         return Response(formatted_data, status=status.HTTP_200_OK)


from django.http import JsonResponse
from django.utils import timezone
from django.db.models import Sum, Count, Q, Case, When, IntegerField, Avg
from django.db.models.functions import ExtractMonth, ExtractYear
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status


# ==============================================================================
# HELPER: Build hierarchy filters for SkillMatrix and AdvanceManpowerDashboard
# ==============================================================================
def get_hierarchy_filters(request):
    """
    Returns two dicts:
      - skill_filters    → for SkillMatrix queryset  (uses hierarchy__ prefix)
      - dashboard_filters → for AdvanceManpowerDashboard queryset (direct FK fields)
    """
    skill_filters = {}
    dashboard_filters = {}

    mapping = {
        'hq':         ('hierarchy__hq_id',         'hq_id'),
        'factory':    ('hierarchy__factory_id',     'factory_id'),
        'department': ('hierarchy__department_id',  'department_id'),
        'line':       ('hierarchy__line_id',        'line_id'),
        'subline':    ('hierarchy__subline_id',     'subline_id'),
        'station':    ('hierarchy__station_id',     'station_id'),
    }

    for param, (skill_key, dashboard_key) in mapping.items():
        val = request.GET.get(param)
        if val and val.strip() and val != 'null':
            skill_filters[skill_key] = val
            dashboard_filters[dashboard_key] = val

    return skill_filters, dashboard_filters


# ==============================================================================
# VIEW 1: bifurcation_stats_view  →  /chart/bifurcation-statslive/
# PROBLEM WAS: 4 separate COUNT queries for L1/L2/L3/L4
# FIX: 1 query using values + annotate grouping by level
# ==============================================================================
def bifurcation_stats_view(request):
    now = timezone.now()
    today_date = now.date()
    current_month = now.month
    current_year = now.year

    skill_filters, dashboard_filters = get_hierarchy_filters(request)

    # ------------------------------------------------------------------
    # SECTION A: AVAILABLE  (today's biometric → SkillMatrix)
    # card_no in BiometricAttendance matches emp_id in MasterTable
    # ------------------------------------------------------------------

    # Step 1: Get present employee card_nos as a SUBQUERY (stays in SQL)
    present_cards = BiometricAttendance.objects.filter(
        attendance_date=today_date
    ).values('card_no')  # kept as subquery, NOT .values_list(..., flat=True)

    # Step 2: Filter SkillMatrix where employee's emp_id is in present cards
    available_qs = SkillMatrix.objects.filter(
        employee__emp_id__in=present_cards,
        **skill_filters
    )

    # Step 3: ONE query — group by level_name, count distinct employees
    level_counts_qs = (
        available_qs
        .values('level__level_name')
        .annotate(total=Count('employee_id', distinct=True))
    )

    # Step 4: Build a dict  { 'Level 1': 5, 'Level 2': 3, ... }
    level_count_map = {}
    for row in level_counts_qs:
        level_count_map[row['level__level_name']] = row['total']

    # Step 5: Match level names flexibly (handles "Level 1", "L1", "1" etc.)
    def get_level_count(n):
        for key, val in level_count_map.items():
            if str(n) in str(key):
                return val
        return 0

    l1_available = get_level_count(1)
    l2_available = get_level_count(2)
    l3_available = get_level_count(3)
    l4_available = get_level_count(4)

    # ------------------------------------------------------------------
    # SECTION B: REQUIRED  (AdvanceManpowerDashboard — same as before)
    # ------------------------------------------------------------------
    dashboard_qs = AdvanceManpowerDashboard.objects.filter(
        month=current_month,
        year=current_year,
        **dashboard_filters
    ).order_by('created_at')

    unique_records = {}
    for row in dashboard_qs:
        key = (row.station_id, row.subline_id, row.line_id, row.department_id)
        unique_records[key] = row

    l1_required = sum(r.l1_required or 0 for r in unique_records.values())
    l2_required = sum(r.l2_required or 0 for r in unique_records.values())
    l3_required = sum(r.l3_required or 0 for r in unique_records.values())
    l4_required = sum(r.l4_required or 0 for r in unique_records.values())

    return JsonResponse({
        "l1_required": l1_required, "l1_available": l1_available,
        "l2_required": l2_required, "l2_available": l2_available,
        "l3_required": l3_required, "l3_available": l3_available,
        "l4_required": l4_required, "l4_available": l4_available,
    })


# ==============================================================================
# VIEW 2: current_manpower_card_view  →  /chart/current-stats/
# PROBLEM WAS: Same 4-query pattern as bifurcation
# FIX: Same single-query grouping approach
# ==============================================================================
def current_manpower_card_view(request):
    now = timezone.now()
    today_date = now.date()
    current_month = now.month
    current_year = now.year

    skill_filters, dashboard_filters = get_hierarchy_filters(request)

    # PART A: AVAILABLE
    present_cards = BiometricAttendance.objects.filter(
        attendance_date=today_date
    ).values('card_no')

    available_qs = SkillMatrix.objects.filter(
        employee__emp_id__in=present_cards,
        **skill_filters
    )

    level_counts_qs = (
        available_qs
        .values('level__level_name')
        .annotate(total=Count('employee_id', distinct=True))
    )

    level_count_map = {row['level__level_name']: row['total'] for row in level_counts_qs}

    def get_level_count(n):
        for key, val in level_count_map.items():
            if str(n) in str(key):
                return val
        return 0

    operators_available_total = (
        get_level_count(1) + get_level_count(2) +
        get_level_count(3) + get_level_count(4)
    )

    # PART B: REQUIRED
    dashboard_qs = AdvanceManpowerDashboard.objects.filter(
        month=current_month,
        year=current_year,
        **dashboard_filters
    ).order_by('created_at')

    unique_records = {}
    for row in dashboard_qs:
        key = (row.station_id, row.subline_id, row.line_id, row.department_id, row.factory_id)
        unique_records[key] = row

    total_stations = 0
    buffer_manpower_required = 0
    buffer_manpower_available = 0
    operators_required_total = 0

    for row in unique_records.values():
        total_stations            += (row.total_stations or 0)
        buffer_manpower_required  += (row.buffer_manpower_required or 0)
        buffer_manpower_available += (row.buffer_manpower_available or 0)
        operators_required_total  += (
            (row.l1_required or 0) + (row.l2_required or 0) +
            (row.l3_required or 0) + (row.l4_required or 0)
        )

    return JsonResponse({
        "total_stations": total_stations,
        "operators_required": operators_required_total,
        "operators_available": operators_available_total,
        "buffer_manpower_required": buffer_manpower_required,
        "buffer_manpower_available": buffer_manpower_available,
    })


# ==============================================================================
# VIEW 3: total_manpower_stats_view  →  /chart/total-stats/
# PROBLEM WAS: 12-iteration loop = 60 DB queries
# FIX: 3 total DB queries for the full financial year
def get_fy_date_range(fy_start):
    """
    Returns (date_from, date_to) as Python date objects.
    FY runs April 1 of fy_start  →  March 31 of fy_start+1
    SQLite can use its index on a plain DateField when filtering
    with __gte / __lte on actual date objects (no EXTRACT needed).
    """
    date_from = datetime.date(fy_start, 4, 1)
    date_to   = datetime.date(fy_start + 1, 3, 31)
    return date_from, date_to


# ==============================================================================
# VIEW 1: total_manpower_stats_view  →  /chart/total-stats/
# ==============================================================================
def total_manpower_stats_view(request):
    now = timezone.now()
    try:
        target_fy_start = int(request.GET.get("year", 0)) or (
            now.year - 1 if now.month <= 3 else now.year
        )
    except (ValueError, TypeError):
        target_fy_start = now.year - 1 if now.month <= 3 else now.year

    fy_end = target_fy_start + 1
    skill_filters, dashboard_filters = get_hierarchy_filters(request)

    # KEY FIX: use __gte/__lte on real date objects
    # SQLite indexes DateField directly — no EXTRACT, no full table scan
    date_from, date_to = get_fy_date_range(target_fy_start)

    # ------------------------------------------------------------------
    # QUERY 1: All present (card_no, attendance_date) in FY — one query
    # Filter by date range directly, exclude absent statuses
    # Then extract month/year in Python (cheaper than doing it in SQLite)
    # ------------------------------------------------------------------
    present_rows = (
        BiometricAttendance.objects
        .filter(
            attendance_date__gte=date_from,
            attendance_date__lte=date_to,
        )
        .exclude(status__in=['A', 'Absent', 'Leave', 'WO', 'HD'])
        .values('card_no', 'attendance_date')   # only fetch 2 columns
        .distinct()
    )

    # Build { (month, year): set(card_nos) } in Python — fast
    monthly_cards = defaultdict(set)
    for row in present_rows:
        d = row['attendance_date']
        monthly_cards[(d.month, d.year)].add(row['card_no'])

    # ------------------------------------------------------------------
    # QUERY 2: All skill-matrix emp_ids passing hierarchy filters
    # ------------------------------------------------------------------
    skill_emp_ids = set(
        SkillMatrix.objects
        .filter(**skill_filters)
        .values_list('employee__emp_id', flat=True)
        .distinct()
    )

    # ------------------------------------------------------------------
    # QUERY 3: Required from dashboard — single aggregation
    # ------------------------------------------------------------------
    req_qs = (
        AdvanceManpowerDashboard.objects
        .filter(
            Q(year=target_fy_start, month__gte=4) |
            Q(year=fy_end,          month__lte=3),
            **dashboard_filters
        )
        .values('month', 'year')
        .annotate(total_required=Sum('operators_required'))
    )
    req_map = {
        (row['month'], row['year']): (row['total_required'] or 0)
        for row in req_qs
    }

    # ------------------------------------------------------------------
    # Assemble — pure Python, zero extra DB calls
    # ------------------------------------------------------------------
    monthly_data = []
    for month in range(1, 13):
        year = target_fy_start if month >= 4 else fy_end
        present_this_month = monthly_cards.get((month, year), set())
        available = len(present_this_month & skill_emp_ids)
        required  = req_map.get((month, year), 0)
        monthly_data.append({
            "month": month,
            "year": year,
            "total_required": required,
            "total_available": available,
            "gap": required - available,
        })

    return JsonResponse(monthly_data, safe=False)
# ==============================================================================
# VIEW 4: AbsenteeismTrendView  →  /chart/absenteeism-trendlive/
# PROBLEM WAS: Loop doing biometric queries per month + complex hierarchy join
# FIX: Single aggregation query, hierarchy filter via department name match
# ==============================================================================
class AbsenteeismTrendView(APIView):
    def get(self, request):
        try:
            fy_start = int(request.query_params.get('year', 0))
        except (ValueError, TypeError):
            fy_start = 0

        if not fy_start:
            now = timezone.now()
            fy_start = now.year - 1 if now.month <= 3 else now.year

        date_from, date_to = get_fy_date_range(fy_start)

        # Base queryset — date range filter (SQLite-friendly, no EXTRACT)
        qs = BiometricAttendance.objects.filter(
            attendance_date__gte=date_from,
            attendance_date__lte=date_to,
        )

        # Department filter via name lookup
        dept_id = request.query_params.get('department')
        if dept_id and dept_id != 'null':
            try:
                from .models import Department
                dept_obj = Department.objects.get(pk=dept_id)
                qs = qs.filter(department__iexact=dept_obj.department_name)
            except Exception:
                pass

        # Fetch only the 2 columns we need — reduces data transfer from DB
        rows = qs.values('attendance_date', 'status')

        # Aggregate in Python — avoids SQLite EXTRACT slowness entirely
        monthly_totals  = defaultdict(int)
        monthly_absents = defaultdict(int)

        for row in rows:
            d = row['attendance_date']
            if d is None:
                continue
            key = d.month
            monthly_totals[key]  += 1
            if row['status'] and row['status'].upper().startswith('A'):
                monthly_absents[key] += 1

        # Build result for all 12 months
        result = []
        for month in range(1, 13):
            year  = fy_start if month >= 4 else fy_start + 1
            total  = monthly_totals.get(month, 0)
            absent = monthly_absents.get(month, 0)
            rate   = round((absent / total * 100), 2) if total > 0 else 0.0
            result.append({
                "month": month,
                "year": year,
                "absenteeism_rate": rate,
            })

        return Response(result, status=status.HTTP_200_OK)
        
class BufferManpowerChartView(APIView):
    def get(self, request):
        now = timezone.now()
        try:
            target_fy_start = int(request.GET.get("year", 0)) or (
                now.year - 1 if now.month <= 3 else now.year
            )
        except (ValueError, TypeError):
            target_fy_start = now.year - 1 if now.month <= 3 else now.year

        fy_end = target_fy_start + 1
        skill_filters, dashboard_filters = get_hierarchy_filters(request)

        date_from, date_to = get_fy_date_range(target_fy_start)

        # ------------------------------------------------------------------
        # QUERY 1: Present rows — date range filter (SQLite-friendly)
        # ------------------------------------------------------------------
        present_rows = (
            BiometricAttendance.objects
            .filter(
                attendance_date__gte=date_from,
                attendance_date__lte=date_to,
            )
            .exclude(status__in=['A', 'Absent', 'Leave', 'WO', 'HD'])
            .values('card_no', 'attendance_date')
            .distinct()
        )

        monthly_cards = defaultdict(set)
        for row in present_rows:
            d = row['attendance_date']
            monthly_cards[(d.month, d.year)].add(row['card_no'])

        # ------------------------------------------------------------------
        # QUERY 2: Skill-matrix emp_ids
        # ------------------------------------------------------------------
        skill_emp_ids = set(
            SkillMatrix.objects
            .filter(**skill_filters)
            .values_list('employee__emp_id', flat=True)
            .distinct()
        )

        # ------------------------------------------------------------------
        # QUERY 3: Dashboard buffer data
        # ------------------------------------------------------------------
        dashboard_qs = (
            AdvanceManpowerDashboard.objects
            .filter(
                Q(year=target_fy_start, month__gte=4) |
                Q(year=fy_end,          month__lte=3),
                **dashboard_filters
            )
            .values('month', 'year')
            .annotate(
                total_required=Sum('operators_required'),
                buffer_req=Sum('buffer_manpower_required'),
            )
        )
        dashboard_map = {
            (row['month'], row['year']): row for row in dashboard_qs
        }

        # ------------------------------------------------------------------
        # Assemble
        # ------------------------------------------------------------------
        months_sequence = [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3]
        monthly_data = []

        for month in months_sequence:
            year = target_fy_start if month >= 4 else fy_end
            present_this_month = monthly_cards.get((month, year), set())
            total_available    = len(present_this_month & skill_emp_ids)
            dash               = dashboard_map.get((month, year), {})
            total_required     = dash.get('total_required', 0) or 0
            buffer_req         = dash.get('buffer_req', 0) or 0
            gap                = max(0, total_required - total_available)

            monthly_data.append({
                "month": month,
                "year": year,
                "buffer_manpower_required": buffer_req,
                "buffer_manpower_available": gap,
            })

        return Response(monthly_data, status=status.HTTP_200_OK)




from django.db.models import Q, Sum
from django.utils import timezone
from django.http import JsonResponse

# def current_manpower_card_view(request):
#     now = timezone.now()
#     current_month = now.month
#     current_year = now.year

#     # Filters
#     hq_id = request.GET.get('hq')
#     factory_id = request.GET.get('factory')
#     department_id = request.GET.get('department')
#     line_id = request.GET.get('line')
#     subline_id = request.GET.get('subline')
#     station_id = request.GET.get('station')

#     # ====================================================
#     # PART 1: CALCULATE AVAILABLE (Real-time from Biometric)
#     # ====================================================
    
#     # 1. Get people present THIS month
#     biometric_qs = BiometricAttendance.objects.filter(
#         attendance_date__month=current_month,
#         attendance_date__year=current_year
#     ).exclude(status__istartswith='A') # Exclude Absent

#     present_pay_codes = biometric_qs.values_list('pay_code', flat=True).distinct()

#     # 2. Filter SkillMatrix for these people
#     available_qs = SkillMatrix.objects.filter(
#         employee__emp_id__in=present_pay_codes
#     )

#     # 3. Apply Hierarchy Filters to SkillMatrix
#     if hq_id and hq_id != 'null': available_qs = available_qs.filter(hierarchy__hq_id=hq_id)
#     if factory_id and factory_id != 'null': available_qs = available_qs.filter(hierarchy__factory_id=factory_id)
#     if department_id and department_id != 'null': available_qs = available_qs.filter(hierarchy__department_id=department_id)
#     if line_id and line_id != 'null': available_qs = available_qs.filter(hierarchy__line_id=line_id)
#     if subline_id and subline_id != 'null': available_qs = available_qs.filter(hierarchy__subline_id=subline_id)
#     if station_id and station_id != 'null': available_qs = available_qs.filter(hierarchy__station_id=station_id)

#     # 4. Get the Count
#     calculated_available = available_qs.values('employee__emp_id').distinct().count()


#     # ====================================================
#     # PART 2: GET REQUIRED (From Dashboard Plan)
#     # ====================================================
#     dashboard_qs = AdvanceManpowerDashboard.objects.filter(month=current_month, year=current_year)

#     if hq_id and hq_id != 'null': dashboard_qs = dashboard_qs.filter(hq_id=hq_id)
#     if factory_id and factory_id != 'null': dashboard_qs = dashboard_qs.filter(factory_id=factory_id)
#     if department_id and department_id != 'null': dashboard_qs = dashboard_qs.filter(department_id=department_id)
#     if line_id and line_id != 'null': dashboard_qs = dashboard_qs.filter(line_id=line_id)
#     if subline_id and subline_id != 'null': dashboard_qs = dashboard_qs.filter(subline_id=subline_id)
#     if station_id and station_id != 'null': dashboard_qs = dashboard_qs.filter(station_id=station_id)

#     # Deduplication Logic (To avoid double counting stations)
#     rows = dashboard_qs.order_by('created_at')
#     unique_records = {}
#     for row in rows:
#         key = (row.station_id, row.subline_id, row.line_id, row.department_id, row.factory_id)
#         unique_records[key] = row

#     # Summing up the required fields
#     total_stations = 0
#     operators_required = 0
#     buffer_manpower_required = 0
    
#     # Note: We fetch 'buffer available' from the dashboard if you manually enter it there.
#     # If you want to calculate it from Biometric, you need specific logic (e.g., Designation check).
#     buffer_manpower_available = 0 

#     for row in unique_records.values():
#         total_stations += (row.total_stations or 0)
#         operators_required += (row.operators_required or 0)
#         buffer_manpower_required += (row.buffer_manpower_required or 0)
#         buffer_manpower_available += (row.buffer_manpower_available or 0)

#     # ====================================================
#     # PART 3: RETURN JSON OBJECT
#     # ====================================================
#     data = {
#         "total_stations": total_stations,
#         "operators_required": operators_required,
#         "operators_available": calculated_available, # From Biometric Calc
#         "buffer_manpower_required": buffer_manpower_required,
#         "buffer_manpower_available": buffer_manpower_available, # From Dashboard Entry (or 0)
#     }

#     return JsonResponse(data)

# from django.http import JsonResponse
# from django.utils import timezone
# from django.db.models import Sum
# # Ensure you import your models: BiometricAttendance, SkillMatrix, AdvanceManpowerDashboard

# def current_manpower_card_view(request):
#     # ==============================================================================
#     # 1. Setup Dates & Parameters
#     # ==============================================================================
#     now = timezone.now()
#     today_date = now.date()
#     current_month = now.month
#     current_year = now.year

#     # Get Query Parameters
#     hq_id = request.GET.get('hq')
#     factory_id = request.GET.get('factory')
#     department_id = request.GET.get('department')
#     line_id = request.GET.get('line')
#     subline_id = request.GET.get('subline')
#     station_id = request.GET.get('station')

#     # ==============================================================================
#     # PART A: CALCULATE AVAILABLE (Live Data -> Sum of L1+L2+L3+L4)
#     # ==============================================================================
    
#     # 1. Get IDs of employees present TODAY (Live Logic)
#     present_emp_ids = BiometricAttendance.objects.filter(
#         created_at__date=today_date
#     ).values_list('card_no', flat=True).distinct()

#     # 2. Base QuerySet for SkillMatrix based on attendance
#     available_qs = SkillMatrix.objects.filter(employee__emp_id__in=present_emp_ids)

#     # 3. Apply Hierarchy Filters to Available QS
#     if hq_id and hq_id != 'null': 
#         available_qs = available_qs.filter(hierarchy__hq_id=hq_id)
#     if factory_id and factory_id != 'null': 
#         available_qs = available_qs.filter(hierarchy__factory_id=factory_id)
#     if department_id and department_id != 'null': 
#         available_qs = available_qs.filter(hierarchy__department_id=department_id)
#     if line_id and line_id != 'null': 
#         available_qs = available_qs.filter(hierarchy__line_id=line_id)
#     if subline_id and subline_id != 'null': 
#         available_qs = available_qs.filter(hierarchy__subline_id=subline_id)
#     if station_id and station_id != 'null': 
#         available_qs = available_qs.filter(hierarchy__station_id=station_id)

#     # 4. Calculate Counts for L1, L2, L3, L4 individually
#     l1_available = available_qs.filter(level__level_name__icontains="1").values('employee').distinct().count()
#     l2_available = available_qs.filter(level__level_name__icontains="2").values('employee').distinct().count()
#     l3_available = available_qs.filter(level__level_name__icontains="3").values('employee').distinct().count()
#     l4_available = available_qs.filter(level__level_name__icontains="4").values('employee').distinct().count()

#     # 5. Sum them up for Total Operators Available
#     operators_available_total = l1_available + l2_available + l3_available + l4_available

#     # ==============================================================================
#     # PART B: CALCULATE REQUIRED (Budget Data -> Sum of L1+L2+L3+L4)
#     # ==============================================================================
#     dashboard_qs = AdvanceManpowerDashboard.objects.filter(
#         month=current_month, 
#         year=current_year
#     )

#     # Apply Hierarchy Filters to Required QS
#     if hq_id and hq_id != 'null': 
#         dashboard_qs = dashboard_qs.filter(hq_id=hq_id)
#     if factory_id and factory_id != 'null': 
#         dashboard_qs = dashboard_qs.filter(factory_id=factory_id)
#     if department_id and department_id != 'null': 
#         dashboard_qs = dashboard_qs.filter(department_id=department_id)
#     if line_id and line_id != 'null': 
#         dashboard_qs = dashboard_qs.filter(line_id=line_id)
#     if subline_id and subline_id != 'null': 
#         dashboard_qs = dashboard_qs.filter(subline_id=subline_id)
#     if station_id and station_id != 'null': 
#         dashboard_qs = dashboard_qs.filter(station_id=station_id)

#     # Deduplication Logic (To avoid double counting stations due to multiple uploads)
#     # We order by created_at so the last item in the loop is the latest record.
#     rows = dashboard_qs.order_by('created_at')
#     unique_records = {}
    
#     for row in rows:
#         key = (row.station_id, row.subline_id, row.line_id, row.department_id, row.factory_id)
#         unique_records[key] = row

#     # Initialize counters
#     total_stations = 0
#     buffer_manpower_required = 0
#     buffer_manpower_available = 0
    
#     # Individual Required Levels
#     l1_required_sum = 0
#     l2_required_sum = 0
#     l3_required_sum = 0
#     l4_required_sum = 0

#     for row in unique_records.values():
#         total_stations += (row.total_stations or 0)
#         buffer_manpower_required += (row.buffer_manpower_required or 0)
#         buffer_manpower_available += (row.buffer_manpower_available or 0)
        
#         # Calculate Required Sums (handling NoneTypes)
#         l1_required_sum += (row.l1_required or 0)
#         l2_required_sum += (row.l2_required or 0)
#         l3_required_sum += (row.l3_required or 0)
#         l4_required_sum += (row.l4_required or 0)

#     # Sum them up for Total Operators Required
#     operators_required_total = l1_required_sum + l2_required_sum + l3_required_sum + l4_required_sum

#     # ==============================================================================
#     # PART C: RETURN JSON OBJECT
#     # ==============================================================================
#     data = {
#         "total_stations": total_stations,
#         "operators_required": operators_required_total, # Sum of L1-L4 Required
#         "operators_available": operators_available_total, # Sum of L1-L4 Available
#         "buffer_manpower_required": buffer_manpower_required,
#         "buffer_manpower_available": buffer_manpower_available,
#     }

#     return JsonResponse(data)




# =================================================== end ===============================  

# ===== ACTION PLAN ===================

# app/views.py
from rest_framework import generics
from .models import ActionItem, ActionItemRejection
from .serializers import ActionItemSerializer, ActionItemRejectionSerializer

# Handles GET (List) and POST (Add)
class ActionItemListCreateView(generics.ListCreateAPIView):
    queryset = ActionItem.objects.all().order_by('-date')
    serializer_class = ActionItemSerializer

# Handles PUT (Update) and DELETE (Remove) for a specific ID
class ActionItemDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ActionItem.objects.all()
    serializer_class = ActionItemSerializer



# Handles GET (List) and POST (Add)
class ActionItemRejectionListCreateView(generics.ListCreateAPIView):
    queryset = ActionItemRejection.objects.all().order_by('-date')
    serializer_class = ActionItemRejectionSerializer

# Handles PUT (Update) and DELETE (Remove) for a specific ID
class ActionItemRejectionDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ActionItemRejection.objects.all()
    serializer_class = ActionItemRejectionSerializer

# ============== END ==================

class SkillMatrixDashboardUnifiedView(APIView):
    """
    Unified API to fetch ALL data needed for Skill Matrix Page in one go.
    Replaces 6 separate API calls.
    """
    def get(self, request, *args, **kwargs):
        try:
            logger.info("🚀 SkillMatrixDashboardUnifiedView: Starting fetch...")
            start_time = timezone.now()

            # 1. Fetch Hierarchy (Optimized)
            # ---------------------------------------------------
            hierarchy_query = HierarchyStructure.objects.all().select_related(
                'hq', 'factory', 'department', 'line', 'subline', 'station'
            )
            
            # Group hierarchy by department for frontend (matches existing structure)
            hierarchy_data = []
            departments_map = {} # Helper to group stations/lines by department

            for item in hierarchy_query:
                dept_id = item.department.department_id if item.department else None
                if not dept_id: continue

                if dept_id not in departments_map:
                    departments_map[dept_id] = {
                        "department_id": dept_id,
                        "department_name": item.department.department_name,
                        "lines": [],
                        "sublines": [],
                        "stations": []
                    }
                
                dept = departments_map[dept_id]
                
                # Add Line
                if item.line:
                    line_exists = any(l['line_id'] == item.line.line_id for l in dept['lines'])
                    if not line_exists:
                        dept['lines'].append({
                            "line_id": item.line.line_id,
                            "line_name": item.line.line_name,
                            "sublines": [], # Populated if needed
                            "stations": []
                        })

                # Add Subline
                if item.subline:
                    subline_exists = any(sl['subline_id'] == item.subline.subline_id for sl in dept['sublines'])
                    if not subline_exists:
                        dept['sublines'].append({
                            "subline_id": item.subline.subline_id,
                            "subline_name": item.subline.subline_name,
                            "stations": []
                        })

                # Add Station
                if item.station:
                    station_exists = any(st['station_id'] == item.station.station_id for st in dept['stations'])
                    if not station_exists:
                        dept['stations'].append({
                            "station_id": item.station.station_id,
                            "station_name": item.station.station_name
                        })

            hierarchy_data = list(departments_map.values())


            # 2. Fetch Skill Matrices (Optimized)
            # ---------------------------------------------------
            # Fetch all skill matrices
            skill_matrix_query = SkillMatrix.objects.all().select_related(
                'employee', 'level', 'hierarchy__station', 'hierarchy__department'
            )
            
            skill_matrices_data = []
            for sm in skill_matrix_query:
                skill_matrices_data.append({
                    "id": sm.id,
                    "employee_name": sm.employee_name,
                    "emp_id": sm.emp_id,
                    "doj": sm.doj,
                    "level": sm.level.level_id if sm.level else 0, # Assuming level_id is the numeric value
                    "department": sm.hierarchy.department.department_name if sm.hierarchy and sm.hierarchy.department else "",
                    "station_id": sm.hierarchy.station.station_id if sm.hierarchy and sm.hierarchy.station else None,
                    "station_name": sm.hierarchy.station.station_name if sm.hierarchy and sm.hierarchy.station else "",
                })


            # 3. Fetch Employees (Master List)
            # ---------------------------------------------------
            # Replaces fetchMonthlySkill
            employees_query = MasterTable.objects.all().select_related('department')
            employees_data = []
            for emp in employees_query:
                employees_data.append({
                    "employee_code": emp.emp_id,
                    "full_name": f"{emp.first_name} {emp.last_name}",
                    "designation": emp.designation,
                    "date_of_join": emp.date_of_joining,
                    "department": emp.department.department_name if emp.department else "",
                    # "section": emp.sub_department_name # Or link to line
                })


            # 4. Fetch Operations (Stations List)
            # ---------------------------------------------------
            # Simple list of all stations for dropdowns/mapping
            stations_query = Station.objects.all()
            operations_data = [{"id": s.station_id, "name": s.station_name} for s in stations_query]


            # 5. Fetch Sections (Departments List)
            # ---------------------------------------------------
            sections_query = Department.objects.all()
            sections_data = [{"id": d.department_id, "name": d.department_name} for d in sections_query]


            # 6. Fetch Station Requirements (Optimized)
            # ---------------------------------------------------
            requirements_query = StationManager.objects.all().select_related('station', 'station__station')
            requirements_data = []
            level_mapping = {'Level 1': 1, 'Level 2': 2, 'Level 3': 3, 'Level 4': 4}
            
            for req in requirements_query:
                # Direct link via StationManager logic (reused from get_station_requirements)
                current_station_pk = req.station.station.pk if req.station and req.station.station else None
                if not current_station_pk: continue

                requirements_data.append({
                    'id': req.id,
                    'station_id': current_station_pk,
                    'minimum_operators': req.minimum_operators,
                    'minimum_level_required': req.minimum_level_required,
                    'minimum_level_name': level_mapping.get(req.minimum_level_required, 0)
                })


            # 7. Metadata (Colors, Settings)
            # ---------------------------------------------------
            colors = list(LevelColour.objects.all().values('level__level_id', 'colour_code')) # Assuming level_id links to number
            display_setting = SkillMatrixDisplaySetting.objects.first()
            metadata = {
                "level_colors": {c['level__level_id']: c['colour_code'] for c in colors},
                "display_shape": display_setting.display_shape if display_setting else 'piechart'
            }


            logger.info(f"✅ Unified View completed in {(timezone.now() - start_time).total_seconds()}s")

            return Response({
                "hierarchy": hierarchy_data,
                "skill_matrices": skill_matrices_data,
                "employees": employees_data,
                "operations": operations_data,
                "sections": sections_data,
                "station_requirements": requirements_data,
                "metadata": metadata
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"❌ Error in Unified View: {str(e)}", exc_info=True)
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)





# ==================================poison cake test=================================
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Q, Count, Sum
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from datetime import timedelta
from rest_framework.permissions import AllowAny

from .models import (
    DefectCategory,
    DefectType,
    PoisonCakeTest,
    PlantedDefect,
    MasterTable,
    Department,
    Station
)
from .serializers import (
    DefectCategorySerializer,
    DefectTypeSerializer,
    DefectTypeListSerializer,
    OperatorSerializer,
    OperatorSearchSerializer,
    DepartmentSerializer,
    StationSerializer,
    PoisonCakeTestListSerializer,
    PoisonCakeTestDetailSerializer,
    PoisonCakeTestCreateUpdateSerializer,
    TestStatisticsSerializer
)


# ================================================================
# DEFECT MANAGEMENT VIEWSETS (Settings Page)
# ================================================================

class DefectCategoryViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Defect Categories
    - List all categories
    - Create new category
    - Update/Delete category
    """
    queryset = DefectCategory.objects.all()
    serializer_class = DefectCategorySerializer
    permission_classes = [AllowAny]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['category_name', 'description']
    ordering_fields = ['category_name', 'created_at']
    ordering = ['category_name']
    
    def get_queryset(self):
        """Filter by active status if requested"""
        queryset = super().get_queryset()
        is_active = self.request.query_params.get('is_active', None)
        
        if is_active is not None:
            is_active_bool = is_active.lower() in ['true', '1', 'yes']
            queryset = queryset.filter(is_active=is_active_bool)
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def toggle_active(self, request, pk=None):
        """Toggle active status of a category"""
        category = self.get_object()
        category.is_active = not category.is_active
        category.save()
        serializer = self.get_serializer(category)
        return Response(serializer.data)


class DefectTypeViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Defect Types
    - List defect types (can filter by category)
    - Create new defect type
    - Update/Delete defect type
    """
    queryset = DefectType.objects.select_related('category').all()
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'is_active']
    search_fields = ['defect_name', 'description']
    ordering_fields = ['defect_name', 'created_at']
    ordering = ['category__category_name', 'defect_name']
    
    def get_serializer_class(self):
        """Use list serializer for list action"""
        if self.action == 'list':
            return DefectTypeListSerializer
        return DefectTypeSerializer
    
    def get_queryset(self):
        """Custom filtering by category"""
        queryset = super().get_queryset()
        
        # Filter by category if provided
        category_id = self.request.query_params.get('category_id', None)
        if category_id:
            queryset = queryset.filter(category_id=category_id)
        
        # Filter by active status
        is_active = self.request.query_params.get('is_active', None)
        if is_active is not None:
            is_active_bool = is_active.lower() in ['true', '1', 'yes']
            queryset = queryset.filter(is_active=is_active_bool)
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def toggle_active(self, request, pk=None):
        """Toggle active status of a defect type"""
        defect_type = self.get_object()
        defect_type.is_active = not defect_type.is_active
        defect_type.save()
        serializer = self.get_serializer(defect_type)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def by_category(self, request):
        """Get defect types grouped by category"""
        categories = DefectCategory.objects.filter(is_active=True).prefetch_related('defect_types')
        
        result = []
        for category in categories:
            result.append({
                'category_id': category.category_id,
                'category_name': category.category_name,
                'defect_types': DefectTypeListSerializer(
                    category.defect_types.filter(is_active=True),
                    many=True
                ).data
            })
        
        return Response(result)


# ================================================================
# EMPLOYEE/OPERATOR VIEWSETS
# ================================================================

class OperatorViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for Operators (Employees)
    - List operators
    - Get operator details
    - Search operators
    """
    queryset = MasterTable.objects.select_related(
        'department', 'station', 'sub_department'
    ).all()
    permission_classes = [AllowAny]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['emp_id', 'first_name', 'last_name', 'designation']
    ordering_fields = ['emp_id', 'first_name', 'last_name']
    ordering = ['emp_id']
    
    def get_serializer_class(self):
        """Use search serializer for list, full serializer for detail"""
        if self.action == 'list':
            return OperatorSearchSerializer
        return OperatorSerializer
    
    def get_queryset(self):
        """Filter operators by department, station, designation"""
        queryset = super().get_queryset()
        
        # Filter by department
        department_id = self.request.query_params.get('department', None)
        if department_id:
            queryset = queryset.filter(department_id=department_id)
        
        # Filter by station
        station_id = self.request.query_params.get('station', None)
        if station_id:
            queryset = queryset.filter(station_id=station_id)
        
        # Filter by designation
        designation = self.request.query_params.get('designation', None)
        if designation:
            queryset = queryset.filter(designation=designation)
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def search(self, request):
        """
        Quick search for operators (for autocomplete)
        Query param: q (search term)
        """
        search_term = request.query_params.get('q', '')
        
        if len(search_term) < 2:
            return Response({'results': []})
        
        queryset = self.get_queryset().filter(
            Q(emp_id__icontains=search_term) |
            Q(first_name__icontains=search_term) |
            Q(last_name__icontains=search_term)
        )[:10]  # Limit to 10 results
        
        serializer = OperatorSearchSerializer(queryset, many=True)
        return Response({'results': serializer.data})


# ================================================================
# DEPARTMENT & STATION VIEWSETS
# ================================================================


# ================================================================
# POISON CAKE TEST VIEWSET
# ================================================================


class PoisonCakeTestViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Poison Cake Tests
    - List all tests
    - Create new test
    - Get test details
    - Update test
    - Delete test
    """
    queryset = PoisonCakeTest.objects.select_related(
        'operator', 'department', 'station', 'level'  # Added 'level'
    ).prefetch_related(
        'planted_defects__defect_category',
        'planted_defects__defect_type'
    ).all()
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['test_judgment', 'department', 'operator', 'test_date','operator', 'department', 'level', 'station']  # Added 'level'
    search_fields = ['model_name', 'operator__emp_id', 'operator__first_name', 'operator__last_name']
    ordering_fields = ['test_date', 'test_id', 'test_judgment']
    ordering = ['-test_date', '-test_id']
    
    
    def get_serializer_class(self):
        """Use appropriate serializer based on action"""
        if self.action == 'list':
            return PoisonCakeTestListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return PoisonCakeTestCreateUpdateSerializer
        else:
            return PoisonCakeTestDetailSerializer
    
    def get_queryset(self):
        """Custom filtering"""
        queryset = super().get_queryset()
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date', None)
        end_date = self.request.query_params.get('end_date', None)
        
        if start_date:
            queryset = queryset.filter(test_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(test_date__lte=end_date)
        
        # Filter by model name
        model_name = self.request.query_params.get('model_name', None)
        if model_name:
            queryset = queryset.filter(model_name__icontains=model_name)
        
        # Filter by level (NEW)
        level_id = self.request.query_params.get('level_id', None)
        if level_id:
            queryset = queryset.filter(level_id=level_id)
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """
        Get test statistics
        - Total tests
        - Pass/Fail counts
        - Pass rate
        - Tests by department
        - Tests by level (NEW)
        - Recent tests
        """
        queryset = self.filter_queryset(self.get_queryset())
        
        # Basic counts
        total_tests = queryset.count()
        passed_tests = queryset.filter(test_judgment='OK').count()
        failed_tests = queryset.filter(test_judgment='NOT OK').count()
        pass_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
        
        # Tests by judgment
        tests_by_judgment = {
            'OK': passed_tests,
            'NOT OK': failed_tests
        }
        
        # Tests by department
        tests_by_department = {}
        dept_counts = queryset.values(
            'department__department_name'
        ).annotate(
            count=Count('test_id')
        ).order_by('-count')
        
        for item in dept_counts:
            dept_name = item['department__department_name'] or 'Unknown'
            tests_by_department[dept_name] = item['count']
        
        # Tests by level (NEW)
        tests_by_level = {}
        level_counts = queryset.values(
            'level__level_name'
        ).annotate(
            count=Count('test_id')
        ).order_by('level__level_id')
        
        for item in level_counts:
            level_name = item['level__level_name'] or 'Unknown'
            tests_by_level[level_name] = item['count']
        
        # Recent tests (last 10)
        recent_tests = queryset.order_by('-test_date', '-test_id')[:10]
        
        statistics_data = {
            'total_tests': total_tests,
            'passed_tests': passed_tests,
            'failed_tests': failed_tests,
            'pass_rate': round(pass_rate, 2),
            'tests_by_judgment': tests_by_judgment,
            'tests_by_department': tests_by_department,
            'tests_by_level': tests_by_level,  # NEW
            'recent_tests': PoisonCakeTestListSerializer(recent_tests, many=True).data
        }
        
        serializer = TestStatisticsSerializer(statistics_data)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def by_level(self, request):
        """Get tests grouped by level with their pass rates (NEW)"""
        level_id = request.query_params.get('level_id', None)
        
        queryset = self.get_queryset()
        if level_id:
            queryset = queryset.filter(level_id=level_id)
        
        # Aggregate by level
        level_stats = queryset.values(
            'level__level_id',
            'level__level_name'
        ).annotate(
            total_tests=Count('test_id'),
            passed_tests=Count('test_id', filter=Q(test_judgment='OK')),
            failed_tests=Count('test_id', filter=Q(test_judgment='NOT OK'))
        ).order_by('level__level_id')
        
        # Calculate pass rates
        result = []
        for stat in level_stats:
            pass_rate = (stat['passed_tests'] / stat['total_tests'] * 100) if stat['total_tests'] > 0 else 0
            result.append({
                'level_id': stat['level__level_id'],
                'level_name': stat['level__level_name'],
                'total_tests': stat['total_tests'],
                'passed_tests': stat['passed_tests'],
                'failed_tests': stat['failed_tests'],
                'pass_rate': round(pass_rate, 2)
            })
        
        return Response(result)
    
    @action(detail=False, methods=['get'])
    def recent(self, request):
        """Get recent tests (last 7 days by default)"""
        days = int(request.query_params.get('days', 7))
        start_date = timezone.now().date() - timedelta(days=days)
        
        queryset = self.get_queryset().filter(test_date__gte=start_date)
        serializer = PoisonCakeTestListSerializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def judgment_breakdown(self, request, pk=None):
        """Get detailed breakdown of judgment for a specific test"""
        test = self.get_object()
        judgment_details = test.get_judgment_details()
        return Response(judgment_details)
    
    @action(detail=False, methods=['get'])
    def failed_tests(self, request):
        """Get all failed tests that need re-inspection"""
        queryset = self.get_queryset().filter(test_judgment='NOT OK')
        serializer = PoisonCakeTestListSerializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def by_operator(self, request):
        """Get tests grouped by operator with their pass rates"""
        operator_id = request.query_params.get('operator_id', None)
        
        queryset = self.get_queryset()
        if operator_id:
            queryset = queryset.filter(operator_id=operator_id)
        
        # Aggregate by operator
        operator_stats = queryset.values(
            'operator__emp_id',
            'operator__first_name',
            'operator__last_name'
        ).annotate(
            total_tests=Count('test_id'),
            passed_tests=Count('test_id', filter=Q(test_judgment='OK')),
            failed_tests=Count('test_id', filter=Q(test_judgment='NOT OK'))
        ).order_by('-total_tests')
        
        # Calculate pass rates
        result = []
        for stat in operator_stats:
            pass_rate = (stat['passed_tests'] / stat['total_tests'] * 100) if stat['total_tests'] > 0 else 0
            result.append({
                'operator_id': stat['operator__emp_id'],
                'operator_name': f"{stat['operator__first_name'] or ''} {stat['operator__last_name'] or ''}".strip(),
                'total_tests': stat['total_tests'],
                'passed_tests': stat['passed_tests'],
                'failed_tests': stat['failed_tests'],
                'pass_rate': round(pass_rate, 2)
            })
        
        return Response(result)
    
    @action(detail=False, methods=['get'])
    def common_defects(self, request):
        """Get most common defect types found in tests"""
        start_date = request.query_params.get('start_date', None)
        end_date = request.query_params.get('end_date', None)
        
        queryset = PlantedDefect.objects.select_related(
            'defect_category', 'defect_type', 'test'
        )
        
        if start_date:
            queryset = queryset.filter(test__test_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(test__test_date__lte=end_date)
        
        # Aggregate defect types
        defect_stats = queryset.values(
            'defect_category__category_name',
            'defect_type__defect_name'
        ).annotate(
            total_quantity=Sum('quantity'),
            occurrence_count=Count('planted_defect_id')
        ).order_by('-total_quantity')[:20]  # Top 20
        
        result = []
        for stat in defect_stats:
            result.append({
                'category': stat['defect_category__category_name'],
                'defect_type': stat['defect_type__defect_name'],
                'total_quantity': stat['total_quantity'],
                'occurrence_count': stat['occurrence_count']
            })
        
        return Response(result)
    @action(detail=True, methods=['patch'])  # ← must be 'patch' lowercase
    def schedule_reeval(self, request, pk=None):
        test = self.get_object()
        test.reeval_scheduled_date = request.data.get('reeval_scheduled_date')
        test.reeval_extension_weeks = request.data.get('reeval_extension_weeks')
        test.save(update_fields=[
            'reeval_scheduled_date', 
            'reeval_extension_weeks', 
            'updated_at'
        ])
        return Response(PoisonCakeTestDetailSerializer(test).data)

# ================================================================
# OPTIONAL: Function-Based Views (if you prefer)
# ================================================================

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated

@api_view(['GET'])
@permission_classes([AllowAny])
def dashboard_summary(request):
    """
    Get overall dashboard summary
    """
    # Tests overview
    total_tests = PoisonCakeTest.objects.count()
    passed_tests = PoisonCakeTest.objects.filter(test_judgment='OK').count()
    failed_tests = PoisonCakeTest.objects.filter(test_judgment='NOT OK').count()
    
    # Recent activity (last 7 days)
    last_week = timezone.now().date() - timedelta(days=7)
    recent_tests = PoisonCakeTest.objects.filter(test_date__gte=last_week).count()
    
    # Pass rate
    pass_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
    
    # Most active departments
    top_departments = PoisonCakeTest.objects.values(
        'department__department_name'
    ).annotate(
        test_count=Count('test_id')
    ).order_by('-test_count')[:5]
    
    return Response({
        'total_tests': total_tests,
        'passed_tests': passed_tests,
        'failed_tests': failed_tests,
        'pass_rate': round(pass_rate, 2),
        'recent_tests_count': recent_tests,
        'top_departments': list(top_departments)
    })



# Add these to your existing views.py

class RecurringTestScheduleViewSet(viewsets.ModelViewSet):
    """ViewSet for Recurring Test Schedule"""
    queryset = RecurringTestSchedule.objects.select_related(
        'employee', 'station', 'level', 'last_test'
    ).all()
    serializer_class = RecurringTestScheduleSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'station', 'level', 'employee']
    search_fields = ['employee__emp_id', 'employee__first_name', 'employee__last_name']
    ordering_fields = ['next_test_date', 'last_test_date', 'status']
    ordering = ['next_test_date']
    
    def get_queryset(self):
        """Update all statuses before returning"""
        queryset = super().get_queryset()
        
        # Update statuses for all records
        for schedule in queryset:
            schedule.update_status()
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def due_soon(self, request):
        """Get tests due in next 30 days"""
        from datetime import date, timedelta
        today = date.today()
        thirty_days = today + timedelta(days=30)
        
        queryset = self.get_queryset().filter(
            next_test_date__gte=today,
            next_test_date__lte=thirty_days,
            status__in=['SCHEDULED', 'DUE_SOON']
        ).order_by('next_test_date')
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def overdue(self, request):
        """Get overdue tests"""
        queryset = self.get_queryset().filter(status='OVERDUE')
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class ReInspectionPlanViewSet(viewsets.ModelViewSet):
    """ViewSet for Re-Inspection Plan"""
    queryset = ReInspectionPlan.objects.select_related(
        'employee', 'station', 'level', 'failed_test', 'completed_test'
    ).all()
    serializer_class = ReInspectionPlanSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'station', 'level', 'employee']
    search_fields = ['employee__emp_id', 'employee__first_name', 'employee__last_name']
    ordering_fields = ['failed_date', 'scheduled_date', 'status']
    ordering = ['-failed_date']
    
    @action(detail=False, methods=['get'])
    def pending(self, request):
        """Get pending re-inspections"""
        queryset = self.get_queryset().filter(
            status__in=['PENDING', 'SCHEDULED', 'IN_PROGRESS']
        )
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def schedule(self, request, pk=None):
        """Schedule a re-inspection"""
        plan = self.get_object()
        scheduled_date = request.data.get('scheduled_date')
        notes = request.data.get('notes', '')
        
        if not scheduled_date:
            return Response(
                {'error': 'scheduled_date is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        plan.scheduled_date = scheduled_date
        plan.status = 'SCHEDULED'
        plan.notes = notes
        plan.save()
        
        serializer = self.get_serializer(plan)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Mark re-inspection as completed"""
        plan = self.get_object()
        test_id = request.data.get('test_id')
        
        if not test_id:
            return Response(
                {'error': 'test_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            test = PoisonCakeTest.objects.get(test_id=test_id)
            plan.completed_test = test
            plan.status = 'COMPLETED'
            plan.save()
            
            serializer = self.get_serializer(plan)
            return Response(serializer.data)
        except PoisonCakeTest.DoesNotExist:
            return Response(
                {'error': 'Test not found'},
                status=status.HTTP_404_NOT_FOUND
            )
# ==================================poison cake test=================================

# ==========attrition======================================
import datetime
from django.db.models import Sum, Avg
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response

from .models import AttritionRecord
from .serializers import AttritionRecordSerializer


class AttritionRecordViewSet(viewsets.ModelViewSet):
    """
    CRUD + Financial Year endpoints
    """
    queryset = AttritionRecord.objects.all()
    serializer_class = AttritionRecordSerializer

    # ── Helper: FY range (April - March) ──
    def _get_fy_range(self, fy=None):
        today = datetime.date.today()
        if fy:
            start_year = int(fy.split('-')[0])
        else:
            start_year = today.year if today.month >= 4 else today.year - 1

        end_year = start_year + 1
        label = f"{start_year}-{str(end_year)[-2:]}"
        start = datetime.date(start_year, 4, 1)
        end = datetime.date(end_year, 3, 31)
        return label, start, end

    # ── Filter list by ?fy=2024-25 ──
    def get_queryset(self):
        qs = super().get_queryset()
        fy = self.request.query_params.get('fy')
        if fy:
            _, fy_start, fy_end = self._get_fy_range(fy)
            qs = qs.filter(date__gte=fy_start, date__lte=fy_end)
        return qs

    # ── FY Summary ──
    @action(detail=False, methods=['get'], url_path='fy-summary')
    def fy_summary(self, request):
        fy = request.query_params.get('fy')
        label, fy_start, fy_end = self._get_fy_range(fy)

        records = AttritionRecord.objects.filter(
            date__gte=fy_start, date__lte=fy_end
        ).order_by('date')

        if not records.exists():
            return Response(
                {"error": f"No records for FY {label}"},
                status=status.HTTP_404_NOT_FOUND
            )

        agg = records.aggregate(
            total_oet=Sum('oet'),
            total_associate=Sum('associate'),
            total_overall=Sum('overall'),
            avg_oet=Avg('oet'),
            avg_associate=Avg('associate'),
            avg_overall=Avg('overall'),
        )

        data = {
            "financial_year": label,
            "fy_start": fy_start,
            "fy_end": fy_end,
            "total_months": records.count(),
            "total_oet": agg['total_oet'] or 0,
            "total_associate": agg['total_associate'] or 0,
            "total_overall": agg['total_overall'] or 0,
            "avg_oet": round(agg['avg_oet'] or 0, 2),
            "avg_associate": round(agg['avg_associate'] or 0, 2),
            "avg_overall": round(agg['avg_overall'] or 0, 2),
            "monthly_data": AttritionRecordSerializer(records, many=True).data,
        }
        return Response(data)

    # ── Available FYs ──
    @action(detail=False, methods=['get'], url_path='available-fy')
    def available_fy(self, request):
        dates = AttritionRecord.objects.values_list('date', flat=True).order_by('date')
        fy_set = set()
        for d in dates:
            start_year = d.year if d.month >= 4 else d.year - 1
            fy_set.add(f"{start_year}-{str(start_year + 1)[-2:]}")
        return Response(sorted(fy_set))

# ==========attrition======================================

